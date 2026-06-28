"""Teste route-level pentru T1.12 project nested tenant guard."""

from io import BytesIO

import pytest
from openpyxl import load_workbook

from tests.unit.test_tenant_access_project_nested import _creeaza_date, _curata_date


@pytest.fixture(autouse=True)
def curata_project_nested_routes(app):
    _curata_date(app)
    yield
    _curata_date(app)


def test_mode_off_detalii_pastreaza_agregari_legacy(authenticated_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'off'

    raspuns = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}?luna=1&anul=2026'
    )

    assert raspuns.status_code == 200
    assert b'TPN Angajat A' in raspuns.data
    assert b'TPN Angajat B' in raspuns.data
    assert b'TPN Doc Conflict' in raspuns.data


def test_strict_detalii_filtreaza_copii_conflict(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}?luna=1&anul=2026'
    )

    assert raspuns.status_code == 200
    assert b'TPN Angajat A' in raspuns.data
    assert b'TPN Doc A' in raspuns.data
    assert b'TPN Angajat B' not in raspuns.data
    assert b'TPN Doc B' not in raspuns.data
    assert b'TPN Doc Conflict' not in raspuns.data


def test_strict_hub_si_utilaje_filtreaza_nested(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    hub = authenticated_client.get(f'/proiecte/{ids["proiect_a"]}/hub')
    utilaje = authenticated_client.get(f'/proiecte/{ids["proiect_a"]}/utilaje')

    assert hub.status_code == 200
    assert b'TPN-SITE-A' in hub.data
    assert b'TPN-SITE-B' not in hub.data
    assert utilaje.status_code == 200
    assert b'TPN Consum A' in utilaje.data
    assert b'TPN-A' in utilaje.data
    assert b'TPN Consum Conflict' not in utilaje.data
    assert b'TPN-B' not in utilaje.data


def test_strict_blocheaza_mutatii_nested_straine(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    add_angajat = authenticated_client.post(
        f'/proiecte/{ids["proiect_a"]}/adauga-angajat',
        data={'angajat_id': str(ids['angajat_b']), 'functie_pe_proiect': 'Strain'},
    )
    add_utilaj = authenticated_client.post(
        f'/proiecte/{ids["proiect_a"]}/utilaje/adauga',
        data={
            'denumire': 'TPN Utilaj Nou',
            'masina_id': str(ids['masina_b']),
            'ore': '1',
            'tarif_ora': '10',
            'cost': '',
        },
    )
    link_site = authenticated_client.post(
        f'/proiecte/{ids["proiect_a"]}/leaga-santier',
        data={'santier_id': str(ids['site_b'])},
    )

    assert add_angajat.status_code == 404
    assert add_utilaj.status_code == 404
    assert link_site.status_code == 404
    with app.app_context():
        from models import ConsumUtilaj, ProiectSantier

        assert ConsumUtilaj.query.filter_by(denumire='TPN Utilaj Nou').count() == 0
        assert ProiectSantier.query.filter_by(
            proiect_id=ids['proiect_a'],
            santier_id=ids['site_b'],
        ).count() == 0


def test_strict_export_excel_exclude_copii_conflict(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(f'/proiecte/{ids["proiect_a"]}/export-excel')

    assert raspuns.status_code == 200
    wb = load_workbook(BytesIO(raspuns.data), read_only=True)
    valori = {
        str(cell.value)
        for sheet in wb.worksheets
        for row in sheet.iter_rows()
        for cell in row
        if cell.value is not None
    }
    text = '\n'.join(valori)
    assert 'TPN Angajat A Test' in text
    assert 'TPN Angajat B Test' not in text


def test_strict_export_proiect_strain_returneaza_404(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(f'/proiecte/{ids["proiect_b"]}/export-excel')

    assert raspuns.status_code == 404


def test_optional_fara_tenant_ramane_migration_friendly(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, None)
    app.config['MULTI_TENANT_MODE'] = 'optional'

    raspuns = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}?luna=1&anul=2026'
    )

    assert raspuns.status_code == 200
    assert b'TPN Angajat B' in raspuns.data


def _seteaza_tenant_user(app, user_id, tenant_id):
    from models import Utilizator, db

    with app.app_context():
        user = db.session.get(Utilizator, user_id)
        user.tenant_id = tenant_id
        db.session.commit()
