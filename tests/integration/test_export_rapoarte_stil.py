"""
Export rapoarte activitati 'mod edifico' - stilizat cu paleta brand (navy/gold/cream).
Verifica ca exportul se genereaza si ca paleta veche (rosu/albastru/portocaliu) a fost
inlocuita cu navy 0B1426 + gold C9A961.
"""
import io
from datetime import date, timedelta

from openpyxl import load_workbook


def test_export_paleta_edifico(app, authenticated_client):
    from models import db, Angajat
    with app.app_context():
        a = Angajat.query.filter_by(cnp='1990909090909').first()
        if not a:
            a = Angajat(cnp='1990909090909', nume='Export', prenume='Test',
                        status='activ', data_angajare=date(2020, 1, 1))
            db.session.add(a); db.session.commit()
        aid = a.id

    luna = date.today().strftime('%Y-%m')
    r = authenticated_client.get(f'/activitati/export?angajat_id={aid}&luna={luna}')
    assert r.status_code == 200, r.status_code
    assert r.data[:2] == b'PK'  # xlsx = zip

    wb = load_workbook(io.BytesIO(r.data))
    culori = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                fc = getattr(getattr(c.fill, 'start_color', None), 'rgb', None)
                tc = getattr(getattr(c.font, 'color', None), 'rgb', None)
                if fc:
                    culori.add(str(fc))
                if tc:
                    culori.add(str(tc))
    blob = ' '.join(culori)
    # paleta noua prezenta
    assert 'C9A961' in blob, 'gold lipseste'
    assert '0B1426' in blob, 'navy lipseste'
    # paleta veche disparuta
    assert 'C62828' not in blob, 'rosu vechi inca prezent'
    assert '283593' not in blob, 'albastru vechi inca prezent'
    assert 'FFE0B2' not in blob, 'portocaliu vechi inca prezent'


def test_export_continut_proiect_ore_activitate(app, authenticated_client):
    """Exportul contine coloanele Proiect + Ore + activitatea principala pe zi."""
    from models import db, Angajat, Proiect, RaportActivitate
    today = date.today()
    with app.app_context():
        p = Proiect.query.filter_by(cod_proiect='RA-COD').first()
        if not p:
            p = Proiect(cod_proiect='RA-COD', nume='Reabilitare Spital Tulcea',
                        data_start=date(2026, 1, 1), status='activ')
            db.session.add(p); db.session.commit()
        a = Angajat.query.filter_by(cnp='1991010101010').first()
        if not a:
            a = Angajat(cnp='1991010101010', nume='Raport', prenume='Test',
                        status='activ', data_angajare=date(2020, 1, 1), functie='Inginer')
            db.session.add(a); db.session.commit()
        zi = today.replace(day=1)
        while zi.weekday() >= 5:
            zi += timedelta(days=1)
        if not RaportActivitate.query.filter_by(angajat_id=a.id, data=zi,
                                                tip_activitate='zilnica').first():
            db.session.add(RaportActivitate(
                angajat_id=a.id, proiect_id=p.id, data=zi, tip_activitate='zilnica',
                activitate_principala='Montaj armatura fundatii (etaj 4)', ore_lucrate=8,
                status='aprobat'))
            db.session.commit()
        aid = a.id

    luna = today.strftime('%Y-%m')
    r = authenticated_client.get(f'/activitati/export?angajat_id={aid}&luna={luna}')
    assert r.status_code == 200
    from routes.activitati import LUNI_RO_SCURT
    wb = load_workbook(io.BytesIO(r.data))
    blob = []
    col_d = []   # Data
    col_e = []   # Proiect
    col_g = []   # Activitati desfasurate
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                if c.value is None:
                    continue
                blob.append(str(c.value))
                if c.column == 4:
                    col_d.append(str(c.value))
                if c.column == 5:
                    col_e.append(str(c.value))
                if c.column == 7:
                    col_g.append(str(c.value))
    blob = ' | '.join(blob)
    d_txt = ' | '.join(col_d)
    e_txt = ' | '.join(col_e)
    g_txt = ' | '.join(col_g)
    assert 'Proiect' in blob                          # header coloana
    assert 'Ore' in blob                              # header coloana
    assert 'Reabilitare Spital Tulcea' in e_txt       # NUMELE proiectului, nu codul
    assert 'RA-COD' not in blob                       # codul nu apare nicaieri
    assert 'Montaj armatura fundatii' in g_txt        # activitatea principala
    assert 'etaj 4' in g_txt                          # continutul parantezei pastrat
    assert '(' not in g_txt and ')' not in g_txt      # fara paranteze in activitati
    assert '•' not in g_txt                           # fara bullet
    # data in romana, fara engleza si fara paranteze
    assert LUNI_RO_SCURT[today.month] in d_txt        # ex 'ian', 'feb'...
    assert '(' not in d_txt
    for eng in ('Jan', 'Feb', 'Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sun'):
        assert eng not in d_txt


def test_export_preview_html_inainte_de_download(app, authenticated_client):
    """Preview HTML inainte de export: contine datele (proiect per zi) + link de
    descarcare xlsx cu aceiasi parametri. NU returneaza fisierul direct."""
    import json as _json
    from models import db, Angajat, Proiect, RaportActivitate
    with app.app_context():
        for cod, nume in (('PV-A', 'Proiect Preview Alfa'),
                          ('PV-B', 'Proiect Preview Beta')):
            if not Proiect.query.filter_by(cod_proiect=cod).first():
                db.session.add(Proiect(cod_proiect=cod, nume=nume,
                                       data_start=date(2026, 1, 1), status='activ'))
        db.session.commit()
        p1 = Proiect.query.filter_by(cod_proiect='PV-A').first()
        p2 = Proiect.query.filter_by(cod_proiect='PV-B').first()
        a = Angajat.query.filter_by(cnp='1994040404040').first()
        if not a:
            a = Angajat(cnp='1994040404040', nume='Preview', prenume='Test',
                        status='activ', data_angajare=date(2020, 1, 1), functie='Inginer')
            db.session.add(a); db.session.commit()
        if not RaportActivitate.query.filter_by(angajat_id=a.id,
                                                tip_activitate='lunara').first():
            det = [
                {'data': '2026-01-05', 'proiect_id': p1.id, 'text': 'Montaj', 'ore': 8},
                {'data': '2026-01-06', 'proiect_id': p2.id, 'text': 'Turnare', 'ore': 6},
            ]
            db.session.add(RaportActivitate(
                angajat_id=a.id, proiect_id=p1.id, tip_activitate='lunara',
                data=date(2026, 1, 5), data_sfarsit=date(2026, 1, 6), luna_an='2026-01',
                activitate_principala='Coordonare',
                detalii_pe_zi=_json.dumps(det), status='aprobat'))
            db.session.commit()
        aid = a.id

    r = authenticated_client.get(
        f'/activitati/export/preview?angajat_id={aid}&luna_start=2026-01&luna_end=2026-01')
    assert r.status_code == 200
    assert r.data[:2] != b'PK'                          # e HTML, nu xlsx
    html = r.get_data(as_text=True)
    assert 'Preview Export EDIFICO' in html
    assert 'Proiect Preview Alfa' in html               # proiect per zi (din detalii)
    assert 'Proiect Preview Beta' in html
    assert 'Lu 05 ian' in html                          # data in romana
    assert '/activitati/export?' in html                # link spre exportul real
    assert 'luna_start=2026-01' in html                 # cu aceiasi parametri
    assert 'Descarca Excel' in html


def test_export_nume_proiect_complet_si_rand_inalt(app, authenticated_client):
    """Numele proiectului apare COMPLET (fara trunchiere la 60) + randul e inaltat
    ca textul wrapuit sa fie vizibil + coloana Proiect e lata."""
    import json as _json
    from models import db, Angajat, Proiect, RaportActivitate
    nume_lung = ('LUCRARI DE INTRETINERE SI REPARATII CURENTE - AMENAJARE SPATII '
                 'LABORATOR HIPOXIE-HIPOBARISM SI HIPERBARISM CORP B')
    assert len(nume_lung) > 60
    with app.app_context():
        p = Proiect.query.filter_by(cod_proiect='NL-1').first()
        if not p:
            p = Proiect(cod_proiect='NL-1', nume=nume_lung,
                        data_start=date(2026, 1, 1), status='activ')
            db.session.add(p); db.session.commit()
        a = Angajat.query.filter_by(cnp='1993030303030').first()
        if not a:
            a = Angajat(cnp='1993030303030', nume='Nume', prenume='Lung',
                        status='activ', data_angajare=date(2020, 1, 1), functie='Inginer')
            db.session.add(a); db.session.commit()
        if not RaportActivitate.query.filter_by(angajat_id=a.id,
                                                tip_activitate='lunara').first():
            det = [{'data': '2026-01-05', 'proiect_id': p.id, 'text': 'Montaj', 'ore': 8}]
            db.session.add(RaportActivitate(
                angajat_id=a.id, proiect_id=p.id, tip_activitate='lunara',
                data=date(2026, 1, 5), data_sfarsit=date(2026, 1, 5), luna_an='2026-01',
                activitate_principala='Coordonare',
                detalii_pe_zi=_json.dumps(det), status='aprobat'))
            db.session.commit()
        aid = a.id

    r = authenticated_client.get(f'/activitati/export?angajat_id={aid}&luna=2026-01')
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.data)).worksheets[0]
    full_present = False
    tall_row = False
    for row in ws.iter_rows():
        for c in row:
            if c.column == 5 and c.value and 'CORP B' in str(c.value):
                full_present = True
                assert len(str(c.value)) == len(nume_lung)   # NU trunchiat la 60
                h = ws.row_dimensions[c.row].height
                if h and h > 16.5:
                    tall_row = True
    assert full_present, 'numele complet al proiectului lipseste din coloana E'
    assert tall_row, 'randul nu a fost inaltat pentru textul wrapuit'
    assert ws.column_dimensions['E'].width >= 40, 'coloana Proiect prea ingusta'


def test_export_detalii_pe_zi_proiect_per_zi(app, authenticated_client):
    """Raport saptamanal cu detalii_pe_zi: proiect + text + ore DIFERITE pe zi."""
    import json as _json
    from models import db, Angajat, Proiect, RaportActivitate
    with app.app_context():
        for cod, nume in (('DZ-A', 'Proiect Alfa'), ('DZ-B', 'Proiect Beta')):
            if not Proiect.query.filter_by(cod_proiect=cod).first():
                db.session.add(Proiect(cod_proiect=cod, nume=nume,
                                       data_start=date(2026, 1, 1), status='activ'))
        db.session.commit()
        p1 = Proiect.query.filter_by(cod_proiect='DZ-A').first()
        p2 = Proiect.query.filter_by(cod_proiect='DZ-B').first()
        a = Angajat.query.filter_by(cnp='1992020202020').first()
        if not a:
            a = Angajat(cnp='1992020202020', nume='Detalii', prenume='Zi',
                        status='activ', data_angajare=date(2020, 1, 1), functie='Inginer')
            db.session.add(a); db.session.commit()
        if not RaportActivitate.query.filter_by(angajat_id=a.id,
                                                tip_activitate='saptamanala').first():
            det = [
                {'data': '2026-01-05', 'proiect_id': p1.id, 'text': 'Montaj armatura', 'ore': 8},
                {'data': '2026-01-06', 'proiect_id': p2.id, 'text': 'Turnare beton', 'ore': 6},
            ]
            db.session.add(RaportActivitate(
                angajat_id=a.id, proiect_id=p1.id, tip_activitate='saptamanala',
                data=date(2026, 1, 5), data_sfarsit=date(2026, 1, 9), numar_saptamana=2,
                activitate_principala='Lucrari structura',
                detalii_pe_zi=_json.dumps(det), status='aprobat'))
            db.session.commit()
        aid = a.id

    r = authenticated_client.get(f'/activitati/export?angajat_id={aid}&luna=2026-01')
    assert r.status_code == 200
    ws = load_workbook(io.BytesIO(r.data)).worksheets[0]
    rows = []
    for row in ws.iter_rows():
        e = ws.cell(row=row[0].row, column=5).value
        g = ws.cell(row=row[0].row, column=7).value
        f = ws.cell(row=row[0].row, column=6).value
        rows.append((str(e or ''), str(g or ''), f))
    e_all = ' | '.join(r[0] for r in rows)
    g_all = ' | '.join(r[1] for r in rows)
    # proiecte DIFERITE pe zile (din detalii_pe_zi, nu un singur proiect)
    assert 'Proiect Alfa' in e_all and 'Proiect Beta' in e_all
    # textul pe zi (din detalii_pe_zi)
    assert 'Montaj armatura' in g_all and 'Turnare beton' in g_all
    # orele pe zi
    assert any(rv[2] in (8, 8.0) for rv in rows) and any(rv[2] in (6, 6.0) for rv in rows)
