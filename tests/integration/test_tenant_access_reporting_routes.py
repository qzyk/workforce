"""Teste tenant access pentru dashboard si rutele de rapoarte."""

import os
from datetime import date, timedelta
from io import BytesIO

import pytest
from openpyxl import load_workbook


@pytest.fixture(autouse=True)
def curata_reporting_tenant_routes(app):
    _curata_date(app)
    yield
    _curata_date(app)


def test_off_mode_download_raport_ramane_compatibil(authenticated_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'off'

    response = authenticated_client.get(f'/rapoarte/descarca/{ids["raport_b"]}')

    assert response.status_code == 200
    assert response.data == b'TA REPORT B'


def test_strict_dashboard_stats_scopeaza_tenant(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    response = authenticated_client.get('/api/dashboard-stats')

    assert response.status_code == 200
    payload = response.get_json()
    assert payload['angajati_activi'] == 1
    assert payload['proiecte_active'] == 1
    assert payload['ore_luna'] == 8.0
    assert payload['doc_expirate'] == 1
    assert payload['pontaje_pending'] == 0


def test_strict_dashboard_executiv_si_cautare_nu_afiseaza_tenant_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    executiv = authenticated_client.get('/dashboard/executiv')
    cautare = authenticated_client.get('/cauta?q=TA-REPORT-ROUTE-PRJ')

    assert executiv.status_code == 200
    assert b'TA-REPORT-ROUTE-PRJ-A' in executiv.data
    assert b'TA-REPORT-ROUTE-PRJ-B' not in executiv.data

    assert cautare.status_code == 200
    rezultate = cautare.get_json()
    labels = ' '.join(r['label'] for r in rezultate)
    assert 'TA-REPORT-ROUTE-PRJ-A' in labels
    assert 'TA-REPORT-ROUTE-PRJ-B' not in labels


def test_strict_istoric_download_si_stergere_blocheaza_raport_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    istoric = authenticated_client.get('/rapoarte/istoric')
    download_strain = authenticated_client.get(f'/rapoarte/descarca/{ids["raport_b"]}')
    sterge_strain = authenticated_client.post(f'/rapoarte/sterge/{ids["raport_b"]}')
    download_propriu = authenticated_client.get(f'/rapoarte/descarca/{ids["raport_a"]}')

    assert istoric.status_code == 200
    assert b'TA-REPORT-ROUTE-A' in istoric.data
    assert b'TA-REPORT-ROUTE-B' not in istoric.data
    assert download_strain.status_code == 404
    assert sterge_strain.status_code == 404
    assert download_propriu.status_code == 200

    with app.app_context():
        from models import Raport, db

        assert db.session.get(Raport, ids['raport_b']) is not None


def test_strict_generare_raport_proiect_strain_esueaza(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    response = authenticated_client.post(
        '/rapoarte/situatie-proiect',
        data={
            'proiect_id': str(ids['proiect_b']),
            'nivel': 'detaliat',
            'format': 'xlsx',
        },
    )

    assert response.status_code == 404
    with app.app_context():
        from models import Raport

        assert Raport.query.filter_by(titlu='Situatie TA-REPORT-ROUTE-PRJ-B').first() is None


def test_strict_stat_plata_toti_exclude_angajat_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    response = authenticated_client.post(
        '/rapoarte/stat-plata',
        data={
            'luna': date.today().month,
            'an': date.today().year,
            'format': 'xlsx',
        },
    )

    assert response.status_code == 200
    text = _text_xlsx(response.data)
    assert 'ReportRouteA Test' in text
    assert 'ReportRouteB Test' not in text


def test_strict_centralizator_ore_exclude_tenant_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    response = authenticated_client.post(
        '/rapoarte/centralizator-ore',
        data={
            'luna': date.today().month,
            'an': date.today().year,
            'grupare': 'angajat',
        },
    )

    assert response.status_code == 200
    text = _text_xlsx(response.data)
    assert 'ReportRouteA Test' in text
    assert 'ReportRouteB Test' not in text


def test_strict_user_fara_tenant_esueaza_inchis(operator_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'strict'

    download = operator_client.get(f'/rapoarte/descarca/{ids["raport_a"]}')
    stats = operator_client.get('/api/dashboard-stats')

    assert download.status_code == 404
    assert stats.status_code == 200
    payload = stats.get_json()
    assert payload['angajati_activi'] == 0
    assert payload['proiecte_active'] == 0
    assert payload['ore_luna'] == 0.0
    assert payload['doc_expirate'] == 0


def test_optional_fara_tenant_ramane_permisiv(operator_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'optional'

    download = operator_client.get(f'/rapoarte/descarca/{ids["raport_b"]}')
    panou = operator_client.get('/rapoarte/')

    assert download.status_code == 200
    assert panou.status_code == 200
    assert b'TA-REPORT-ROUTE-A' in panou.data
    assert b'TA-REPORT-ROUTE-B' in panou.data


def _creeaza_date(app):
    from models import Angajat, Document, Pontaj, Proiect, Raport, Tenant, Utilizator, db

    today = date.today()
    prima_zi = today.replace(day=1)
    with app.app_context():
        tenant_a = Tenant(cod='test-ta-report-route-a', nume='Tenant Reporting Route A')
        tenant_b = Tenant(cod='test-ta-report-route-b', nume='Tenant Reporting Route B')
        db.session.add_all([tenant_a, tenant_b])
        db.session.commit()

        user_b = Utilizator(
            tenant_id=tenant_b.id,
            nume='Report',
            prenume='RouteB',
            email='ta-report-route-b@test.local',
            rol='manager',
            activ=True,
        )
        user_b.set_password('test_pass_123')

        proiect_a = _proiect(tenant_a.id, 'TA-REPORT-ROUTE-PRJ-A')
        proiect_b = _proiect(tenant_b.id, 'TA-REPORT-ROUTE-PRJ-B')
        angajat_a = _angajat(tenant_a.id, '7990101010101', 'ReportRouteA')
        angajat_b = _angajat(tenant_b.id, '8990101010101', 'ReportRouteB')
        db.session.add_all([user_b, proiect_a, proiect_b, angajat_a, angajat_b])
        db.session.commit()

        pontaj_a = Pontaj(
            angajat_id=angajat_a.id,
            proiect_id=proiect_a.id,
            data=prima_zi,
            ore_lucrate=8,
            ore_normale=8,
            status='aprobat',
        )
        pontaj_b = Pontaj(
            angajat_id=angajat_b.id,
            proiect_id=proiect_b.id,
            data=prima_zi,
            ore_lucrate=7,
            ore_normale=7,
            status='aprobat',
        )
        doc_a = Document(
            angajat_id=angajat_a.id,
            tip='alte',
            nume_document='TA-REPORT-ROUTE-DOC-A',
            data_expirare=today - timedelta(days=1),
            status='expirat',
        )
        doc_b = Document(
            angajat_id=angajat_b.id,
            tip='alte',
            nume_document='TA-REPORT-ROUTE-DOC-B',
            data_expirare=today - timedelta(days=1),
            status='expirat',
        )
        db.session.add_all([pontaj_a, pontaj_b, doc_a, doc_b])
        db.session.commit()

        export_dir = app.config['EXPORT_FOLDER']
        os.makedirs(export_dir, exist_ok=True)
        path_a = os.path.join(export_dir, 'ta_report_route_a.xlsx')
        path_b = os.path.join(export_dir, 'ta_report_route_b.xlsx')
        with open(path_a, 'wb') as handle:
            handle.write(b'TA REPORT A')
        with open(path_b, 'wb') as handle:
            handle.write(b'TA REPORT B')

        admin = Utilizator.query.filter_by(email='admin_test@test.local').first()
        raport_a = Raport(
            tip_raport='situatie_proiect',
            titlu='TA-REPORT-ROUTE-A',
            parametri=f'{{"proiect_id": {proiect_a.id}}}',
            fisier_path=path_a,
            format='xlsx',
            generat_de=admin.id if admin else None,
            dimensiune_fisier=os.path.getsize(path_a),
        )
        raport_b = Raport(
            tip_raport='situatie_proiect',
            titlu='TA-REPORT-ROUTE-B',
            parametri=f'{{"proiect_id": {proiect_b.id}}}',
            fisier_path=path_b,
            format='xlsx',
            generat_de=user_b.id,
            dimensiune_fisier=os.path.getsize(path_b),
        )
        db.session.add_all([raport_a, raport_b])
        db.session.commit()

        return {
            'tenant_a': tenant_a.id,
            'tenant_b': tenant_b.id,
            'proiect_a': proiect_a.id,
            'proiect_b': proiect_b.id,
            'angajat_a': angajat_a.id,
            'angajat_b': angajat_b.id,
            'raport_a': raport_a.id,
            'raport_b': raport_b.id,
        }


def _proiect(tenant_id, cod):
    from models import Proiect

    return Proiect(
        tenant_id=tenant_id,
        cod_proiect=cod,
        nume=cod,
        data_start=date(2026, 1, 1),
        status='activ',
    )


def _angajat(tenant_id, cnp, nume):
    from models import Angajat

    return Angajat(
        tenant_id=tenant_id,
        nume=nume,
        prenume='Test',
        cnp=cnp,
        functie='Muncitor',
        data_angajare=date(2026, 1, 1),
        salariu_baza=1680,
        status='activ',
    )


def _seteaza_tenant_user(app, user_id, tenant_id):
    from models import Utilizator, db

    with app.app_context():
        user = db.session.get(Utilizator, user_id)
        user.tenant_id = tenant_id
        db.session.commit()


def _text_xlsx(data):
    wb = load_workbook(BytesIO(data), read_only=True)
    valori = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            valori.extend(str(value) for value in row if value is not None)
    return '\n'.join(valori)


def _curata_date(app):
    from models import Angajat, Document, Pontaj, Proiect, Raport, Tenant, Utilizator, db

    with app.app_context():
        app.config['MULTI_TENANT_MODE'] = 'off'

        for user in Utilizator.query.filter(Utilizator.email.in_([
            'admin_test@test.local',
            'operator_test@test.local',
        ])).all():
            user.tenant_id = None

        Raport.query.filter(
            Raport.titlu.like('TA-REPORT-ROUTE-%')
        ).delete(synchronize_session=False)
        Document.query.filter(
            Document.nume_document.like('TA-REPORT-ROUTE-%')
        ).delete(synchronize_session=False)

        angajat_ids = [
            a.id for a in Angajat.query.filter(
                Angajat.cnp.in_(['7990101010101', '8990101010101'])
            ).all()
        ]
        proiect_ids = [
            p.id for p in Proiect.query.filter(
                Proiect.cod_proiect.like('TA-REPORT-ROUTE-PRJ-%')
            ).all()
        ]
        if angajat_ids or proiect_ids:
            query = Pontaj.query
            if angajat_ids and proiect_ids:
                query = query.filter(
                    (Pontaj.angajat_id.in_(angajat_ids))
                    | (Pontaj.proiect_id.in_(proiect_ids))
                )
            elif angajat_ids:
                query = query.filter(Pontaj.angajat_id.in_(angajat_ids))
            else:
                query = query.filter(Pontaj.proiect_id.in_(proiect_ids))
            query.delete(synchronize_session=False)

        Angajat.query.filter(
            Angajat.cnp.in_(['7990101010101', '8990101010101'])
        ).delete(synchronize_session=False)
        Proiect.query.filter(
            Proiect.cod_proiect.like('TA-REPORT-ROUTE-PRJ-%')
        ).delete(synchronize_session=False)
        Utilizator.query.filter(
            Utilizator.email.like('ta-report-route-%')
        ).delete(synchronize_session=False)
        Tenant.query.filter(
            Tenant.cod.like('test-ta-report-route-%')
        ).delete(synchronize_session=False)
        db.session.commit()

        _sterge_fisiere_test(app.config['EXPORT_FOLDER'])


def _sterge_fisiere_test(export_folder):
    if not export_folder or not os.path.isdir(export_folder):
        return
    for filename in os.listdir(export_folder):
        if filename.startswith('ta_report_route_'):
            try:
                os.remove(os.path.join(export_folder, filename))
            except OSError:
                pass
