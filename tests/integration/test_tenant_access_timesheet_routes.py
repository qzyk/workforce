"""Teste tenant access pentru rutele principale de pontaje."""

from datetime import date
from io import BytesIO

import pytest


CNP_A = '7900404010101'
CNP_B = '7900404010102'
COD_PROIECT_A = 'TA-PONT-R-P-A'
COD_PROIECT_B = 'TA-PONT-R-P-B'
IMPORT_DATE = date(2026, 2, 10)


@pytest.fixture(autouse=True)
def curata_timesheet_tenant_routes(app):
    _curata_date(app)
    yield
    _curata_date(app)


def test_mode_off_lista_arata_toate_pontajele(authenticated_client, app):
    _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'off'

    raspuns = authenticated_client.get('/pontaje/?luna=2&anul=2026')

    assert raspuns.status_code == 200
    assert COD_PROIECT_A.encode() in raspuns.data
    assert COD_PROIECT_B.encode() in raspuns.data


def test_mode_off_editare_si_export_functioneaza(authenticated_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'off'

    editare = authenticated_client.get(f'/pontaje/{ids["pontaj_b_draft"]}/editeaza')
    export = authenticated_client.get('/pontaje/export-lunar?luna=2&anul=2026')

    assert editare.status_code == 200
    assert export.status_code == 200
    assert export.mimetype == 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'


def test_strict_tenant_vede_doar_pontajele_sale_in_lista(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/pontaje/?luna=2&anul=2026')

    assert raspuns.status_code == 200
    assert COD_PROIECT_A.encode() in raspuns.data
    assert COD_PROIECT_B.encode() not in raspuns.data


def test_strict_tenant_poate_edita_pontaj_propriu(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(f'/pontaje/{ids["pontaj_a_draft"]}/editeaza')

    assert raspuns.status_code == 200
    assert COD_PROIECT_A.encode() in raspuns.data


def test_strict_tenant_nu_poate_edita_pontaj_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(f'/pontaje/{ids["pontaj_b_draft"]}/editeaza')

    assert raspuns.status_code == 404


def test_strict_tenant_nu_poate_trimite_pontaj_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post(f'/pontaje/{ids["pontaj_b_draft"]}/trimite')

    assert raspuns.status_code == 404


def test_strict_tenant_nu_poate_aproba_pontaj_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post(f'/pontaje/{ids["pontaj_b_trimis"]}/aproba')

    assert raspuns.status_code == 404


def test_strict_tenant_nu_poate_respinge_pontaj_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post(
        f'/pontaje/{ids["pontaj_b_trimis"]}/respinge',
        data={'motiv': 'Tenant strain'},
    )

    assert raspuns.status_code == 404


def test_strict_tenant_nu_poate_sterge_pontaj_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post(f'/pontaje/{ids["pontaj_b_draft"]}/sterge')

    assert raspuns.status_code == 404


def test_strict_panou_aprobare_arata_doar_pontajele_tenantului(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/pontaje/aprobare')

    assert raspuns.status_code == 200
    assert COD_PROIECT_A.encode() in raspuns.data
    assert COD_PROIECT_B.encode() not in raspuns.data


def test_strict_aprobare_multipla_mixta_este_respinsa_integral(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post('/pontaje/aproba-multiplu', data={
        'pontaj_ids': [str(ids['pontaj_a_trimis']), str(ids['pontaj_b_trimis'])],
    })

    assert raspuns.status_code == 404
    with app.app_context():
        from models import Pontaj

        pontaj_a = Pontaj.query.get(ids['pontaj_a_trimis'])
        pontaj_b = Pontaj.query.get(ids['pontaj_b_trimis'])
        assert pontaj_a.status == 'trimis'
        assert pontaj_b.status == 'trimis'


def test_strict_user_normal_fara_tenant_nu_acceseaza_pontaj(
    operator_client, app
):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = operator_client.get(f'/pontaje/{ids["pontaj_a_draft"]}/editeaza')

    assert raspuns.status_code == 404


def test_optional_cu_tenant_filtreaza_pontajele(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'optional'

    lista = authenticated_client.get('/pontaje/?luna=2&anul=2026')
    editare_straina = authenticated_client.get(f'/pontaje/{ids["pontaj_b_draft"]}/editeaza')

    assert lista.status_code == 200
    assert COD_PROIECT_A.encode() in lista.data
    assert COD_PROIECT_B.encode() not in lista.data
    assert editare_straina.status_code == 404


def test_optional_fara_tenant_pastreaza_comportamentul_de_migrare(
    authenticated_client, app
):
    _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'optional'

    raspuns = authenticated_client.get('/pontaje/?luna=2&anul=2026')

    assert raspuns.status_code == 200
    assert COD_PROIECT_A.encode() in raspuns.data
    assert COD_PROIECT_B.encode() in raspuns.data


def test_strict_creare_respinge_proiect_strain(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post('/pontaje/adauga', data=_form_pontaj(
        ids['angajat_a'],
        ids['proiect_b'],
        '2026-02-11',
    ))

    assert raspuns.status_code == 404
    _assert_nu_exista_pontaj(app, ids['angajat_a'], date(2026, 2, 11))


def test_strict_creare_respinge_angajat_strain(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post('/pontaje/adauga', data=_form_pontaj(
        ids['angajat_b'],
        ids['proiect_a'],
        '2026-02-12',
    ))

    assert raspuns.status_code == 404
    _assert_nu_exista_pontaj(app, ids['angajat_b'], date(2026, 2, 12))


def test_strict_situatie_zilnica_returneaza_doar_pontajele_tenantului(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/pontaje/situatie-zilnica?data=2026-02-02')
    payload = raspuns.get_json()

    assert raspuns.status_code == 200
    assert len(payload) == 1
    assert payload[0]['proiect'] == COD_PROIECT_A


def test_strict_verificare_duplicat_nu_dezvaluie_angajat_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(
        f'/pontaje/verificare-duplicat?angajat_id={ids["angajat_b"]}&data=2026-02-02'
    )

    assert raspuns.status_code == 200
    assert raspuns.get_json() == {'exists': False}


def test_strict_export_lunar_include_doar_pontajele_tenantului(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/pontaje/export-lunar?luna=2&anul=2026')
    texte = _texte_workbook(raspuns.data)

    assert raspuns.status_code == 200
    assert 'TenantA RutaPontaj' in texte
    assert 'TenantB RutaPontaj' not in texte


def test_strict_export_lunar_respinge_proiect_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get(
        f'/pontaje/export-lunar?luna=2&anul=2026&proiect_id={ids["proiect_b"]}'
    )

    assert raspuns.status_code == 404


def test_strict_import_nu_rezolva_angajat_sau_proiect_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    fisier = _fisier_import([
        [CNP_B, COD_PROIECT_A, '10.02.2026', '08:00', '16:00', 'lucratoare', 'angajat strain'],
        [CNP_A, COD_PROIECT_B, '11.02.2026', '08:00', '16:00', 'lucratoare', 'proiect strain'],
    ])
    raspuns = authenticated_client.post('/pontaje/import-excel', data={
        'fisier': (fisier, 'pontaje_tenant.xlsx'),
    })

    assert raspuns.status_code == 302
    _assert_nu_exista_pontaj(app, ids['angajat_b'], IMPORT_DATE)
    _assert_nu_exista_pontaj(app, ids['angajat_a'], date(2026, 2, 11))


def test_strict_teren_pontaj_respinge_proiect_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post('/teren/pontaj', data={
        'proiect_id': ids['proiect_b'],
        'angajat_id': ids['angajat_a'],
        'ore': '8',
        'data': '2026-02-13',
    })

    assert raspuns.status_code == 404
    _assert_nu_exista_pontaj(app, ids['angajat_a'], date(2026, 2, 13))


def test_strict_super_admin_are_acces_explicit_nefiltrat(authenticated_client, app):
    _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/pontaje/?luna=2&anul=2026')

    assert raspuns.status_code == 200
    assert COD_PROIECT_A.encode() in raspuns.data
    assert COD_PROIECT_B.encode() in raspuns.data


def _creeaza_date(app):
    from models import db, Angajat, AngajatProiect, Pontaj, Proiect, Tenant

    with app.app_context():
        tenant_a = Tenant(cod='test-ta-pont-route-a', nume='Tenant Pont Route A')
        tenant_b = Tenant(cod='test-ta-pont-route-b', nume='Tenant Pont Route B')
        db.session.add_all([tenant_a, tenant_b])
        db.session.commit()

        proiect_a = Proiect(
            tenant_id=tenant_a.id,
            cod_proiect=COD_PROIECT_A,
            nume='Tenant Pont Route Project A',
            data_start=date(2026, 1, 1),
            status='activ',
        )
        proiect_b = Proiect(
            tenant_id=tenant_b.id,
            cod_proiect=COD_PROIECT_B,
            nume='Tenant Pont Route Project B',
            data_start=date(2026, 1, 1),
            status='activ',
        )
        angajat_a = Angajat(
            tenant_id=tenant_a.id,
            nume='TenantA',
            prenume='RutaPontaj',
            cnp=CNP_A,
            functie='Inginer',
            data_angajare=date(2026, 1, 1),
            status='activ',
        )
        angajat_b = Angajat(
            tenant_id=tenant_b.id,
            nume='TenantB',
            prenume='RutaPontaj',
            cnp=CNP_B,
            functie='Inginer',
            data_angajare=date(2026, 1, 1),
            status='activ',
        )
        db.session.add_all([proiect_a, proiect_b, angajat_a, angajat_b])
        db.session.commit()

        db.session.add_all([
            AngajatProiect(angajat_id=angajat_a.id, proiect_id=proiect_a.id, data_start=date(2026, 1, 1)),
            AngajatProiect(angajat_id=angajat_b.id, proiect_id=proiect_b.id, data_start=date(2026, 1, 1)),
        ])
        db.session.commit()

        pontaj_a_trimis = Pontaj(
            angajat_id=angajat_a.id,
            proiect_id=proiect_a.id,
            data=date(2026, 2, 2),
            ore_lucrate=8,
            ore_normale=8,
            status='trimis',
        )
        pontaj_b_trimis = Pontaj(
            angajat_id=angajat_b.id,
            proiect_id=proiect_b.id,
            data=date(2026, 2, 2),
            ore_lucrate=8,
            ore_normale=8,
            status='trimis',
        )
        pontaj_a_draft = Pontaj(
            angajat_id=angajat_a.id,
            proiect_id=proiect_a.id,
            data=date(2026, 2, 3),
            ore_lucrate=8,
            ore_normale=8,
            status='draft',
        )
        pontaj_b_draft = Pontaj(
            angajat_id=angajat_b.id,
            proiect_id=proiect_b.id,
            data=date(2026, 2, 3),
            ore_lucrate=8,
            ore_normale=8,
            status='draft',
        )
        db.session.add_all([
            pontaj_a_trimis,
            pontaj_b_trimis,
            pontaj_a_draft,
            pontaj_b_draft,
        ])
        db.session.commit()

        return {
            'tenant_a': tenant_a.id,
            'tenant_b': tenant_b.id,
            'proiect_a': proiect_a.id,
            'proiect_b': proiect_b.id,
            'angajat_a': angajat_a.id,
            'angajat_b': angajat_b.id,
            'pontaj_a_trimis': pontaj_a_trimis.id,
            'pontaj_b_trimis': pontaj_b_trimis.id,
            'pontaj_a_draft': pontaj_a_draft.id,
            'pontaj_b_draft': pontaj_b_draft.id,
        }


def _seteaza_tenant_user(app, user_id, tenant_id):
    from models import db, Utilizator

    with app.app_context():
        user = db.session.get(Utilizator, user_id)
        user.tenant_id = tenant_id
        db.session.commit()


def _form_pontaj(angajat_id, proiect_id, data_pontaj):
    return {
        'angajat_id': str(angajat_id),
        'proiect_id': str(proiect_id),
        'data': data_pontaj,
        'ora_start': '08:00',
        'ora_sfarsit': '16:00',
        'tip_zi': 'lucratoare',
        'actiune': 'draft',
    }


def _fisier_import(randuri):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append([
        'CNP Angajat*',
        'Cod Proiect*',
        'Data (ZZ.LL.AAAA)*',
        'Ora Start (HH:MM)*',
        'Ora Sfarsit (HH:MM)*',
        'Tip Zi',
        'Observatii',
    ])
    for rand in randuri:
        ws.append(rand)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def _texte_workbook(data):
    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(data), data_only=True)
    texte = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    texte.add(str(value))
    return texte


def _assert_nu_exista_pontaj(app, angajat_id, data_pontaj):
    from models import Pontaj

    with app.app_context():
        assert Pontaj.query.filter_by(angajat_id=angajat_id, data=data_pontaj).first() is None


def _curata_date(app):
    from models import db, Angajat, AngajatProiect, Pontaj, Proiect, Tenant, Utilizator

    with app.app_context():
        app.config['MULTI_TENANT_MODE'] = 'off'

        for user in Utilizator.query.filter(Utilizator.email.in_([
            'admin_test@test.local',
            'operator_test@test.local',
        ])).all():
            user.tenant_id = None

        angajat_ids = [
            a.id for a in Angajat.query.filter(Angajat.cnp.in_([CNP_A, CNP_B])).all()
        ]
        proiect_ids = [
            p.id for p in Proiect.query.filter(Proiect.cod_proiect.like('TA-PONT-R-%')).all()
        ]
        if angajat_ids:
            Pontaj.query.filter(Pontaj.angajat_id.in_(angajat_ids)).delete(synchronize_session=False)
            AngajatProiect.query.filter(AngajatProiect.angajat_id.in_(angajat_ids)).delete(synchronize_session=False)
        if proiect_ids:
            Pontaj.query.filter(Pontaj.proiect_id.in_(proiect_ids)).delete(synchronize_session=False)
            AngajatProiect.query.filter(AngajatProiect.proiect_id.in_(proiect_ids)).delete(synchronize_session=False)
        Proiect.query.filter(Proiect.cod_proiect.like('TA-PONT-R-%')).delete(synchronize_session=False)
        Angajat.query.filter(Angajat.cnp.in_([CNP_A, CNP_B])).delete(synchronize_session=False)
        Tenant.query.filter(Tenant.cod.like('test-ta-pont-route-%')).delete(synchronize_session=False)
        db.session.commit()
