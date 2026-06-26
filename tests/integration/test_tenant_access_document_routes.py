"""Teste tenant access pentru rutele de documente."""

from datetime import date
from io import BytesIO
import os

import pytest


@pytest.fixture(autouse=True)
def curata_document_tenant_access(app):
    _curata_date(app)
    yield
    _curata_date(app)


def test_off_mode_documente_legacy_si_export_proiect_ramane_compatibil(
    authenticated_client, app
):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'off'

    lista = authenticated_client.get('/documente/lista')
    download = authenticated_client.get(f'/documente/{ids["legacy_b"]}/descarca')
    export = authenticated_client.get(
        f'/proiecte/{ids["proiect_b"]}/documente/export-index'
    )

    assert lista.status_code == 200
    assert b'TA-DOC-ROUTE-LEG-A' in lista.data
    assert b'TA-DOC-ROUTE-LEG-B' in lista.data
    assert download.status_code == 200
    assert export.status_code == 200
    assert export.mimetype == (
        'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


def test_strict_tenant_poate_accesa_document_proiect_propriu(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    detaliu = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}/documente/{ids["doc_proiect_a"]}'
    )
    download = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}/documente/{ids["doc_proiect_a"]}/descarca'
    )

    assert detaliu.status_code == 200
    assert b'TA-DOC-ROUTE-PROJ-A' in detaliu.data
    assert download.status_code == 200


def test_strict_blocheaza_document_proiect_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    detaliu = authenticated_client.get(
        f'/proiecte/{ids["proiect_b"]}/documente/{ids["doc_proiect_b"]}'
    )
    download = authenticated_client.get(
        f'/proiecte/{ids["proiect_b"]}/documente/{ids["doc_proiect_b"]}/descarca'
    )
    preview = authenticated_client.get(
        f'/proiecte/{ids["proiect_b"]}/documente/{ids["doc_proiect_b"]}/preview'
    )
    sterge = authenticated_client.post(
        f'/proiecte/{ids["proiect_b"]}/documente/{ids["doc_proiect_b"]}/sterge'
    )

    assert detaliu.status_code == 404
    assert download.status_code == 404
    assert preview.status_code == 404
    assert sterge.status_code == 404

    with app.app_context():
        from models import DocumentProiect, db

        assert db.session.get(DocumentProiect, ids['doc_proiect_b']) is not None


def test_strict_blocheaza_revizie_straina(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    download = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}/documente/revizie/{ids["revizie_b"]}/descarca'
    )

    assert download.status_code == 404


def test_strict_blocheaza_document_legacy_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    lista = authenticated_client.get('/documente/lista')
    download = authenticated_client.get(f'/documente/{ids["legacy_b"]}/descarca')
    preview = authenticated_client.get(f'/documente/{ids["legacy_b"]}/preview')
    edit = authenticated_client.get(f'/documente/{ids["legacy_b"]}/editeaza')

    assert lista.status_code == 200
    assert b'TA-DOC-ROUTE-LEG-A' in lista.data
    assert b'TA-DOC-ROUTE-LEG-B' not in lista.data
    assert download.status_code == 404
    assert preview.status_code == 404
    assert edit.status_code == 404


def test_strict_user_fara_tenant_esueaza_inchis(operator_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'strict'

    legacy = operator_client.get(f'/documente/{ids["legacy_a"]}/descarca')
    proiect_doc = operator_client.get(
        f'/proiecte/{ids["proiect_a"]}/documente/{ids["doc_proiect_a"]}'
    )

    assert legacy.status_code == 404
    assert proiect_doc.status_code == 404


def test_optional_mode_scopeaza_cu_tenant_si_ramane_permisiv_fara_tenant(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'optional'

    scoped = authenticated_client.get('/documente/lista')

    _seteaza_tenant_user(app, admin_user.id, None)
    permisiv = authenticated_client.get('/documente/lista')

    assert scoped.status_code == 200
    assert b'TA-DOC-ROUTE-LEG-A' in scoped.data
    assert b'TA-DOC-ROUTE-LEG-B' not in scoped.data
    assert permisiv.status_code == 200
    assert b'TA-DOC-ROUTE-LEG-A' in permisiv.data
    assert b'TA-DOC-ROUTE-LEG-B' in permisiv.data


def test_strict_upload_nu_accepta_proiect_sau_angajat_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    proiect_upload = authenticated_client.post(
        f'/proiecte/{ids["proiect_b"]}/documente/adauga',
        data={
            'tip_instalatie_id': str(ids['instalatie_id']),
            'denumire_document': 'TA-DOC-ROUTE-UPLOAD-PROJ-B',
        },
    )
    hr_upload = authenticated_client.post(
        '/documente/upload',
        data={
            'angajat_id': str(ids['angajat_b']),
            'proiect_id': '0',
            'tip': 'alte',
            'nume_document': 'TA-DOC-ROUTE-UPLOAD-HR-B',
            'fisier': (BytesIO(b'%PDF-1.4 test'), 'ta-doc-route-upload.pdf'),
        },
        content_type='multipart/form-data',
    )

    assert proiect_upload.status_code == 404
    assert hr_upload.status_code == 200

    with app.app_context():
        from models import Document

        assert Document.query.filter_by(
            nume_document='TA-DOC-ROUTE-UPLOAD-HR-B'
        ).first() is None


def test_strict_export_index_scopeaza_documentele(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    export = authenticated_client.get(
        f'/proiecte/{ids["proiect_a"]}/documente/export-index'
    )
    export_strain = authenticated_client.get(
        f'/proiecte/{ids["proiect_b"]}/documente/export-index'
    )

    assert export.status_code == 200
    assert export_strain.status_code == 404

    from openpyxl import load_workbook

    wb = load_workbook(BytesIO(export.data), read_only=True)
    valori = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            valori.extend(str(value) for value in row if value is not None)
    text = '\n'.join(valori)

    assert 'TA-DOC-ROUTE-PROJ-A' in text
    assert 'TA-DOC-ROUTE-PROJ-B' not in text


def _creeaza_date(app):
    from models import (
        db, Angajat, Document, DocumentProiect, Proiect, RevizieDocument,
        Tenant, TipInstalatie,
    )

    with app.app_context():
        tenant_a = Tenant(cod='test-ta-doc-route-a', nume='Tenant Doc Route A')
        tenant_b = Tenant(cod='test-ta-doc-route-b', nume='Tenant Doc Route B')
        db.session.add_all([tenant_a, tenant_b])
        db.session.commit()

        proiect_a = _proiect(tenant_a.id, 'TA-DOC-ROUTE-PRJ-A')
        proiect_b = _proiect(tenant_b.id, 'TA-DOC-ROUTE-PRJ-B')
        angajat_a = _angajat(tenant_a.id, '3990101010101', 'DocRouteA')
        angajat_b = _angajat(tenant_b.id, '4990101010101', 'DocRouteB')
        inst = TipInstalatie(
            cod='TA-DOC-ROUTE-INST',
            denumire='TA Document Route Instalatie',
            activ=True,
        )
        db.session.add_all([proiect_a, proiect_b, angajat_a, angajat_b, inst])
        db.session.commit()

        upload_folder = app.config['UPLOAD_FOLDER']
        legacy_a_path = _write_abs_file(
            upload_folder, 'ta_doc_route_legacy_a.pdf', b'%PDF legacy A'
        )
        legacy_b_path = _write_abs_file(
            upload_folder, 'ta_doc_route_legacy_b.pdf', b'%PDF legacy B'
        )
        proiect_a_rel = _write_project_file(
            upload_folder, proiect_a.id, 'ta_doc_route_project_a.pdf', b'%PDF project A'
        )
        proiect_b_rel = _write_project_file(
            upload_folder, proiect_b.id, 'ta_doc_route_project_b.pdf', b'%PDF project B'
        )
        rev_a_rel = _write_project_file(
            upload_folder, proiect_a.id, 'ta_doc_route_revision_a.pdf', b'%PDF rev A'
        )
        rev_b_rel = _write_project_file(
            upload_folder, proiect_b.id, 'ta_doc_route_revision_b.pdf', b'%PDF rev B'
        )

        legacy_a = _doc_legacy(
            angajat_a.id, None, 'TA-DOC-ROUTE-LEG-A', legacy_a_path
        )
        legacy_b = _doc_legacy(
            angajat_b.id, None, 'TA-DOC-ROUTE-LEG-B', legacy_b_path
        )
        doc_proiect_a = _doc_proiect(
            proiect_a.id, inst.id, 'TA-DOC-ROUTE-PROJ-A', proiect_a_rel
        )
        doc_proiect_b = _doc_proiect(
            proiect_b.id, inst.id, 'TA-DOC-ROUTE-PROJ-B', proiect_b_rel
        )
        db.session.add_all([legacy_a, legacy_b, doc_proiect_a, doc_proiect_b])
        db.session.commit()

        revizie_a = RevizieDocument(
            document_proiect_id=doc_proiect_a.id,
            nr_revizie=1,
            fisier_path=rev_a_rel,
        )
        revizie_b = RevizieDocument(
            document_proiect_id=doc_proiect_b.id,
            nr_revizie=1,
            fisier_path=rev_b_rel,
        )
        db.session.add_all([revizie_a, revizie_b])
        db.session.commit()

        return {
            'tenant_a': tenant_a.id,
            'tenant_b': tenant_b.id,
            'proiect_a': proiect_a.id,
            'proiect_b': proiect_b.id,
            'angajat_a': angajat_a.id,
            'angajat_b': angajat_b.id,
            'instalatie_id': inst.id,
            'legacy_a': legacy_a.id,
            'legacy_b': legacy_b.id,
            'doc_proiect_a': doc_proiect_a.id,
            'doc_proiect_b': doc_proiect_b.id,
            'revizie_a': revizie_a.id,
            'revizie_b': revizie_b.id,
        }


def _proiect(tenant_id, cod):
    from models import Proiect

    return Proiect(
        tenant_id=tenant_id,
        cod_proiect=cod,
        nume=cod,
        data_start=date(2026, 1, 1),
        status='activ',
    )


def _angajat(tenant_id, cnp, nume):
    from models import Angajat

    return Angajat(
        tenant_id=tenant_id,
        nume=nume,
        prenume='Test',
        cnp=cnp,
        functie='Muncitor',
        data_angajare=date(2026, 1, 1),
        status='activ',
    )


def _doc_legacy(angajat_id, proiect_id, denumire, path):
    from models import Document

    return Document(
        angajat_id=angajat_id,
        proiect_id=proiect_id,
        tip='alte',
        nume_document=denumire,
        fisier_path=path,
        status='valabil',
    )


def _doc_proiect(proiect_id, instalatie_id, denumire, rel_path):
    from models import DocumentProiect

    return DocumentProiect(
        proiect_id=proiect_id,
        tip_instalatie_id=instalatie_id,
        denumire_document=denumire,
        status='draft',
        versiune_curenta=True,
        fisier_path=rel_path,
        tip_fisier='pdf',
    )


def _write_abs_file(upload_folder, filename, content):
    os.makedirs(upload_folder, exist_ok=True)
    path = os.path.join(upload_folder, filename)
    with open(path, 'wb') as handle:
        handle.write(content)
    return path


def _write_project_file(upload_folder, proiect_id, filename, content):
    folder = os.path.join(upload_folder, f'proiect_{proiect_id}')
    os.makedirs(folder, exist_ok=True)
    path = os.path.join(folder, filename)
    with open(path, 'wb') as handle:
        handle.write(content)
    return os.path.join(f'proiect_{proiect_id}', filename)


def _seteaza_tenant_user(app, user_id, tenant_id):
    from models import db, Utilizator

    with app.app_context():
        user = db.session.get(Utilizator, user_id)
        user.tenant_id = tenant_id
        db.session.commit()


def _curata_date(app):
    from models import (
        db, Angajat, Document, DocumentProiect, Proiect, RevizieDocument,
        Tenant, TipInstalatie, Utilizator,
    )

    with app.app_context():
        app.config['MULTI_TENANT_MODE'] = 'off'

        for user in Utilizator.query.filter(Utilizator.email.in_([
            'admin_test@test.local',
            'operator_test@test.local',
        ])).all():
            user.tenant_id = None

        Document.query.filter(
            Document.nume_document.like('TA-DOC-ROUTE-%')
        ).delete(synchronize_session=False)
        doc_ids = [
            d.id for d in DocumentProiect.query.filter(
                DocumentProiect.denumire_document.like('TA-DOC-ROUTE-%')
            ).all()
        ]
        if doc_ids:
            RevizieDocument.query.filter(
                RevizieDocument.document_proiect_id.in_(doc_ids)
            ).delete(synchronize_session=False)
        DocumentProiect.query.filter(
            DocumentProiect.denumire_document.like('TA-DOC-ROUTE-%')
        ).delete(synchronize_session=False)
        TipInstalatie.query.filter_by(cod='TA-DOC-ROUTE-INST').delete()
        Angajat.query.filter(Angajat.cnp.in_([
            '3990101010101', '4990101010101',
        ])).delete(synchronize_session=False)
        Proiect.query.filter(
            Proiect.cod_proiect.like('TA-DOC-ROUTE-PRJ-%')
        ).delete(synchronize_session=False)
        Tenant.query.filter(
            Tenant.cod.like('test-ta-doc-route-%')
        ).delete(synchronize_session=False)
        db.session.commit()

        _sterge_fisiere_test(app.config['UPLOAD_FOLDER'])


def _sterge_fisiere_test(upload_folder):
    if not upload_folder or not os.path.isdir(upload_folder):
        return
    for root, _dirs, files in os.walk(upload_folder):
        for filename in files:
            if filename.startswith('ta_doc_route_') or filename.startswith('ta-doc-route-'):
                try:
                    os.remove(os.path.join(root, filename))
                except OSError:
                    pass
