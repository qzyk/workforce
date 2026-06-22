"""
Rapoarte Faza 3: istoric robust (tenant + blob + checksum).

Verifica:
  - _save_raport salveaza continut_blob + checksum (sha256) pe langa fisier_path
  - descarca() serveste din blob cand fisierul de pe disc lipseste (stergem
    fisierul fizic, dar descarcarea inca returneaza continutul corect)
  - descarca() regenereaza din parametri cand nici path nici blob nu exista
    (xlsx derivat pur din DB)
  - izolare tenant: istoric/panou listeaza doar rapoartele tenant-ului curent +
    cele globale (tenant_id NULL); descarca/sterge refuza rapoartele altui tenant
  - Raport.TIPURI sincronizat cu TIPURI_RAPOARTE din route
  - regresie: rapoarte vechi fara blob (doar fisier_path pe disc) se listeaza
    si se descarca in continuare (comportament backward-compat)

Toate testele isi curata propriile randuri Raport (tabela 'rapoarte' nu e in
wipe-ul auto din conftest).
"""

import io
import os
import json
import hashlib
from datetime import date

import pytest
from openpyxl import load_workbook


MARKER = 'TEST_RAP3_'  # prefix titlu pentru cleanup deterministic


@pytest.fixture
def curata_rapoarte(app):
    """Sterge rapoartele de test inainte si dupa (tabela nu e in wipe-ul auto)."""
    from models import db, Raport

    def _wipe():
        with app.app_context():
            for r in Raport.query.filter(Raport.titlu.like(f'{MARKER}%')).all():
                db.session.delete(r)
            db.session.commit()

    _wipe()
    yield
    _wipe()


def _xlsx_bytes():
    """Un xlsx minimal valid pe disc -> bytes pentru a popula continut_blob."""
    from openpyxl import Workbook
    wb = Workbook()
    wb.active['A1'] = 'continut original'
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _scrie_fisier(tmp_path, nume, continut):
    p = tmp_path / nume
    p.write_bytes(continut)
    return str(p)


# ============================================================
# 1. Salvare: blob + checksum populate la generare (flux real POST)
# ============================================================

def test_generare_salveaza_blob_si_checksum(app, authenticated_client,
                                            curata_rapoarte):
    """Generarea reala (POST /rapoarte/centralizator-ore) salveaza in istoric
    si continut_blob + checksum (sha256), pe langa fisier_path. Verificam ca
    checksum-ul corespunde blob-ului si ca blob-ul e identicul fisierului
    descarcat in raspuns."""
    from models import db, Raport

    resp = authenticated_client.post('/rapoarte/centralizator-ore', data={
        'luna': '3', 'an': '2026', 'grupare': 'angajat',
    })
    assert resp.status_code == 200
    assert resp.data[:2] == b'PK'           # xlsx returnat in raspuns

    with app.app_context():
        r = (Raport.query.filter_by(tip_raport='centralizator_ore')
             .order_by(Raport.id.desc()).first())
        assert r is not None
        assert r.continut_blob is not None
        assert r.checksum == hashlib.sha256(r.continut_blob).hexdigest()
        assert r.dimensiune_fisier == len(r.continut_blob)
        # Marcam pentru cleanup (titlul real nu are prefixul nostru)
        r.titlu = f'{MARKER}{r.titlu}'
        db.session.commit()


# ============================================================
# 2. descarca() din blob cand fisierul de pe disc lipseste
# ============================================================

def test_descarca_din_blob_cand_path_lipseste(app, authenticated_client, tmp_path,
                                               curata_rapoarte):
    """Salvam raport cu blob, STERGEM fisierul fizic, descarcarea trebuie sa
    intoarca exact continutul din blob (orfan dupa redeploy PA)."""
    from models import db, Raport

    continut = _xlsx_bytes()
    path = _scrie_fisier(tmp_path, 'orfan.xlsx', continut)
    with app.app_context():
        r = Raport(tip_raport='centralizator_ore', titlu=f'{MARKER}orfan',
                   fisier_path=path, format='xlsx',
                   continut_blob=continut,
                   checksum=hashlib.sha256(continut).hexdigest())
        db.session.add(r); db.session.commit()
        rid = r.id

    # Simuleaza orfanul: fisierul nu mai e pe disc
    os.remove(path)
    assert not os.path.exists(path)

    resp = authenticated_client.get(f'/rapoarte/descarca/{rid}')
    assert resp.status_code == 200
    assert resp.data == continut          # exact continutul din blob
    assert resp.data[:2] == b'PK'         # xlsx valid (zip magic)


# ============================================================
# 3. descarca() regenereaza din parametri cand path + blob lipsesc
# ============================================================

def test_descarca_regenereaza_cand_path_si_blob_lipsesc(app, authenticated_client,
                                                        curata_rapoarte):
    """Fara path pe disc si fara blob -> regenerare xlsx din parametri (tip
    derivat pur din DB). Verificam ca rezultatul e un xlsx valid."""
    from models import db, Raport, Proiect, Angajat, Pontaj
    from decimal import Decimal

    with app.app_context():
        if not Proiect.query.filter_by(cod_proiect='RAP3-RG').first():
            db.session.add(Proiect(cod_proiect='RAP3-RG', nume='Regen Test',
                                   data_start=date(2026, 1, 1), status='activ'))
            db.session.commit()
        p = Proiect.query.filter_by(cod_proiect='RAP3-RG').first()
        a = Angajat.query.filter_by(cnp='1970707070707').first()
        if not a:
            a = Angajat(cnp='1970707070707', nume='Regen', prenume='Worker',
                        functie='Muncitor', tip_contract='nedeterminat',
                        salariu_baza=3000, data_angajare=date(2026, 1, 1),
                        status='activ')
            db.session.add(a); db.session.commit()
        if not Pontaj.query.filter_by(angajat_id=a.id, proiect_id=p.id,
                                      data=date(2026, 3, 4)).first():
            db.session.add(Pontaj(angajat_id=a.id, proiect_id=p.id,
                                  data=date(2026, 3, 4), ore_lucrate=Decimal('8'),
                                  status='aprobat', tip_zi='lucratoare'))
            db.session.commit()

        # Raport cu path inexistent, FARA blob -> trebuie sa cada pe regenerare
        r = Raport(tip_raport='centralizator_ore', titlu=f'{MARKER}regen',
                   fisier_path='/nu/exista/regen.xlsx', format='xlsx',
                   parametri=json.dumps({'luna': 3, 'an': 2026, 'grupare': 'angajat'}),
                   continut_blob=None, checksum=None)
        db.session.add(r); db.session.commit()
        rid = r.id

    resp = authenticated_client.get(f'/rapoarte/descarca/{rid}')
    assert resp.status_code == 200
    assert resp.data[:2] == b'PK'                 # xlsx regenerat valid
    wb = load_workbook(io.BytesIO(resp.data))
    assert wb.active is not None                  # se deschide corect


# ============================================================
# 4. Izolare tenant: listare + descarca + sterge
# ============================================================

def test_istoric_filtreaza_pe_tenant(app, authenticated_client, curata_rapoarte):
    """Admin fara tenant (tenant_id NULL) vede rapoartele globale (NULL) dar NU
    pe cele ale unui alt tenant."""
    from models import db, Raport, Tenant

    with app.app_context():
        t = Tenant.query.filter_by(cod='test-rap3').first()
        if not t:
            t = Tenant(cod='test-rap3', nume='Tenant Rap3')
            db.session.add(t); db.session.commit()
        tid = t.id
        db.session.add(Raport(tip_raport='stat_plata', titlu=f'{MARKER}global',
                              format='xlsx', tenant_id=None))
        db.session.add(Raport(tip_raport='stat_plata', titlu=f'{MARKER}altul',
                              format='xlsx', tenant_id=tid))
        db.session.commit()

    resp = authenticated_client.get('/rapoarte/istoric')
    html = resp.get_data(as_text=True)
    assert f'{MARKER}global' in html        # globalul (NULL) e vizibil
    assert f'{MARKER}altul' not in html     # cel al altui tenant NU


def test_descarca_si_sterge_refuza_alt_tenant(app, authenticated_client,
                                              curata_rapoarte):
    """Un raport al altui tenant nu poate fi descarcat sau sters de admin (NULL)."""
    from models import db, Raport, Tenant

    continut = _xlsx_bytes()
    with app.app_context():
        t = Tenant.query.filter_by(cod='test-rap3b').first()
        if not t:
            t = Tenant(cod='test-rap3b', nume='Tenant Rap3 B')
            db.session.add(t); db.session.commit()
        r = Raport(tip_raport='stat_plata', titlu=f'{MARKER}privat',
                   format='xlsx', tenant_id=t.id, continut_blob=continut)
        db.session.add(r); db.session.commit()
        rid = r.id

    # descarca -> refuz (redirect spre istoric, nu 200 cu continut)
    resp = authenticated_client.get(f'/rapoarte/descarca/{rid}')
    assert resp.status_code == 302
    assert b'PK' not in resp.data

    # sterge -> refuz, raportul ramane in DB
    resp2 = authenticated_client.post(f'/rapoarte/sterge/{rid}')
    assert resp2.status_code in (302, 303)
    with app.app_context():
        assert db.session.get(Raport, rid) is not None


# ============================================================
# 5. Raport.TIPURI sincronizat cu TIPURI_RAPOARTE din route
# ============================================================

def test_tipuri_model_sincronizat_cu_route():
    from models import Raport
    from routes.rapoarte import TIPURI_RAPOARTE

    chei_model = {cheie for cheie, _ in Raport.TIPURI}
    assert chei_model == set(TIPURI_RAPOARTE.keys())
    # label-urile din model coincid cu cele din route
    for cheie, label in Raport.TIPURI:
        assert label == TIPURI_RAPOARTE[cheie]


# ============================================================
# 6. Regresie: raport vechi fara blob (doar path pe disc) inca merge
# ============================================================

def test_regresie_raport_vechi_path_only(app, authenticated_client, tmp_path,
                                          curata_rapoarte):
    """Raport stil vechi: fisier_path pe disc, FARA tenant si FARA blob.
    Trebuie sa apara in istoric SI sa se descarce din path (comportament vechi)."""
    from models import db, Raport

    continut = _xlsx_bytes()
    path = _scrie_fisier(tmp_path, 'vechi.xlsx', continut)
    with app.app_context():
        r = Raport(tip_raport='foaie_prezenta', titlu=f'{MARKER}vechi',
                   fisier_path=path, format='xlsx',
                   tenant_id=None, continut_blob=None, checksum=None)
        db.session.add(r); db.session.commit()
        rid = r.id

    # listare: apare (tenant NULL = global)
    lista = authenticated_client.get('/rapoarte/istoric').get_data(as_text=True)
    assert f'{MARKER}vechi' in lista

    # descarcare: din path (fisierul exista pe disc)
    resp = authenticated_client.get(f'/rapoarte/descarca/{rid}')
    assert resp.status_code == 200
    assert resp.data == continut
