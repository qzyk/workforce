"""Teste tenant access pentru rutele HR si flota."""

from datetime import date
from io import BytesIO

import pytest
from openpyxl import load_workbook


@pytest.fixture(autouse=True)
def curata_hr_fleet_routes(app):
    _curata_date(app)
    yield
    _curata_date(app)


def test_mode_off_pastreaza_detalii_hr_si_flota(authenticated_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'off'

    angajat = authenticated_client.get(f'/angajati/{ids["angajat_b"]}')
    masina = authenticated_client.get(f'/masini/{ids["masina_b"]}')

    assert angajat.status_code == 200
    assert masina.status_code == 200
    assert b'TAHR Route Angajat B' in angajat.data
    assert b'TAHR-R-B' in masina.data


def test_strict_lista_hr_arata_doar_angajatii_tenantului(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/angajati/')

    assert raspuns.status_code == 200
    assert b'TAHR Route Angajat A' in raspuns.data
    assert b'TAHR Route Angajat B' not in raspuns.data


def test_strict_blocheaza_detalii_editare_si_dezactivare_angajat_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    detalii = authenticated_client.get(f'/angajati/{ids["angajat_b"]}')
    editare = authenticated_client.get(f'/angajati/{ids["angajat_b"]}/editeaza')
    dezactivare = authenticated_client.post(f'/angajati/{ids["angajat_b"]}/dezactiveaza')

    assert detalii.status_code == 404
    assert editare.status_code == 404
    assert dezactivare.status_code == 404


def test_strict_export_hr_contine_doar_angajatii_tenantului(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/angajati/export-excel')

    assert raspuns.status_code == 200
    texte = _xlsx_texts(raspuns.data)
    assert 'TAHR Route Angajat A' in texte
    assert 'TAHR Route Angajat B' not in texte


def test_strict_creare_angajat_asigneaza_tenantul_curent(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post('/angajati/adauga', data=_form_angajat('TAHR Nou'))

    assert raspuns.status_code == 302
    with app.app_context():
        from models import Angajat

        angajat = Angajat.query.filter_by(nume='TAHR Nou').one()
        assert angajat.tenant_id == ids['tenant_a']


def test_strict_user_normal_fara_tenant_nu_acceseaza_si_nu_creeaza_hr(operator_client, app):
    ids = _creeaza_date(app)
    app.config['MULTI_TENANT_MODE'] = 'strict'

    detalii = operator_client.get(f'/angajati/{ids["angajat_a"]}')
    creare = operator_client.post('/angajati/adauga', data=_form_angajat('TAHR Fara Tenant'))

    assert detalii.status_code == 404
    assert creare.status_code == 403


def test_strict_lista_flota_arata_doar_masinile_tenantului(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/masini/')

    assert raspuns.status_code == 200
    assert b'TAHR-R-A' in raspuns.data
    assert b'TAHR-R-B' not in raspuns.data


def test_strict_blocheaza_detalii_editare_status_masina_straina(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    detalii = authenticated_client.get(f'/masini/{ids["masina_b"]}')
    editare = authenticated_client.get(f'/masini/{ids["masina_b"]}/editeaza')
    status = authenticated_client.post(
        f'/masini/{ids["masina_b"]}/status',
        data={'new_status': 'service'},
    )

    assert detalii.status_code == 404
    assert editare.status_code == 404
    assert status.status_code == 404


def test_strict_export_flota_contine_doar_masinile_tenantului(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.get('/masini/export-excel')

    assert raspuns.status_code == 200
    texte = _xlsx_texts(raspuns.data)
    assert 'TAHR-R-A' in texte
    assert 'TAHR-R-B' not in texte


def test_strict_blocheaza_document_si_defectiune_masina_straina(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    doc = authenticated_client.post(f'/masini/document/{ids["document_b"]}/sterge')
    defect = authenticated_client.post(
        f'/masini/defectiune/{ids["defect_b"]}/status',
        data={'status_defectiune': 'rezolvata'},
    )

    assert doc.status_code == 404
    assert defect.status_code == 404


def test_strict_blocheaza_atribuire_cu_angajat_sau_proiect_strain(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    angajat_strain = authenticated_client.post(
        f'/masini/{ids["masina_a"]}/atribuie',
        data={
            'angajat_id': str(ids['angajat_b']),
            'proiect_id': str(ids['proiect_a']),
            'data_atribuire': '2026-03-01',
        },
    )
    proiect_strain = authenticated_client.post(
        f'/masini/{ids["masina_a"]}/atribuie',
        data={
            'angajat_id': str(ids['angajat_a']),
            'proiect_id': str(ids['proiect_b']),
            'data_atribuire': '2026-03-01',
        },
    )

    assert angajat_strain.status_code == 404
    assert proiect_strain.status_code == 404


def test_strict_blocheaza_defectiune_pe_masina_straina(authenticated_client, app, admin_user):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post(
        f'/masini/{ids["masina_b"]}/defectiune',
        data={
            'raportat_de': str(ids['angajat_a']),
            'data_raportare': '2026-03-01',
            'descriere_defectiune': 'Nu ar trebui salvat',
            'gravitate': 'medie',
        },
    )

    assert raspuns.status_code == 404


def test_strict_creare_masina_fara_owner_tenant_safe_este_blocata(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    raspuns = authenticated_client.post('/masini/adauga', data=_form_masina('TAHR-R-NO'))

    assert raspuns.status_code == 403


def test_strict_creare_masina_cu_owner_tenant_safe_functioneaza(
    authenticated_client, app, admin_user
):
    ids = _creeaza_date(app)
    _seteaza_tenant_user(app, admin_user.id, ids['tenant_a'])
    app.config['MULTI_TENANT_MODE'] = 'strict'

    form = _form_masina('TAHR-R-NEW')
    form['proiect_id'] = str(ids['proiect_a'])
    form['angajat_responsabil_id'] = str(ids['angajat_a'])
    raspuns = authenticated_client.post('/masini/adauga', data=form)

    assert raspuns.status_code == 302
    with app.app_context():
        from models import Masina

        masina = Masina.query.filter_by(numar_inmatriculare='TAHR-R-NEW').one()
        assert masina.proiect_id == ids['proiect_a']
        assert masina.angajat_responsabil_id == ids['angajat_a']


def _creeaza_date(app):
    from models import (
        Angajat, AtribuireMasina, Concediu, DefectiuneMasina, DocumentMasina,
        Masina, Proiect, Tenant, db,
    )

    with app.app_context():
        tenant_a = Tenant(cod='test-tahr-route-a', nume='Tenant HR Route A')
        tenant_b = Tenant(cod='test-tahr-route-b', nume='Tenant HR Route B')
        db.session.add_all([tenant_a, tenant_b])
        db.session.commit()

        proiect_a = Proiect(
            tenant_id=tenant_a.id,
            cod_proiect='TAHR-R-PA',
            nume='TAHR Route Proiect A',
            data_start=date(2026, 1, 1),
            status='activ',
        )
        proiect_b = Proiect(
            tenant_id=tenant_b.id,
            cod_proiect='TAHR-R-PB',
            nume='TAHR Route Proiect B',
            data_start=date(2026, 1, 1),
            status='activ',
        )
        angajat_a = Angajat(
            tenant_id=tenant_a.id,
            nume='TAHR Route Angajat A',
            prenume='Test',
            functie='Muncitor',
            data_angajare=date(2026, 1, 1),
            status='activ',
            cnp='1900101010201',
        )
        angajat_b = Angajat(
            tenant_id=tenant_b.id,
            nume='TAHR Route Angajat B',
            prenume='Test',
            functie='Muncitor',
            data_angajare=date(2026, 1, 1),
            status='activ',
            cnp='1900101010202',
        )
        db.session.add_all([proiect_a, proiect_b, angajat_a, angajat_b])
        db.session.commit()

        masina_a = Masina(
            numar_inmatriculare='TAHR-R-A',
            marca='Dacia',
            model='A',
            proiect_id=proiect_a.id,
            angajat_responsabil_id=angajat_a.id,
            status='atribuita',
        )
        masina_b = Masina(
            numar_inmatriculare='TAHR-R-B',
            marca='Dacia',
            model='B',
            proiect_id=proiect_b.id,
            angajat_responsabil_id=angajat_b.id,
            status='atribuita',
        )
        db.session.add_all([masina_a, masina_b])
        db.session.commit()

        document_b = DocumentMasina(masina_id=masina_b.id, tip='itp', nume_document='Doc B')
        defect_b = DefectiuneMasina(
            masina_id=masina_b.id,
            raportat_de=angajat_b.id,
            descriere='Defect B',
        )
        db.session.add_all([
            document_b,
            defect_b,
            AtribuireMasina(masina_id=masina_a.id, angajat_id=angajat_a.id, proiect_id=proiect_a.id),
            AtribuireMasina(masina_id=masina_b.id, angajat_id=angajat_b.id, proiect_id=proiect_b.id),
            Concediu(
                angajat_id=angajat_b.id,
                tip='CO',
                data_start=date(2026, 2, 1),
                data_sfarsit=date(2026, 2, 2),
                nr_zile=2,
            ),
        ])
        db.session.commit()

        return {
            'tenant_a': tenant_a.id,
            'tenant_b': tenant_b.id,
            'proiect_a': proiect_a.id,
            'proiect_b': proiect_b.id,
            'angajat_a': angajat_a.id,
            'angajat_b': angajat_b.id,
            'masina_a': masina_a.id,
            'masina_b': masina_b.id,
            'document_b': document_b.id,
            'defect_b': defect_b.id,
        }


def _form_angajat(nume):
    return {
        'nume': nume,
        'prenume': 'Test',
        'cnp': '',
        'telefon': '',
        'email': '',
        'adresa': '',
        'data_nasterii': '',
        'functie': 'Muncitor',
        'specializari': '',
        'data_angajare': '2026-01-01',
        'data_incetare': '',
        'tip_contract': 'nedeterminat',
        'salariu_baza': '',
        'nr_contract': '',
        'serie_bi': '',
        'nr_bi': '',
        'status': 'activ',
        'observatii': '',
        'proiecte_asignate': [],
    }


def _form_masina(numar):
    return {
        'numar_inmatriculare': numar,
        'marca': 'Dacia',
        'model': 'Logan',
        'an_fabricatie': '',
        'vin': '',
        'culoare': '',
        'tip_combustibil': 'motorina',
        'capacitate_cilindrica': '',
        'putere_kw': '',
        'tip_vehicul': 'autoturism',
        'nr_locuri': '5',
        'masa_maxima': '',
        'categorie_permis': 'B',
        'km_bord': '0',
        'consum_mediu': '',
        'serie_civ': '',
        'nr_carte_identitate': '',
        'proiect_id': '',
        'angajat_responsabil_id': '',
        'status': 'disponibila',
        'observatii': '',
        'data_achizitie': '',
        'data_prima_inmatriculare': '',
        'data_itp_expirare': '',
        'data_rca_expirare': '',
        'data_casco_expirare': '',
        'data_rovinieta_expirare': '',
    }


def _seteaza_tenant_user(app, user_id, tenant_id):
    from models import Utilizator, db

    with app.app_context():
        user = db.session.get(Utilizator, user_id)
        user.tenant_id = tenant_id
        db.session.commit()


def _xlsx_texts(data):
    wb = load_workbook(BytesIO(data), read_only=True)
    texte = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if value is not None:
                    texte.add(str(value))
    return texte


def _curata_date(app):
    from models import (
        Angajat, AtribuireMasina, Concediu, DefectiuneMasina, DocumentMasina,
        Masina, Proiect, Tenant, Utilizator, db,
    )

    with app.app_context():
        app.config['MULTI_TENANT_MODE'] = 'off'

        for user in Utilizator.query.filter(Utilizator.email.in_([
            'admin_test@test.local',
            'operator_test@test.local',
        ])).all():
            user.tenant_id = None

        for cls in (DocumentMasina, AtribuireMasina, DefectiuneMasina):
            for obj in cls.query.join(Masina).filter(Masina.numar_inmatriculare.like('TAHR-R-%')).all():
                db.session.delete(obj)
        for concediu in Concediu.query.join(Angajat).filter(Angajat.nume.like('TAHR Route%')).all():
            db.session.delete(concediu)
        for masina in Masina.query.filter(Masina.numar_inmatriculare.like('TAHR-R-%')).all():
            db.session.delete(masina)
        for angajat in Angajat.query.filter(
            (Angajat.nume.like('TAHR Route%')) | (Angajat.nume.like('TAHR Nou%'))
            | (Angajat.nume.like('TAHR Fara Tenant%'))
        ).all():
            db.session.delete(angajat)
        for proiect in Proiect.query.filter(Proiect.cod_proiect.like('TAHR-R-%')).all():
            db.session.delete(proiect)
        for tenant in Tenant.query.filter(Tenant.cod.like('test-tahr-route-%')).all():
            db.session.delete(tenant)

        db.session.commit()
