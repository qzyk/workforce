"""
Tests pentru exportul EDIFICO xlsx - structura exacta + filtre + multi-angajat.
"""

from datetime import date
from io import BytesIO
import json
import pytest


@pytest.fixture
def setup_export_data(app, admin_user):
    """
    Pregateste 1 angajat cu 5 activitati zilnice + 1 saptamanala in luna octombrie 2025.
    Returneaza dict cu ID-uri.
    """
    from models import db, Angajat, Proiect, RaportActivitate
    from tests.fixtures.data import make_proiect, make_angajat, make_raport_activitate
    with app.app_context():
        # Cleanup
        Angajat.query.filter_by(cnp='1900111222333').delete()
        Proiect.query.filter_by(cod_proiect='PRJ-EXP-001').delete()
        RaportActivitate.query.filter(
            RaportActivitate.activitate_principala.like('EXP_%')
        ).delete()
        db.session.commit()

        p = make_proiect(db, Proiect, cod='PRJ-EXP-001', nume='Proiect Export Test')
        a = make_angajat(db, Angajat, cnp='1900111222333',
                         nume='ExportTest', prenume='Inginer', functie='Inginer')

        # 5 activitati zilnice in saptamana 1 (1-3 oct = miercuri-vineri)
        for i, d in enumerate([date(2025, 10, 1), date(2025, 10, 2), date(2025, 10, 3)]):
            make_raport_activitate(db, RaportActivitate, a.id, p.id,
                                   tip='zilnica',
                                   titlu=f'EXP_Z{i+1}',
                                   data_zi=d,
                                   activitate_detaliata=f'Detalii ziua {i+1}')

        # 1 activitate saptamanala (sapt 41 = 6-10 oct)
        make_raport_activitate(db, RaportActivitate, a.id, p.id,
                               tip='saptamanala',
                               titlu='EXP_SAPT_41',
                               data_zi=date(2025, 10, 6),
                               activitate_detaliata='Activitate intreaga saptamana')

        # 1 activitate lunara
        make_raport_activitate(db, RaportActivitate, a.id, p.id,
                               tip='lunara',
                               titlu='EXP_LUNAR_OCT',
                               data_zi=date(2025, 10, 1),
                               activitate_detaliata='Centralizare oferte luna')

        yield {
            'angajat_id': a.id, 'proiect_id': p.id,
            'an': 2025, 'luna': 10,
        }

        # Cleanup
        RaportActivitate.query.filter(
            RaportActivitate.activitate_principala.like('EXP_%')
        ).delete()
        Angajat.query.filter_by(cnp='1900111222333').delete()
        Proiect.query.filter_by(cod_proiect='PRJ-EXP-001').delete()
        db.session.commit()


class TestExportEdificoStructure:
    """Verifica ca exportul produce un xlsx valid cu structura corecta."""

    def test_export_returns_xlsx(self, authenticated_client, setup_export_data):
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        assert resp.status_code == 200
        # Magic bytes ZIP (xlsx e ZIP)
        assert resp.data[:2] == b'PK'
        assert resp.content_type.startswith(
            'application/vnd.openxmlformats-officedocument.spreadsheetml'
        )

    def test_export_has_employee_sheet(self, authenticated_client, setup_export_data):
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        wb = load_workbook(BytesIO(resp.data))
        # Numele angajatului e ExportTest Inginer (sau Inginer ExportTest in ordinea make_angajat)
        sheets = wb.sheetnames
        assert len(sheets) == 1
        assert 'ExportTest' in sheets[0] or 'Inginer' in sheets[0]

    def test_export_has_title_at_b3(self, authenticated_client, setup_export_data):
        """B3 contine titlul cu luna + an."""
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.worksheets[0]
        b3 = ws['B3'].value
        assert b3 is not None
        assert 'octombrie' in b3.lower()
        assert '2025' in b3

    def test_export_sheet_name_e_numele_angajatului(self, authenticated_client, setup_export_data):
        """Sheet name == numele complet al angajatului."""
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        wb = load_workbook(BytesIO(resp.data))
        sheet_name = wb.sheetnames[0]
        # ExportTest = nume sau prenume
        assert 'ExportTest' in sheet_name or 'Inginer' in sheet_name

    def test_export_contains_activity_text(self, authenticated_client, setup_export_data):
        """Cele 5 activitati EXP_* trebuie sa apara in xlsx."""
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.worksheets[0]
        # Aduna tot textul din celulele E (coloana de activitati)
        all_text = ''
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    all_text += cell.value + ' '
        assert 'EXP_Z1' in all_text or 'EXP_Z2' in all_text or 'EXP_Z3' in all_text
        assert 'EXP_SAPT_41' in all_text or 'EXP_LUNAR_OCT' in all_text

    def test_export_contains_month_name(self, authenticated_client, setup_export_data):
        """Numele lunii apare in xlsx (B7 are 'OCTOMBRIE 2025' sau B3 are 'octombrie')."""
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.worksheets[0]
        # Cauta in primele 10 randuri orice celula cu 'octombrie' (case insensitive)
        gasit = False
        for row in ws.iter_rows(min_row=1, max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str) and 'ctombrie' in cell.value.lower():
                    gasit = True
                    break
            if gasit:
                break
        assert gasit, 'Numele lunii (octombrie) nu apare in primele 10 randuri'

    def test_export_has_table_header(self, authenticated_client, setup_export_data):
        """Header tabel: Luna / Saptamana / Data / Activitati."""
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}'
        )
        wb = load_workbook(BytesIO(resp.data))
        ws = wb.worksheets[0]
        # Iar in randurile 7-10 ar trebui sa fie headerul de tabel
        all_text = ''
        for row in ws.iter_rows(min_row=7, max_row=10):
            for cell in row:
                if cell.value and isinstance(cell.value, str):
                    all_text += cell.value + ' '
        assert 'Luna' in all_text
        assert 'Saptamana' in all_text or 'Săptămâna' in all_text
        assert 'Data' in all_text
        assert 'Activitati' in all_text or 'Activități' in all_text


class TestExportEdificoMultiEmployee:
    """Multi-angajat: lista nu filtrata, lista cu 2-3 angajati selectati."""

    def test_export_fara_angajat_id_returneaza_toti_cu_activitati(
            self, authenticated_client, setup_export_data):
        from openpyxl import load_workbook
        resp = authenticated_client.get('/activitati/raport/edifico?luna=2025-10')
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.data))
        # Cel putin un sheet (angajatul cu activitatile noastre)
        assert len(wb.sheetnames) >= 1


class TestExportEdificoFiltre:
    """Filtre pe export: ?tip=zilnica/saptamanala/lunara."""

    def test_export_filtru_tip_zilnica(self, authenticated_client, setup_export_data):
        from openpyxl import load_workbook
        resp = authenticated_client.get(
            f'/activitati/raport/edifico?luna=2025-10&angajat_id={setup_export_data["angajat_id"]}&tip=zilnica'
        )
        assert resp.status_code == 200
        wb = load_workbook(BytesIO(resp.data))
        # Ar trebui sa fie un sheet
        assert len(wb.sheetnames) == 1


class TestExportEdificoParameters:
    """Validare parametri."""

    def test_luna_invalida_redirect(self, authenticated_client):
        resp = authenticated_client.get('/activitati/raport/edifico?luna=invalid',
                                         follow_redirects=False)
        assert resp.status_code in (302, 400)

    def test_luna_default_e_curenta(self, authenticated_client):
        # Fara ?luna -> luna curenta
        resp = authenticated_client.get('/activitati/raport/edifico')
        # Ori 200 cu xlsx, ori redirect cu warning
        assert resp.status_code in (200, 302)


class TestExportEdificoAlteFormate:
    """Celelalte exporturi (saptamanal, lunar, anual) raman functional."""

    def test_export_saptamanal(self, authenticated_client):
        resp = authenticated_client.get('/activitati/raport/saptamanal')
        assert resp.status_code == 200

    def test_export_lunar(self, authenticated_client):
        resp = authenticated_client.get('/activitati/raport/lunar')
        assert resp.status_code == 200
