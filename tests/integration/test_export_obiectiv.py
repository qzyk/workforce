"""Teste pentru exportul centralizator obiectiv (xlsx + pdf)."""

from decimal import Decimal
from io import BytesIO

import pytest

from models import db, Obiectiv, Obiect, GanttPlan
from services.ingestie_obiectiv import construieste_arbore
from services.export_obiectiv import export_xlsx, export_pdf


DATE = {
    'nume': 'Obiectiv Export Test',
    'valoare_constructii': Decimal('100.00'),
    'valoare_totala': Decimal('120.00'),
    'obiecte': [
        {'cod': '001', 'nume': 'Arhitectura', 'disciplina': 'arhitectura',
         'valoare_f2': Decimal('70.00'), 'valoare_f1': Decimal('70.00'), 'planuri': [
             {'nume': 'Lucrari noi', 'nume_fisier': '001_002_F3.xls',
              'continut': b'x', 'cost_total': Decimal('60.00')},
         ]},
        {'cod': '002', 'nume': 'Structura', 'disciplina': 'structural',
         'valoare_f2': Decimal('30.00'), 'valoare_f1': Decimal('30.00'), 'planuri': [
             {'nume': 'Corp A', 'nume_fisier': '002_001_F3.xls',
              'continut': b'z', 'cost_total': Decimal('30.00')},
         ]},
    ],
}


@pytest.fixture(autouse=True)
def _curata(app):
    with app.app_context():
        GanttPlan.query.filter(GanttPlan.obiect_id.isnot(None)).delete()
        Obiect.query.delete()
        Obiectiv.query.delete()
        db.session.commit()
    yield


def test_export_xlsx_structura(app):
    import openpyxl
    with app.app_context():
        construieste_arbore(DATE)
        ob = Obiectiv.query.first()
        data = export_xlsx(ob.id)
        assert isinstance(data, (bytes, bytearray)) and len(data) > 0
        wb = openpyxl.load_workbook(BytesIO(data))
        assert 'Centralizator obiectiv' in wb.sheetnames
        assert 'Liste F3' in wb.sheetnames
        # numele obiectivului prezent
        ws = wb['Centralizator obiectiv']
        assert ws['A2'].value == 'Obiectiv Export Test'
        # listele F3: 2 randuri de date
        ws2 = wb['Liste F3']
        valori = [ws2.cell(row=r, column=4).value for r in (2, 3)]
        assert 60.0 in valori and 30.0 in valori


def test_export_pdf_bytes(app):
    with app.app_context():
        construieste_arbore(DATE)
        ob = Obiectiv.query.first()
        data = export_pdf(ob.id)
        assert isinstance(data, (bytes, bytearray))
        assert data[:4] == b'%PDF'
