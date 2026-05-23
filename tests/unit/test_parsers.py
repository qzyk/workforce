"""
Teste unit pentru cele 3 parsere Faza 11:
  - MSProjectXMLParser
  - EDevizeXMLParser
  - ExcelBoQParser

Foloseste fixturile din tests/fixtures/imports/.
ExcelBoQParser primeste fisierul XLSX construit programatic in test
(evitam binari committed in repo).
"""

import os
from decimal import Decimal
from pathlib import Path

import openpyxl
import pytest

from services.parsers import (
    MSProjectXMLParser, MSProjectMPPParser, EDevizeXMLParser, ExcelBoQParser,
    ParseError, ParseResult,
)


FIXTURES_DIR = Path(__file__).parent.parent / 'fixtures' / 'imports'


# ============================================================
# MSProjectXMLParser
# ============================================================

class TestMSProjectXMLParser:
    def _parse(self):
        return MSProjectXMLParser().parse(str(FIXTURES_DIR / 'msproject_minimal.xml'))

    def test_parses_without_errors(self):
        r = self._parse()
        assert not r.has_errors, f'Errors: {r.errors}'
        assert r.sursa == 'msproject_xml'

    def test_skips_root_task_uid_0(self):
        r = self._parse()
        codes = [e['cod_extern'] for e in r.entities]
        assert '0' not in codes, 'Task root UID=0 nu trebuie inclus.'

    def test_correct_count_and_metadata(self):
        r = self._parse()
        # Fixture: 5 taskuri (UID 1-5)
        assert len(r.entities) == 5
        assert r.stats['project_name'] == 'Proiect Test Edifico'

    def test_summary_milestone_task_types(self):
        r = self._parse()
        by_uid = {e['cod_extern']: e for e in r.entities}
        assert by_uid['1']['tip_task'] == 'summary'   # Faza proiectare
        assert by_uid['2']['tip_task'] == 'task'      # Releveu
        assert by_uid['4']['tip_task'] == 'summary'   # Faza executie
        assert by_uid['5']['tip_task'] == 'milestone' # Receptie finala

    def test_hierarchy_outline_level(self):
        r = self._parse()
        by_uid = {e['cod_extern']: e for e in r.entities}
        assert by_uid['1']['nivel_ierarhie'] == 1
        assert by_uid['2']['nivel_ierarhie'] == 2
        assert by_uid['3']['nivel_ierarhie'] == 2

    def test_predecessors_extracted(self):
        r = self._parse()
        by_uid = {e['cod_extern']: e for e in r.entities}
        # Task UID 3 (Proiect tehnic) are 1 predecesor: UID 2
        assert len(by_uid['3']['predecesori']) == 1
        pred = by_uid['3']['predecesori'][0]
        assert pred['uid_extern'] == '2'
        assert pred['tip'] == 'FS'  # Type=1 in XML

    def test_duration_in_days(self):
        r = self._parse()
        by_uid = {e['cod_extern']: e for e in r.entities}
        # Task UID 2: Duration=PT80H = 80/8 = 10 zile
        assert by_uid['2']['durata_zile'] == 10
        # Milestone (PT0H) = 0 zile
        assert by_uid['5']['durata_zile'] == 0

    def test_percent_complete_decimal(self):
        r = self._parse()
        by_uid = {e['cod_extern']: e for e in r.entities}
        assert by_uid['1']['procent_realizare'] == Decimal('100')
        assert by_uid['4']['procent_realizare'] == Decimal('40')

    def test_invalid_xml_raises_parse_error(self, tmp_path):
        bad = tmp_path / 'bad.xml'
        bad.write_text('<not xml')
        with pytest.raises(ParseError):
            MSProjectXMLParser().parse(str(bad))

    def test_missing_tasks_element(self, tmp_path):
        empty = tmp_path / 'empty.xml'
        empty.write_text('<?xml version="1.0"?><Project><Name>X</Name></Project>')
        r = MSProjectXMLParser().parse(str(empty))
        assert r.has_errors


# ============================================================
# EDevizeXMLParser
# ============================================================

class TestEDevizeXMLParser:
    def _parse(self):
        return EDevizeXMLParser().parse(str(FIXTURES_DIR / 'edevize_minimal.xml'))

    def test_parses_without_errors(self):
        r = self._parse()
        assert not r.has_errors, f'Errors: {r.errors}'
        assert r.sursa == 'edevize_xml'

    def test_count_and_metadata(self):
        r = self._parse()
        # Fixture: 5 articole (2 in CA01, 3 in CB01)
        assert len(r.entities) == 5
        assert r.stats['numar'] == 'OF-2025-0001'
        assert r.stats['moneda'] == 'RON'

    def test_cod_capitol_set_from_parent(self):
        r = self._parse()
        by_cod = {e['cod_articol']: e for e in r.entities}
        assert by_cod['CA01A1']['cod_capitol'] == 'CA01'
        assert by_cod['CB01A1']['cod_capitol'] == 'CB01'

    def test_categorii_normalized(self):
        r = self._parse()
        by_cod = {e['cod_articol']: e for e in r.entities}
        assert by_cod['CA01A1']['categorie'] == 'manopera'
        assert by_cod['CA01A2']['categorie'] == 'utilaje'
        assert by_cod['CB01A1']['categorie'] == 'materiale'

    def test_decimal_with_comma_handled(self):
        """pret_unitar='650,00' (virgula RO) -> Decimal('650.00')"""
        r = self._parse()
        by_cod = {e['cod_articol']: e for e in r.entities}
        assert by_cod['CB01A1']['pret_unitar'] == Decimal('650.00')

    def test_detalii_valori_unitar(self):
        r = self._parse()
        by_cod = {e['cod_articol']: e for e in r.entities}
        # CA01A1 are <Detalii> in XML
        assert by_cod['CA01A1']['valoare_manopera_unitar'] == Decimal('30.00')
        assert by_cod['CA01A1']['valoare_materiale_unitar'] == Decimal('10.00')
        # CA01A2 nu are <Detalii> -> None
        assert by_cod['CA01A2']['valoare_manopera_unitar'] is None

    def test_invalid_xml_raises(self, tmp_path):
        bad = tmp_path / 'bad.xml'
        bad.write_text('<broken')
        with pytest.raises(ParseError):
            EDevizeXMLParser().parse(str(bad))

    def test_missing_root_returns_error(self, tmp_path):
        bad = tmp_path / 'wrong_root.xml'
        bad.write_text('<?xml version="1.0"?><Altceva><X/></Altceva>')
        r = EDevizeXMLParser().parse(str(bad))
        assert r.has_errors

    def test_empty_no_articole(self, tmp_path):
        empty = tmp_path / 'empty.xml'
        empty.write_text(
            '<?xml version="1.0"?><Oferta>'
            '<Antet><Numar>X</Numar></Antet><Capitole></Capitole></Oferta>'
        )
        r = EDevizeXMLParser().parse(str(empty))
        assert r.has_errors  # no entities -> error


# ============================================================
# ExcelBoQParser
# ============================================================

class TestExcelBoQParser:
    @pytest.fixture
    def boq_xlsx(self, tmp_path):
        """Construieste un BoQ XLSX de test cu 4 rânduri valide + 1 gol."""
        wb = openpyxl.Workbook()
        ws = wb.active
        # Header rand 1
        ws.append(['cod_articol', 'cod_capitol', 'denumire', 'um',
                   'cantitate', 'pret_unitar', 'categorie'])
        # Date
        ws.append(['B01-001', 'B01', 'Beton C25/30', 'mc', 50.0, 650.00, 'materiale'])
        ws.append(['B01-002', 'B01', 'Cofraj metalic', 'mp', 280.0, 55.00, 'manopera'])
        # Rand gol (skip silent)
        ws.append([None, None, None, None, None, None, None])
        ws.append(['B02-001', 'B02', 'Armatura PC52', 'kg', 3200.0, 4.80, 'materiale'])
        # Rand cu UM lipsa -> warning + skip
        ws.append(['B02-002', 'B02', 'Articol fara UM', None, 100, 10, 'mixt'])
        # Rand cu categorie invalida -> warning + default mixt
        ws.append(['B03-001', 'B03', 'Articol cu cat invalida', 'buc', 5, 25, 'altceva'])

        path = tmp_path / 'boq.xlsx'
        wb.save(path)
        wb.close()
        return path

    def test_parses_without_errors(self, boq_xlsx):
        r = ExcelBoQParser().parse(str(boq_xlsx))
        assert not r.has_errors, f'Errors: {r.errors}'
        assert r.sursa == 'excel_xlsx'

    def test_correct_entity_count(self, boq_xlsx):
        r = ExcelBoQParser().parse(str(boq_xlsx))
        # Parser-ul robust (cu auto-detect) e mai TOLERANT: include si
        # B02-002 cu UM lipsa (default 'buc' + warning), nu il pierde.
        # 5 articole: B01-001, B01-002, B02-001, B02-002, B03-001.
        # Randul gol -> skip silent.
        assert len(r.entities) == 5

    def test_warnings_for_invalid_rows(self, boq_xlsx):
        r = ExcelBoQParser().parse(str(boq_xlsx))
        # >=2 warnings: UM lipsa (B02-002) + categorie necunoscuta (B03-001)
        assert len(r.warnings) >= 2

    def test_invalid_category_falls_back_to_mixt(self, boq_xlsx):
        r = ExcelBoQParser().parse(str(boq_xlsx))
        by_cod = {e['cod_articol']: e for e in r.entities}
        assert by_cod['B03-001']['categorie'] == 'mixt'

    def test_decimals_correct(self, boq_xlsx):
        r = ExcelBoQParser().parse(str(boq_xlsx))
        by_cod = {e['cod_articol']: e for e in r.entities}
        assert by_cod['B01-001']['cantitate_oferta'] == Decimal('50.0')
        assert by_cod['B01-001']['pret_unitar'] == Decimal('650.00')

    def test_ordine_assigned(self, boq_xlsx):
        r = ExcelBoQParser().parse(str(boq_xlsx))
        # Ordinea incrementala (5 articole acum)
        ordini = [e['ordine'] for e in r.entities]
        assert ordini == [1, 2, 3, 4, 5]

    def test_corrupt_xlsx_raises(self, tmp_path):
        bad = tmp_path / 'bad.xlsx'
        bad.write_bytes(b'not a real xlsx file')
        with pytest.raises(ParseError):
            ExcelBoQParser().parse(str(bad))

    def test_auto_detect_header_not_on_row_1(self, tmp_path):
        """Header pe rand 4 (cu antet deasupra) -> auto-detectie."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['PROIECT TEST'])           # R1 antet
        ws.append([])                          # R2 gol
        ws.append(['Lista cantitati'])         # R3 subtitlu
        ws.append(['Nr.', 'DENUMIRE', 'U.M.', 'CANTITATE', 'PRET UNITAR'])  # R4 HEADER
        ws.append(['1', 'Grupa A'])            # R5 grup (fara UM)
        ws.append(['1.1', 'Beton C25/30', 'mc', '50', '650'])
        ws.append(['1.2', 'Armatura', 'kg', '1200', '6'])
        path = tmp_path / 'header_row4.xlsx'
        wb.save(path)

        r = ExcelBoQParser().parse(str(path))
        assert not r.has_errors, f'Errors: {r.errors}'
        assert len(r.entities) == 2
        by_cod = {e['cod_articol']: e for e in r.entities}
        assert by_cod['1.1']['denumire'] == 'Beton C25/30'
        assert by_cod['1.1']['cod_capitol'] == 'Grupa A'  # grupul mostenit

    def test_multi_sheet_skips_empty(self, tmp_path):
        """Multi-sheet: primul sheet gol (title), al doilea cu date."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = 'TITLE PAGE'
        ws1.append(['Doar antet, fara tabel'])
        ws2 = wb.create_sheet('Obiect 1')
        ws2.append(['Nr.', 'DENUMIRE', 'U.M.', 'CANTITATE'])
        ws2.append(['1', 'Sapatura', 'mc', '100'])
        ws2.append(['2', 'Beton', 'mc', '50'])
        path = tmp_path / 'multi.xlsx'
        wb.save(path)

        r = ExcelBoQParser().parse(str(path))
        assert len(r.entities) == 2
        # Un singur sheet procesat (TITLE PAGE skip)
        assert len(r.stats['sheets_procesate']) == 1
        assert r.stats['sheets_procesate'][0]['sheet'] == 'Obiect 1'

    def test_skip_note_rows(self, tmp_path):
        """Randurile NOTE/NOTES nu sunt articole."""
        import openpyxl
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['Nr.', 'DENUMIRE', 'U.M.', 'CANTITATE'])
        ws.append(['1', 'Beton', 'mc', '50'])
        ws.append(['NOTE /NOTES', None, None, None])  # nota -> skip
        ws.append(['TOTAL', None, None, None])         # total -> skip
        path = tmp_path / 'note.xlsx'
        wb.save(path)

        r = ExcelBoQParser().parse(str(path))
        assert len(r.entities) == 1
        assert r.entities[0]['denumire'] == 'Beton'


REAL_XLS_PATH = os.path.expanduser(
    '~/Downloads/PT DE Hala Campina/02.Rezistenta/Parti scrise/Editabil/'
    '2404_AEN_PTh+DE_STR_PS_09_00-Liste cantitati.xls'
)


@pytest.mark.skipif(not os.path.exists(REAL_XLS_PATH),
                    reason='Fisier .xls real Hala Campina absent - test optional')
class TestExcelBoQParserRealXLS:
    """Test pe deviz .xls real (Hala Campina) - confirma parser pe format binar."""

    def test_parses_real_xls(self):
        r = ExcelBoQParser().parse(REAL_XLS_PATH)
        assert not r.has_errors, f'Errors: {r.errors}'
        # 6 sheet-uri cu date (al 7-lea TITLE PAGE skip); ~76 articole reale
        # (randurile NOTE/TOTAL excluse corect)
        assert len(r.entities) >= 70

    def test_real_xls_multi_sheet(self):
        r = ExcelBoQParser().parse(REAL_XLS_PATH)
        # Minim 5 sheet-uri obiect procesate
        assert len(r.stats['sheets_procesate']) >= 5

    def test_real_xls_capitole_din_grupe(self):
        r = ExcelBoQParser().parse(REAL_XLS_PATH)
        capitole = {e['cod_capitol'] for e in r.entities if e['cod_capitol']}
        # Grupele Infrastructura / Suprastructura detectate ca cod_capitol
        assert any('frastructura' in (c or '').lower() for c in capitole)

    def test_real_xls_material_in_denumire(self):
        r = ExcelBoQParser().parse(REAL_XLS_PATH)
        # Materialul (C25/30, S500C) prefixat in denumire
        cu_material = [e for e in r.entities
                       if '(' in e['denumire'] and ')' in e['denumire']]
        assert len(cu_material) > 5


# ============================================================
# MSProjectMPPParser (.mpp binar via MPXJ/JVM)
# ============================================================

REAL_MPP_PATH = os.path.expanduser('~/Downloads/GRAFIC TURDA-.mpp')


def _mpp_toolchain_ok() -> bool:
    """True daca jpype + jar-uri MPXJ + un JVM sunt disponibile local."""
    try:
        import jpype  # noqa: F401
    except ImportError:
        return False
    from services.parsers.msproject_mpp_parser import (
        _mpxj_jars, _resolve_jvm_path,
    )
    if not _mpxj_jars():
        return False
    lib, _ = _resolve_jvm_path()
    return bool(lib)


class TestMSProjectMPPParserUnit:
    """Teste fara JVM: importabilitate + erori grafioase (nu pornesc JVM)."""

    def test_importable_and_sursa_cod(self):
        # Modulul trebuie importabil chiar daca jpype/JDK lipsesc (import lazy).
        assert MSProjectMPPParser.SURSA_COD == 'msproject_mpp'

    def test_missing_file_raises_parse_error(self):
        # Verificarea de existenta ruleaza INAINTE de pornirea JVM.
        with pytest.raises(ParseError):
            MSProjectMPPParser().parse('/nu/exista/plan.mpp')


@pytest.mark.skipif(
    not (os.path.exists(REAL_MPP_PATH) and _mpp_toolchain_ok()),
    reason='Fisier .mpp real sau toolchain MPXJ/JVM absent - test optional',
)
class TestMSProjectMPPParserReal:
    """Test pe .mpp real (GRAFIC TURDA) via MPXJ -> MSPDI -> XML parser."""

    def test_parses_real_mpp(self):
        r = MSProjectMPPParser().parse(REAL_MPP_PATH)
        assert not r.has_errors, f'Errors: {r.errors}'
        assert r.sursa == 'msproject_mpp'
        # GRAFIC TURDA: ~208 taskuri
        assert len(r.entities) > 100

    def test_via_chain_metadata(self):
        r = MSProjectMPPParser().parse(REAL_MPP_PATH)
        assert r.stats.get('format_sursa') == 'mpp'
        assert 'mspdi' in r.stats.get('via', '')
        # Namespace MSPDI standard (reutilizat de XML parser)
        assert r.stats.get('namespace') == 'http://schemas.microsoft.com/project'

    def test_entities_have_required_fields(self):
        r = MSProjectMPPParser().parse(REAL_MPP_PATH)
        e = r.entities[0]
        for k in ('cod_extern', 'denumire', 'nivel_ierarhie',
                  'data_start_planificat', 'data_sfarsit_planificat',
                  'tip_task', 'procent_realizare'):
            assert k in e, f'Camp lipsa: {k}'

    def test_project_name_artifact_removed(self):
        """MSPDI writer pune 'project.xml' ca Name; trebuie suprascris."""
        r = MSProjectMPPParser().parse(REAL_MPP_PATH)
        assert r.stats.get('project_name') != 'project.xml'
