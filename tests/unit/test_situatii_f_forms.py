"""
Teste pentru formularele F1/F2/F3 (HG 907/2016) pe situatii de lucrari - dz-4.

Verifica VALORI (nu doar existenta), cu reconciliere matematica:
  - F3 (lista cantitati): articole detaliate grupate pe obiect -> categorie,
    Sigma valoare articole == valoare luna; valoare articol == cant_luna * pret.
  - F2 (centralizator categorii): subtotaluri pe obiect; Sigma subtotaluri ==
    total general == valoare luna.
  - F1 (centralizator obiectiv): un rand per obiect; TVA 21% calculat o singura
    data la final; total_fara_tva == valoare luna; total_cu_tva == total + TVA.
  - Doar lucrarea EXECUTATA in luna intra in formulare (pozitiile fara cant_luna
    sunt excluse), dar raman in oferta (nu se sterg).
  - Excel-ul salvat pe disc se reciteste si totalul lui == totalul calculat
    (freeze_panes sub randul de antet real, fara TVA pe pozitii).
  - Gating: rutele de export sunt gated pe flag 'situatii-f-forms' (404 cu OFF).

Setup: pret 1 RON/unitate -> cantitate == valoare (mai usor de verificat).
Doua obiecte (disciplina): structural (cod_capitol '1 REZISTENTA') si
electrice (cod_capitol '5 INSTALATII ELECTRICE').
"""
from datetime import date
from decimal import Decimal

import pytest

from services.feature_flags import set_flag


def _curata(db, Proiect, Contract, OfertaContract, PozitieBoQ,
            CantitateExecutataLunara):
    """Curata seed-ul FF-* (copii inainte de parinti). Idempotent."""
    from models import SituatieLunara
    pr = Proiect.query.filter_by(cod_proiect='FF-PRJ').first()
    if pr is not None:
        SituatieLunara.query.filter_by(proiect_id=pr.id).delete()
    CantitateExecutataLunara.query.filter(
        CantitateExecutataLunara.pozitie_boq_id.in_(
            db.session.query(PozitieBoQ.id).filter(
                PozitieBoQ.cod_articol.like('FF-%')))).delete(
        synchronize_session=False)
    PozitieBoQ.query.filter(PozitieBoQ.cod_articol.like('FF-%')).delete(
        synchronize_session=False)
    OfertaContract.query.filter(
        OfertaContract.contract_id.in_(
            db.session.query(Contract.id).filter(
                Contract.nr_contract.like('FF-%')))).delete(
        synchronize_session=False)
    Contract.query.filter(Contract.nr_contract.like('FF-%')).delete(
        synchronize_session=False)
    Proiect.query.filter(Proiect.cod_proiect == 'FF-PRJ').delete()
    db.session.commit()


def _seed(db, Proiect, Contract, OfertaContract, PozitieBoQ,
          CantitateExecutataLunara, uid):
    """
    Construieste un contract cu 4 pozitii pe 2 obiecte/categorii + o pozitie
    FARA activitate lunara (trebuie exclusa din formulare).

    Cantitati validate luna 3/2026 (pret 1 RON/unit -> valoare == cantitate):
      structural / beton    : 10000
      structural / armatura :  4000
      electrice  / cabluri  :  3000
      electrice  / tablouri :  1000
      structural / cofraje  :     0  (fara cant_luna -> exclus din F1/F2/F3)
    Total executat luna = 18000.
    """
    _curata(db, Proiect, Contract, OfertaContract, PozitieBoQ,
            CantitateExecutataLunara)
    p = Proiect(cod_proiect='FF-PRJ', nume='Formulare F test',
                data_start=date(2026, 1, 1), status='activ')
    db.session.add(p); db.session.flush()
    c = Contract(proiect_id=p.id, nr_contract='FF-CTR-001',
                 data_semnare=date(2026, 1, 15), status='activ',
                 valoare_totala=Decimal('500000'), moneda='RON')
    db.session.add(c); db.session.flush()
    o = OfertaContract(contract_id=c.id, proiect_id=p.id, versiune=1,
                       data_emitere=date(2026, 1, 20),
                       valoare_totala=Decimal('500000'),
                       sursa_import='manual', aprobata=True)
    db.session.add(o); db.session.flush()

    # (cod_articol, cod_capitol, categorie_lucrare, cant_luna)
    specs = [
        ('FF-001', '1 REZISTENTA', 'beton', 10000),
        ('FF-002', '1 REZISTENTA', 'armatura', 4000),
        ('FF-003', '5 INSTALATII ELECTRICE', 'cabluri', 3000),
        ('FF-004', '5 INSTALATII ELECTRICE', 'tablouri', 1000),
        ('FF-005', '1 REZISTENTA', 'cofraje', 0),  # fara activitate luna
    ]
    ordine = 1
    for cod, cap, catl, cant_luna in specs:
        pz = PozitieBoQ(oferta_id=o.id, proiect_id=p.id,
                        cod_articol=cod, cod_capitol=cap,
                        denumire=f'Lucrare {catl}', um='mc',
                        cantitate_oferta=Decimal('50000'),
                        pret_unitar=Decimal('1'), categorie='mixt',
                        categorie_lucrare=catl, ordine=ordine)
        db.session.add(pz); db.session.flush()
        ordine += 1
        if cant_luna > 0:
            cant = CantitateExecutataLunara(
                pozitie_boq_id=pz.id, proiect_id=p.id, an=2026, luna=3,
                cantitate_executata=Decimal(str(cant_luna)),
                valoare_calculata=Decimal(str(cant_luna)),
                validat=True, validat_de_id=uid)
            db.session.add(cant)
    db.session.commit()
    return c.id


def _genereaza(app, db, models, uid):
    from services.situatii import genereaza_situatie
    (Proiect, Contract, OfertaContract, PozitieBoQ,
     CantitateExecutataLunara) = models
    cid = _seed(db, Proiect, Contract, OfertaContract, PozitieBoQ,
                CantitateExecutataLunara, uid)
    s = genereaza_situatie(cid, 2026, 3, uid)
    return s


@pytest.fixture
def _models():
    from models import (
        Proiect, Contract, OfertaContract, PozitieBoQ,
        CantitateExecutataLunara,
    )
    return (Proiect, Contract, OfertaContract, PozitieBoQ,
            CantitateExecutataLunara)


def _uid():
    from models import Utilizator
    return Utilizator.query.filter_by(email='admin_test@test.local').first().id


def test_f3_reconciliaza_cu_valoarea_lunii(app, admin_user, _models):
    """F3: Sigma articole == total general == valoare luna (18000); pe articol
    valoare == cant_luna * pret_unitar; pozitia fara activitate e exclusa."""
    from models import db
    from services.situatii import genereaza_f3
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        f3 = genereaza_f3(s.id)

        assert f3['total_general'] == Decimal('18000.00')
        assert s.valoare_totala_luna == Decimal('18000.00')

        # Doua obiecte: structural + electrice
        disc = {g['disciplina'] for g in f3['grupe']}
        assert disc == {'structural', 'electrice'}

        # Suma articolelor pe toate categoriile == total general
        suma_art = Decimal('0')
        n_art = 0
        coduri = set()
        for g in f3['grupe']:
            sub = Decimal('0')
            for cat in g['categorii']:
                for a in cat['articole']:
                    # valoare articol == cant_luna * pret_unitar
                    assert a['valoare'] == (a['cant_luna'] * a['pret_unitar']).quantize(Decimal('0.01'))
                    suma_art += a['valoare']
                    sub += a['valoare']
                    n_art += 1
                    coduri.add(a['cod_articol'])
            # subtotal obiect == suma categoriilor lui
            assert g['subtotal'] == sub
        assert suma_art == Decimal('18000.00')
        # 4 articole cu activitate; FF-005 (cofraje, 0) exclus
        assert n_art == 4
        assert 'FF-005' not in coduri


def test_f3_subtotal_obiect_structural(app, admin_user, _models):
    """Subtotal obiect structural = beton 10000 + armatura 4000 = 14000;
    electrice = cabluri 3000 + tablouri 1000 = 4000."""
    from models import db
    from services.situatii import genereaza_f3
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        f3 = genereaza_f3(s.id)
        subt = {g['disciplina']: g['subtotal'] for g in f3['grupe']}
        assert subt['structural'] == Decimal('14000.00')
        assert subt['electrice'] == Decimal('4000.00')


def test_f2_centralizator_categorii(app, admin_user, _models):
    """F2: subtotaluri pe obiect + categorie; total general == 18000."""
    from models import db
    from services.situatii import genereaza_f2
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        f2 = genereaza_f2(s.id)
        assert f2['total_general'] == Decimal('18000.00')

        # Map (obiect, categorie) -> valoare
        vals = {}
        for g in f2['grupe']:
            for cat in g['categorii']:
                vals[(g['disciplina'], cat['categorie'])] = cat['valoare']
        assert vals[('structural', 'beton')] == Decimal('10000.00')
        assert vals[('structural', 'armatura')] == Decimal('4000.00')
        assert vals[('electrice', 'cabluri')] == Decimal('3000.00')
        assert vals[('electrice', 'tablouri')] == Decimal('1000.00')
        # F2 nu are articole detaliate
        assert all('articole' not in cat for g in f2['grupe'] for cat in g['categorii'])


def test_f1_tva_la_final(app, admin_user, _models):
    """F1: un rand per obiect; total fara TVA == 18000; TVA 21% == 3780;
    total cu TVA == 21780. TVA aplicat o singura data la final."""
    from models import db
    from services.situatii import genereaza_f1
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        f1 = genereaza_f1(s.id)

        # Un rand per obiect (2)
        assert len(f1['randuri']) == 2
        rv = {r['obiect']: r['valoare'] for r in f1['randuri']}
        assert rv['structural'] == Decimal('14000.00')
        assert rv['electrice'] == Decimal('4000.00')

        assert f1['total_fara_tva'] == Decimal('18000.00')
        assert f1['cota_tva'] == Decimal('21')
        assert f1['tva'] == Decimal('3780.00')  # 18000 * 21% = 3780
        assert f1['total_cu_tva'] == Decimal('21780.00')
        # Reconciliere: total cu TVA = total fara TVA + TVA
        assert f1['total_cu_tva'] == f1['total_fara_tva'] + f1['tva']


def test_f1_f2_f3_acelasi_total_fara_tva(app, admin_user, _models):
    """Cele trei formulare au acelasi total fara TVA (consistenta intre niveluri)."""
    from models import db
    from services.situatii import genereaza_f1, genereaza_f2, genereaza_f3
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        assert genereaza_f3(s.id)['total_general'] == Decimal('18000.00')
        assert genereaza_f2(s.id)['total_general'] == Decimal('18000.00')
        assert genereaza_f1(s.id)['total_fara_tva'] == Decimal('18000.00')


def _citeste_total_xlsx(path):
    """Reciteste un Excel F si intoarce valorile numerice din coloana totalurilor."""
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    vals = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, (int, float)):
                vals.append(Decimal(str(c.value)))
    return ws, vals


def test_export_f3_xlsx_se_reciteste_cu_total_corect(app, admin_user, _models):
    """Excel F3 salvat pe disc se reciteste; contine valoarea total general 18000
    si freeze_panes e fixat sub randul de antet real (A6)."""
    from models import db
    from services.situatii import export_f3
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        path = export_f3(s.id)
        ws, vals = _citeste_total_xlsx(path)
        # Total general 18000 apare in foaie (ca total, ca subtotal nu se duble-numara la verificare)
        assert Decimal('18000.00') in vals
        # freeze_panes sub antetul real (rand 5 header -> A6)
        assert ws.freeze_panes == 'A6'


def test_export_f1_xlsx_contine_total_cu_tva(app, admin_user, _models):
    """Excel F1 salvat se reciteste; contine total fara TVA 18000, TVA 3780 si
    total cu TVA 21780."""
    from models import db
    from services.situatii import export_f1
    with app.app_context():
        s = _genereaza(app, db, _models, _uid())
        path = export_f1(s.id)
        _ws, vals = _citeste_total_xlsx(path)
        assert Decimal('18000.00') in vals
        assert Decimal('3780.00') in vals
        assert Decimal('21780.00') in vals


def test_ruta_export_f3_gated_pe_flag(app, admin_user, _models, authenticated_client):
    """Cu flag 'situatii-f-forms' OFF: 404. Cu ON: 200 + content-type Excel."""
    from models import db
    from services.situatii import genereaza_situatie
    with app.app_context():
        uid = _uid()
        cid = _seed(db, *_models, uid)
        s = genereaza_situatie(cid, 2026, 3, uid)
        sid = s.id

    client = authenticated_client

    # Tot modulul contracte e gated pe 'controale-contract' (before_request).
    # Il tinem ON aici ca sa ajungem la ruta; gatingul fin testat e 'situatii-f-forms'.
    with app.app_context():
        set_flag('controale-contract', True, commit=True)
        set_flag('situatii-f-forms', False, commit=True)
    r_off = client.get(f'/contracte/situatie/{sid}/export/f3')
    assert r_off.status_code == 404

    with app.app_context():
        set_flag('situatii-f-forms', True, commit=True)
    r_on = client.get(f'/contracte/situatie/{sid}/export/f3')
    assert r_on.status_code == 200
    assert 'spreadsheetml' in r_on.headers.get('Content-Type', '')

    # Curatare flaguri (lasam OFF, default)
    with app.app_context():
        set_flag('situatii-f-forms', False, commit=True)
        set_flag('controale-contract', False, commit=True)


# ============================================================
# REGRESIE (fix review dz-4) - prinde bug-urile pe care suita verde le rata:
#   1. cantitati NEvalidate intra in F1/F2/F3 (divergenta vs valoare_totala_luna)
#   2. acumulare de rotunjire: F1/F2/F3 diferite + articole F3 != subtotal
#   3. F3 Excel: subtotaluri + articole pe ACEEASI coloana -> SUM da ~4x
# ============================================================

def _seed_custom(db, models, uid, specs, pret=Decimal('1')):
    """
    Seed parametrizat: specs = [(cod, cod_capitol, categorie, cant_luna,
    validat)]. Pret unitar comun. Reutilizeaza _curata pe prefixul FF-.
    Returneaza contract_id.
    """
    (Proiect, Contract, OfertaContract, PozitieBoQ,
     CantitateExecutataLunara) = models
    _curata(db, Proiect, Contract, OfertaContract, PozitieBoQ,
            CantitateExecutataLunara)
    p = Proiect(cod_proiect='FF-PRJ', nume='Formulare F test',
                data_start=date(2026, 1, 1), status='activ')
    db.session.add(p); db.session.flush()
    c = Contract(proiect_id=p.id, nr_contract='FF-CTR-001',
                 data_semnare=date(2026, 1, 15), status='activ',
                 valoare_totala=Decimal('500000'), moneda='RON')
    db.session.add(c); db.session.flush()
    o = OfertaContract(contract_id=c.id, proiect_id=p.id, versiune=1,
                       data_emitere=date(2026, 1, 20),
                       valoare_totala=Decimal('500000'),
                       sursa_import='manual', aprobata=True)
    db.session.add(o); db.session.flush()
    ordine = 1
    for cod, cap, catl, cant_luna, validat in specs:
        pz = PozitieBoQ(oferta_id=o.id, proiect_id=p.id,
                        cod_articol=cod, cod_capitol=cap,
                        denumire=f'Lucrare {catl}', um='mc',
                        cantitate_oferta=Decimal('50000'),
                        pret_unitar=pret, categorie='mixt',
                        categorie_lucrare=catl, ordine=ordine)
        db.session.add(pz); db.session.flush()
        ordine += 1
        if cant_luna is not None:
            cant = CantitateExecutataLunara(
                pozitie_boq_id=pz.id, proiect_id=p.id, an=2026, luna=3,
                cantitate_executata=Decimal(str(cant_luna)),
                validat=validat, validat_de_id=uid if validat else None)
            db.session.add(cant)
    db.session.commit()
    return c.id


def test_f_forms_exclud_cantitatile_nevalidate(app, admin_user, _models):
    """
    Bug #1/#3: o cantitate validat=False NU trebuie sa apara in F1/F2/F3 si
    F.total trebuie sa se reconcilieze cu situatie.valoare_totala_luna (care e
    calculat doar pe validate). Seed: 3 validate (10000+4000+3000=17000) + 1
    NEvalidata (cabluri 999, validat=False) care trebuie exclusa.
    """
    from models import db
    from services.situatii import (genereaza_situatie, genereaza_f1,
                                    genereaza_f2, genereaza_f3)
    specs = [
        ('FF-001', '1 REZISTENTA', 'beton', 10000, True),
        ('FF-002', '1 REZISTENTA', 'armatura', 4000, True),
        ('FF-003', '5 INSTALATII ELECTRICE', 'cabluri', 3000, True),
        ('FF-099', '5 INSTALATII ELECTRICE', 'cabluri', 999, False),  # NEvalidat
    ]
    with app.app_context():
        uid = _uid()
        cid = _seed_custom(db, _models, uid, specs)
        s = genereaza_situatie(cid, 2026, 3, uid)  # default doar_validate=True

        # valoare_totala_luna ignora deja nevalidatul -> 17000
        assert s.valoare_totala_luna == Decimal('17000.00')

        f3 = genereaza_f3(s.id)
        f2 = genereaza_f2(s.id)
        f1 = genereaza_f1(s.id)

        # Toate cele 3 formulare se reconciliaza cu valoarea (validata) a lunii
        assert f3['total_general'] == Decimal('17000.00')
        assert f2['total_general'] == Decimal('17000.00')
        assert f1['total_fara_tva'] == Decimal('17000.00')

        # Articolul nevalidat (FF-099, 999) NU apare nicaieri in F3
        coduri = set()
        suma_art = Decimal('0')
        for g in f3['grupe']:
            for cat in g['categorii']:
                for a in cat['articole']:
                    coduri.add(a['cod_articol'])
                    suma_art += a['valoare']
        assert 'FF-099' not in coduri
        assert suma_art == Decimal('17000.00')


def test_f_forms_reconciliaza_pe_valori_fractionare(app, admin_user, _models):
    """
    Bug #2: cu pret/cantitate fractionare, cele 3 formulare trebuie sa aiba
    acelasi total (quantize GLOBAL, nu acumulare de subtotaluri rotunjite), si
    suma articolelor AFISATE pe fiecare categorie == subtotalul afisat.

    Construim un caz unde rotunjirea per-articol diverge de rotunjirea globala:
    4 articole de cant 1, pret 0.014 -> raw 0.014 fiecare. Pe categorie (4 buc):
    suma raw = 0.056 -> subtotal 0.06; dar 4 x quantize(0.014)=4x0.01=0.04.
    Verificam ca articolele afisate insumeaza EXACT subtotalul (0.06).
    """
    from models import db
    from services.situatii import (genereaza_situatie, genereaza_f1,
                                    genereaza_f2, genereaza_f3)
    # Toate pe acelasi obiect+categorie ca sa fortam divergenta de rotunjire.
    specs = [
        ('FF-001', '1 REZISTENTA', 'beton', 1, True),
        ('FF-002', '1 REZISTENTA', 'beton', 1, True),
        ('FF-003', '1 REZISTENTA', 'beton', 1, True),
        ('FF-004', '1 REZISTENTA', 'beton', 1, True),
    ]
    with app.app_context():
        uid = _uid()
        cid = _seed_custom(db, _models, uid, specs, pret=Decimal('0.014'))
        s = genereaza_situatie(cid, 2026, 3, uid)

        # raw total = 4 * 0.014 = 0.056; persistat pe Numeric -> 0.06
        val_luna = s.valoare_totala_luna

        f3 = genereaza_f3(s.id)
        f2 = genereaza_f2(s.id)
        f1 = genereaza_f1(s.id)

        # Sursa unica de adevar: totalul global quantizat (0.06) pe toate 3,
        # identic cu valoarea lunii (reconciliere F == situatie)
        assert f3['total_general'] == Decimal('0.06')
        assert f2['total_general'] == Decimal('0.06')
        assert f1['total_fara_tva'] == Decimal('0.06')
        assert f1['total_fara_tva'] == f2['total_general'] == f3['total_general']
        assert f3['total_general'] == val_luna

        # Articolele afisate pe fiecare categorie insumeaza EXACT subtotalul
        for g in f3['grupe']:
            for cat in g['categorii']:
                suma_art = sum((a['valoare'] for a in cat['articole']),
                               Decimal('0'))
                assert suma_art == cat['valoare']
            # si subtotalul obiect == suma categoriilor
            sub_obj = sum((c['valoare'] for c in g['categorii']), Decimal('0'))
            assert g['subtotal'] == sub_obj


def test_export_f3_xlsx_suma_coloana_articole(app, admin_user, _models):
    """
    Bug #4: in F3 Excel subtotalurile obiect/categorie trebuie sa fie pe coloana
    separata 'Subtotal', NU amestecate cu valorile de articol. Verificam ca
    SUM-ul pe coloana 'Valoare (fara TVA)' (col 6, doar articole) == total
    general - nu ~4x din dublarea subtotaluri+articole+total pe aceeasi coloana.
    """
    from openpyxl import load_workbook
    from models import db
    from services.situatii import genereaza_situatie, export_f3
    specs = [
        ('FF-001', '1 REZISTENTA', 'beton', 10000, True),
        ('FF-002', '1 REZISTENTA', 'armatura', 4000, True),
        ('FF-003', '5 INSTALATII ELECTRICE', 'cabluri', 3000, True),
        ('FF-004', '5 INSTALATII ELECTRICE', 'tablouri', 1000, True),
    ]
    with app.app_context():
        uid = _uid()
        cid = _seed_custom(db, _models, uid, specs)
        s = genereaza_situatie(cid, 2026, 3, uid)
        path = export_f3(s.id)

        wb = load_workbook(path, data_only=True)
        ws = wb.active
        # Header tabel la randul 5; articolele de la randul 6 in jos. Coloana 6
        # = 'Valoare (fara TVA)' (doar articole). Randul TOTAL F3 isi pune
        # valoarea pe coloana 7 (Subtotal), deci col 6 contine STRICT articole.
        suma_col_articole = Decimal('0')
        for row in ws.iter_rows(min_row=6):
            c = row[5]  # coloana 6 (0-indexat 5)
            if isinstance(c.value, (int, float)):
                suma_col_articole += Decimal(str(c.value))
        assert suma_col_articole == Decimal('18000.00')

        # Coloana 7 (Subtotal): contine subtotaluri + total. Header la randul 5.
        assert ws.cell(row=5, column=7).value == 'Subtotal (fara TVA)'
