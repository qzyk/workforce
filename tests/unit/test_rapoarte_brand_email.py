"""
Rapoarte Faza 1: branding centralizat al exporturilor + email cu atasament.

Verifica:
  - modulul rapoarte.brand expune paleta Edifico (navy/gold/cream) si helperele
  - Excel generat foloseste paleta noua (gold C9A961 in header, navy 0B1426 text);
    paleta veche (indigo 1A237E / rosu C62828) a disparut
  - PDF se genereaza fara crash chiar daca fontul Cinzel lipseste (fallback Times)
  - email-ul cu atasament construieste MIME corect (attachment cu filename)
  - fara SMTP configurat -> trimite_raport_email returneaza False, fara crash
  - mock SMTP -> trimiterea reuseste
  - flag-ul 'rapoarte-email' exista in catalog si e default OFF
  - regresie: generatoarele Excel vechi inca produc fisiere valide
"""

import io
import os
from datetime import date

import pytest
from openpyxl import load_workbook


# ============================================================
# 1. Modulul brand
# ============================================================

def test_brand_paleta_constante():
    from rapoarte import brand
    assert brand.NAVY == '0B1426'
    assert brand.GOLD == 'C9A961'
    assert brand.CREAM == 'F5F1E8'
    assert brand.NAVY_HEX == '#0B1426'
    assert brand.GOLD_HEX == '#C9A961'


def test_brand_pdf_fonts_fallback_fara_crash():
    """Fara fisier Cinzel in repo -> fallback documentat la Times, fara exceptie."""
    from rapoarte import brand
    normal, bold = brand.get_pdf_fonts()
    assert normal and bold
    # Fie Cinzel inregistrat (EdificoSerif), fie fallback Times.
    assert normal in ('EdificoSerif', 'Times-Roman')
    assert bold in ('EdificoSerif', 'EdificoSerif-Bold', 'Times-Bold')


def test_brand_excel_styles_paleta():
    from rapoarte import brand
    st = brand.excel_styles()
    assert st['header_fill'].start_color.rgb.endswith('C9A961')   # gold
    assert st['title_font'].color.rgb.endswith('0B1426')          # navy
    assert st['total_fill'].start_color.rgb.endswith('F5F1E8')    # cream


def test_brand_pdf_table_style_header_gold():
    from rapoarte import brand
    ts = brand.pdf_table_style()
    bg = [c for c in ts.getCommands() if c[0] == 'BACKGROUND']
    # primul BACKGROUND e headerul; trebuie sa fie gold
    header_color = bg[0][3]
    # ReportLab Color -> hexval() da '0xrrggbbaa'
    assert 'c9a961' in header_color.hexval().lower()


def test_brand_pdf_header_elements():
    from rapoarte import brand
    els = brand.pdf_header_elements('FOAIE COLECTIVA', subtitlu='Test')
    # wordmark + titlu + subtitlu + spacer
    assert len(els) >= 3


# ============================================================
# 2. Excel generat real foloseste paleta (cu DB)
# ============================================================

def _culori_din_workbook(wb):
    culori = set()
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for c in row:
                fc = getattr(getattr(c.fill, 'start_color', None), 'rgb', None)
                tc = getattr(getattr(c.font, 'color', None), 'rgb', None)
                if fc:
                    culori.add(str(fc))
                if tc:
                    culori.add(str(tc))
    return ' '.join(culori)


def test_foaie_prezenta_excel_paleta_edifico(app):
    """Foaia colectiva de prezenta foloseste gold+navy, nu paleta indigo veche."""
    from models import db, Proiect, Angajat, AngajatProiect
    from tests.fixtures.data import make_proiect, make_angajat
    from rapoarte.excel_generator import generate_foaie_prezenta

    with app.app_context():
        p = make_proiect(db, Proiect, cod='BR-FOAIE-1', nume='Proiect Brand Foaie')
        a = make_angajat(db, Angajat, cnp='1988080808081', nume='Brand', prenume='Foaie')
        db.session.add(AngajatProiect(angajat_id=a.id, proiect_id=p.id))
        db.session.commit()

        wb = generate_foaie_prezenta(p.id, date.today().month, date.today().year)
        blob = _culori_din_workbook(wb)

    assert 'C9A961' in blob, 'gold lipseste din foaia de prezenta'
    assert '0B1426' in blob, 'navy lipseste din foaia de prezenta'
    # paleta veche disparuta
    assert '1A237E' not in blob, 'indigo vechi inca prezent'
    assert 'C62828' not in blob, 'rosu vechi inca prezent'


def test_stat_plata_excel_se_genereaza_valid(app):
    """Regresie: stat de plata produce un xlsx valid (PK magic bytes) cu paleta noua."""
    from models import db, Proiect, Angajat, AngajatProiect
    from tests.fixtures.data import make_proiect, make_angajat
    from rapoarte.excel_generator import generate_stat_plata

    with app.app_context():
        p = make_proiect(db, Proiect, cod='BR-STAT-1', nume='Proiect Brand Stat')
        a = make_angajat(db, Angajat, cnp='1988080808082', nume='Stat', prenume='Brand')
        db.session.add(AngajatProiect(angajat_id=a.id, proiect_id=p.id))
        db.session.commit()

        wb = generate_stat_plata(p.id, date.today().month, date.today().year)
        buf = io.BytesIO()
        wb.save(buf)

    data = buf.getvalue()
    assert data[:2] == b'PK'                       # xlsx valid
    wb2 = load_workbook(io.BytesIO(data))          # re-deschidere fara eroare
    assert 'C9A961' in _culori_din_workbook(wb2)


# ============================================================
# 3. PDF generat real (fallback font fara crash)
# ============================================================

def test_pdf_foaie_prezenta_se_genereaza(app):
    """PDF brandat se genereaza fara crash chiar fara Cinzel (fallback Times)."""
    reportlab = pytest.importorskip('reportlab')  # noqa: F841
    from models import db, Proiect, Angajat, AngajatProiect
    from tests.fixtures.data import make_proiect, make_angajat
    from rapoarte.pdf_generator import generate_pdf_foaie_prezenta

    with app.app_context():
        p = make_proiect(db, Proiect, cod='BR-PDF-1', nume='Proiect Brand PDF')
        a = make_angajat(db, Angajat, cnp='1988080808083', nume='Pdf', prenume='Brand')
        db.session.add(AngajatProiect(angajat_id=a.id, proiect_id=p.id))
        db.session.commit()

        filepath, filename = generate_pdf_foaie_prezenta(
            p.id, date.today().month, date.today().year)

    assert os.path.exists(filepath)
    assert filename.endswith('.pdf')
    with open(filepath, 'rb') as fh:
        head = fh.read(5)
    assert head[:4] == b'%PDF'                       # PDF valid


# ============================================================
# 4. Email cu atasament (MIME)
# ============================================================

def _scrie_fisier_temp(tmp_path, nume='raport_test.xlsx', continut=b'PKtest-bytes'):
    f = tmp_path / nume
    f.write_bytes(continut)
    return str(f)


def test_construieste_mesaj_raport_cu_atasament(tmp_path):
    from services.email_notif import construieste_mesaj_raport
    path = _scrie_fisier_temp(tmp_path)

    msg = construieste_mesaj_raport(
        'sef@edifico.space', 'Raport lunar', 'Vezi atasamentul.',
        fisier_path=path)

    assert msg['Subject'] == 'Raport lunar'
    assert msg['To'] == 'sef@edifico.space'
    # gasim partea de atasament cu filename-ul corect
    nume_atasamente = []
    for part in msg.walk():
        disp = part.get('Content-Disposition', '')
        if disp and 'attachment' in disp:
            nume_atasamente.append(part.get_filename())
    assert 'raport_test.xlsx' in nume_atasamente


def test_construieste_mesaj_raport_lista_destinatari(tmp_path):
    from services.email_notif import construieste_mesaj_raport
    msg = construieste_mesaj_raport(
        ['a@edifico.space', 'b@edifico.space'], 'Test', 'corp')
    assert 'a@edifico.space' in msg['To'] and 'b@edifico.space' in msg['To']


def test_atasament_inexistent_nu_crapa(tmp_path):
    """Fisier inexistent -> mesajul se construieste oricum, fara atasament, fara crash."""
    from services.email_notif import construieste_mesaj_raport
    msg = construieste_mesaj_raport(
        'x@edifico.space', 'Test', 'corp',
        fisier_path=str(tmp_path / 'nu_exista.xlsx'))
    has_attach = any('attachment' in (p.get('Content-Disposition', '') or '')
                     for p in msg.walk())
    assert not has_attach


def test_trimite_raport_email_fara_smtp_returneaza_false(tmp_path, monkeypatch):
    """Fara SMTP configurat -> False, fara crash si fara trimitere reala."""
    from services import email_notif
    for var in ('SMTP_HOST', 'SMTP_FROM'):
        monkeypatch.delenv(var, raising=False)
    path = _scrie_fisier_temp(tmp_path)
    rezultat = email_notif.trimite_raport_email(
        'sef@edifico.space', 'Raport', 'corp', fisier_path=path)
    assert rezultat is False


def test_trimite_raport_email_cu_smtp_mock(tmp_path, monkeypatch):
    """SMTP configurat + mock -> True; mesajul contine atasamentul cu filename."""
    from services import email_notif

    monkeypatch.setenv('SMTP_HOST', 'smtp.test.local')
    monkeypatch.setenv('SMTP_FROM', 'noreply@edifico.space')
    monkeypatch.setenv('SMTP_PORT', '587')
    monkeypatch.setenv('SMTP_USE_TLS', 'false')

    capturat = {}

    class _FakeSMTP:
        def __init__(self, host, port, timeout=None):
            capturat['host'] = host
            capturat['port'] = port

        def __enter__(self):
            return self

        def __exit__(self, *a):
            return False

        def starttls(self):
            capturat['starttls'] = True

        def login(self, user, password):
            capturat['login'] = (user, password)

        def sendmail(self, from_addr, to_addrs, msg_str):
            capturat['from'] = from_addr
            capturat['to'] = to_addrs
            capturat['body'] = msg_str

    monkeypatch.setattr(email_notif.smtplib, 'SMTP', _FakeSMTP)

    path = _scrie_fisier_temp(tmp_path, nume='situatie_proiect.pdf', continut=b'%PDF-1.4')
    rezultat = email_notif.trimite_raport_email(
        'sef@edifico.space', 'Situatie proiect', 'Atasat raportul.',
        fisier_path=path)

    assert rezultat is True
    assert capturat['host'] == 'smtp.test.local'
    assert 'sef@edifico.space' in capturat['to']
    # atasamentul cu filename apare in corpul mesajului serializat
    assert 'situatie_proiect.pdf' in capturat['body']
    # TLS dezactivat -> starttls nu a fost apelat
    assert 'starttls' not in capturat


# ============================================================
# 5. Flag rapoarte-email
# ============================================================

def test_flag_rapoarte_email_in_catalog():
    from services.feature_flags import KNOWN_FLAGS
    assert 'rapoarte-email' in KNOWN_FLAGS
    assert 'email' in KNOWN_FLAGS['rapoarte-email'].lower()


def test_flag_rapoarte_email_default_off(app):
    """Default OFF: flag necunoscut in DB -> is_enabled False."""
    from services.feature_flags import is_enabled
    with app.app_context():
        assert is_enabled('rapoarte-email') is False
