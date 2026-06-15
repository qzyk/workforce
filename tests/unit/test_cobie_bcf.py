"""
Teste unit pentru COBie export + BCF round-trip.
"""

import io
import json
import zipfile

import pytest
from openpyxl import load_workbook

from models import (db, Santier, Cladire, Nivel, Spatiu, ElementBIM,
                    IssueBIM, BIMComment, Utilizator)
from services import cobie_export, bcf_io
from services import feature_flags as ff


@pytest.fixture
def admin(app):
    with app.app_context():
        u = Utilizator.query.filter_by(email='cb_admin@test.local').first()
        if not u:
            u = Utilizator(nume='CB', prenume='A', email='cb_admin@test.local',
                           rol='admin', activ=True)
            u.set_password('x'); db.session.add(u); db.session.commit()
        yield u


@pytest.fixture
def hierarchy(app):
    with app.app_context():
        s = Santier(cod='S-COBIE', nume='Test Santier'); db.session.add(s); db.session.flush()
        c = Cladire(santier_id=s.id, cod='C1', nume='Cladire 1'); db.session.add(c); db.session.flush()
        n = Nivel(cladire_id=c.id, cod='P', nume='Parter', elevatie_m=0,
                  inaltime_m=3.0)
        db.session.add(n); db.session.flush()
        sp = Spatiu(nivel_id=n.id, cod='SP1', nume='Birou', tip_spatiu='Office')
        db.session.add(sp); db.session.flush()
        el = ElementBIM(cladire_id=c.id, spatiu_id=sp.id,
                        cod='W001', tip_element='wall', status='construit',
                        nume='Perete 1', ifc_global_id='1xY2zA3B4C5D6E7F8GHIJK')
        db.session.add(el); db.session.commit()
        yield {'santier_id': s.id, 'cladire_id': c.id,
               'nivel_id': n.id, 'spatiu_id': sp.id, 'el_id': el.id}


# ====================================================
# COBie EXPORT
# ====================================================

def test_cobie_export_has_required_sheets(app, hierarchy, admin):
    with app.app_context():
        buf = cobie_export.generate_cobie_workbook(hierarchy['santier_id'],
                                                    generated_by='test@local')
        wb = load_workbook(buf)
        expected = {'Facility', 'Floor', 'Space', 'Type', 'Component', 'Contact'}
        assert set(wb.sheetnames) == expected


def test_cobie_facility_row(app, hierarchy, admin):
    with app.app_context():
        buf = cobie_export.generate_cobie_workbook(hierarchy['santier_id'])
        wb = load_workbook(buf)
        ws = wb['Facility']
        # Header + 1 row = 2 randuri total
        assert ws.max_row == 2
        # Verific ca numele santierului apare in coloana A randul 2
        assert ws.cell(row=2, column=1).value == 'Test Santier'


def test_cobie_component_includes_ifc_guid(app, hierarchy, admin):
    with app.app_context():
        buf = cobie_export.generate_cobie_workbook(hierarchy['santier_id'])
        wb = load_workbook(buf)
        ws = wb['Component']
        assert ws.max_row >= 2  # header + 1 component minim
        # ExtIdentifier ar trebui sa contina ifc_global_id-ul
        row = list(ws.iter_rows(min_row=2, max_row=2, values_only=True))[0]
        assert '1xY2zA3B4C5D6E7F8GHIJK' in str(row)


def test_cobie_inexistent_santier_raises(app):
    with app.app_context():
        with pytest.raises(ValueError):
            cobie_export.generate_cobie_workbook(99999)


# ====================================================
# BCF EXPORT / IMPORT round-trip
# ====================================================

def test_bcf_export_creates_zip(app, hierarchy, admin):
    with app.app_context():
        issue = IssueBIM(titlu='Test BCF',
                         descriere='Test description',
                         tip='neconformitate', severitate='medie', status='deschis',
                         element_bim_id=hierarchy['el_id'],
                         cladire_id=hierarchy['cladire_id'],
                         raportat_de_id=admin.id)
        db.session.add(issue); db.session.commit()

        buf = bcf_io.export_bcfzip([issue.id])
        # E un zip valid
        zf = zipfile.ZipFile(buf, 'r')
        names = zf.namelist()
        assert 'bcf.version' in names
        # Cel putin un markup.bcf
        markups = [n for n in names if n.endswith('markup.bcf')]
        assert len(markups) == 1


def test_bcf_export_empty_raises(app):
    with app.app_context():
        # Clear all issues
        IssueBIM.query.delete()
        db.session.commit()
        with pytest.raises(ValueError):
            bcf_io.export_bcfzip()


def test_bcf_round_trip(app, hierarchy, admin):
    """Export + import = acelasi issue."""
    with app.app_context():
        # Sterg vechile issues
        IssueBIM.query.delete(); db.session.commit()

        issue = IssueBIM(titlu='Round-trip test',
                         descriere='Test desc',
                         tip='neconformitate', severitate='mare', status='in_lucru',
                         element_bim_id=hierarchy['el_id'],
                         raportat_de_id=admin.id)
        db.session.add(issue); db.session.commit()
        original_id = issue.id
        original_guid = issue.bcf_topic_guid  # poate fi None initial

        # Export
        buf = bcf_io.export_bcfzip([original_id])
        # Acum issue ar trebui sa aibă bcf_topic_guid asignat
        db.session.refresh(issue)
        assert issue.bcf_topic_guid is not None
        export_guid = issue.bcf_topic_guid

        # Modific titlul si re-import - ar trebui sa se updateze
        issue.titlu = 'Original modified'
        db.session.commit()

        stats = bcf_io.import_bcfzip(buf, user=admin)
        assert stats['updated'] == 1
        assert stats['created'] == 0

        # Re-fetch si verific titlul revenit la original
        db.session.refresh(issue)
        assert issue.titlu == 'Round-trip test'


def test_bcf_import_new_creates_issue(app, admin):
    """Import BCF cu un GUID nou creeaza un IssueBIM nou."""
    with app.app_context():
        IssueBIM.query.delete(); db.session.commit()

        # Construiesc un .bcfzip manual minimal
        markup_xml = '''<?xml version="1.0" encoding="utf-8"?>
<Markup>
    <Topic Guid="abcd-1234-5678-90ab" TopicType="Issue" TopicStatus="Open">
        <Title>Imported issue</Title>
        <Priority>High</Priority>
        <CreationDate>2026-05-11T10:00:00</CreationDate>
        <Description>Imported description text</Description>
    </Topic>
</Markup>
'''
        buf = io.BytesIO()
        with zipfile.ZipFile(buf, 'w', zipfile.ZIP_DEFLATED) as zf:
            zf.writestr('bcf.version', '<Version VersionId="2.1"/>')
            zf.writestr('abcd-1234-5678-90ab/markup.bcf', markup_xml)
        buf.seek(0)

        stats = bcf_io.import_bcfzip(buf, user=admin)
        assert stats['created'] == 1
        assert stats['updated'] == 0

        iss = IssueBIM.query.filter_by(bcf_topic_guid='abcd-1234-5678-90ab').first()
        assert iss is not None
        assert iss.titlu == 'Imported issue'
        assert iss.severitate == 'mare'  # High -> mare
        assert iss.status == 'deschis'  # Open -> deschis


def test_bcf_status_mapping():
    assert bcf_io._map_status('deschis') == 'Open'
    assert bcf_io._map_status('in_lucru') == 'InProgress'
    assert bcf_io._reverse_status('Open') == 'deschis'
    assert bcf_io._reverse_status('InProgress') == 'in_lucru'


def test_bcf_severity_mapping():
    assert bcf_io._map_severity('critica') == 'Critical'
    assert bcf_io._reverse_severity('Low') == 'mica'
    assert bcf_io._reverse_severity('High') == 'mare'


# ====================================================
# BCF VIEWPOINT round-trip (Faza 4) - flag bim-bcf-full
# ====================================================

# Valori ASIMETRICE intentionat: prind erori de mapare axe (x/y/z) + distanta.
_CAMERA = {
    'eye': [10.0, 20.0, 30.0],
    'look': [1.0, 2.0, 3.0],
    'up': [0.0, 0.0, 1.0],
    'fov': 55.0,
}


def _vec_aprox(a, b, tol=1e-4):
    return all(abs(x - y) <= tol for x, y in zip(a, b))


def test_bcf_viewpoint_roundtrip(app, hierarchy, admin):
    """
    Issue cu camera cunoscuta -> export (flag ON) -> .bcfzip contine viewpoint.bcfv
    cu CameraViewPoint corect -> import in issue nou -> camera (eye/look/up)
    pastrata in toleranta float. Miezul fazei.
    """
    with app.app_context():
        ff.set_flag('bim-bcf-full', True)
        IssueBIM.query.delete(); db.session.commit()

        issue = IssueBIM(titlu='Viewpoint test', tip='neconformitate',
                         severitate='mare', status='deschis',
                         element_bim_id=hierarchy['el_id'],
                         raportat_de_id=admin.id,
                         viewpoint_json=json.dumps({
                             'camera': _CAMERA,
                             'visible_guids': ['1xY2zA3B4C5D6E7F8GHIJK'],
                             'clipping': [{'pos': [0, 0, 1.5], 'dir': [0, 0, -1]}],
                         }))
        db.session.add(issue); db.session.commit()

        buf = bcf_io.export_bcfzip([issue.id])
        db.session.refresh(issue)
        export_guid = issue.bcf_topic_guid

        # 1. .bcfzip contine viewpoint.bcfv si markup-ul il refera
        zf = zipfile.ZipFile(buf, 'r')
        names = zf.namelist()
        bcfv = [n for n in names if n.endswith('viewpoint.bcfv')]
        assert len(bcfv) == 1, f'lipseste viewpoint.bcfv: {names}'
        markup = zf.read(f'{export_guid}/markup.bcf').decode('utf-8')
        assert 'Viewpoints' in markup and 'viewpoint.bcfv' in markup

        # 2. CameraViewPoint din .bcfv reflecta eye (10,20,30)
        vp_xml = zf.read(bcfv[0]).decode('utf-8')
        assert 'PerspectiveCamera' in vp_xml
        assert 'CameraViewPoint' in vp_xml
        parsed = bcf_io._parse_viewpoint_xml(vp_xml)
        assert _vec_aprox(parsed['camera']['eye'], _CAMERA['eye'])

        # 3. Import intr-un issue nou (guid nou) -> camera pastrata
        IssueBIM.query.delete(); db.session.commit()
        buf.seek(0)
        stats = bcf_io.import_bcfzip(buf, user=admin)
        assert stats['created'] == 1

        iss2 = IssueBIM.query.filter_by(bcf_topic_guid=export_guid).first()
        assert iss2 is not None and iss2.viewpoint_json
        vp = json.loads(iss2.viewpoint_json)
        # eye, look, up TOATE verificate (nu doar ca fisierul exista)
        assert _vec_aprox(vp['camera']['eye'], _CAMERA['eye'])
        assert _vec_aprox(vp['camera']['look'], _CAMERA['look'])  # distanta pastrata
        assert _vec_aprox(vp['camera']['up'], _CAMERA['up'])
        assert abs(vp['camera']['fov'] - _CAMERA['fov']) < 1e-3
        # visibility + clipping pastrate
        assert vp.get('visible_guids') == ['1xY2zA3B4C5D6E7F8GHIJK']
        assert vp['clipping'][0]['pos'] == [0, 0, 1.5]

        ff.set_flag('bim-bcf-full', False)


def test_bcf_export_regresie_flag_off(app, hierarchy, admin):
    """
    Cu flag 'bim-bcf-full' OFF, chiar daca issue are viewpoint_json, .bcfzip e ca
    azi: doar markup.bcf, FARA viewpoint.bcfv, FARA element Viewpoints in markup.
    """
    with app.app_context():
        ff.set_flag('bim-bcf-full', False)
        IssueBIM.query.delete(); db.session.commit()

        issue = IssueBIM(titlu='Regresie flag off', tip='neconformitate',
                         severitate='medie', status='deschis',
                         element_bim_id=hierarchy['el_id'],
                         raportat_de_id=admin.id,
                         viewpoint_json=json.dumps({'camera': _CAMERA}))
        db.session.add(issue); db.session.commit()

        buf = bcf_io.export_bcfzip([issue.id])
        zf = zipfile.ZipFile(buf, 'r')
        names = zf.namelist()
        assert 'bcf.version' in names
        assert not any(n.endswith('viewpoint.bcfv') for n in names), \
            'cu flag OFF nu trebuie sa existe viewpoint.bcfv'
        db.session.refresh(issue)
        markup = zf.read(f'{issue.bcf_topic_guid}/markup.bcf').decode('utf-8')
        assert 'Viewpoints' not in markup, 'markup-ul nu trebuie sa refere viewpoint'


def test_bcf_viewpoint_xml_helpers():
    """_make_viewpoint_xml + _parse_viewpoint_xml: round-trip pur la nivel XML."""
    xml = bcf_io._make_viewpoint_xml({'camera': _CAMERA})
    assert xml is not None
    parsed = bcf_io._parse_viewpoint_xml(xml)
    assert _vec_aprox(parsed['camera']['eye'], _CAMERA['eye'])
    assert _vec_aprox(parsed['camera']['look'], _CAMERA['look'])
    assert _vec_aprox(parsed['camera']['up'], _CAMERA['up'])
    # camera fara look/eye -> None (degradare gratioasa)
    assert bcf_io._make_viewpoint_xml({'camera': {'eye': [1, 2, 3]}}) is None
