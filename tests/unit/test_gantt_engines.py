"""
Teste unitare pentru motoarele de planificare Gantt (services/gantt/*).
Nu necesita aplicatia Flask - testeaza direct functiile pure.
"""
import io
import xml.etree.ElementTree as ET

import pytest

from services.gantt.normalizare import fara_diacritice, normalizeaza
from services.gantt import import_engine
from services.gantt.clasificare import Clasificator
from services.gantt.durate import estimeaza_durata
from services.gantt.wbs import genereaza_wbs
from services.gantt.dependinte import genereaza_dependinte
from services.gantt.validare import valideaza
from services.gantt import export as export_engine
from services.gantt.modele import ArticolF3, Activitate, Dependenta, RezultatPlanificare
from services.gantt.pipeline import MotorPlanificare
from services.gantt import config_loader as cfg


SAMPLE_CSV = (
    "cod_articol;denumire;um;cantitate;obiect;tronson;categorie\n"
    "ART001;Trasare traseu;m;800;Retea apa;Strada A;Terasamente\n"
    "ART002;Sapatura mecanizata;mc;1200;Retea apa;Strada A;Terasamente\n"
    "ART003;Pozare conducta PEHD;m;800;Retea apa;Strada A;Conducte\n"
    "ART004;Umplutura compactare;mc;900;Retea apa;Strada A;Terasamente\n"
    "ART005;Refacere asfalt;mp;640;Retea apa;Strada A;Drumuri\n"
    "ART006;Sapatura mecanizata;mc;1100;Retea apa;Strada B;Terasamente\n"
    "ART007;Pozare conducta;m;750;Retea apa;Strada B;Conducte\n"
).encode('utf-8')


# ------------------------------------------------------------------ normalizare
def test_fara_diacritice():
    assert fara_diacritice('săpătură') == 'sapatura'
    assert fara_diacritice('tronson și țeavă') == 'tronson si teava'
    assert normalizeaza('  Săpătură   MECANIZATĂ ') == 'sapatura mecanizata'


# ---------------------------------------------------------------------- import
def test_import_csv_mapeaza_coloane_si_dedup():
    articole, raport = import_engine.importa(SAMPLE_CSV, '.csv')
    assert raport['nr_articole'] == 7
    assert 'cod_articol' in raport['coloane_mapate']
    assert articole[0].obiect == 'Retea apa'
    assert articole[1].cantitate == 1200.0


def test_import_dedup_si_null():
    csv = (
        "cod;denumire;cantitate;obiect;tronson\n"
        "X1;Sapatura;10;O;T\n"
        "X1;Sapatura duplicat cod;20;O;T\n"      # cod duplicat -> redenumit, pastrat
        ";Fara cod dar cu denumire;5;O;T\n"      # fara cod -> cod auto
        "X2;;1;O;T\n"                             # fara denumire -> ignorat
    ).encode('utf-8')
    articole, raport = import_engine.importa(csv, '.csv')
    assert raport['nr_articole'] == 3
    assert raport['nr_duplicate_redenumite'] == 1
    assert raport['nr_randuri_ignorate'] == 1


def test_import_format_necunoscut():
    with pytest.raises(import_engine.EroareImport):
        import_engine.importa(b'x', '.pdf')


# -- detectie format dupa CONTINUT (magic bytes), nu dupa extensie ------------
_RANDURI_F3 = [
    ('cod_articol', 'denumire', 'um', 'cantitate', 'obiect', 'tronson', 'categorie'),
    ('ART001', 'Trasare traseu', 'm', '800', 'Retea apa', 'Strada A', 'Terasamente'),
    ('ART002', 'Sapatura mecanizata', 'mc', '1200', 'Retea apa', 'Strada A', 'Terasamente'),
    ('ART003', 'Pozare conducta PEHD', 'm', '800', 'Retea apa', 'Strada A', 'Conducte'),
]


def _xlsx_bytes():
    from openpyxl import Workbook
    wb = Workbook(); ws = wb.active
    for r in _RANDURI_F3:
        ws.append(list(r))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def _html_bytes():
    h = "<html><head><meta charset='utf-8'></head><body><table border='1'>"
    for r in _RANDURI_F3:
        h += '<tr>' + ''.join(f'<td>{c}</td>' for c in r) + '</tr>'
    return (h + '</table></body></html>').encode('utf-8')


def _spreadsheetml_bytes():
    x = ('<?xml version="1.0"?>\n<Workbook '
         'xmlns="urn:schemas-microsoft-com:office:spreadsheet" '
         'xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet">'
         '<Worksheet ss:Name="F3"><Table>')
    for r in _RANDURI_F3:
        x += '<Row>' + ''.join(
            f'<Cell><Data ss:Type="String">{c}</Data></Cell>' for c in r) + '</Row>'
    return (x + '</Table></Worksheet></Workbook>').encode('utf-8')


def test_import_xlsx_real():
    articole, raport = import_engine.importa(_xlsx_bytes(), '.xlsx')
    assert raport['nr_articole'] == 3
    assert articole[0].obiect == 'Retea apa'


def test_import_html_deghizat_in_xlsx():
    # export tipic de la softuri de devize: continut HTML cu extensia .xlsx
    articole, raport = import_engine.importa(_html_bytes(), '.xlsx')
    assert raport['nr_articole'] == 3
    assert articole[1].denumire == 'Sapatura mecanizata'
    assert articole[1].cantitate == 1200.0


def test_import_spreadsheetml_xml():
    articole, raport = import_engine.importa(_spreadsheetml_bytes(), '.xls')
    assert raport['nr_articole'] == 3
    assert articole[2].categorie == 'Conducte'


def test_import_xls_binar_fara_continut_da_mesaj_clar():
    # antet OLE2 valid dar continut bogus -> EroareImport prietenos (nu crash brut)
    fake_xls = import_engine._MAGIC_OLE2 + b'\x00' * 256
    with pytest.raises(import_engine.EroareImport) as ei:
        import_engine.importa(fake_xls, '.xlsx')   # extensia minte, magic spune .xls
    assert '.xls' in str(ei.value)


def test_import_fisier_corupt_mesaj_prietenos():
    with pytest.raises(import_engine.EroareImport) as ei:
        import_engine.importa(b'doar text aleator, nu e excel \x00\x01', '.xlsx')
    assert 'xlsx' in str(ei.value).lower()


# -- structura F3 reala: multi-sheet, antet pe rand 4, obiect din nume sheet,
#    randuri-titlu de sectiune, articole fara cod, randuri NOTA -----------------
def _xlsx_f3_real():
    """Construieste un xlsx ca un F3 real de devize (Hala Campina-style)."""
    from openpyxl import Workbook
    wb = Workbook()
    # sheet 1: pagina de titlu, fara tabel -> trebuie sarit
    ws0 = wb.active
    ws0.title = 'TITLE PAGE'
    ws0.append(['PROIECT TEHNIC'])
    ws0.append(['Beneficiar: X'])
    ws0.append(['F3 - Lista de cantitati'])

    def _adauga_obiect(ws):
        ws.append(['Formular F3', '', '', '', ''])          # rand 1: titlu liber
        ws.append(['Lista cu cantitati de lucrari', '', '', '', ''])  # rand 2
        ws.append(['', '', '', '', ''])                      # rand 3: gol
        ws.append(['Nr./No.', 'DENUMIRE/ NAME', 'U.M.', 'CANTITATE', 'PRET UNITAR'])  # rand 4 = antet
        ws.append(['0', '1', '2', '3', '4'])                 # rand 5: numerotare coloane
        ws.append(['', 'INSTALATII SANITARE', '', '', ''])   # subtitlu disciplina (col 1)
        ws.append(['1.', 'HIDRANTI INTERIORI', '', 0, 0])    # titlu sectiune (fara um/cant)
        ws.append(['1.0', 'Hidrant interior cu furtun', 'set', 7, 0])    # articol
        ws.append(['2.0', 'Teava din otel zincata', 'm', 50, 0])         # articol
        ws.append(['', 'idem DN65', 'm', 30, 0])             # articol FARA cod -> AUTO
        ws.append(['2.', 'OBIECTE SANITARE', '', 0, 0])      # alt titlu sectiune
        ws.append(['1.0', 'Lavoar portelan', 'buc', 4, 0])   # cod "1.0" se repeta -> dedup
        ws.append(['NOTA', '', '', '', ''])                  # incepe disclaimerul
        ws.append(['', 'Orice nume de produs e cu titlu informativ si...', '', '', ''])

    _adauga_obiect(wb.create_sheet('2.1 Obiect 1 - Anexa'))
    _adauga_obiect(wb.create_sheet('2.3 Obiect 2 - Hala'))
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_import_f3_real_multi_sheet():
    articole, raport = import_engine.importa(_xlsx_f3_real(), '.xlsx')
    # TITLE PAGE sarit, 2 obiecte procesate
    assert raport['nr_sheeturi'] == 3
    assert 'TITLE PAGE' in raport['sheeturi_sarite']
    # antet gasit pe randul 4 (nu pe primul), coloanele mapate corect
    assert raport['rand_antet'] == 4
    assert raport['coloane_mapate']['denumire'] == 'DENUMIRE/ NAME'
    assert raport['coloane_mapate']['um'] == 'U.M.'
    # 4 articole reale per sheet x 2 sheet-uri = 8
    assert raport['nr_articole'] == 8
    # obiect derivat din numele sheet-ului (fara prefixul "2.1 ")
    obiecte = {a.obiect for a in articole}
    assert obiecte == {'Obiect 1 - Anexa', 'Obiect 2 - Hala'}


def test_import_f3_real_titlu_articol_si_cod_auto():
    articole, raport = import_engine.importa(_xlsx_f3_real(), '.xlsx')
    o1 = [a for a in articole if a.obiect == 'Obiect 1 - Anexa']
    # articolul fara cod a primit un cod AUTO (nu s-a pierdut)
    assert any(a.cod_articol.startswith('AUTO') and a.denumire == 'idem DN65' for a in o1)
    # randul-titlu de sectiune a devenit tronson, nu articol
    tronsoane = {a.tronson for a in o1}
    assert 'HIDRANTI INTERIORI' in tronsoane and 'OBIECTE SANITARE' in tronsoane
    # randul NOTA si disclaimerul de dupa nu au generat articole
    assert all('titlu informativ' not in a.denumire for a in articole)
    # cod-ul "1.0" repetat in a doua sectiune a fost dedup-uit, nu pierdut
    assert raport['nr_duplicate_redenumite'] >= 1


# ------------------------------------------------------------------ clasificare
def test_clasificare_exacta_si_diacritice():
    c = Clasificator(cfg.CLASIFICARE_IMPLICITA, cfg.SETARI_IMPLICITE['sinonime'])
    cat, scor = c.clasifica('Săpătură mecanizată în teren tare')
    assert cat == 'SAPATURA' and scor == 1.0
    cat, _ = c.clasifica('Pozare conducta PEHD De160')
    assert cat == 'POZARE_CONDUCTA'


def test_clasificare_sinonim():
    c = Clasificator(cfg.CLASIFICARE_IMPLICITA, cfg.SETARI_IMPLICITE['sinonime'])
    # 'montaj teava' -> sinonim teava->conducta -> 'montaj conducta' = POZARE_CONDUCTA
    cat, _ = c.clasifica('Montaj teava otel')
    assert cat == 'POZARE_CONDUCTA'


def test_clasificare_fuzzy_typo():
    c = Clasificator({'SAPATURA': ['sapatura']}, prag_fuzzy=0.8)
    cat, scor = c.clasifica('sapatraa in teren')  # greseala de scriere
    assert cat == 'SAPATURA' and 0.8 <= scor < 1.0


def test_clasificare_neclasificat():
    c = Clasificator(cfg.CLASIFICARE_IMPLICITA)
    cat, _ = c.clasifica('Articol complet aiurea zzz')
    assert cat is None


def test_clasificare_categorii_noi_sanitare():
    # dictionarul real (config/gantt/clasificare.json) - calibrat pe instalatii sanitare
    c = Clasificator(cfg.incarca('clasificare', cfg.CLASIFICARE_IMPLICITA),
                     cfg.incarca('setari', cfg.SETARI_IMPLICITE).get('sinonime'))
    cazuri = {
        'Robinet cu obturator sferic Dn15': 'ARMATURI',
        'Reductie PPR 25-20': 'ARMATURI',
        'Clapeta de sens': 'ARMATURI',
        'Hidrant interior cu furtun plat': 'ARMATURI',
        'Instalatie completa pentru WC': 'OBIECTE_SANITARE',
        'Vas Pisoar din portelan sanitar': 'OBIECTE_SANITARE',
        'Protejare la foc a strapungerilor': 'IZOLATII',
        'Banda izolatoare': 'IZOLATII',
        'Demontare conducte de canalizare': 'DEMONTARI',
        'Transport moloz': 'TRANSPORT',
        'Efectuarea probei de etanseitate': 'PROBE',
        'Spalarea conductelor de apa': 'PROBE',
        'Teava din otel zincata inclusiv fitinguri': 'POZARE_CONDUCTA',
        'Termostat de ambianta': 'APARATURA_AMC',
    }
    for d, cat in cazuri.items():
        assert c.clasifica(d)[0] == cat, d


def test_clasificare_prefix_cod():
    c = Clasificator({'SAPATURA': ['sapatura']},
                     reguli_prefix=[('TSA', 'SAPATURA', 100),
                                    ('ACE', 'POZARE_CONDUCTA', 100),
                                    ('SD', 'OBIECTE_SANITARE', 100)])
    # prefixul de cod are prioritate, indiferent de denumire
    assert c.clasifica('text irelevant', 'TSA02B1>') == ('SAPATURA', 1.0)
    assert c.clasifica('orice', 'ACE10A')[0] == 'POZARE_CONDUCTA'
    # cod fara prefix cunoscut (ex: numar de rand '1.0') -> cade pe denumire
    assert c.clasifica('sapatura manuala', '1.0')[0] == 'SAPATURA'
    assert c.clasifica('necunoscut zzz', '1.0')[0] is None


# ---------------------------------------------------------------------- durate
def test_durata_din_randament():
    setari = cfg.SETARI_IMPLICITE
    assert estimeaza_durata(1200, 'SAPATURA', setari) == 6   # 1200/200
    assert estimeaza_durata(0, 'SAPATURA', setari) == 1      # min 1
    assert estimeaza_durata(100, None, setari) == 1          # neclasificat -> implicit


def test_durata_categorii_cladire():
    """Regresie: categoriile de cladire (clasificator extins) au randamente -
    fara ele, articole mari (ex. schela 13120 mp) cadeau pe durata implicita 1 zi."""
    for setari in (cfg.SETARI_IMPLICITE, cfg.incarca('setari', cfg.SETARI_IMPLICITE)):
        assert estimeaza_durata(600, 'BETON', setari) == 20          # 600/30
        assert estimeaza_durata(4000, 'COFRAJE', setari) == 50       # 4000/80
        assert estimeaza_durata(45000, 'ARMATURA_BETON', setari) == 30
        assert estimeaza_durata(13120, 'CONFECTII_METALICE', setari) == 53
        assert estimeaza_durata(120, 'TAMPLARIE', setari) == 12


def test_randamente_sincronizate_json_si_implicite():
    """Cele doua surse de config (setari.json + fallback-ul SETARI_IMPLICITE)
    trebuie tinute identice - vezi bug-ul similar pe 'coloane' (053b18c).
    In plus, orice categorie din clasificator trebuie sa aiba randament,
    altfel durata cade pe implicit (1 zi) indiferent de cantitate."""
    json_setari = cfg.incarca('setari', cfg.SETARI_IMPLICITE)
    assert json_setari['randamente'] == cfg.SETARI_IMPLICITE['randamente']

    clasificare = cfg.incarca('clasificare', cfg.CLASIFICARE_IMPLICITA)
    fara_randament = set(clasificare) - set(json_setari['randamente'])
    assert not fara_randament, f'categorii fara randament: {sorted(fara_randament)}'


# ------------------------------------------------------------------------- wbs
def test_wbs_ierarhie():
    articole, _ = import_engine.importa(SAMPLE_CSV, '.csv')
    motor = MotorPlanificare()
    acts = motor.clasifica_articole(articole)
    noduri = genereaza_wbs(acts, motor.dependinte['ordine_categorii'])
    niveluri = {n.nivel for n in noduri}
    assert niveluri == {1, 2, 3, 4}
    obiecte = [n for n in noduri if n.nivel == 1]
    assert obiecte[0].wbs_id == '1'
    # activitatile au wbs_id de forma a.b.c.d
    act = next(a for a in acts if a.tronson == 'Strada A' and a.categorie_tehnologica == 'SAPATURA')
    assert act.wbs_id.count('.') == 3


# ------------------------------------------------------------------ dependinte
def test_dependinte_lant_si_multiplicare_pe_tronson():
    articole, _ = import_engine.importa(SAMPLE_CSV, '.csv')
    motor = MotorPlanificare()
    rez = motor.proceseaza(articole)
    by_id = {a.id: a for a in rez.activitati}

    # Pozarea de pe Strada A are ca predecesor sapatura de pe Strada A (FS)
    poz_a = next(a for a in rez.activitati if a.tronson == 'Strada A' and a.categorie_tehnologica == 'POZARE_CONDUCTA')
    assert poz_a.predecesori, 'pozarea trebuie sa aiba predecesor'
    pred = by_id[poz_a.predecesori[0].predecesor_id]
    assert pred.categorie_tehnologica == 'SAPATURA' and pred.tronson == 'Strada A'

    # template multiplicat: si Strada B are lant propriu
    poz_b = next(a for a in rez.activitati if a.tronson == 'Strada B' and a.categorie_tehnologica == 'POZARE_CONDUCTA')
    pred_b = by_id[poz_b.predecesori[0].predecesor_id]
    assert pred_b.tronson == 'Strada B'

    # decalaj UMPLUTURA->REFACERE = 2 (din config implicit)
    refacere = next(a for a in rez.activitati if a.tronson == 'Strada A' and a.categorie_tehnologica == 'REFACERE')
    assert any(d.decalaj == 2 for d in refacere.predecesori)


# -------------------------------------------------------------------- validare
def test_validare_ciclu():
    a = Activitate(id='A1', cod='c1', nume='A', categorie_tehnologica='SAPATURA')
    b = Activitate(id='A2', cod='c2', nume='B', categorie_tehnologica='POZARE_CONDUCTA')
    a.predecesori.append(Dependenta('A2', 'FS', 0))
    b.predecesori.append(Dependenta('A1', 'FS', 0))   # ciclu A1<->A2
    rap = valideaza([a, b])
    assert not rap.valid
    assert any(p.cod == 'ciclu' for p in rap.erori)


def test_validare_predecesor_lipsa_si_neclasificat():
    a = Activitate(id='A1', cod='c1', nume='A', categorie_tehnologica=None)
    a.predecesori.append(Dependenta('INEXISTENT', 'FS', 0))
    rap = valideaza([a])
    coduri = {p.cod for p in rap.probleme}
    assert 'predecesor_lipsa' in coduri
    assert 'neclasificat' in coduri


def test_validare_id_duplicat():
    a1 = Activitate(id='DUP', cod='c1', nume='A', categorie_tehnologica='SAPATURA')
    a2 = Activitate(id='DUP', cod='c2', nume='B', categorie_tehnologica='SAPATURA')
    rap = valideaza([a1, a2])
    assert any(p.cod == 'id_duplicat' for p in rap.erori)


# --------------------------------------------------------------------- export
def _rezultat_demo():
    articole, _ = import_engine.importa(SAMPLE_CSV, '.csv')
    return MotorPlanificare().proceseaza(articole)


def test_export_csv_contine_predecesori():
    rez = _rezultat_demo()
    data, mime, ext = export_engine.exporta('csv', rez)
    text = data.decode('utf-8-sig')
    assert text.splitlines()[0].startswith('ID,WBS,Activity Name,Duration,Predecessors')
    assert 'FS' in text  # exista cel putin o relatie FS


def test_export_msproject_xml_valid():
    rez = _rezultat_demo()
    data, _, _ = export_engine.exporta('msproject', rez)
    root = ET.fromstring(data)               # trebuie sa fie XML valid
    ns = '{http://schemas.microsoft.com/project}'
    tasks = root.find(f'{ns}Tasks')
    assert tasks is not None and len(tasks) > 0
    # exista cel putin un PredecessorLink
    assert root.find(f'.//{ns}PredecessorLink') is not None


def test_export_primavera_xml_valid():
    rez = _rezultat_demo()
    data, _, _ = export_engine.exporta('primavera', rez)
    root = ET.fromstring(data)
    assert root.tag == 'APIBusinessObjects'
    assert root.find('.//Activity') is not None
    assert root.find('.//Relationship') is not None


def test_export_format_invalid():
    rez = _rezultat_demo()
    with pytest.raises(ValueError):
        export_engine.exporta('dxf', rez)


# --------------------------------------------------------------------- cost 5D
def test_cost_estimat_din_tarif():
    from services.gantt.cost import calculeaza_cost
    art = ArticolF3('1.0', 'Sapatura', um='mc', cantitate=100)
    tarife = {'SAPATURA': {'tarif': 35, 'um': 'mc', 'material': 0.20}}
    val, mat, man, uti, est = calculeaza_cost(art, 'SAPATURA', tarife)
    assert val == 3500.0 and est is True
    assert abs(mat - 700.0) < 0.01 and abs(man - 2800.0) < 0.01 and uti == 0.0


def test_cost_utilaj_din_pondere():
    """Pondere utilaj pe categorie -> split pe 3 resurse (material/utilaj/manopera)."""
    from services.gantt.cost import calculeaza_cost
    art = ArticolF3('TS', 'Sapatura mecanizata', um='mc', cantitate=100)
    tarife = {'SAPATURA': {'tarif': 35, 'um': 'mc', 'material': 0.10, 'utilaj': 0.60}}
    val, mat, man, uti, est = calculeaza_cost(art, 'SAPATURA', tarife)
    assert val == 3500.0
    assert abs(mat - 350.0) < 0.01 and abs(uti - 2100.0) < 0.01 and abs(man - 1050.0) < 0.01


def test_cost_din_pret_f3_nu_estimeaza():
    from services.gantt.cost import calculeaza_cost
    art = ArticolF3('1', 'X', cantitate=10, pret_total=5000)
    val, _mat, _man, _uti, est = calculeaza_cost(art, 'ARMATURI', {'ARMATURI': {'tarif': 250}})
    assert val == 5000.0 and est is False


def test_program_forward_pass_si_lag():
    from services.gantt.program import programeaza
    a = Activitate(id='A1', cod='1', nume='a', categorie_tehnologica='SAPATURA', durata=3)
    b = Activitate(id='A2', cod='2', nume='b', categorie_tehnologica='POZARE_CONDUCTA', durata=2)
    b.predecesori.append(Dependenta('A1', 'FS', 1))  # start dupa finish(A1) + 1
    total = programeaza([a, b])
    assert (a.start_zi, a.finish_zi) == (0, 3)
    assert (b.start_zi, b.finish_zi) == (4, 6)
    assert total == 6


def test_curba_s_si_cost_in_pipeline():
    articole, _ = import_engine.importa(SAMPLE_CSV, '.csv')
    st = MotorPlanificare().proceseaza(articole).statistici
    assert st['cost_total'] > 0
    assert st['durata_totala_zile'] > 0
    assert st['curba_s'] and st['curba_s'][-1]['procent'] == 100.0
    # material + manopera ~ total
    assert abs((st['cost_material'] + st['cost_manopera']) - st['cost_total']) < 1.0


# CSV sintetic: deviz de cladire (structura + arhitectura + fatada), cantitati
# tipice pentru un corp mic. Inainte de randamentele de cladire, fiecare articol
# cadea pe durata implicita (1 zi) -> plan total nerealist de scurt; iar pe
# devize reale articolele mari (ex. schela 13120 mp) dadeau durate aberante.
CSV_CLADIRE = (
    "cod_articol;denumire;um;cantitate;obiect;tronson\n"
    "1;Turnare beton C25/30 in fundatii;mc;600;Cladire C1;Corp A\n"
    "2;Cofraje stalpi si grinzi;mp;4000;Cladire C1;Corp A\n"
    "3;Otel beton BST500 fasonat si montat;kg;45000;Cladire C1;Corp A\n"
    "4;Zidarie din caramida la pereti;mc;300;Cladire C1;Corp A\n"
    "5;Termosistem polistiren expandat 10 cm;mp;2400;Cladire C1;Corp A\n"
    "6;Tencuieli interioare la pereti;mp;8000;Cladire C1;Corp A\n"
    "7;Pardoseli din parchet laminat;mp;3000;Cladire C1;Corp A\n"
    "8;Tamplarie PVC cu geam termopan;buc;120;Cladire C1;Corp A\n"
    "9;Placaje din gresie si faianta;mp;500;Cladire C1;Corp A\n"
).encode('utf-8')


def test_plan_cladire_durata_realista():
    """Planul unui deviz de cladire are durata totala intr-un interval realist
    (luni, nu zile si nu zeci de ani) - regresie pentru randamentele de cladire."""
    articole, _ = import_engine.importa(CSV_CLADIRE, '.csv')
    rez = MotorPlanificare().proceseaza(articole)
    st = rez.statistici

    assert st['nr_neclasificate'] == 0          # toate denumirile au categorie
    pe_categorie = {a.categorie_tehnologica: a for a in rez.activitati}
    assert pe_categorie['BETON'].durata == 20            # 600 mc / 30 pe zi
    assert pe_categorie['FINISAJE'].durata == 100        # 8000 mp / 80 pe zi
    assert pe_categorie['TAMPLARIE'].durata == 12        # 120 buc / 10 pe zi
    # nicio activitate nu mai cade pe durata implicita (1 zi la cantitati mari)
    assert all(a.durata > 1 for a in rez.activitati)

    # interval realist pentru un corp mic de cladire: ~2 luni .. ~1.5 ani
    assert 60 <= st['durata_totala_zile'] <= 400


# ----------------------------------------------------------------- 4D / vizual
def test_drum_critic_marja():
    from services.gantt.program import programeaza, drum_critic
    a = Activitate(id='A', cod='1', nume='a', categorie_tehnologica='X', durata=3)
    c = Activitate(id='C', cod='2', nume='c', categorie_tehnologica='Y', durata=1)
    d = Activitate(id='D', cod='3', nume='d', categorie_tehnologica='Z', durata=2)
    d.predecesori = [Dependenta('A', 'FS', 0), Dependenta('C', 'FS', 0)]
    total = programeaza([a, c, d])
    nr = drum_critic([a, c, d], total)
    assert a.critic and d.critic               # lantul lung A->D e critic
    assert (not c.critic) and c.marja == 2     # C are 2 zile marja totala
    assert nr == 2


def test_diagrama_sarcini_gantt_zile_lucratoare():
    from datetime import date as _date
    from services.gantt import diagrama
    articole, _ = import_engine.importa(SAMPLE_CSV, '.csv')
    rez = MotorPlanificare().proceseaza(articole)
    d = diagrama.sarcini_gantt(rez, _date(2026, 6, 1))
    assert d['sarcini'] and d['total'] == len(rez.activitati)
    s0 = d['sarcini'][0]
    assert {'id', 'name', 'start', 'end', 'custom_class'}.issubset(s0.keys())
    for s in d['sarcini'][:15]:                # toate datele cad in zile lucratoare
        y, m, dd = map(int, s['start'].split('-'))
        assert _date(y, m, dd).weekday() < 5
