"""
EDIFICO WORKFORCE - Generator Rapoarte Excel
Toate rapoartele Excel cu openpyxl: foaie prezenta, stat plata, situatie proiect, centralizator ore.
"""

import os
import calendar
from datetime import date, datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter


# ============================================================
# STILURI COMUNE
# ============================================================

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
HEADER_FONT = Font(bold=True, color='FFFFFF', size=10, name='Arial')
HEADER_FILL = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
TITLE_FONT = Font(bold=True, size=14, name='Arial', color='1A237E')
SUBTITLE_FONT = Font(bold=True, size=11, name='Arial')
DATA_FONT = Font(size=9, name='Arial')
TOTAL_FONT = Font(bold=True, size=10, name='Arial')
TOTAL_FILL = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')

WEEKEND_FILL = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
SARBATOARE_FILL = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
SUPL_FILL = PatternFill(start_color='FFF3E0', end_color='FFF3E0', fill_type='solid')
RED_FILL = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
YELLOW_FILL = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
GREEN_FILL = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

CENTER = Alignment(horizontal='center', vertical='center')
LEFT = Alignment(horizontal='left', vertical='center')
WRAP_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)


def _auto_width(ws, min_w=5, max_w=40):
    for col in ws.columns:
        max_len = 0
        col_letter = None
        for cell in col:
            if hasattr(cell, 'column_letter'):
                col_letter = cell.column_letter
            if cell.value and not isinstance(cell, type(None)):
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
        if col_letter:
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, min_w), max_w)


def _style_header_row(ws, row, col_start, col_end):
    for c in range(col_start, col_end + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def _style_data_cell(ws, row, col, value, font=None, fill=None, alignment=None):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font or DATA_FONT
    cell.border = THIN_BORDER
    cell.alignment = alignment or CENTER
    if fill:
        cell.fill = fill
    return cell


# ============================================================
# 1. FOAIE COLECTIVA DE PREZENTA
# ============================================================

def generate_foaie_prezenta(proiect_id, luna, an, include_supl=True, app=None):
    """Genereaza Foaie Colectiva de Prezenta format constructii."""
    from models import db, Proiect, Angajat, AngajatProiect, Pontaj, SarbatoareLegala

    proiect = Proiect.query.get(proiect_id)
    if not proiect:
        raise ValueError('Proiectul nu a fost gasit.')

    # Angajati pe proiect
    asocieri = AngajatProiect.query.filter_by(proiect_id=proiect_id).all()
    angajati = [a.angajat for a in asocieri]
    if not angajati:
        angajati = []

    nr_zile = calendar.monthrange(an, luna)[1]
    month_names = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
                   'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    # Sarbatori in luna
    sarbatori = set()
    sarb_q = SarbatoareLegala.query.filter_by(an=an).all()
    for s in sarb_q:
        if s.data.month == luna:
            sarbatori.add(s.data.day)

    # Weekend-uri
    weekends = set()
    for zi in range(1, nr_zile + 1):
        d = date(an, luna, zi)
        if d.weekday() >= 5:
            weekends.add(zi)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Foaie Colectiva'
    ws.page_setup.orientation = 'landscape'
    ws.page_setup.paperSize = ws.PAPERSIZE_A3
    ws.page_setup.fitToWidth = 1

    # === HEADER ===
    row = 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nr_zile + 5)
    cell = ws.cell(row=row, column=1, value='EDIFICO WORKFORCE SRL')
    cell.font = Font(bold=True, size=16, name='Arial', color='1A237E')
    cell.alignment = CENTER

    row = 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nr_zile + 5)
    cell = ws.cell(row=row, column=1, value='FOAIE COLECTIVA DE PREZENTA')
    cell.font = TITLE_FONT
    cell.alignment = CENTER

    row = 3
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=nr_zile + 5)
    cell = ws.cell(row=row, column=1,
                   value=f'Proiect: {proiect.cod_proiect} - {proiect.nume} | Locatie: {proiect.locatie or "-"} | Luna: {month_names[luna]} {an}')
    cell.font = SUBTITLE_FONT
    cell.alignment = CENTER

    # === HEADER TABEL ===
    row = 5
    headers = ['Nr.', 'Nume si Prenume', 'Functie']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, row, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    for zi in range(1, nr_zile + 1):
        col = zi + 3
        d = date(an, luna, zi)
        day_name = ['Lu', 'Ma', 'Mi', 'Jo', 'Vi', 'Sa', 'Du'][d.weekday()]
        cell = ws.cell(row=row, column=col, value=f'{zi}\n{day_name}')
        cell.font = Font(bold=True, size=8, color='FFFFFF', name='Arial')
        cell.fill = HEADER_FILL
        cell.alignment = WRAP_CENTER
        cell.border = THIN_BORDER

        if zi in weekends:
            cell.fill = PatternFill(start_color='455A64', end_color='455A64', fill_type='solid')
        if zi in sarbatori:
            cell.fill = PatternFill(start_color='C62828', end_color='C62828', fill_type='solid')

    total_col = nr_zile + 4
    semn_col = nr_zile + 5
    _style_data_cell(ws, row, total_col, 'TOTAL', font=HEADER_FONT, fill=HEADER_FILL)
    _style_data_cell(ws, row, semn_col, 'Semnatura', font=HEADER_FONT, fill=HEADER_FILL)

    ws.row_dimensions[row].height = 28

    # === DATE ANGAJATI ===
    totale_zile = {zi: 0 for zi in range(1, nr_zile + 1)}

    for i, ang in enumerate(angajati):
        r = row + 1 + i
        _style_data_cell(ws, r, 1, i + 1)
        _style_data_cell(ws, r, 2, ang.nume_complet, alignment=LEFT)
        _style_data_cell(ws, r, 3, ang.functie)

        # Pontaje luna
        pontaje = Pontaj.query.filter(
            Pontaj.angajat_id == ang.id,
            Pontaj.proiect_id == proiect_id,
            db.extract('month', Pontaj.data) == luna,
            db.extract('year', Pontaj.data) == an
        ).all()

        pontaje_dict = {p.data.day: p for p in pontaje}
        total_ore = 0

        for zi in range(1, nr_zile + 1):
            col = zi + 3
            fill = None
            val = ''

            if zi in sarbatori:
                val = 'SL'
                fill = SARBATOARE_FILL
            elif zi in weekends:
                fill = WEEKEND_FILL

            if zi in pontaje_dict:
                p = pontaje_dict[zi]
                if p.tip_zi == 'co':
                    val = 'CO'
                elif p.tip_zi == 'cm':
                    val = 'CM'
                elif p.tip_zi == 'invoiere':
                    val = 'INV'
                else:
                    ore = float(p.ore_lucrate or 0)
                    val = int(ore) if ore == int(ore) else ore
                    total_ore += ore
                    totale_zile[zi] += ore
                    supl = float(p.ore_suplimentare_50 or 0) + float(p.ore_suplimentare_100 or 0)
                    if supl > 0 and include_supl:
                        fill = SUPL_FILL

            cell = _style_data_cell(ws, r, col, val, fill=fill)
            if zi in weekends and not fill:
                cell.fill = WEEKEND_FILL

        _style_data_cell(ws, r, total_col, total_ore, font=TOTAL_FONT, fill=TOTAL_FILL)
        _style_data_cell(ws, r, semn_col, '')

    # === ROW TOTAL ===
    r_total = row + 1 + len(angajati)
    _style_data_cell(ws, r_total, 1, '', font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 2, 'TOTAL ORE / ZI', font=TOTAL_FONT, fill=TOTAL_FILL, alignment=LEFT)
    _style_data_cell(ws, r_total, 3, '', font=TOTAL_FONT, fill=TOTAL_FILL)

    grand_total = 0
    for zi in range(1, nr_zile + 1):
        col = zi + 3
        val = totale_zile[zi]
        grand_total += val
        _style_data_cell(ws, r_total, col, val if val > 0 else '', font=TOTAL_FONT, fill=TOTAL_FILL)

    _style_data_cell(ws, r_total, total_col, grand_total, font=TOTAL_FONT, fill=TOTAL_FILL)

    # === FOOTER ===
    r_footer = r_total + 2
    ws.cell(row=r_footer, column=1, value=f'Data generare: {datetime.now().strftime("%d.%m.%Y %H:%M")}').font = Font(size=8, italic=True, name='Arial')
    ws.cell(row=r_footer + 1, column=1, value='Semnatura Maistru: ___________________').font = Font(size=9, name='Arial')
    ws.cell(row=r_footer + 1, column=nr_zile // 2 + 3, value='Semnatura Manager Proiect: ___________________').font = Font(size=9, name='Arial')

    # Column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    for zi in range(1, nr_zile + 1):
        ws.column_dimensions[get_column_letter(zi + 3)].width = 4.5
    ws.column_dimensions[get_column_letter(total_col)].width = 8
    ws.column_dimensions[get_column_letter(semn_col)].width = 12

    ws.freeze_panes = 'D6'

    return wb


# ============================================================
# 2. STAT DE PLATA SIMPLIFICAT
# ============================================================

def generate_stat_plata(proiect_id, luna, an, include_bonusuri=False):
    """Genereaza Stat de Plata simplificat."""
    from models import db, Proiect, Angajat, AngajatProiect, Pontaj

    month_names = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
                   'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    wb = Workbook()
    ws = wb.active
    ws.title = 'Stat de Plata'
    ws.page_setup.orientation = 'landscape'

    # Header
    ws.merge_cells('A1:M1')
    ws.cell(row=1, column=1, value='EDIFICO WORKFORCE SRL').font = Font(bold=True, size=14, name='Arial', color='1A237E')
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:M2')
    titlu = f'STAT DE PLATA - {month_names[luna]} {an}'
    if proiect_id:
        proiect = Proiect.query.get(proiect_id)
        if proiect:
            titlu += f' | Proiect: {proiect.cod_proiect}'
    ws.cell(row=2, column=1, value=titlu).font = TITLE_FONT
    ws['A2'].alignment = CENTER

    # Headers tabel
    row = 4
    headers = ['Nr.', 'Nume si Prenume', 'Functie', 'Ore Normale', 'Ore Supl. 50%',
               'Ore Supl. 100%', 'Total Ore', 'Tarif Orar', 'Salariu Baza',
               'Spor 50%', 'Spor 100%', 'BRUT', 'Semnatura']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, row, c, h, font=HEADER_FONT, fill=HEADER_FILL)
    ws.row_dimensions[row].height = 30

    # Query angajati
    if proiect_id:
        asocieri = AngajatProiect.query.filter_by(proiect_id=proiect_id).all()
        angajati = [a.angajat for a in asocieri]
    else:
        angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()

    totals = {'ore_n': 0, 'ore_50': 0, 'ore_100': 0, 'ore_total': 0,
              'sal_baza': 0, 'spor_50': 0, 'spor_100': 0, 'brut': 0}

    for i, ang in enumerate(angajati):
        r = row + 1 + i
        query = Pontaj.query.filter(
            Pontaj.angajat_id == ang.id,
            db.extract('month', Pontaj.data) == luna,
            db.extract('year', Pontaj.data) == an,
            Pontaj.status == 'aprobat'
        )
        if proiect_id:
            query = query.filter(Pontaj.proiect_id == proiect_id)

        pontaje = query.all()

        ore_n = sum(float(p.ore_normale or 0) for p in pontaje)
        ore_50 = sum(float(p.ore_suplimentare_50 or 0) for p in pontaje)
        ore_100 = sum(float(p.ore_suplimentare_100 or 0) for p in pontaje)
        ore_total = ore_n + ore_50 + ore_100
        tarif = ang.tarif_orar

        sal_baza = round(ore_n * tarif, 2)
        spor_50 = round(ore_50 * tarif * 1.5, 2)
        spor_100 = round(ore_100 * tarif * 2.0, 2)
        brut = sal_baza + spor_50 + spor_100

        totals['ore_n'] += ore_n
        totals['ore_50'] += ore_50
        totals['ore_100'] += ore_100
        totals['ore_total'] += ore_total
        totals['sal_baza'] += sal_baza
        totals['spor_50'] += spor_50
        totals['spor_100'] += spor_100
        totals['brut'] += brut

        data_row = [i + 1, ang.nume_complet, ang.functie, ore_n, ore_50, ore_100,
                    ore_total, tarif, sal_baza, spor_50, spor_100, brut, '']
        for c, val in enumerate(data_row, 1):
            cell = _style_data_cell(ws, r, c, val)
            if c in (9, 10, 11, 12) and isinstance(val, (int, float)):
                cell.number_format = '#,##0.00'

    # Total row
    r_total = row + 1 + len(angajati)
    _style_data_cell(ws, r_total, 1, '', font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 2, 'TOTAL', font=TOTAL_FONT, fill=TOTAL_FILL, alignment=LEFT)
    _style_data_cell(ws, r_total, 3, '', font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 4, totals['ore_n'], font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 5, totals['ore_50'], font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 6, totals['ore_100'], font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 7, totals['ore_total'], font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 8, '', font=TOTAL_FONT, fill=TOTAL_FILL)
    for c, key in [(9, 'sal_baza'), (10, 'spor_50'), (11, 'spor_100'), (12, 'brut')]:
        cell = _style_data_cell(ws, r_total, c, totals[key], font=TOTAL_FONT, fill=TOTAL_FILL)
        cell.number_format = '#,##0.00'
    _style_data_cell(ws, r_total, 13, '', font=TOTAL_FONT, fill=TOTAL_FILL)

    # Footer semnaturi
    r_footer = r_total + 3
    ws.cell(row=r_footer, column=1, value='Intocmit: ___________________').font = Font(size=9, name='Arial')
    ws.cell(row=r_footer, column=5, value='Verificat: ___________________').font = Font(size=9, name='Arial')
    ws.cell(row=r_footer, column=9, value='Aprobat: ___________________').font = Font(size=9, name='Arial')

    _auto_width(ws, min_w=8, max_w=25)
    ws.column_dimensions['B'].width = 22
    ws.freeze_panes = 'A5'

    return wb


# ============================================================
# 3. SITUATIE PROIECT
# ============================================================

def generate_situatie_proiect(proiect_id, data_start=None, data_sfarsit=None, nivel='detaliat'):
    """Genereaza raport complet situatie proiect (multi-sheet)."""
    from models import db, Proiect, Angajat, AngajatProiect, Pontaj

    proiect = Proiect.query.get(proiect_id)
    if not proiect:
        raise ValueError('Proiectul nu a fost gasit.')

    if not data_start:
        data_start = proiect.data_start
    if not data_sfarsit:
        data_sfarsit = date.today()

    wb = Workbook()

    # === SHEET 1: SUMAR ===
    ws1 = wb.active
    ws1.title = 'Sumar Proiect'

    ws1.merge_cells('A1:F1')
    ws1.cell(row=1, column=1, value='SITUATIE PROIECT').font = TITLE_FONT
    ws1['A1'].alignment = CENTER

    info = [
        ('Cod Proiect:', proiect.cod_proiect),
        ('Denumire:', proiect.nume),
        ('Locatie:', proiect.locatie or '-'),
        ('Beneficiar:', proiect.beneficiar or '-'),
        ('Status:', proiect.status.upper()),
        ('Data Start:', proiect.data_start.strftime('%d.%m.%Y') if proiect.data_start else '-'),
        ('Data Sfarsit Plan:', proiect.data_sfarsit_planificat.strftime('%d.%m.%Y') if proiect.data_sfarsit_planificat else '-'),
        ('Buget Total:', f'{float(proiect.buget_total):,.2f} RON' if proiect.buget_total else '-'),
        ('Buget Manopera:', f'{float(proiect.buget_manopera):,.2f} RON' if proiect.buget_manopera else '-'),
        ('Perioada Raport:', f'{data_start.strftime("%d.%m.%Y")} - {data_sfarsit.strftime("%d.%m.%Y")}'),
    ]

    for i, (label, val) in enumerate(info):
        r = 3 + i
        ws1.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, name='Arial')
        ws1.cell(row=r, column=2, value=val).font = DATA_FONT

    # Totale
    pontaje = Pontaj.query.filter(
        Pontaj.proiect_id == proiect_id,
        Pontaj.data >= data_start,
        Pontaj.data <= data_sfarsit
    ).all()

    total_ore = sum(float(p.ore_lucrate or 0) for p in pontaje)
    total_normale = sum(float(p.ore_normale or 0) for p in pontaje)
    total_supl50 = sum(float(p.ore_suplimentare_50 or 0) for p in pontaje)
    total_supl100 = sum(float(p.ore_suplimentare_100 or 0) for p in pontaje)

    r_stats = 15
    ws1.cell(row=r_stats, column=1, value='STATISTICI PONTAJE').font = SUBTITLE_FONT
    stats_data = [
        ('Total Ore Lucrate:', total_ore),
        ('Ore Normale:', total_normale),
        ('Ore Supl. 50%:', total_supl50),
        ('Ore Supl. 100%:', total_supl100),
        ('Nr. Pontaje:', len(pontaje)),
    ]
    for i, (label, val) in enumerate(stats_data):
        r = r_stats + 1 + i
        ws1.cell(row=r, column=1, value=label).font = Font(bold=True, size=10, name='Arial')
        ws1.cell(row=r, column=2, value=val).font = DATA_FONT

    _auto_width(ws1)

    # === SHEET 2: ECHIPA ===
    ws2 = wb.create_sheet('Echipa')
    ws2.cell(row=1, column=1, value=f'ECHIPA PROIECT - {proiect.cod_proiect}').font = TITLE_FONT
    ws2.merge_cells('A1:H1')
    ws2['A1'].alignment = CENTER

    headers2 = ['Nr.', 'Nume', 'Functie', 'Total Ore', 'Ore Normale', 'Ore Supl.', 'Tarif Orar', 'Cost Total']
    for c, h in enumerate(headers2, 1):
        _style_data_cell(ws2, 3, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    asocieri = AngajatProiect.query.filter_by(proiect_id=proiect_id).all()
    grand_cost = 0
    for i, asoc in enumerate(asocieri):
        ang = asoc.angajat
        r = 4 + i
        p_ang = [p for p in pontaje if p.angajat_id == ang.id]
        ore = sum(float(p.ore_lucrate or 0) for p in p_ang)
        ore_n = sum(float(p.ore_normale or 0) for p in p_ang)
        ore_s = sum(float(p.ore_suplimentare_50 or 0) + float(p.ore_suplimentare_100 or 0) for p in p_ang)
        tarif = ang.tarif_orar
        cost = round(ore_n * tarif + sum(float(p.ore_suplimentare_50 or 0) for p in p_ang) * tarif * 1.5 +
                      sum(float(p.ore_suplimentare_100 or 0) for p in p_ang) * tarif * 2.0, 2)
        grand_cost += cost

        data_row = [i + 1, ang.nume_complet, ang.functie, ore, ore_n, ore_s, tarif, cost]
        for c, val in enumerate(data_row, 1):
            cell = _style_data_cell(ws2, r, c, val)
            if c in (7, 8):
                cell.number_format = '#,##0.00'

    r_total = 4 + len(asocieri)
    _style_data_cell(ws2, r_total, 2, 'TOTAL', font=TOTAL_FONT, fill=TOTAL_FILL, alignment=LEFT)
    cell = _style_data_cell(ws2, r_total, 8, grand_cost, font=TOTAL_FONT, fill=TOTAL_FILL)
    cell.number_format = '#,##0.00'
    for c in [1, 3, 4, 5, 6, 7]:
        _style_data_cell(ws2, r_total, c, '', font=TOTAL_FONT, fill=TOTAL_FILL)

    _auto_width(ws2)
    ws2.freeze_panes = 'A4'

    # === SHEET 3: PONTAJE DETALIATE ===
    if nivel == 'detaliat':
        ws3 = wb.create_sheet('Pontaje Detaliate')
        ws3.cell(row=1, column=1, value=f'PONTAJE DETALIATE - {proiect.cod_proiect}').font = TITLE_FONT
        ws3.merge_cells('A1:I1')
        ws3['A1'].alignment = CENTER

        headers3 = ['Nr.', 'Data', 'Angajat', 'Ora Start', 'Ora Sfarsit',
                     'Ore Lucrate', 'Ore Supl.', 'Tip Zi', 'Status']
        for c, h in enumerate(headers3, 1):
            _style_data_cell(ws3, 3, c, h, font=HEADER_FONT, fill=HEADER_FILL)

        pontaje_sorted = sorted(pontaje, key=lambda p: (p.data, p.angajat.nume if p.angajat else ''))
        for i, p in enumerate(pontaje_sorted):
            r = 4 + i
            supl = float(p.ore_suplimentare_50 or 0) + float(p.ore_suplimentare_100 or 0)
            data_row = [
                i + 1,
                p.data.strftime('%d.%m.%Y') if p.data else '-',
                p.angajat.nume_complet if p.angajat else '-',
                p.ora_start or '-',
                p.ora_sfarsit or '-',
                float(p.ore_lucrate or 0),
                supl,
                p.tip_zi or '-',
                p.status or '-',
            ]
            for c, val in enumerate(data_row, 1):
                _style_data_cell(ws3, r, c, val)

        _auto_width(ws3)
        ws3.freeze_panes = 'A4'

    return wb


# ============================================================
# 4. CENTRALIZATOR ORE LUNARE
# ============================================================

def generate_centralizator_ore(luna, an, grupare='angajat'):
    """Genereaza centralizator ore lunare cu pivot angajati vs zile."""
    from models import db, Angajat, Proiect, Pontaj, AngajatProiect

    nr_zile = calendar.monthrange(an, luna)[1]
    month_names = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
                   'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    wb = Workbook()

    if grupare == 'proiect':
        proiecte = Proiect.query.filter(Proiect.status.in_(['activ', 'planificat'])).all()
        for proj in proiecte:
            ws = wb.create_sheet(title=proj.cod_proiect[:31])
            _build_centralizator_sheet(ws, proj.cod_proiect, luna, an, nr_zile,
                                       month_names, proiect_id=proj.id)
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
    else:
        ws = wb.active
        ws.title = f'Centralizator {month_names[luna][:3]} {an}'
        _build_centralizator_sheet(ws, 'TOATE PROIECTELE', luna, an, nr_zile, month_names)

    # Sheet Total
    ws_total = wb.create_sheet('Total General')
    ws_total.cell(row=1, column=1, value=f'CENTRALIZATOR ORE - {month_names[luna]} {an}').font = TITLE_FONT
    ws_total.merge_cells('A1:F1')
    ws_total['A1'].alignment = CENTER

    from models import Angajat, Pontaj
    headers_t = ['Nr.', 'Angajat', 'Functie', 'Ore Normale', 'Ore Suplimentare', 'TOTAL']
    for c, h in enumerate(headers_t, 1):
        _style_data_cell(ws_total, 3, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    gt_n, gt_s, gt_t = 0, 0, 0
    for i, ang in enumerate(angajati):
        pontaje = Pontaj.query.filter(
            Pontaj.angajat_id == ang.id,
            db.extract('month', Pontaj.data) == luna,
            db.extract('year', Pontaj.data) == an
        ).all()
        ore_n = sum(float(p.ore_normale or 0) for p in pontaje)
        ore_s = sum(float(p.ore_suplimentare_50 or 0) + float(p.ore_suplimentare_100 or 0) for p in pontaje)
        total = ore_n + ore_s
        gt_n += ore_n
        gt_s += ore_s
        gt_t += total

        _style_data_cell(ws_total, 4 + i, 1, i + 1)
        _style_data_cell(ws_total, 4 + i, 2, ang.nume_complet, alignment=LEFT)
        _style_data_cell(ws_total, 4 + i, 3, ang.functie)
        _style_data_cell(ws_total, 4 + i, 4, ore_n)
        _style_data_cell(ws_total, 4 + i, 5, ore_s)
        _style_data_cell(ws_total, 4 + i, 6, total, font=TOTAL_FONT)

    r_tot = 4 + len(angajati)
    _style_data_cell(ws_total, r_tot, 2, 'TOTAL', font=TOTAL_FONT, fill=TOTAL_FILL, alignment=LEFT)
    _style_data_cell(ws_total, r_tot, 4, gt_n, font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws_total, r_tot, 5, gt_s, font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws_total, r_tot, 6, gt_t, font=TOTAL_FONT, fill=TOTAL_FILL)
    for c in [1, 3]:
        _style_data_cell(ws_total, r_tot, c, '', font=TOTAL_FONT, fill=TOTAL_FILL)

    _auto_width(ws_total)

    return wb


def _build_centralizator_sheet(ws, titlu, luna, an, nr_zile, month_names, proiect_id=None):
    """Helper: construieste un sheet centralizator cu pivot angajati vs zile."""
    from models import db, Angajat, Pontaj

    ws.cell(row=1, column=1, value=f'CENTRALIZATOR ORE - {titlu} - {month_names[luna]} {an}').font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=nr_zile + 4)
    ws['A1'].alignment = CENTER

    row = 3
    headers = ['Nr.', 'Angajat', 'Functie']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, row, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    for zi in range(1, nr_zile + 1):
        col = zi + 3
        d = date(an, luna, zi)
        cell = _style_data_cell(ws, row, col, zi, font=HEADER_FONT, fill=HEADER_FILL)
        if d.weekday() >= 5:
            cell.fill = PatternFill(start_color='455A64', end_color='455A64', fill_type='solid')

    _style_data_cell(ws, row, nr_zile + 4, 'TOTAL', font=HEADER_FONT, fill=HEADER_FILL)

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()

    for i, ang in enumerate(angajati):
        r = row + 1 + i
        _style_data_cell(ws, r, 1, i + 1)
        _style_data_cell(ws, r, 2, ang.nume_complet, alignment=LEFT)
        _style_data_cell(ws, r, 3, ang.functie)

        query = Pontaj.query.filter(
            Pontaj.angajat_id == ang.id,
            db.extract('month', Pontaj.data) == luna,
            db.extract('year', Pontaj.data) == an
        )
        if proiect_id:
            query = query.filter(Pontaj.proiect_id == proiect_id)

        pontaje = query.all()
        pontaje_dict = {p.data.day: p for p in pontaje}
        total = 0

        for zi in range(1, nr_zile + 1):
            col = zi + 3
            if zi in pontaje_dict:
                ore = float(pontaje_dict[zi].ore_lucrate or 0)
                total += ore
                _style_data_cell(ws, r, col, ore if ore > 0 else '')
            else:
                _style_data_cell(ws, r, col, '')
                d = date(an, luna, zi)
                if d.weekday() >= 5:
                    ws.cell(row=r, column=col).fill = WEEKEND_FILL

        _style_data_cell(ws, r, nr_zile + 4, total, font=TOTAL_FONT)

    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 14
    for zi in range(1, nr_zile + 1):
        ws.column_dimensions[get_column_letter(zi + 3)].width = 4.5
    ws.column_dimensions[get_column_letter(nr_zile + 4)].width = 8
    ws.freeze_panes = 'D4'


# ============================================================
# 5. RAPORT DOCUMENTE
# ============================================================

def generate_raport_documente(tip_raport='toate', functie_filter=None):
    """Genereaza raport documente cu status si alerte."""
    from models import db, Document, Angajat

    wb = Workbook()
    ws = wb.active
    ws.title = 'Documente'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value='RAPORT DOCUMENTE ANGAJATI').font = TITLE_FONT
    ws['A1'].alignment = CENTER

    ws.merge_cells('A2:J2')
    ws.cell(row=2, column=1, value=f'Generat: {datetime.now().strftime("%d.%m.%Y %H:%M")} | Filtru: {tip_raport}').font = Font(size=9, italic=True, name='Arial')
    ws['A2'].alignment = CENTER

    headers = ['Nr.', 'Angajat', 'Functie', 'Tip Document', 'Denumire', 'Serie/Nr',
               'Emitent', 'Data Emitere', 'Data Expirare', 'Status']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, 4, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    query = Document.query.join(Angajat)
    if functie_filter:
        query = query.filter(Angajat.functie == functie_filter)
    if tip_raport == 'expirate':
        query = query.filter(Document.status == 'expirat')
    elif tip_raport == 'in_curand':
        query = query.filter(Document.status == 'in_curand')

    docs = query.order_by(Angajat.nume, Document.tip).all()

    TIPURI_DICT = dict(Document.TIPURI)

    for i, doc in enumerate(docs):
        r = 5 + i
        status = doc.status_calculat
        data_row = [
            i + 1,
            doc.angajat.nume_complet if doc.angajat else '-',
            doc.angajat.functie if doc.angajat else '-',
            TIPURI_DICT.get(doc.tip, doc.tip),
            doc.nume_document,
            doc.serie_numar or '-',
            doc.emitent or '-',
            doc.data_emitere.strftime('%d.%m.%Y') if doc.data_emitere else '-',
            doc.data_expirare.strftime('%d.%m.%Y') if doc.data_expirare else 'Permanent',
            status.upper().replace('_', ' '),
        ]

        for c, val in enumerate(data_row, 1):
            fill = None
            if status == 'expirat':
                fill = RED_FILL
            elif status == 'in_curand':
                fill = YELLOW_FILL
            elif status == 'valabil':
                fill = GREEN_FILL
            _style_data_cell(ws, r, c, val, fill=fill)

    _auto_width(ws)
    ws.freeze_panes = 'A5'

    return wb


# ============================================================
# 6. PONTAJ INDIVIDUAL
# ============================================================

def generate_pontaj_individual(angajat_id, data_start, data_sfarsit):
    """Genereaza fisa pontaj individual per angajat."""
    from models import db, Angajat, Pontaj, Proiect

    angajat = Angajat.query.get(angajat_id)
    if not angajat:
        raise ValueError('Angajatul nu a fost gasit.')

    wb = Workbook()
    ws = wb.active
    ws.title = 'Pontaj Individual'

    ws.merge_cells('A1:I1')
    ws.cell(row=1, column=1, value='FISA PONTAJ INDIVIDUAL').font = TITLE_FONT
    ws['A1'].alignment = CENTER

    info = [
        ('Angajat:', angajat.nume_complet),
        ('Functie:', angajat.functie),
        ('Perioada:', f'{data_start.strftime("%d.%m.%Y")} - {data_sfarsit.strftime("%d.%m.%Y")}'),
    ]
    for i, (label, val) in enumerate(info):
        ws.cell(row=3 + i, column=1, value=label).font = Font(bold=True, size=10, name='Arial')
        ws.cell(row=3 + i, column=2, value=val).font = DATA_FONT

    row = 7
    headers = ['Nr.', 'Data', 'Proiect', 'Ora Start', 'Ora Sfarsit', 'Ore Lucrate',
               'Ore Supl. 50%', 'Ore Supl. 100%', 'Tip Zi']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, row, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    pontaje = Pontaj.query.filter(
        Pontaj.angajat_id == angajat_id,
        Pontaj.data >= data_start,
        Pontaj.data <= data_sfarsit
    ).order_by(Pontaj.data).all()

    totals = {'ore': 0, 'ore_n': 0, 'ore_50': 0, 'ore_100': 0}

    for i, p in enumerate(pontaje):
        r = row + 1 + i
        ore = float(p.ore_lucrate or 0)
        ore_50 = float(p.ore_suplimentare_50 or 0)
        ore_100 = float(p.ore_suplimentare_100 or 0)
        totals['ore'] += ore
        totals['ore_50'] += ore_50
        totals['ore_100'] += ore_100

        data_row = [
            i + 1,
            p.data.strftime('%d.%m.%Y') if p.data else '-',
            p.proiect.cod_proiect if p.proiect else '-',
            p.ora_start or '-',
            p.ora_sfarsit or '-',
            ore, ore_50, ore_100,
            p.tip_zi or '-',
        ]
        for c, val in enumerate(data_row, 1):
            _style_data_cell(ws, r, c, val)

    # Total
    r_total = row + 1 + len(pontaje)
    _style_data_cell(ws, r_total, 2, 'TOTAL', font=TOTAL_FONT, fill=TOTAL_FILL, alignment=LEFT)
    _style_data_cell(ws, r_total, 6, totals['ore'], font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 7, totals['ore_50'], font=TOTAL_FONT, fill=TOTAL_FILL)
    _style_data_cell(ws, r_total, 8, totals['ore_100'], font=TOTAL_FONT, fill=TOTAL_FILL)
    for c in [1, 3, 4, 5, 9]:
        _style_data_cell(ws, r_total, c, '', font=TOTAL_FONT, fill=TOTAL_FILL)

    _auto_width(ws)
    ws.freeze_panes = 'A8'

    return wb


# ============================================================
# 7. PREZENTA ZILNICA
# ============================================================

def generate_prezenta_zilnica(data_zi, proiect_id=None):
    """Genereaza raport prezenta zilnica."""
    from models import db, Pontaj, Angajat, Proiect

    wb = Workbook()
    ws = wb.active
    ws.title = 'Prezenta Zilnica'

    day_names = ['Luni', 'Marti', 'Miercuri', 'Joi', 'Vineri', 'Sambata', 'Duminica']
    zi_sapt = day_names[data_zi.weekday()]

    ws.merge_cells('A1:H1')
    ws.cell(row=1, column=1, value=f'RAPORT PREZENTA ZILNICA - {data_zi.strftime("%d.%m.%Y")} ({zi_sapt})').font = TITLE_FONT
    ws['A1'].alignment = CENTER

    row = 3
    headers = ['Nr.', 'Angajat', 'Functie', 'Proiect', 'Ora Start', 'Ora Sfarsit', 'Ore Lucrate', 'Status']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, row, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    query = Pontaj.query.filter(Pontaj.data == data_zi)
    if proiect_id:
        query = query.filter(Pontaj.proiect_id == proiect_id)

    pontaje = query.join(Angajat).order_by(Angajat.nume).all()
    total_ore = 0

    for i, p in enumerate(pontaje):
        r = row + 1 + i
        ore = float(p.ore_lucrate or 0)
        total_ore += ore
        data_row = [
            i + 1,
            p.angajat.nume_complet if p.angajat else '-',
            p.angajat.functie if p.angajat else '-',
            p.proiect.cod_proiect if p.proiect else '-',
            p.ora_start or '-',
            p.ora_sfarsit or '-',
            ore,
            p.status or '-',
        ]
        for c, val in enumerate(data_row, 1):
            _style_data_cell(ws, r, c, val)

    r_total = row + 1 + len(pontaje)
    _style_data_cell(ws, r_total, 2, f'TOTAL: {len(pontaje)} angajati', font=TOTAL_FONT, fill=TOTAL_FILL, alignment=LEFT)
    _style_data_cell(ws, r_total, 7, total_ore, font=TOTAL_FONT, fill=TOTAL_FILL)
    for c in [1, 3, 4, 5, 6, 8]:
        _style_data_cell(ws, r_total, c, '', font=TOTAL_FONT, fill=TOTAL_FILL)

    _auto_width(ws)

    return wb


# ============================================================
# 8. RAPORT SSM
# ============================================================

def generate_raport_ssm(tip_document=None, status_filter=None):
    """Genereaza raport SSM - instructaje si autorizatii."""
    from models import db, Document, Angajat

    TIPURI_SSM = ['instructaj_SSM', 'fisa_aptitudini', 'autorizatie_ISCIR', 'permis_inaltime', 'certificat_calificare']
    TIPURI_DICT = dict(Document.TIPURI)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Raport SSM'

    ws.merge_cells('A1:J1')
    ws.cell(row=1, column=1, value='RAPORT SSM - INSTRUCTAJE SI AUTORIZATII').font = TITLE_FONT
    ws['A1'].alignment = CENTER

    headers = ['Nr.', 'Angajat', 'Functie', 'Tip Document', 'Denumire', 'Serie/Nr',
               'Emitent', 'Data Emitere', 'Data Expirare', 'Status']
    for c, h in enumerate(headers, 1):
        _style_data_cell(ws, 3, c, h, font=HEADER_FONT, fill=HEADER_FILL)

    query = Document.query.join(Angajat).filter(Document.tip.in_(TIPURI_SSM))
    if tip_document and tip_document in TIPURI_SSM:
        query = query.filter(Document.tip == tip_document)
    if status_filter == 'expirat':
        query = query.filter(Document.status == 'expirat')
    elif status_filter == 'valabil':
        query = query.filter(Document.status == 'valabil')

    docs = query.order_by(Angajat.nume, Document.tip).all()

    for i, doc in enumerate(docs):
        r = 4 + i
        status = doc.status_calculat
        fill = RED_FILL if status == 'expirat' else (YELLOW_FILL if status == 'in_curand' else GREEN_FILL)
        data_row = [
            i + 1,
            doc.angajat.nume_complet if doc.angajat else '-',
            doc.angajat.functie if doc.angajat else '-',
            TIPURI_DICT.get(doc.tip, doc.tip),
            doc.nume_document,
            doc.serie_numar or '-',
            doc.emitent or '-',
            doc.data_emitere.strftime('%d.%m.%Y') if doc.data_emitere else '-',
            doc.data_expirare.strftime('%d.%m.%Y') if doc.data_expirare else 'Permanent',
            status.upper().replace('_', ' '),
        ]
        for c, val in enumerate(data_row, 1):
            _style_data_cell(ws, r, c, val, fill=fill)

    _auto_width(ws)
    ws.freeze_panes = 'A4'

    return wb
