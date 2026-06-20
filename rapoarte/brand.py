"""
EDIFICO WORKFORCE - Branding centralizat pentru exporturi (Rapoarte Faza 1).

Un singur loc pentru paleta de brand Edifico + stilurile partajate intre
generatoarele Excel (openpyxl) si PDF (reportlab). Scopul: exporturile vechi
(rapoarte/excel_generator.py, rapoarte/pdf_generator.py) si cele noi
(services/situatii.py etc) folosesc EXACT aceleasi culori si acelasi header.

Paleta (din brand identity Edifico):
  - Navy obsidian  #0B1426  (titluri, fundal header tabel)
  - Champagne gold #C9A961  (accent header tabel, linii)
  - Cream           #F5F1E8  (fundal randuri totale / zebra deschisa)

Tipografie:
  - PDF: incearca sa inregistreze un font serif "imperial" (Cinzel) daca un
    fisier .ttf e gasit in repo / sistem. Daca nu, fallback DOCUMENTAT la
    'Times-Roman' / 'Times-Bold' (serif standard ReportLab, fara crash).
  - Excel: nu suporta embedding de font; folosim numele 'Cinzel' pentru titlu
    (cade gratios pe serif-ul implicit daca fontul nu e instalat pe masina ce
    deschide fisierul) si 'Arial' pentru corp.

Aditiv: nu rupe semnaturile functiilor existente. Generatoarele pot importa
constantele si helperele de aici fara modificari de structura.
"""

from __future__ import annotations

import logging
import os

_logger = logging.getLogger(__name__)


# ============================================================
# PALETA EDIFICO (hex fara '#', pt openpyxl; cu '#' pt reportlab)
# ============================================================

NAVY = '0B1426'      # navy obsidian
GOLD = 'C9A961'      # champagne gold
CREAM = 'F5F1E8'     # cream / paper white
WHITE = 'FFFFFF'

# Nuante derivate (zebra / stari) - tinute neutre, in armonie cu paleta
ZEBRA_LIGHT = 'FAF8F2'   # crem foarte deschis pentru randuri alternante
GREY_LINE = 'CCCCCC'     # linii subtiri de grila
WEEKEND = 'E8E4DA'       # weekend (crem-gri, nu albastrul vechi)
SARBATOARE = 'A8893D'    # capatul inchis al gradientului gold (sarbatori)

# Stari documente (pastram semantica rosu/galben/verde, dar tonuri calme)
STARE_EXPIRAT = 'F0C9C2'
STARE_IN_CURAND = 'F2E2B8'
STARE_VALABIL = 'CFE3C9'

# Cu prefix '#' pentru reportlab.lib.colors.HexColor
NAVY_HEX = '#' + NAVY
GOLD_HEX = '#' + GOLD
CREAM_HEX = '#' + CREAM
ZEBRA_LIGHT_HEX = '#' + ZEBRA_LIGHT
GREY_LINE_HEX = '#' + GREY_LINE
WEEKEND_HEX = '#' + WEEKEND
SARBATOARE_HEX = '#' + SARBATOARE

# Fonturi (nume logice; rezolvarea reala pt PDF se face in get_pdf_fonts)
EXCEL_TITLE_FONT = 'Cinzel'
EXCEL_BODY_FONT = 'Arial'


# ============================================================
# PDF: inregistrare font serif (Cinzel daca exista, altfel Times)
# ============================================================

# Locatii candidate pentru un .ttf serif "imperial". Daca gasim Cinzel il
# folosim; altfel ramanem pe fallback-ul ReportLab built-in 'Times-*'.
_CINZEL_CANDIDATE_PATHS = [
    # in repo (daca cineva adauga fontul ulterior)
    'static/fonts/Cinzel-Regular.ttf',
    'static/fonts/Cinzel-Bold.ttf',
    'static/img/fonts/Cinzel-Regular.ttf',
    # sistem macOS / linux uzual
    '/Library/Fonts/Cinzel-Regular.ttf',
    '/usr/share/fonts/truetype/cinzel/Cinzel-Regular.ttf',
]

# Cache rezultat (nume_font_normal, nume_font_bold). Calculat o singura data.
_pdf_fonts_cache: tuple[str, str] | None = None


def _basedir() -> str:
    """Radacina proiectului (un nivel peste pachetul rapoarte/)."""
    return os.path.dirname(os.path.dirname(os.path.abspath(__file__)))


def get_pdf_fonts() -> tuple[str, str]:
    """
    Returneaza (font_normal, font_bold) pentru titluri/headere PDF.

    Incearca sa inregistreze Cinzel (regular + bold) daca gaseste fisierele.
    Daca nu reuseste (lipsa fisier sau reportlab indisponibil), cade pe
    serif-ul standard ReportLab ('Times-Roman' / 'Times-Bold') - garantat
    prezent, fara crash. Idempotent (rezultat cache-uit).
    """
    global _pdf_fonts_cache
    if _pdf_fonts_cache is not None:
        return _pdf_fonts_cache

    fallback = ('Times-Roman', 'Times-Bold')

    try:
        from reportlab.pdfbase import pdfmetrics
        from reportlab.pdfbase.ttfonts import TTFont
    except Exception:
        _pdf_fonts_cache = fallback
        return fallback

    base = _basedir()
    reg_path = None
    bold_path = None
    for rel in _CINZEL_CANDIDATE_PATHS:
        path = rel if os.path.isabs(rel) else os.path.join(base, rel)
        if os.path.exists(path):
            low = path.lower()
            if 'bold' in low and bold_path is None:
                bold_path = path
            elif reg_path is None:
                reg_path = path

    if not reg_path:
        _pdf_fonts_cache = fallback
        return fallback

    try:
        pdfmetrics.registerFont(TTFont('EdificoSerif', reg_path))
        bold_name = 'EdificoSerif'
        if bold_path:
            pdfmetrics.registerFont(TTFont('EdificoSerif-Bold', bold_path))
            bold_name = 'EdificoSerif-Bold'
        _pdf_fonts_cache = ('EdificoSerif', bold_name)
        _logger.info('Font brand PDF inregistrat din %s', reg_path)
    except Exception as e:
        _logger.warning('Inregistrare font Cinzel a esuat (%s); fallback Times.', e)
        _pdf_fonts_cache = fallback

    return _pdf_fonts_cache


# ============================================================
# EXCEL: stiluri partajate (openpyxl)
# ============================================================

def excel_styles():
    """
    Returneaza un dict cu stiluri openpyxl gata de folosit, in paleta Edifico.

    Chei: title_font, header_font, header_fill, subtitle_font, data_font,
    total_font, total_fill, thin_border, center, left, wrap_center.
    Import lenes (openpyxl poate lipsi in unele contexte de test).
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    thin = Side(style='thin', color=GREY_LINE)
    return {
        'title_font': Font(bold=True, size=14, name=EXCEL_TITLE_FONT, color=NAVY),
        'header_font': Font(bold=True, color=NAVY, size=10, name=EXCEL_BODY_FONT),
        'header_fill': PatternFill('solid', fgColor=GOLD),
        'subtitle_font': Font(bold=True, size=11, name=EXCEL_BODY_FONT, color=NAVY),
        'data_font': Font(size=9, name=EXCEL_BODY_FONT),
        'total_font': Font(bold=True, size=10, name=EXCEL_BODY_FONT, color=NAVY),
        'total_fill': PatternFill('solid', fgColor=CREAM),
        'thin_border': Border(left=thin, right=thin, top=thin, bottom=thin),
        'center': Alignment(horizontal='center', vertical='center'),
        'left': Alignment(horizontal='left', vertical='center'),
        'wrap_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
    }


# ============================================================
# PDF: stiluri + header brandat (reportlab)
# ============================================================

def pdf_paragraph_styles():
    """
    Returneaza un getSampleStyleSheet ReportLab imbogatit cu stilurile de brand
    Edifico: TitleCustom, SubtitleCustom, InfoLabel, Footer.

    Reutilizat de generatoarele PDF. Foloseste fontul serif brand (Cinzel daca
    e disponibil, altfel Times).
    """
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER

    serif, serif_bold = get_pdf_fonts()
    styles = getSampleStyleSheet()

    def _add(name, **kw):
        # Evita KeyError daca un stil cu acelasi nume e deja inregistrat.
        if name in styles.byName:
            st = styles[name]
            for k, v in kw.items():
                setattr(st, k, v)
        else:
            styles.add(ParagraphStyle(name=name, **kw))

    _add('TitleCustom', fontSize=16, fontName=serif_bold,
         textColor=colors.HexColor(NAVY_HEX), alignment=TA_CENTER, spaceAfter=12)
    _add('SubtitleCustom', fontSize=11, fontName=serif_bold,
         textColor=colors.HexColor(NAVY_HEX), alignment=TA_CENTER, spaceAfter=8)
    _add('InfoLabel', fontSize=9, fontName='Helvetica-Bold',
         textColor=colors.HexColor(NAVY_HEX))
    _add('Footer', fontSize=7, fontName='Helvetica-Oblique',
         textColor=colors.gray, alignment=TA_CENTER)
    return styles


def pdf_table_style():
    """
    TableStyle ReportLab partajat in paleta Edifico:
    - header: fundal gold, text navy, font serif bold
    - corp: zebra crem deschis, grila subtire
    Returneaza un TableStyle nou (nu-l muta intre tabele).
    """
    from reportlab.lib import colors
    from reportlab.platypus import TableStyle

    serif, serif_bold = get_pdf_fonts()
    return TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor(GOLD_HEX)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor(NAVY_HEX)),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), serif_bold),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor(GREY_LINE_HEX)),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1),
         [colors.white, colors.HexColor(ZEBRA_LIGHT_HEX)]),
        ('TOPPADDING', (0, 0), (-1, -1), 3),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
    ])


# Locatii candidate pentru un logo PNG (reportlab embed-uie PNG nativ prin
# Pillow, deja disponibil - fara dependinta noua). SVG-ul de brand NU poate fi
# embed-uit fara svglib, deci preferam un PNG. Daca niciunul nu exista,
# pdf_logo_flowable() returneaza None (header doar text, fara crash).
_LOGO_CANDIDATE_PATHS = [
    'static/img/edifico-logo.png',
    'static/img/edifico-mark.png',
    'static/img/pwa/icon-192.png',
    'static/img/pwa/icon-144.png',
]


def pdf_logo_flowable(max_mm: float = 16.0):
    """
    Returneaza un flowable Image cu logoul Edifico (PNG) la o inaltime de ~max_mm,
    sau None daca nu exista PNG / reportlab indisponibil.

    Defensiv: orice eroare la deschiderea/scalarea imaginii -> None (header
    cade pe wordmark text, fara crash). Nu introduce dependinte noi (PNG e
    suportat nativ de reportlab via Pillow).
    """
    try:
        from reportlab.platypus import Image
        from reportlab.lib.units import mm
    except Exception:
        return None

    base = _basedir()
    path = None
    for rel in _LOGO_CANDIDATE_PATHS:
        cand = rel if os.path.isabs(rel) else os.path.join(base, rel)
        if os.path.exists(cand):
            path = cand
            break
    if not path:
        return None

    try:
        img = Image(path)
        # Scaleaza proportional la inaltimea ceruta.
        iw, ih = img.imageWidth, img.imageHeight
        if ih <= 0:
            return None
        target_h = max_mm * mm
        ratio = target_h / float(ih)
        img.drawHeight = target_h
        img.drawWidth = iw * ratio
        img.hAlign = 'CENTER'
        return img
    except Exception as e:
        _logger.warning('Logo PDF a esuat (%s); header doar text.', e)
        return None


def pdf_header_elements(titlu, subtitlu=None, detaliu=None, cu_logo=False):
    """
    Construieste lista de elemente flowable pentru un header PDF brandat Edifico
    (logo optional + wordmark + titlu raport + linie de detaliu optionala).

    `cu_logo=True` adauga logoul PNG inaintea wordmark-ului daca un fisier exista
    (altfel e sarit gratios, fara crash).

    Returneaza o lista de Paragraph/Spacer gata de extins in `elements`.
    Generatoarele fac: elements += pdf_header_elements(...).
    """
    from reportlab.platypus import Paragraph, Spacer
    from reportlab.lib.units import mm

    styles = pdf_paragraph_styles()
    out = []
    if cu_logo:
        logo = pdf_logo_flowable()
        if logo is not None:
            out.append(logo)
            out.append(Spacer(1, 2 * mm))
    out.append(Paragraph('EDIFICO WORKFORCE SRL', styles['TitleCustom']))
    if titlu:
        out.append(Paragraph(titlu, styles['SubtitleCustom']))
    if subtitlu:
        out.append(Paragraph(subtitlu, styles['SubtitleCustom']))
    if detaliu:
        out.append(Paragraph(detaliu, styles['SubtitleCustom']))
    out.append(Spacer(1, 6 * mm))
    return out
