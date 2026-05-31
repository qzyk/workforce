"""
Motor de import F3 (XLSX / XLS binar / CSV / HTML / SpreadsheetML) - TOLERANT.

Obiective (Faza 1 - import robust pe devize reale):
- citeste TOATE sheet-urile (un F3 real are adesea cate un sheet per Obiect)
- detecteaza formatul dupa CONTINUT (magic bytes), nu doar dupa extensie
  (rezolva "File is not a zip file": .xls binar / HTML / XML deghizat in .xlsx)
- gaseste randul de antet ORIUNDE in primele randuri (nu doar primul rand),
  tolerant: are nevoie de `denumire` + (`um` SAU `cantitate`); `cod` e OPTIONAL
- mapare coloane order-independent, pe scor (specific inainte de generic) ->
  rezolva coliziunea "Capitol de lucrari" (=denumire) vs "categorie"
- `obiect` din numele sheet-ului daca nu e coloana; `tronson`/`categorie` din
  randurile-titlu de sectiune (ex: "1. HIDRANTI INTERIORI") cand nu sunt coloane
- ignora randurile de numerotare / pret / total / nota / disclaimer; pastreaza
  articolele neclasificabile (NECLASIFICAT) -> importul NU mai esueaza

Intoarce (articole: list[ArticolF3], raport: dict cu statistici si avertismente).
"""
from __future__ import annotations

import csv
import io
import os
import re
from typing import Iterable, Optional

from .modele import ArticolF3, _to_float
from .normalizare import normalizeaza, normalizeaza_cheie
from .config_loader import SETARI_IMPLICITE


class EroareImport(Exception):
    """Eroare fatala de import (format necunoscut, lipsa antet in toate sheet-urile)."""


# ----------------------------------------------------------------------------
# Maparea coloanelor: scor best-match, order-independent.
# Fiecare coloana primeste campul logic cu cel mai mare scor; intre campuri,
# fiecare camp pastreaza cea mai buna coloana a lui. Sinonimele mai lungi
# (mai specifice) bat sinonimele scurte la egalitate de scor -> "capitol de
# lucrari" (denumire) nu mai e furat de "categorie".
# ----------------------------------------------------------------------------
def _scor_potrivire(h: str, sin: str) -> int:
    """0=fara, 1=contine, 2=prefix/sufix, 3=egalitate (pe text normalizat)."""
    if not h or not sin:
        return 0
    if h == sin:
        return 3
    if h.startswith(sin) or sin.startswith(h):
        return 2
    if sin in h or h in sin:
        return 1
    return 0


def _mapeaza_coloane(antet: list, coloane_cfg: dict) -> dict:
    """Construieste {camp_logic: index_coloana} pe baza sinonimelor din config."""
    antet_norm = [normalizeaza(c) for c in antet]
    # pentru fiecare coloana: (camp_castigator, scor, lungime_sinonim)
    camp_col: dict = {}
    for col, h in enumerate(antet_norm):
        if not h:
            continue
        best_camp, best_scor, best_len = None, 0, 0
        for camp, sinonime in coloane_cfg.items():
            for s in sinonime:
                sn = normalizeaza(s)
                sc = _scor_potrivire(h, sn)
                if sc == 0:
                    continue
                if sc > best_scor or (sc == best_scor and len(sn) > best_len):
                    best_camp, best_scor, best_len = camp, sc, len(sn)
        if best_camp is not None:
            camp_col[col] = (best_camp, best_scor)
    # intoarce, pentru fiecare camp, coloana cu scorul cel mai mare
    harta: dict = {}
    scor_camp: dict = {}
    for col, (camp, sc) in camp_col.items():
        if camp not in harta or sc > scor_camp[camp]:
            harta[camp] = col
            scor_camp[camp] = sc
    return harta


# ----------------------------------------------------------------------------
# Citirea randurilor pe formate (multi-sheet unde are sens).
# ----------------------------------------------------------------------------
_MAGIC_OLE2 = b'\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1'  # .xls binar (OLE2/BIFF)


def _decode_best(continut: bytes) -> str:
    """Decodeaza bytes incercand mai multe codari (utf-8, Windows-1250 RO, latin-1)."""
    if continut[:3] == b'\xef\xbb\xbf':
        continut = continut[3:]
    for enc in ('utf-8', 'cp1250', 'iso-8859-2', 'latin-1'):
        try:
            return continut.decode(enc)
        except UnicodeDecodeError:
            continue
    return continut.decode('utf-8', errors='replace')


def _sheeturi_xlsx(continut: bytes) -> Iterable[tuple]:
    """(nume_sheet, randuri) pentru fiecare foaie dintr-un xlsx (read_only)."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(continut), read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            randuri = [list(r) for r in ws.iter_rows(values_only=True)]
            yield (ws.title, randuri)
    finally:
        wb.close()


def _sheeturi_xls_binar(continut: bytes) -> Iterable[tuple]:
    """(nume_sheet, randuri) pentru fiecare foaie dintr-un .xls binar (xlrd)."""
    try:
        import xlrd
    except ImportError:
        raise EroareImport(
            'Fisierul e un Excel binar vechi (.xls). Instaleaza "xlrd" '
            '(pe PythonAnywhere: ~/.virtualenvs/workforce-env/bin/pip install xlrd) '
            'sau re-salveaza fisierul ca .xlsx (Salvare ca -> Registru Excel) ori .csv.'
        )
    wb = xlrd.open_workbook(file_contents=continut)
    for sh in wb.sheets():
        randuri = [[sh.cell_value(r, c) for c in range(sh.ncols)]
                   for r in range(sh.nrows)]
        yield (sh.name, randuri)


def _randuri_csv(continut: bytes) -> Iterable[list]:
    """Genereaza randuri dintr-un CSV (detecteaza separatorul ; sau ,)."""
    text = continut.decode('utf-8-sig', errors='replace')
    proba = text[:4096]
    sep = ';' if proba.count(';') >= proba.count(',') else ','
    for row in csv.reader(io.StringIO(text), delimiter=sep):
        yield row


def _randuri_html(continut: bytes) -> Iterable[list]:
    """Parseaza un 'Excel' care e de fapt HTML (<table>) - export tipic de devize."""
    from html.parser import HTMLParser

    class _Tabel(HTMLParser):
        def __init__(self):
            super().__init__(convert_charrefs=True)
            self.randuri: list = []
            self._rand = None
            self._cel: list = []
            self._in_cel = False

        def handle_starttag(self, tag, attrs):
            t = tag.lower()
            if t == 'tr':
                self._rand = []
            elif t in ('td', 'th'):
                self._cel = []
                self._in_cel = True
            elif t == 'br' and self._in_cel:
                self._cel.append(' ')

        def handle_endtag(self, tag):
            t = tag.lower()
            if t in ('td', 'th') and self._rand is not None:
                self._rand.append(''.join(self._cel).strip())
                self._in_cel = False
            elif t == 'tr' and self._rand is not None:
                self.randuri.append(self._rand)
                self._rand = None

        def handle_data(self, data):
            if self._in_cel:
                self._cel.append(data)

    p = _Tabel()
    p.feed(_decode_best(continut))
    for r in p.randuri:
        yield r


def _sheeturi_spreadsheetml(continut: bytes) -> Iterable[tuple]:
    """(nume_sheet, randuri) din 'XML Spreadsheet 2003' (SpreadsheetML)."""
    import xml.etree.ElementTree as ET
    ns = '{urn:schemas-microsoft-com:office:spreadsheet}'
    root = ET.fromstring(_decode_best(continut))
    for ws in root.iter(f'{ns}Worksheet'):
        nume = ws.get(f'{ns}Name') or ''
        table = ws.find(f'{ns}Table')
        if table is None:
            continue
        randuri = []
        for row in table.findall(f'{ns}Row'):
            celule: list = []
            col = 0
            for cell in row.findall(f'{ns}Cell'):
                idx = cell.get(f'{ns}Index')  # poate sari peste coloane goale (1-based)
                if idx:
                    target = int(idx) - 1
                    while col < target:
                        celule.append('')
                        col += 1
                data = cell.find(f'{ns}Data')
                celule.append(data.text if (data is not None and data.text) else '')
                col += 1
            randuri.append(celule)
        yield (nume, randuri)


def _citeste_sheeturi(continut: bytes, ext: str) -> list:
    """Alege parserul potrivit dupa CONTINUT (magic bytes), apoi dupa extensie.
    Intoarce o lista de (nume_sheet, randuri)."""
    if not continut:
        raise EroareImport('Fisier gol.')
    cap = continut[:8]
    raw = continut[3:] if continut[:3] == b'\xef\xbb\xbf' else continut
    proba = raw[:8192].lstrip().lower()

    if cap[:2] == b'PK':                                     # zip -> .xlsx/.xlsm real
        return list(_sheeturi_xlsx(continut))
    if cap == _MAGIC_OLE2:                                   # OLE2 -> .xls binar vechi
        return list(_sheeturi_xls_binar(continut))
    if proba[:5] == b'<?xml' and b'spreadsheet' in proba:    # SpreadsheetML 2003
        return list(_sheeturi_spreadsheetml(continut))
    if proba[:1] == b'<' and (b'<table' in proba or b'<html' in proba
                              or b'<!doctype html' in proba):  # HTML <table> deghizat
        return [('', list(_randuri_html(continut)))]

    # Fara semnatura clara -> dupa extensie
    if ext == 'csv':
        return [('', list(_randuri_csv(continut)))]
    if ext in ('xlsx', 'xlsm'):
        return list(_sheeturi_xlsx(continut))
    if ext == 'xls':
        return list(_sheeturi_xls_binar(continut))
    return [('', list(_randuri_csv(continut)))]  # ultim fallback: text delimitat


def _mesaj_format(continut: bytes, e: Exception) -> str:
    """Mesaj prietenos in functie de formatul real detectat (pentru flash UI)."""
    cap = continut[:8] if continut else b''
    proba = (continut[3:] if continut[:3] == b'\xef\xbb\xbf' else continut)[:64].lstrip().lower()
    if cap == _MAGIC_OLE2:
        return ('Fisierul e un Excel binar vechi (.xls), nu .xlsx. Deschide-l in Excel '
                'si foloseste "Salvare ca -> Registru Excel (.xlsx)" sau salveaza ca .csv. '
                f'(detaliu tehnic: {e})')
    if proba[:1] == b'<':
        return ('Fisierul pare HTML/XML, nu un Excel real (export tipic de la softuri de '
                'devize). Deschide-l in Excel si "Salveaza ca -> .xlsx" sau .csv. '
                f'(detaliu tehnic: {e})')
    if cap[:2] != b'PK':
        return ('Formatul fisierului nu e un Excel valid (.xlsx trebuie sa fie arhiva zip). '
                f'Re-salveaza ca .xlsx (Registru Excel) sau .csv. (detaliu tehnic: {e})')
    return f'Nu pot citi fisierul Excel: {e}. Re-salveaza ca .xlsx sau .csv.'


# ----------------------------------------------------------------------------
# Heuristici de structura F3 (rand-articol vs rand-titlu de sectiune).
# ----------------------------------------------------------------------------
_UM_VALIDE = {
    'm', 'ml', 'm.l', 'mc', 'mp', 'm2', 'm3', 'mp.', 'mc.', 'km', 'cm', 'mm', 'dm',
    'kg', 'g', 'gr', 't', 'to', 'to.', 'tona', 'tone', 'kt',
    'buc', 'buc.', 'bc', 'bucata', 'bucati',
    'set', 'set.', 'ans', 'ans.', 'ansamblu', 'gar', 'garnitura', 'pereche',
    'l', 'litri', 'mii', 'h', 'ora', 'ore', 'zi', 'zile', 'luna', 'luni',
    'proc', 'cmp', 'dmc', 'cp',
}
_RX_UM = re.compile(
    r'^(m|ml|mp|mc|m2|m3|km|cm|mm|dm|kg|g|gr|t|to|kt|buc|bc|set|ans|gar|l|h|'
    r'ora|ore|zi|zile|mii|proc|cp|cmp|dmc|km)\.?$'
)
_RX_PREFIX_NUMERIC = re.compile(r'^\s*\d+([.\-)]\d*)*[.\-)\s]+')
_RX_NUMEROTARE = re.compile(r'^\d{1,2}$')
_RX_TITLU_NUM = re.compile(r'^[\sIVXLCDM0-9]+[.\)\-:]\s+')


def _um_valida(v) -> bool:
    """True daca valoarea arata ca o unitate de masura reala (nu un titlu)."""
    if v is None:
        return False
    s = normalizeaza(str(v)).replace(' ', '')
    if not s or len(s) > 10:
        return False
    return s.rstrip('.') in _UM_VALIDE or bool(_RX_UM.match(s))


def _are_numar(v) -> bool:
    """True daca v contine o cantitate numerica > 0."""
    if v is None or v == '':
        return False
    if isinstance(v, (int, float)):
        return float(v) > 0
    return _to_float(v) > 0


def _este_rand_numerotare(rand) -> bool:
    """True pentru randul '0 1 2 3 4 5 ...' care numeroteaza coloanele sub antet."""
    vals = [str(c).strip() for c in (rand or []) if c is not None and str(c).strip()]
    if len(vals) < 3:
        return False
    return all(_RX_NUMEROTARE.fullmatch(v) for v in vals)


def _curata_nume_obiect(nume) -> str:
    """Curata numele de sheet -> nume de obiect ('2.4 Obiect 2 - Hala' -> 'Obiect 2 - Hala')."""
    s = str(nume or '').strip()
    if not s:
        return ''
    s2 = _RX_PREFIX_NUMERIC.sub('', s).strip()
    return (s2 or s)[:120]


def _clasa_titlu(den_norm: str) -> str:
    """Clasifica un rand-titlu: 'obiect' | 'categorie' | 'sectiune'."""
    if re.match(r'^(obiectivul|obiectiv|obiectul|obiect|deviz)\b', den_norm):
        return 'obiect'
    if re.match(r'^(categoria|capitolul|capitol|cap\.)\b', den_norm):
        return 'categorie'
    return 'sectiune'


def _curata_titlu(text) -> str:
    """Scoate numerotarea / prefixele 'Obiectul:' etc. dintr-un titlu de sectiune."""
    s = str(text or '').strip()
    s = _RX_TITLU_NUM.sub('', s)
    s = re.sub(r'^(obiectivul|obiectul|obiectiv|obiect|categoria de lucrari|categoria|'
               r'capitolul de lucrari|capitolul|capitol)\b[:\s\-]*', '', s, flags=re.I)
    s = s.strip(' :-.\t')
    return s[:120] or str(text).strip()[:120]


def _combina_randuri(a: list, b: list) -> list:
    """Concateneaza pe coloane doua randuri (antet pe 2 linii)."""
    m = max(len(a or []), len(b or []))
    out = []
    for i in range(m):
        va = a[i] if a and i < len(a) else ''
        vb = b[i] if b and i < len(b) else ''
        va = '' if va is None else str(va).strip()
        vb = '' if vb is None else str(vb).strip()
        out.append((va + ' ' + vb).strip())
    return out


def _este_antet(harta: dict) -> bool:
    """Antet valid = avem 'denumire' + ('um' sau 'cantitate'). `cod` e optional."""
    return 'denumire' in harta and ('um' in harta or 'cantitate' in harta)


def _gaseste_antet(randuri: list, coloane_cfg: dict, max_scan: int = 25):
    """Gaseste randul de antet (tolerant). Intoarce (index_antet, harta) sau None."""
    n = min(len(randuri), max_scan)
    # pasul 1: rand singular
    for i in range(n):
        if not randuri[i]:
            continue
        harta = _mapeaza_coloane(randuri[i], coloane_cfg)
        if _este_antet(harta):
            return i, harta
    # pasul 2: antet pe 2 randuri (etichete sparte pe doua linii)
    for i in range(n - 1):
        harta = _mapeaza_coloane(_combina_randuri(randuri[i], randuri[i + 1]), coloane_cfg)
        if _este_antet(harta):
            return i + 1, harta   # articolele incep dupa al doilea rand de antet
    return None


def _val(rand: list, camp: str, harta: dict) -> str:
    i = harta.get(camp)
    if i is None or i >= len(rand):
        return ''
    v = rand[i]
    return '' if v is None else str(v).strip()


# ----------------------------------------------------------------------------
# Import principal.
# ----------------------------------------------------------------------------
def importa(continut: bytes, extensie: str, setari: Optional[dict] = None):
    """Importa un fisier F3 (toate sheet-urile, antet tolerant).

    Returns:
        (articole: list[ArticolF3], raport: dict cu statistici).
    """
    setari = setari or SETARI_IMPLICITE
    coloane_cfg = setari.get('coloane', SETARI_IMPLICITE['coloane'])
    ext = (extensie or '').lower().lstrip('.')

    try:
        sheeturi = _citeste_sheeturi(continut, ext)
    except EroareImport:
        raise
    except Exception as e:  # BadZipFile, XML parse, xlrd etc. -> mesaj prietenos
        raise EroareImport(_mesaj_format(continut, e)) from e

    if not sheeturi or all(not randuri for _, randuri in sheeturi):
        raise EroareImport('Fisier gol.')

    articole: list[ArticolF3] = []
    avertismente: list[str] = []
    chei_vazute: dict = {}
    nr_duplicate = 0
    nr_ignorate = 0
    nr_randuri_total = 0
    sheeturi_sarite: list = []
    coloane_mapate: Optional[dict] = None
    rand_antet_global: Optional[int] = None

    for nume_sheet, randuri in sheeturi:
        nr_randuri_total += len(randuri)
        if not randuri:
            continue
        rez = _gaseste_antet(randuri, coloane_cfg)
        if rez is None:
            if any(any(c is not None and str(c).strip() for c in r) for r in randuri):
                sheeturi_sarite.append(_curata_nume_obiect(nume_sheet) or '(sheet)')
            continue
        idx_antet, harta = rez
        if coloane_mapate is None:
            coloane_mapate = {k: antet_nume(randuri[idx_antet], v) for k, v in harta.items()}
            rand_antet_global = idx_antet + 1

        obiect_sheet = _curata_nume_obiect(nume_sheet)
        sectiune_curenta = ''   # tronson/sectiune din randuri-titlu
        obiect_titlu = ''       # obiect din randuri-titlu (ex "Obiectul: ...")
        in_note = False

        start = idx_antet + 1
        if start < len(randuri) and _este_rand_numerotare(randuri[start]):
            start += 1

        for off in range(start, len(randuri)):
            if in_note:
                break
            rand = randuri[off]
            nr_rand = off + 1
            if not rand or not any(c is not None and str(c).strip() for c in rand):
                continue  # rand gol

            den = _val(rand, 'denumire', harta)
            if not den:
                nr_ignorate += 1
                if len(avertismente) < 50:
                    avertismente.append(f'{obiect_sheet or "sheet"} rand {nr_rand}: '
                                        f'fara denumire - ignorat.')
                continue

            den_norm = normalizeaza(den)
            if den_norm.startswith('nota') or den_norm.startswith('note') \
                    or den_norm.startswith('observatii'):
                in_note = True          # disclaimerul de la finalul sheet-ului
                continue
            if den_norm.startswith('total') or den_norm.startswith('subtotal') \
                    or den_norm.startswith('valoare totala'):
                continue                # rand de insumare

            um = _val(rand, 'um', harta)
            cant_raw = _val(rand, 'cantitate', harta)
            este_articol = _um_valida(um) or _are_numar(cant_raw)

            if not este_articol:
                # rand-titlu de sectiune (sau paragraf de disclaimer)
                if len(den) > 80:
                    continue
                if _clasa_titlu(den_norm) == 'obiect':
                    obiect_titlu = _curata_titlu(den)
                else:
                    sectiune_curenta = _curata_titlu(den)
                continue

            # --- rand-articol real ---
            cod = _val(rand, 'cod_articol', harta)
            if not cod:
                cod = f'AUTO{nr_randuri_total + off}'
            cheie = normalizeaza_cheie(cod) or f'auto{nr_randuri_total + off}'
            if cheie in chei_vazute:
                chei_vazute[cheie] += 1
                cod = f'{cod}#{chei_vazute[cheie] - 1}'
                nr_duplicate += 1
            else:
                chei_vazute[cheie] = 1

            articole.append(ArticolF3(
                cod_articol=cod,
                denumire=den,
                um=um,
                cantitate=_to_float(cant_raw),
                obiect=(_val(rand, 'obiect', harta) or obiect_titlu
                        or obiect_sheet or '(fara obiect)'),
                tronson=(_val(rand, 'tronson', harta) or sectiune_curenta
                         or '(fara tronson)'),
                categorie=(_val(rand, 'categorie', harta) or sectiune_curenta or ''),
                rand_sursa=nr_rand,
            ))

    if coloane_mapate is None:
        raise EroareImport(
            'Nu am gasit randul de antet in niciun sheet. Asigura-te ca exista o coloana '
            'de denumire si una de unitate de masura sau cantitate (ex: Denumire / U.M. / '
            'Cantitate). Sinonimele de coloane sunt configurabile in config/gantt/setari.json.'
        )

    raport = {
        'nr_randuri_fisier': nr_randuri_total,
        'nr_sheeturi': len(sheeturi),
        'sheeturi_sarite': sheeturi_sarite,
        'rand_antet': rand_antet_global,
        'coloane_mapate': coloane_mapate,
        'nr_articole': len(articole),
        'nr_duplicate_redenumite': nr_duplicate,
        'nr_randuri_ignorate': nr_ignorate,
        'avertismente': avertismente[:50],
    }
    return articole, raport


def antet_nume(rand_antet: list, idx: int) -> str:
    try:
        return str(rand_antet[idx])
    except Exception:
        return f'col{idx}'


def importa_din_cale(cale: str, setari: Optional[dict] = None):
    """Helper: importa direct de pe disc (util pentru teste / batch)."""
    with open(cale, 'rb') as f:
        continut = f.read()
    return importa(continut, os.path.splitext(cale)[1], setari)
