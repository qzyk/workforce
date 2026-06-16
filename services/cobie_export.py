"""
COBie export (Construction Operations Building Information Exchange).

Genereaza un fisier Excel cu structura COBie standard pentru facility
management handover. Coloane conform COBie 2.4.

Tab-uri generate (COBie 2.4, set complet de 10 sheet-uri suportate):
- Facility (santier)
- Floor (nivel)
- Space (spatiu)
- Zone (grupare logica de spatii - din modelul Zona)
- Type (tip element clasificat)
- Component (element_bim) - din modelele BIM
- System (grupare functionala de componente, derivata din categoria tipului)
- Contact (utilizatori implicati)
- Job (mentenanta/operare) - fara sursa de date in modele -> header + 0 randuri
- Resource (resurse de mentenanta) - fara sursa de date -> header + 0 randuri

Job si Resource sunt emise cu antet corect chiar daca 0 randuri: un workbook
COBie valid asteapta sheet-urile prezente. Cand vom avea un model de mentenanta
(planuri PPM / resurse FM) le vom popula din relatiile respective.
"""

from __future__ import annotations

import logging
from datetime import datetime
from io import BytesIO
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from models import (db, Santier, Cladire, Nivel, Zona, Spatiu, ElementBIM,
                     Utilizator, Asset)


_logger = logging.getLogger(__name__)


# COBie required columns per sheet (subset esential)
COBIE_FACILITY_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category',
                       'ProjectName', 'SiteName', 'LinearUnits', 'AreaUnits',
                       'VolumeUnits', 'CurrencyUnit', 'AreaMeasurement',
                       'ExternalSystem', 'ExternalProjectObject',
                       'ExternalProjectIdentifier', 'ExternalSiteObject',
                       'ExternalSiteIdentifier', 'Description']
COBIE_FLOOR_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category',
                    'ExtSystem', 'ExtObject', 'ExtIdentifier',
                    'Description', 'Elevation', 'Height']
COBIE_SPACE_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category', 'FloorName',
                    'Description', 'ExtSystem', 'ExtObject', 'ExtIdentifier',
                    'RoomTag', 'UsableHeight', 'GrossArea', 'NetArea']
COBIE_TYPE_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category', 'Description',
                   'AssetType', 'Manufacturer', 'ModelNumber', 'WarrantyGuarantorParts',
                   'WarrantyDurationParts', 'WarrantyGuarantorLabor',
                   'WarrantyDurationLabor', 'WarrantyDurationUnit',
                   'ExtSystem', 'ExtObject', 'ExtIdentifier']
COBIE_COMPONENT_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'TypeName', 'Space',
                        'Description', 'ExtSystem', 'ExtObject', 'ExtIdentifier',
                        'SerialNumber', 'InstallationDate', 'WarrantyStartDate',
                        'TagNumber', 'BarCode', 'AssetIdentifier']
COBIE_CONTACT_COLS = ['Email', 'CreatedBy', 'CreatedOn', 'Category', 'Company',
                      'Phone', 'ExtSystem', 'ExtObject', 'ExtIdentifier',
                      'GivenName', 'FamilyName', 'Street', 'PostalBox',
                      'Town', 'StateRegion', 'PostalCode', 'Country', 'Category']
# COBie 2.4 - Zone: grupare logica de spatii (SpaceNames = lista membri).
COBIE_ZONE_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category', 'SpaceNames',
                   'ExtSystem', 'ExtObject', 'ExtIdentifier', 'Description']
# COBie 2.4 - System: grupare functionala de componente (ComponentNames = membri).
COBIE_SYSTEM_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category', 'ComponentNames',
                     'ExtSystem', 'ExtObject', 'ExtIdentifier', 'Description']
# COBie 2.4 - Job: activitati de mentenanta/operare planificate.
COBIE_JOB_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category', 'Status', 'TypeName',
                  'Description', 'Duration', 'DurationUnit', 'Start', 'TaskStartUnit',
                  'Frequency', 'FrequencyUnit', 'ExtSystem', 'ExtObject',
                  'ExtIdentifier', 'TaskNumber', 'Priors', 'ResourceNames']
# COBie 2.4 - Resource: resurse (materiale/unelte/personal) folosite la Jobs.
COBIE_RESOURCE_COLS = ['Name', 'CreatedBy', 'CreatedOn', 'Category',
                       'ExtSystem', 'ExtObject', 'ExtIdentifier', 'Description']


# Maparea tip_element -> categorie (din ElementBIM.TIPURI) pentru a deriva
# sistemele COBie (System grupeaza componentele pe functie/categorie).
def _categorie_tip(tip_element: str) -> str:
    """Categoria functionala a unui tip de element (ex. 'mep_hvac', 'structural').

    Cade pe 'general' cand tipul nu e in nomenclatorul ElementBIM.TIPURI."""
    for cod, _label, categorie in ElementBIM.TIPURI:
        if cod == tip_element:
            return categorie
    return 'general'


def _write_header(ws, columns: list[str]):
    """Scrie header cu format bold + fundal."""
    header_fill = PatternFill('solid', fgColor='305496')
    header_font = Font(bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    for col_idx, col_name in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
    ws.row_dimensions[1].height = 22


def _autosize(ws, max_width: int = 50):
    """Auto-size columns la maxim valoarea din celule."""
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                if cell.value is not None:
                    s = str(cell.value)
                    if len(s) > max_len:
                        max_len = len(s)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), max_width)


def generate_cobie_workbook(santier_id: int, *, generated_by: Optional[str] = None) -> BytesIO:
    """
    Genereaza un Workbook Excel cu structura COBie pentru un santier.
    Returneaza BytesIO cu .xlsx.
    """
    santier = Santier.query.get(santier_id)
    if not santier:
        raise ValueError(f'Santier inexistent: {santier_id}')

    wb = Workbook()
    # Delete default sheet
    wb.remove(wb.active)

    now_str = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S')
    created_by = generated_by or 'workforce-bim'

    # =========================
    # FACILITY
    # =========================
    ws = wb.create_sheet('Facility')
    _write_header(ws, COBIE_FACILITY_COLS)
    ws.append([
        santier.nume,
        created_by, now_str,
        'BUILDING_COMPLEX',
        santier.proiect.nume if santier.proiect else santier.nume,
        santier.nume,
        'meters', 'squareMeters', 'cubicMeters',
        'RON',
        'GrossArea',
        'workforce-bim', 'Santier', santier.cod,
        '', '',
        santier.descriere or '',
    ])
    _autosize(ws)

    # =========================
    # FLOOR
    # =========================
    ws = wb.create_sheet('Floor')
    _write_header(ws, COBIE_FLOOR_COLS)
    cladiri = Cladire.query.filter_by(santier_id=santier_id).all()
    for cladire in cladiri:
        for nivel in Nivel.query.filter_by(cladire_id=cladire.id).order_by(Nivel.ordine).all():
            ws.append([
                f'{cladire.cod}-{nivel.cod}',
                created_by, now_str,
                'Floor',
                'workforce-bim', 'Nivel', nivel.extern_id or str(nivel.id),
                nivel.nume,
                float(nivel.elevatie_m or 0),
                float(nivel.inaltime_m or 0),
            ])
    _autosize(ws)

    # =========================
    # SPACE
    # =========================
    ws = wb.create_sheet('Space')
    _write_header(ws, COBIE_SPACE_COLS)
    for cladire in cladiri:
        nivel_ids = [n.id for n in Nivel.query.filter_by(cladire_id=cladire.id).all()]
        if nivel_ids:
            for spatiu in Spatiu.query.filter(Spatiu.nivel_id.in_(nivel_ids)).all():
                nivel = Nivel.query.get(spatiu.nivel_id)
                floor_name = f'{cladire.cod}-{nivel.cod}' if nivel else ''
                ws.append([
                    spatiu.cod,
                    created_by, now_str,
                    spatiu.tip_spatiu or 'Room',
                    floor_name,
                    spatiu.nume,
                    'workforce-bim', 'Spatiu', spatiu.extern_id or str(spatiu.id),
                    spatiu.cod, '',
                    float(spatiu.suprafata_mp or 0) if hasattr(spatiu, 'suprafata_mp') else 0,
                    0,
                ])
    _autosize(ws)

    # =========================
    # ZONE (grupare logica de spatii - din modelul Zona)
    # SpaceNames = lista codurilor spatiilor membre (separator ',').
    # =========================
    ws = wb.create_sheet('Zone')
    _write_header(ws, COBIE_ZONE_COLS)
    for cladire in cladiri:
        for zona in Zona.query.filter_by(cladire_id=cladire.id).order_by(Zona.cod).all():
            spatii_membre = [sp.cod for sp in
                             Spatiu.query.filter_by(zona_id=zona.id).order_by(Spatiu.cod).all()]
            ws.append([
                zona.cod,
                created_by, now_str,
                zona.tip_zona or 'Zone',
                ','.join(spatii_membre),
                'workforce-bim', 'Zona', zona.extern_id or str(zona.id),
                zona.descriere or zona.nume or '',
            ])
    _autosize(ws)

    # =========================
    # TYPE (un Type per tip_element distinct)
    # =========================
    ws = wb.create_sheet('Type')
    _write_header(ws, COBIE_TYPE_COLS)
    cladiri_ids = [c.id for c in cladiri]
    tipuri = set()
    if cladiri_ids:
        for el in ElementBIM.query.filter(ElementBIM.cladire_id.in_(cladiri_ids)).all():
            tipuri.add(el.tip_element)
    for tip in sorted(tipuri):
        ws.append([
            f'TYPE-{tip}',
            created_by, now_str,
            tip,
            f'Type generic pentru {tip}',
            'Fixed', '', '',  # AssetType, Manufacturer, ModelNumber
            '', '', '', '', '',  # Warranty fields
            'workforce-bim', 'ElementBIM.tip_element', tip,
        ])
    _autosize(ws)

    # =========================
    # COMPONENT (un Component per ElementBIM)
    # =========================
    ws = wb.create_sheet('Component')
    _write_header(ws, COBIE_COMPONENT_COLS)
    if cladiri_ids:
        for el in ElementBIM.query.filter(ElementBIM.cladire_id.in_(cladiri_ids)).all():
            spatiu_name = ''
            if el.spatiu_id:
                sp = Spatiu.query.get(el.spatiu_id)
                spatiu_name = sp.cod if sp else ''
            asset = Asset.query.filter_by(element_bim_id=el.id).first()
            ws.append([
                el.cod,
                created_by, now_str,
                f'TYPE-{el.tip_element}',
                spatiu_name,
                el.nume or '',
                'workforce-bim', 'ElementBIM',
                el.ifc_global_id or el.extern_id or str(el.id),
                asset.serial if asset else '',
                '', '',  # InstallationDate, WarrantyStartDate
                el.cod, '',  # TagNumber, BarCode
                f'asset-{asset.id}' if asset else '',
            ])
    _autosize(ws)

    # =========================
    # SYSTEM (grupare functionala de componente, derivata din categoria tipului)
    # ComponentNames = codurile componentelor din acel sistem. Un sistem per
    # categorie distincta gasita pe elemente (ex. mep_hvac, structural, ...).
    # =========================
    ws = wb.create_sheet('System')
    _write_header(ws, COBIE_SYSTEM_COLS)
    sisteme: dict[str, list[str]] = {}
    if cladiri_ids:
        for el in ElementBIM.query.filter(ElementBIM.cladire_id.in_(cladiri_ids)).all():
            categorie = _categorie_tip(el.tip_element)
            sisteme.setdefault(categorie, []).append(el.cod)
    for categorie in sorted(sisteme.keys()):
        componente = sorted(sisteme[categorie])
        ws.append([
            f'SYSTEM-{categorie}',
            created_by, now_str,
            categorie,
            ','.join(componente),
            'workforce-bim', 'ElementBIM.categorie', categorie,
            f'Sistem functional {categorie} ({len(componente)} componente)',
        ])
    _autosize(ws)

    # =========================
    # CONTACT (utilizatorii care au creat asset-uri / incarcat modele)
    # =========================
    ws = wb.create_sheet('Contact')
    _write_header(ws, COBIE_CONTACT_COLS)
    # Selectam doar utilizatorii activi
    users = Utilizator.query.filter_by(activ=True).all()
    for u in users:
        ws.append([
            u.email,
            created_by, now_str,
            getattr(u, 'rol', 'user') or 'user',
            '',  # Company - poate fi populat din Setari firma
            '',  # Phone
            'workforce-bim', 'Utilizator', str(u.id),
            u.prenume or '', u.nume or '',
            '', '', '', '', '', '', '',  # Address fields
        ])
    _autosize(ws)

    # =========================
    # JOB (mentenanta/operare) - fara sursa de date in modelele actuale.
    # Emis cu antet corect + 0 randuri pentru un workbook COBie complet.
    # Se va popula cand vom avea planuri de mentenanta (PPM) in model.
    # =========================
    ws = wb.create_sheet('Job')
    _write_header(ws, COBIE_JOB_COLS)
    _autosize(ws)

    # =========================
    # RESOURCE (resurse pt Jobs) - fara sursa de date in modelele actuale.
    # Emis cu antet corect + 0 randuri (idem Job).
    # =========================
    ws = wb.create_sheet('Resource')
    _write_header(ws, COBIE_RESOURCE_COLS)
    _autosize(ws)

    # Salveaza in memorie
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
