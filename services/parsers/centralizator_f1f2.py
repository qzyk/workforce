"""
Parsere pentru centralizatoarele F1 (pe obiectiv) si F2 (pe obiect).

Ierarhia unui obiectiv de investitie:
  F1 centralizator pe OBIECTIV   -> lista OBIECTE (001 Arhitectura, 002 Structura...)
  F2 centralizator pe OBIECT     -> lista SUB-OBIECTE = liste F3 (001, 002, 003...)
  F3 lista de cantitati          -> articole (parsate de gantt/import_engine)

Nucleul lucreaza pe `randuri` (list[list]) ca sa fie testabil fara fisiere.
Wrapper-ele `parse_f1_file` / `parse_f2_file` citesc .xls (binar) cu xlrd.

Detectia coloanei de valoare e tolerantă: cauta in antet celula care contine
"valoare" (fara TVA) si "C+M" / "din care", deci merge si daca pozitia coloanei
difera intre F1 si F2 (au coloane goale - celule imbinate - diferite).
"""

from __future__ import annotations

import re
from decimal import Decimal, InvalidOperation
from typing import Optional


# Linie de obiect / sub-obiect: incepe cu un cod de 3 cifre (001, 002, ...)
_RE_COD = re.compile(r'^\s*(\d{3})[\s_\-]+(.+?)\s*$')


def _dec(v) -> Optional[Decimal]:
    try:
        if v is None or v == '':
            return None
        d = Decimal(str(v))
        return d
    except (InvalidOperation, ValueError):
        return None


def _norm(s) -> str:
    return str(s or '').strip().lower()


def _gaseste_coloane(randuri: list) -> tuple[Optional[int], Optional[int]]:
    """Scaneaza primele randuri si intoarce (idx_valoare, idx_cm).
    idx_valoare = coloana cu 'valoare'; idx_cm = coloana 'C+M' / 'din care' (sau None)."""
    idx_val = idx_cm = None
    for r in randuri[:8]:
        for c, cell in enumerate(r):
            t = _norm(cell)
            if not t:
                continue
            if idx_val is None and 'valoare' in t:
                idx_val = c
            if idx_cm is None and ('c+m' in t or 'din care' in t or 'c + m' in t):
                idx_cm = c
        if idx_val is not None:
            break
    return idx_val, idx_cm


def _val_din_rand(rand: list, idx: Optional[int]) -> Optional[Decimal]:
    """Valoarea din coloana idx; daca idx None sau gol, ia ultima celula numerica."""
    if idx is not None and idx < len(rand):
        d = _dec(rand[idx])
        if d is not None:
            return d
    # fallback: ultima celula numerica din rand
    for cell in reversed(rand):
        d = _dec(cell)
        if d is not None:
            return d
    return None


def extrage_linii_cod(randuri: list) -> list[dict]:
    """Intoarce liniile care incep cu un cod de 3 cifre (obiecte in F1 / sub-obiecte
    in F2): [{cod, nume, valoare, cm}]. Sare peste TOTAL / capitol / antet."""
    idx_val, idx_cm = _gaseste_coloane(randuri)
    out: list[dict] = []
    for rand in randuri:
        # cauta in oricare din primele celule un cod de 3 cifre + denumire
        cod = nume = None
        for cell in rand[:4]:
            m = _RE_COD.match(str(cell))
            if m:
                cod, nume = m.group(1), m.group(2).strip()
                break
        if not cod:
            continue
        val = _val_din_rand(rand, idx_val)
        if val is None:
            continue
        cm = _dec(rand[idx_cm]) if (idx_cm is not None and idx_cm < len(rand)) else None
        out.append(dict(cod=cod, nume=nume, valoare=val, cm=cm))
    return out


def parse_f1(randuri: list) -> dict:
    """F1 -> {obiecte: [{cod, nume, valoare, cm}], total_4_1, total}.
    obiecte = liniile cu cod de 3 cifre (de regula sub capitolul 4.1 Constructii)."""
    obiecte = extrage_linii_cod(randuri)
    # total constructii (4.1) si total general - din liniile capitol, daca exista
    total_4_1 = _gaseste_valoare_capitol(randuri, '4.1')
    total = _gaseste_valoare_capitol(randuri, '4') or total_4_1
    return dict(obiecte=obiecte, total_4_1=total_4_1, total=total)


def parse_f2(randuri: list) -> dict:
    """F2 -> {sub_obiecte: [{cod, nume, valoare, cm}], total}.
    sub_obiecte = liniile cu cod de 3 cifre (listele F3 ale obiectului)."""
    sub = extrage_linii_cod(randuri)
    total = _gaseste_total(randuri)
    return dict(sub_obiecte=sub, total=total)


def _gaseste_valoare_capitol(randuri: list, cap: str) -> Optional[Decimal]:
    """Valoarea de pe linia al carei cod-capitol (col 0 sau 1) == cap exact."""
    idx_val, _ = _gaseste_coloane(randuri)
    for rand in randuri:
        for cell in rand[:2]:
            if str(cell).strip() == cap:
                return _val_din_rand(rand, idx_val)
    return None


def _gaseste_total(randuri: list) -> Optional[Decimal]:
    """Suma sub-obiectelor (robust) - nu ne bazam pe randul TOTAL care poate lipsi."""
    sub = extrage_linii_cod(randuri)
    if not sub:
        return None
    return sum((s['valoare'] for s in sub), Decimal('0'))


# ============================================================
# Total F3 (lista de cantitati) - doar suma, pentru reconciliere.
# Nu inlocuieste gantt/import_engine (parsare completa de articole);
# aici avem nevoie DOAR de total, robust pe formatul "Capitol de lucrari".
# ============================================================

def total_f3_rows(randuri: list) -> tuple[Decimal, int]:
    """Suma coloanei TOTAL pe randurile de articol (col 0 = Nr numeric).
    Sare peste antet, titluri de sectiune si sub-randurile material:/manopera:."""
    col_tot = None
    for r in randuri[:8]:
        for c, cell in enumerate(r):
            t = _norm(cell)
            if 'totalul' in t or ('total' in t and 'tva' in t):
                col_tot = c
        if col_tot is not None:
            break
    total = Decimal('0')
    n = 0
    for rand in randuri:
        if not rand or _dec(rand[0]) is None:
            continue  # nu e rand de articol (titlu / sub-rand / gol)
        idx = col_tot if (col_tot is not None and col_tot < len(rand)) else len(rand) - 1
        v = _dec(rand[idx])
        if v is None:
            continue   # ex: randul de numerotare "5 = 3 x 4"
        total += v
        n += 1
    return total, n


def total_f3_file(path: str) -> tuple[Decimal, int]:
    return total_f3_rows(_randuri_xls(path))


# ============================================================
# Wrapper-e fisier (.xls binar)
# ============================================================

def _randuri_xls(path: str) -> list:
    """Citeste primul sheet ca list[list]. Suporta .xls binar (xlrd) si
    .xlsx (openpyxl) - xlrd 2.x nu mai citeste xlsx."""
    try:
        import xlrd
        wb = xlrd.open_workbook(path)
        sh = wb.sheet_by_index(0)
        return [[sh.cell_value(r, c) for c in range(sh.ncols)] for r in range(sh.nrows)]
    except Exception:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb.worksheets[0]
        randuri = [['' if v is None else v for v in row]
                   for row in ws.iter_rows(values_only=True)]
        wb.close()
        return randuri


def parse_f1_file(path: str) -> dict:
    return parse_f1(_randuri_xls(path))


def parse_f2_file(path: str) -> dict:
    return parse_f2(_randuri_xls(path))
