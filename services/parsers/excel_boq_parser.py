"""
Parser Excel BoQ (XLSX) generic.

Convenție coloane (modificabilă per import via parametri):
    A: cod_articol         (obligatoriu)
    B: cod_capitol         (optional)
    C: denumire            (obligatoriu)
    D: um                  (obligatoriu)
    E: cantitate_oferta    (obligatoriu, Decimal)
    F: pret_unitar         (obligatoriu, Decimal)
    G: categorie           (optional, default 'mixt')

Header pe rand 1 (configurable via `header_row`); date pornesc de la rand 2.
Randuri goale sau cu cod gol -> skip cu warning.

Foloseste openpyxl (deja in requirements.txt). Read-only mode pentru
memorie eficienta pe fisiere mari.
"""

from __future__ import annotations

from decimal import Decimal, InvalidOperation
from typing import Any, Optional

import openpyxl

from .base import Parser, ParseResult, ParseError


_VALID_CATEGORII = {'materiale', 'manopera', 'utilaje', 'transport', 'mixt'}


class ExcelBoQParser(Parser):
    """Parser Excel XLSX cu BoQ (deviz)."""

    SURSA_COD = 'excel_xlsx'

    # Coloana index (0-based) pentru fiecare camp
    DEFAULT_COLS = {
        'cod_articol': 0,
        'cod_capitol': 1,
        'denumire': 2,
        'um': 3,
        'cantitate_oferta': 4,
        'pret_unitar': 5,
        'categorie': 6,
    }

    def __init__(self, header_row: int = 1,
                 cols: Optional[dict[str, int]] = None,
                 sheet_name: Optional[str] = None):
        """
        Args:
          header_row: Numarul randului cu header (1-based). Datele incep
                      de la `header_row + 1`. Default 1.
          cols:       Override pentru maparea coloanelor (0-based index).
          sheet_name: Numele sheet-ului. Default: primul sheet.
        """
        self.header_row = header_row
        self.cols = {**self.DEFAULT_COLS, **(cols or {})}
        self.sheet_name = sheet_name

    def parse(self, file_path: str) -> ParseResult:
        result = ParseResult(sursa=self.SURSA_COD)

        try:
            wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        except Exception as e:
            raise ParseError(f'Nu pot deschide XLSX: {e}') from e

        try:
            if self.sheet_name:
                if self.sheet_name not in wb.sheetnames:
                    result.add_error(
                        f'Sheet "{self.sheet_name}" nu exista. '
                        f'Disponibile: {wb.sheetnames}'
                    )
                    return result
                ws = wb[self.sheet_name]
            else:
                ws = wb.worksheets[0]

            result.stats['sheet_name'] = ws.title
            result.stats['cols_mapping'] = self.cols

            ordine = 0
            data_start_row = self.header_row + 1
            for row_idx, row in enumerate(ws.iter_rows(min_row=data_start_row,
                                                       values_only=True),
                                          start=data_start_row):
                if self._is_empty_row(row):
                    continue
                entity = self._parse_row(row, row_idx, ordine + 1, result)
                if entity:
                    ordine += 1
                    result.entities.append(entity)
        finally:
            wb.close()

        if not result.entities and not result.errors:
            result.add_error(
                'Nu am gasit randuri valide. Verifica ca fisierul are date '
                f'incepand cu randul {self.header_row + 1} si ca primele '
                '4 coloane (cod, denumire, um, cantitate) sunt completate.'
            )

        result.stats['entities_count'] = len(result.entities)
        result.stats['warnings_count'] = len(result.warnings)
        return result

    # ------------------------------------------------------------
    # Internals
    # ------------------------------------------------------------

    @staticmethod
    def _is_empty_row(row: tuple) -> bool:
        return all(v is None or (isinstance(v, str) and not v.strip())
                   for v in row)

    def _parse_row(self, row: tuple, row_idx: int, ordine: int,
                   result: ParseResult) -> Optional[dict]:
        def _get(field_name: str):
            col = self.cols.get(field_name)
            if col is None or col >= len(row):
                return None
            return row[col]

        cod_articol = _get('cod_articol')
        denumire = _get('denumire')
        um = _get('um')
        cantitate_raw = _get('cantitate_oferta')
        pret_raw = _get('pret_unitar')

        cod_articol = str(cod_articol).strip() if cod_articol is not None else ''
        denumire = str(denumire).strip() if denumire is not None else ''
        um = str(um).strip() if um is not None else ''

        if not cod_articol:
            # Rand incomplet (poate fi un subtotal sau separator) - skip silent
            return None
        if not denumire or not um:
            result.add_warning(
                f'Rand {row_idx}: denumire sau UM lipsa pentru cod '
                f'"{cod_articol}", skip-at.'
            )
            return None

        cantitate = self._decimal(cantitate_raw, result,
                                  f'rand {row_idx} cod {cod_articol} cantitate',
                                  default=Decimal('0'))
        pret_unitar = self._decimal(pret_raw, result,
                                    f'rand {row_idx} cod {cod_articol} pret',
                                    default=Decimal('0'))

        cod_capitol = _get('cod_capitol')
        cod_capitol = str(cod_capitol).strip() if cod_capitol else None

        categorie_raw = _get('categorie')
        if categorie_raw:
            categorie = str(categorie_raw).strip().lower()
        else:
            categorie = 'mixt'
        if categorie not in _VALID_CATEGORII:
            result.add_warning(
                f'Rand {row_idx}: categorie "{categorie}" necunoscuta, default mixt.'
            )
            categorie = 'mixt'

        return {
            'cod_articol': cod_articol,
            'cod_capitol': cod_capitol,
            'denumire': denumire,
            'um': um,
            'cantitate_oferta': cantitate,
            'pret_unitar': pret_unitar,
            'categorie': categorie,
            'ordine': ordine,
            # Excel-ul nu are detalii granulare (materiale/manopera/utilaj/transport)
            'valoare_materiale_unitar': None,
            'valoare_manopera_unitar': None,
            'valoare_utilaj_unitar': None,
            'valoare_transport_unitar': None,
        }

    @staticmethod
    def _decimal(value: Any, result: ParseResult, context: str = '',
                 default: Optional[Decimal] = None) -> Optional[Decimal]:
        if value is None or value == '':
            return default
        try:
            if isinstance(value, str):
                value = value.strip().replace(',', '.')
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            if context:
                result.add_warning(
                    f'Valoare invalida "{value}" ({context}), default aplicat.'
                )
            return default
