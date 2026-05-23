"""
Centralizator + Deviz General la nivel de PROIECT (FEEDBACK §5, §7).

Agrega TOATE ofertele (devizele) unui proiect:
  - clasifica_proiect()       -> ruleaza clasificarea pe toate ofertele
  - dry_run_proiect()         -> distributie categorii + Diverse pe tot proiectul
  - genereaza_centralizator() -> agregare pe disciplina -> categorie
  - genereaza_deviz_general() -> consolidare pe capitole HG907/2016 + TVA
  - export_centralizator_xlsx() / export_deviz_general_xlsx()

Nicio tabela noua: totul derivat la query time. Pozitiile fara TVA;
TVA (21%) adaugat la final (playbook §2.6). Verificare Sigma == total.
"""

from __future__ import annotations

import os
from datetime import datetime
from decimal import Decimal, ROUND_HALF_UP
from typing import Optional

from flask import current_app

from models import (
    db, Proiect, Contract, OfertaContract, PozitieBoQ,
)
from services.deviz_pricing import (
    clasifica_oferta, deduce_disciplina, dry_run_clasificare,
)


COTA_TVA_DEFAULT = Decimal('21')  # % standard RO

# Mapare disciplina -> capitol Deviz General (HG907/2016)
_CAPITOL_HG907 = {
    'structural':  ('4.1', 'Constructii si instalatii'),
    'arhitectura': ('4.1', 'Constructii si instalatii'),
    'electrice':   ('4.1', 'Constructii si instalatii'),
    'hvac':        ('4.1', 'Constructii si instalatii'),
    'sanitare':    ('4.1', 'Constructii si instalatii'),
    'drumuri':     ('4.1', 'Constructii si instalatii'),
    'organizare':  ('5.1.1', 'Organizare de santier'),
    'general':     ('4.1', 'Constructii si instalatii'),
}

_Q2 = Decimal('0.01')


def _oferte_proiect(proiect_id: int) -> list[OfertaContract]:
    """Toate ofertele unui proiect (prin contractele lui)."""
    return OfertaContract.query.join(
        Contract, OfertaContract.contract_id == Contract.id
    ).filter(Contract.proiect_id == proiect_id).order_by(
        OfertaContract.contract_id, OfertaContract.versiune
    ).all()


def _pozitii_proiect(proiect_id: int) -> list[PozitieBoQ]:
    """Toate pozitiile BoQ ale unui proiect (din toate ofertele)."""
    return PozitieBoQ.query.join(
        OfertaContract, PozitieBoQ.oferta_id == OfertaContract.id
    ).join(
        Contract, OfertaContract.contract_id == Contract.id
    ).filter(Contract.proiect_id == proiect_id).order_by(
        PozitieBoQ.oferta_id, PozitieBoQ.ordine
    ).all()


# ============================================================
# Clasificare la nivel proiect
# ============================================================

def clasifica_proiect(proiect_id: int, doar_neclasificate: bool = True) -> dict:
    """
    Clasifica toate ofertele proiectului. Returneaza raport agregat:
      {oferte: N, pozitii: N, distributie: {cat: count}, per_oferta: [...]}
    doar_neclasificate=True protejeaza editarile manuale la re-rulare.
    """
    oferte = _oferte_proiect(proiect_id)
    distributie: dict[str, int] = {}
    per_oferta = []
    total_pozitii = 0

    for of in oferte:
        stats = clasifica_oferta(of, commit=False,
                                 doar_neclasificate=doar_neclasificate)
        n = sum(stats.values())
        total_pozitii += n
        for cat, c in stats.items():
            distributie[cat] = distributie.get(cat, 0) + c
        per_oferta.append({
            'oferta_id': of.id,
            'contract_id': of.contract_id,
            'versiune': of.versiune,
            'sursa': of.sursa_import,
            'pozitii': n,
        })

    db.session.commit()
    return {
        'oferte': len(oferte),
        'pozitii': total_pozitii,
        'distributie': dict(sorted(distributie.items(), key=lambda x: -x[1])),
        'per_oferta': per_oferta,
    }


def dry_run_proiect(proiect_id: int) -> dict:
    """
    Dry-run clasificare pe TOT proiectul (fara persist). Agregat:
      {distributie, diverse: [...], total_pozitii, procent_diverse}.
    Pentru validare (playbook §4: refine pana Diverse near-zero).
    """
    oferte = _oferte_proiect(proiect_id)
    distributie: dict[str, int] = {}
    diverse = []
    total = 0
    for of in oferte:
        r = dry_run_clasificare(of)
        total += r['total_pozitii']
        for cat, c in r['distributie'].items():
            distributie[cat] = distributie.get(cat, 0) + c
        for d in r['diverse']:
            d['oferta_id'] = of.id
            diverse.append(d)
    return {
        'distributie': dict(sorted(distributie.items(), key=lambda x: -x[1])),
        'diverse': diverse,
        'total_pozitii': total,
        'procent_diverse': round(len(diverse) / total * 100, 1) if total else 0,
        'oferte': len(oferte),
    }


# ============================================================
# Centralizator (disciplina -> categorie)
# ============================================================

def genereaza_centralizator(proiect_id: int) -> dict:
    """
    Agregare pe disciplina -> categorie_lucrare cu Sigma valoare + nr pozitii.
    Returneaza structura ierarhica + total general.
    """
    pozitii = _pozitii_proiect(proiect_id)
    # disc -> cat -> {valoare, nr, cantitate-less aggregated}
    grupe: dict[str, dict[str, dict]] = {}
    total_general = Decimal('0')

    for p in pozitii:
        disc = deduce_disciplina(p.cod_capitol)
        cat = p.categorie_lucrare or 'neclasificat'
        val = (p.cantitate_oferta or Decimal('0')) * (p.pret_unitar or Decimal('0'))
        g = grupe.setdefault(disc, {})
        c = g.setdefault(cat, {'valoare': Decimal('0'), 'nr': 0})
        c['valoare'] += val
        c['nr'] += 1
        total_general += val

    # Construiesc lista ordonata + subtotaluri pe disciplina
    rezultat = []
    for disc in sorted(grupe.keys()):
        cats = grupe[disc]
        subtotal = sum((c['valoare'] for c in cats.values()), Decimal('0'))
        rezultat.append({
            'disciplina': disc,
            'subtotal': subtotal.quantize(_Q2),
            'categorii': [
                {'categorie': cat,
                 'valoare': info['valoare'].quantize(_Q2),
                 'nr': info['nr']}
                for cat, info in sorted(cats.items(), key=lambda x: -x[1]['valoare'])
            ],
        })

    return {
        'grupe': rezultat,
        'total_general': total_general.quantize(_Q2),
        'nr_pozitii': len(pozitii),
    }


# ============================================================
# Deviz General (capitole HG907/2016 + TVA)
# ============================================================

def genereaza_deviz_general(proiect_id: int,
                            cota_tva: Decimal = COTA_TVA_DEFAULT) -> dict:
    """
    Consolidare pe capitole HG907/2016 cu subtotaluri + TVA + total cu TVA.
    Pozitiile sunt fara TVA; TVA aplicat la total (playbook §2.6).
    """
    cota_tva = Decimal(str(cota_tva))
    pozitii = _pozitii_proiect(proiect_id)

    capitole: dict[tuple, Decimal] = {}  # (cod_cap, nume_cap) -> valoare
    for p in pozitii:
        disc = deduce_disciplina(p.cod_capitol)
        cap = _CAPITOL_HG907.get(disc, ('4.1', 'Constructii si instalatii'))
        val = (p.cantitate_oferta or Decimal('0')) * (p.pret_unitar or Decimal('0'))
        capitole[cap] = capitole.get(cap, Decimal('0')) + val

    randuri = [
        {'cod': cod, 'denumire': nume, 'valoare': val.quantize(_Q2)}
        for (cod, nume), val in sorted(capitole.items())
    ]
    total_fara_tva = sum((r['valoare'] for r in randuri), Decimal('0'))
    tva = (total_fara_tva * cota_tva / 100).quantize(_Q2, rounding=ROUND_HALF_UP)
    total_cu_tva = total_fara_tva + tva

    return {
        'randuri': randuri,
        'total_fara_tva': total_fara_tva.quantize(_Q2),
        'cota_tva': cota_tva,
        'tva': tva,
        'total_cu_tva': total_cu_tva.quantize(_Q2),
        'nr_pozitii': len(pozitii),
    }


# ============================================================
# Export Excel
# ============================================================

def _upload_dir(subdir: str) -> str:
    base = current_app.config.get(
        'UPLOAD_FOLDER', os.path.join(current_app.root_path, 'uploads'))
    path = os.path.join(base, subdir)
    os.makedirs(path, exist_ok=True)
    return path


def _styled_header(ws, row, headers, gold='C9A961'):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    bold = Font(bold=True, color='0B1426', name='Arial')
    fill = PatternFill('solid', fgColor=gold)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='888888')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = bold; c.fill = fill; c.alignment = center; c.border = bord


def export_centralizator_xlsx(proiect_id: int) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    proiect = Proiect.query.get(proiect_id)
    if proiect is None:
        raise ValueError(f'Proiect {proiect_id} inexistent.')
    data = genereaza_centralizator(proiect_id)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Centralizator'
    ws['A1'] = 'CENTRALIZATOR'
    ws['A1'].font = Font(bold=True, size=14, name='Arial', color='0B1426')
    ws.merge_cells('A1:D1')
    ws['A2'] = f'Proiect: {proiect.cod_proiect} - {proiect.nume}'
    ws.merge_cells('A2:D2')

    r = 4
    _styled_header(ws, r, ['Disciplina', 'Categorie lucrare', 'Nr. pozitii', 'Valoare (fara TVA)'])
    r += 1
    money = '#,##0.00'
    for grup in data['grupe']:
        ws.cell(row=r, column=1, value=grup['disciplina'].upper()).font = Font(bold=True, name='Arial')
        ws.cell(row=r, column=4, value=float(grup['subtotal'])).font = Font(bold=True, name='Arial')
        ws.cell(row=r, column=4).number_format = money
        r += 1
        for cat in grup['categorii']:
            ws.cell(row=r, column=2, value=cat['categorie'])
            ws.cell(row=r, column=3, value=cat['nr'])
            cell = ws.cell(row=r, column=4, value=float(cat['valoare']))
            cell.number_format = money
            r += 1

    r += 1
    ws.cell(row=r, column=1, value='TOTAL GENERAL (fara TVA)').font = Font(bold=True, name='Arial')
    tc = ws.cell(row=r, column=4, value=float(data['total_general']))
    tc.font = Font(bold=True, name='Arial'); tc.number_format = money

    for col, w in zip('ABCD', [22, 32, 12, 18]):
        ws.column_dimensions[col].width = w
    wb.calculation.fullCalcOnLoad = True

    path = os.path.join(_upload_dir('centralizatoare'),
                        f'centralizator_{proiect_id}_{datetime.utcnow():%Y%m%d%H%M%S}.xlsx')
    wb.save(path)
    return path


def export_deviz_general_xlsx(proiect_id: int,
                              cota_tva: Decimal = COTA_TVA_DEFAULT) -> str:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    proiect = Proiect.query.get(proiect_id)
    if proiect is None:
        raise ValueError(f'Proiect {proiect_id} inexistent.')
    data = genereaza_deviz_general(proiect_id, cota_tva)

    wb = Workbook()
    ws = wb.active
    ws.title = 'Deviz General'
    ws['A1'] = 'DEVIZ GENERAL'
    ws['A1'].font = Font(bold=True, size=14, name='Arial', color='0B1426')
    ws.merge_cells('A1:C1')
    ws['A2'] = f'Proiect: {proiect.cod_proiect} - {proiect.nume}'
    ws.merge_cells('A2:C2')

    r = 4
    _styled_header(ws, r, ['Cap.', 'Denumire capitol', 'Valoare (fara TVA)'])
    r += 1
    money = '#,##0.00'
    for rand in data['randuri']:
        ws.cell(row=r, column=1, value=rand['cod'])
        ws.cell(row=r, column=2, value=rand['denumire'])
        c = ws.cell(row=r, column=3, value=float(rand['valoare']))
        c.number_format = money
        r += 1

    bold = Font(bold=True, name='Arial')
    ws.cell(row=r, column=2, value='TOTAL (fara TVA)').font = bold
    ws.cell(row=r, column=3, value=float(data['total_fara_tva'])).number_format = money
    ws.cell(row=r, column=3).font = bold
    r += 1
    ws.cell(row=r, column=2, value=f'TVA ({data["cota_tva"]}%)').font = bold
    ws.cell(row=r, column=3, value=float(data['tva'])).number_format = money
    ws.cell(row=r, column=3).font = bold
    r += 1
    ws.cell(row=r, column=2, value='TOTAL (cu TVA)').font = Font(bold=True, size=12, name='Arial')
    tc = ws.cell(row=r, column=3, value=float(data['total_cu_tva']))
    tc.font = Font(bold=True, size=12, name='Arial'); tc.number_format = money

    for col, w in zip('ABC', [10, 40, 20]):
        ws.column_dimensions[col].width = w
    wb.calculation.fullCalcOnLoad = True

    path = os.path.join(_upload_dir('devize_generale'),
                        f'deviz_general_{proiect_id}_{datetime.utcnow():%Y%m%d%H%M%S}.xlsx')
    wb.save(path)
    return path
