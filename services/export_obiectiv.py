"""
Export centralizator de cheltuieli pe OBIECTIV: xlsx (multi-sheet) + pdf.

Sursa de adevar = arborele ingerat (Obiectiv -> Obiect -> GanttPlan). Optional,
daca se da `director`, se adauga foi cu extrasele de resurse C6/C7/C8/C9 + F4
(citite direct din folder). Toate valorile FARA TVA.

Functii publice:
  - export_xlsx(obiectiv_id, director=None) -> bytes
  - export_pdf(obiectiv_id) -> bytes
"""

from __future__ import annotations

import os
import re
from decimal import Decimal
from io import BytesIO
from typing import Optional

from models import Obiectiv


GOLD = 'C9A961'
NAVY = '0B1426'
MONEY = '#,##0.00'


def _f(v) -> float:
    return float(v or 0)


# ============================================================
# XLSX
# ============================================================

def export_xlsx(obiectiv_id: int, director: Optional[str] = None) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    ob = Obiectiv.query.get(obiectiv_id)
    if ob is None:
        raise ValueError(f'Obiectiv {obiectiv_id} inexistent')

    bold = Font(bold=True, name='Arial', color=NAVY)
    hdr_fill = PatternFill('solid', fgColor=GOLD)
    titlu = Font(bold=True, size=14, name='Arial', color=NAVY)

    wb = Workbook()

    # --- Sheet 1: Centralizator obiectiv (F1) ---
    ws = wb.active
    ws.title = 'Centralizator obiectiv'
    ws['A1'] = f'CENTRALIZATORUL cheltuielilor pe obiectiv (fara TVA)'
    ws['A1'].font = titlu
    ws['A2'] = ob.nume
    ws['A2'].font = Font(bold=True, name='Arial', size=12)
    r = 4
    for c, h in enumerate(['Cod', 'Obiect', 'Disciplina', 'Valoare F2', 'Valoare F1 (4.1)'], start=1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = bold
        cell.fill = hdr_fill
    r += 1
    for o in ob.obiecte:
        ws.cell(row=r, column=1, value=o.cod)
        ws.cell(row=r, column=2, value=o.nume)
        ws.cell(row=r, column=3, value=o.disciplina)
        ws.cell(row=r, column=4, value=_f(o.valoare_f2)).number_format = MONEY
        ws.cell(row=r, column=5, value=_f(o.valoare_f1)).number_format = MONEY
        r += 1
    ws.cell(row=r, column=2, value='TOTAL Constructii (4.1)').font = bold
    tc = ws.cell(row=r, column=5, value=_f(ob.valoare_constructii))
    tc.font = bold
    tc.number_format = MONEY
    for col, w in zip('ABCDE', [8, 42, 14, 18, 18]):
        ws.column_dimensions[col].width = w

    # --- Sheet 2: Liste F3 (pe obiect) ---
    ws2 = wb.create_sheet('Liste F3')
    for c, h in enumerate(['Obiect', 'Lista F3', 'Fisier', 'Cost (fara TVA)'], start=1):
        cell = ws2.cell(row=1, column=c, value=h)
        cell.font = bold
        cell.fill = hdr_fill
    r = 2
    for o in ob.obiecte:
        for pl in o.planuri:
            ws2.cell(row=r, column=1, value=f'{o.cod} {o.nume}')
            ws2.cell(row=r, column=2, value=pl.nume)
            ws2.cell(row=r, column=3, value=pl.nume_fisier)
            ws2.cell(row=r, column=4, value=_f(pl.cost_total)).number_format = MONEY
            r += 1
    for col, w in zip('ABCD', [30, 40, 44, 18]):
        ws2.column_dimensions[col].width = w

    # --- Optional: extrase de resurse C6-C9 + F4 din folder ---
    if director:
        for titlu_sheet, rows, headers in _extrase_pentru_export(director):
            wsx = wb.create_sheet(titlu_sheet[:31])
            for c, h in enumerate(headers, start=1):
                cell = wsx.cell(row=1, column=c, value=h)
                cell.font = bold
                cell.fill = hdr_fill
            for ri, row in enumerate(rows, start=2):
                for c, val in enumerate(row, start=1):
                    cell = wsx.cell(row=ri, column=c, value=val)
                    if isinstance(val, float):
                        cell.number_format = MONEY

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ============================================================
# PDF
# ============================================================

def export_pdf(obiectiv_id: int) -> bytes:
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import mm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle

    ob = Obiectiv.query.get(obiectiv_id)
    if ob is None:
        raise ValueError(f'Obiectiv {obiectiv_id} inexistent')

    def lei(v):
        return f'{_f(v):,.2f}'.replace(',', '@').replace('.', ',').replace('@', '.')

    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, topMargin=18 * mm, bottomMargin=16 * mm,
                            leftMargin=15 * mm, rightMargin=15 * mm)
    styles = getSampleStyleSheet()
    h = ParagraphStyle('t', parent=styles['Title'], textColor=colors.HexColor('#0B1426'), fontSize=15)
    sub = ParagraphStyle('s', parent=styles['Heading2'], textColor=colors.HexColor('#0B1426'), fontSize=11)
    el = [Paragraph('Centralizatorul cheltuielilor pe obiectiv (fara TVA)', h),
          Paragraph(ob.nume, sub), Spacer(1, 6 * mm)]

    navy = colors.HexColor('#0B1426')
    gold = colors.HexColor('#C9A961')
    cream = colors.HexColor('#F5F1E8')

    # Tabel obiecte (F2)
    data = [['Cod', 'Obiect', 'Disciplina', 'Valoare F2', 'Valoare F1 (4.1)']]
    for o in ob.obiecte:
        data.append([o.cod or '', (o.nume or '')[:40], o.disciplina or '',
                     lei(o.valoare_f2), lei(o.valoare_f1)])
    data.append(['', 'TOTAL Constructii (4.1)', '', '', lei(ob.valoare_constructii)])
    t = Table(data, colWidths=[14 * mm, 62 * mm, 26 * mm, 32 * mm, 32 * mm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), gold),
        ('TEXTCOLOR', (0, 0), (-1, 0), navy),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('BACKGROUND', (0, -1), (-1, -1), cream),
        ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.4, colors.HexColor('#CCCCCC')),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.HexColor('#FAF8F2')]),
    ]))
    el += [t, Spacer(1, 6 * mm)]

    # Liste F3 pe obiect
    for o in ob.obiecte:
        el.append(Paragraph(f'{o.cod} {o.nume}', sub))
        d2 = [['Lista F3', 'Cost (fara TVA)']]
        suma = Decimal('0')
        for pl in o.planuri:
            d2.append([(pl.nume or '')[:60], lei(pl.cost_total)])
            suma += (pl.cost_total or Decimal('0'))
        d2.append(['Total liste F3', lei(suma)])
        t2 = Table(d2, colWidths=[130 * mm, 38 * mm])
        t2.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTSIZE', (0, 0), (-1, -1), 7.5),
            ('LINEBELOW', (0, 0), (-1, 0), 0.4, gold),
            ('LINEABOVE', (0, -1), (-1, -1), 0.4, navy),
        ]))
        el += [t2, Spacer(1, 4 * mm)]

    doc.build(el)
    return buf.getvalue()


# ============================================================
# Extrase de resurse C6-C9 + F4 din folder (optional, pt export complet)
# ============================================================

_RE_COD = re.compile(r'^\s*([0-9A-Za-z\[\]\.%#\-/]+?)\s*[-–]\s*(.+)$')


def _split_cod(s):
    m = _RE_COD.match(str(s).strip())
    return (m.group(1).strip(), m.group(2).strip()) if m else ('', str(s).strip())


def _is_nr(v):
    try:
        float(v); return True
    except (TypeError, ValueError):
        return False


def _extrase_pentru_export(director: str):
    """Intoarce [(titlu_sheet, rows, headers)] pt C6/C7/C8/C9/F4 gasite in folder."""
    import xlrd
    specs = [
        ('C6 Materiale', ['c6', 'materiale'], ['Cod', 'Denumire', 'U.M.', 'Pret unitar', 'Furnizor'],
         lambda sh, r: [_split_cod(sh.cell_value(r, 1))[0], _split_cod(sh.cell_value(r, 1))[1],
                        str(sh.cell_value(r, 2)), _num(sh, r, 4), str(sh.cell_value(r, 6)) if sh.ncols > 6 else '']),
        ('C7 Manopera', ['c7', 'manopera'], ['Cod', 'Meserie', 'Tarif lei/ora'],
         lambda sh, r: [_split_cod(sh.cell_value(r, 1))[0], _split_cod(sh.cell_value(r, 1))[1], _num(sh, r, 3)]),
        ('C8 Utilaje', ['c8', 'utilaje'], ['Cod', 'Utilaj', 'Tarif unitar'],
         lambda sh, r: [_split_cod(sh.cell_value(r, 1))[0], _split_cod(sh.cell_value(r, 1))[1], _num(sh, r, 3)]),
        ('C9 Transport', ['c9', 'transport'], ['Cod', 'Tip transport', 'Tarif unitar'],
         lambda sh, r: [_split_cod(sh.cell_value(r, 1))[0], _split_cod(sh.cell_value(r, 1))[1], _num(sh, r, 5)]),
        ('F4 Echipamente', ['f4', 'echipamente'], ['Denumire', 'U.M.', 'Cantitate', 'Pret unitar'],
         lambda sh, r: [_split_cod(sh.cell_value(r, 1))[1], str(sh.cell_value(r, 2)),
                        _num(sh, r, 3), _num(sh, r, 4)]),
    ]
    fisiere = {}
    for root, _d, files in os.walk(director):
        for f in files:
            if f.lower().endswith('.xls'):
                fisiere[f.lower()] = os.path.join(root, f)

    out = []
    for titlu, kws, headers, extrage in specs:
        path = next((p for n, p in fisiere.items() if any(k in n for k in kws)), None)
        if not path:
            continue
        sh = xlrd.open_workbook(path).sheet_by_index(0)
        rows = []
        for r in range(sh.nrows):
            if _is_nr(sh.cell_value(r, 0)) and str(sh.cell_value(r, 1)).strip():
                try:
                    rows.append(extrage(sh, r))
                except Exception:
                    continue
        out.append((titlu, rows, headers))
    return out


def _num(sh, r, c):
    try:
        return float(sh.cell_value(r, c))
    except (ValueError, TypeError, IndexError):
        return 0.0
