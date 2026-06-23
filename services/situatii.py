"""
Serviciu pentru generarea + exportul situatiilor lunare (Faza 12).

3 functii publice:
  - genereaza_situatie(contract_id, an, luna, user_id) -> SituatieLunara
    Agrega cantitatile EXECUTATE+VALIDATE ale lunii X pentru toate pozitiile
    BoQ din contract. Creeaza/actualizeaza SituatieLunara cu valoare totala
    + cumulat la zi + procent avans.

  - export_situatie_xlsx(situatie_id) -> str (path absolut)
    Export Excel format romanesc (Antet + tabel pozitii + recapitulatie).

  - export_situatie_pdf(situatie_id) -> str (path absolut)
    Export PDF cu acelasi continut (reportlab).

Atentie: NU sterge situatii existente. Daca exista deja o situatie pentru
(proiect, an, luna), o reutilizeaza si actualizeaza valorile.
"""

from __future__ import annotations

import os
from datetime import date, datetime
from decimal import Decimal
from typing import Optional

from flask import current_app

from models import (
    db, Contract, OfertaContract, PozitieBoQ, CantitateExecutataLunara,
    SituatieLunara,
)


LUNI_RO = {
    1: 'Ianuarie', 2: 'Februarie', 3: 'Martie', 4: 'Aprilie',
    5: 'Mai', 6: 'Iunie', 7: 'Iulie', 8: 'August',
    9: 'Septembrie', 10: 'Octombrie', 11: 'Noiembrie', 12: 'Decembrie',
}


def genereaza_situatie(contract_id: int, an: int, luna: int,
                       user_id: Optional[int] = None,
                       doar_validate: bool = True) -> SituatieLunara:
    """
    Genereaza (sau regenereaza) o SituatieLunara pentru (contract, an, luna).

    Algoritm:
      1. Verifica existenta SituatieLunara pentru (proiect_id, an, luna).
         Daca exista, o reutilizeaza (update); altfel creeaza una noua.
      2. Pentru fiecare PozitieBoQ asociata contractului (via oferte),
         agreaga cantitatile executate ale lunii X (filtru validat=True
         daca doar_validate=True; altfel toate).
      3. Calculeaza valoare_totala_luna = sum(cant_luna * pret_unitar).
      4. Calculeaza valoare_cumulat_la_zi = sum din toate lunile <= (an, luna).
      5. Procent avans = cumulat / valoare_totala_oferta_aprobata * 100.
      6. Status default: draft (poate fi schimbat ulterior).

    Returneaza SituatieLunara (commit-uit).
    """
    contract = Contract.query.get(contract_id)
    if contract is None:
        raise ValueError(f'Contract id={contract_id} nu exista.')

    # 1. Caut situatia existenta sau creez una noua
    situatie = SituatieLunara.query.filter_by(
        proiect_id=contract.proiect_id, an=an, luna=luna,
    ).first()
    if situatie is None:
        situatie = SituatieLunara(
            proiect_id=contract.proiect_id,
            contract_id=contract.id,
            an=an, luna=luna,
            data_emitere=date.today(),
            status='draft',
            creat_de_id=user_id,
        )
        db.session.add(situatie)
        db.session.flush()

    # 2. Identific TOATE pozitiile BoQ asociate contractului
    pozitie_ids_query = db.session.query(PozitieBoQ.id).join(
        OfertaContract, PozitieBoQ.oferta_id == OfertaContract.id
    ).filter(OfertaContract.contract_id == contract.id)
    pozitie_ids = [pid for (pid,) in pozitie_ids_query.all()]

    if not pozitie_ids:
        situatie.valoare_totala_luna = Decimal('0')
        situatie.valoare_cumulat_la_zi = Decimal('0')
        situatie.procent_avans_total = Decimal('0')
        db.session.commit()
        return situatie

    # 3. Cantitati lunare pentru luna X
    q_luna = CantitateExecutataLunara.query.filter(
        CantitateExecutataLunara.pozitie_boq_id.in_(pozitie_ids),
        CantitateExecutataLunara.an == an,
        CantitateExecutataLunara.luna == luna,
    )
    if doar_validate:
        q_luna = q_luna.filter(CantitateExecutataLunara.validat == True)
    cantitati_luna = q_luna.all()

    valoare_luna = Decimal('0')
    for c in cantitati_luna:
        # Calc + persist valoare_calculata pe cantitate daca lipseste
        pret = c.pozitie_boq.pret_unitar or Decimal('0')
        val = (c.cantitate_executata or Decimal('0')) * pret
        c.valoare_calculata = val
        valoare_luna += val

    # 4. Cumulat la zi: toate cantitatile (an, luna) <= (curentul)
    q_cumul = CantitateExecutataLunara.query.filter(
        CantitateExecutataLunara.pozitie_boq_id.in_(pozitie_ids),
        db.or_(
            CantitateExecutataLunara.an < an,
            db.and_(
                CantitateExecutataLunara.an == an,
                CantitateExecutataLunara.luna <= luna,
            ),
        ),
    )
    if doar_validate:
        q_cumul = q_cumul.filter(CantitateExecutataLunara.validat == True)
    cumul = Decimal('0')
    for c in q_cumul.all():
        pret = c.pozitie_boq.pret_unitar or Decimal('0')
        cumul += (c.cantitate_executata or Decimal('0')) * pret

    # 5. Valoare oferta aprobata (totala contract)
    valoare_oferta = db.session.query(
        db.func.sum(OfertaContract.valoare_totala)
    ).filter(
        OfertaContract.contract_id == contract.id,
        OfertaContract.aprobata == True,
    ).scalar()
    if not valoare_oferta or valoare_oferta == 0:
        # Fallback: cea mai recenta versiune oferta (aprobata sau nu)
        cea_recenta = OfertaContract.query.filter_by(
            contract_id=contract.id
        ).order_by(OfertaContract.versiune.desc()).first()
        valoare_oferta = (cea_recenta.valoare_totala
                          if cea_recenta and cea_recenta.valoare_totala
                          else Decimal('0'))

    procent = Decimal('0')
    if valoare_oferta and valoare_oferta > 0:
        procent = (cumul / valoare_oferta * 100).quantize(Decimal('0.01'))

    situatie.valoare_totala_luna = valoare_luna
    situatie.valoare_cumulat_la_zi = cumul
    situatie.procent_avans_total = procent

    # Deviz Faza 3 - retentii + garantii pe situatie (gated pe flag).
    # Cu OFF: nu atingem coloanele noi (raman NULL), situatia ramane ca azi.
    _aplica_retentii_garantii(situatie, contract, valoare_luna)

    db.session.commit()
    return situatie


def _aplica_retentii_garantii(situatie: SituatieLunara, contract: Contract,
                              valoare_luna: Decimal) -> None:
    """
    Calculeaza retentia, garantia de buna executie si plata neta a lunii.

    Gated pe flag 'situatii-retentii' (default OFF). Cu OFF, coloanele noi raman
    NULL si situatia ramane identica cu cea istorica (zero regresie).

    Formula RO (situatie de lucrari):
        retentie_suma    = valoare_totala_luna * retentie_procent / 100
        garantie_bex_suma= valoare_totala_luna * garantie_bex_procent / 100
        plata_neta       = valoare_totala_luna - retentie_suma
                                               - garantie_bex_suma
                                               - avans_recuperat

    Procentele implicite vin de pe contract (retentie_procent_default,
    garantie_bex_procent). avans_recuperat se pastreaza ca atare daca a fost
    introdus manual (altfel 0) - recuperarea avansului e o decizie comerciala,
    nu o derivam automat.

    Discriminator de editare manuala: coloana explicita
    SituatieLunara.retentii_editate_manual, setata DOAR de ruta
    situatie_retentii. Cand e False/NULL (inclusiv la auto-generarea anterioara),
    recalculam INTOTDEAUNA retentie_suma + garantie_bex_suma din
    valoare_luna * procent, ca sumele sa urmareasca valoarea lunii la fiecare
    regenerare (ex. cand se valideaza cantitati noi). Cand e True, pastram sumele
    introduse manual si recalculam doar plata neta. NU folosim 'sumele sunt
    non-NULL' ca discriminator: prima auto-populare le face non-NULL si s-ar
    masca drept editare manuala (sume inghetate, plata neta desincronizata).
    """
    try:
        from services.feature_flags import is_enabled
        activ = is_enabled('situatii-retentii')
    except Exception:
        activ = False
    if not activ:
        return

    editat_manual = bool(situatie.retentii_editate_manual)

    if editat_manual:
        # Editare manuala reala (via ruta): pastram procentul + sumele introduse,
        # recalculam doar plata neta din valoarea lunii curente.
        retentie_procent = Decimal(situatie.retentie_procent or 0)
        retentie_suma = Decimal(situatie.retentie_suma or 0)
        garantie_bex = Decimal(situatie.garantie_bex_suma or 0)
    else:
        # Auto-generare: recalculam mereu sumele din procent * valoare_luna.
        # Procent retentie: valoarea persistata pe situatie are prioritate (ex.
        # mostenita la o generare anterioara), altfel din contract; fallback 0.
        if situatie.retentie_procent is not None:
            retentie_procent = Decimal(situatie.retentie_procent)
        else:
            retentie_procent = Decimal(contract.retentie_procent_default or 0)
        garantie_bex_procent = Decimal(contract.garantie_bex_procent or 0)
        retentie_suma = (valoare_luna * retentie_procent / 100).quantize(Decimal('0.01'))
        garantie_bex = (valoare_luna * garantie_bex_procent / 100).quantize(Decimal('0.01'))

    avans_recuperat = Decimal(situatie.avans_recuperat or 0)
    plata_neta = (valoare_luna - retentie_suma - garantie_bex - avans_recuperat)

    situatie.retentie_procent = retentie_procent
    situatie.retentie_suma = retentie_suma
    situatie.garantie_bex_suma = garantie_bex
    situatie.avans_recuperat = avans_recuperat
    situatie.plata_neta = plata_neta.quantize(Decimal('0.01'))


def _get_upload_dir(subdir: str) -> str:
    """Returneaza calea catre uploads/<subdir>/ si o creeaza daca lipseste."""
    base = current_app.config.get(
        'UPLOAD_FOLDER',
        os.path.join(current_app.root_path, 'uploads')
    )
    path = os.path.join(base, subdir)
    os.makedirs(path, exist_ok=True)
    return path


def _get_situatie_data(situatie: SituatieLunara,
                       doar_validate: bool = False) -> dict:
    """
    Construieste structura de date completa pentru exporturi.

    doar_validate:
      - False (default): include TOATE cantitatile lunii (validate sau nu).
        Comportament istoric, folosit de exportul/afisarea situatiei clasice.
      - True: filtreaza pe validat==True, IDENTIC cu genereaza_situatie
        (default doar_validate=True). Folosit de formularele F1/F2/F3 ca
        total_luna F sa se reconcilieze exact cu situatie.valoare_totala_luna
        (formularele oficiale HG 907 trimise beneficiarului nu trebuie sa
        contina cantitati neaprobate).
    """
    contract = situatie.contract
    proiect = situatie.proiect
    # Toate pozitiile BoQ pentru contract via oferte
    pozitii = PozitieBoQ.query.join(
        OfertaContract, PozitieBoQ.oferta_id == OfertaContract.id
    ).filter(OfertaContract.contract_id == contract.id).order_by(
        PozitieBoQ.cod_capitol, PozitieBoQ.ordine
    ).all()

    # Cantitati pentru luna situatiei. Cu doar_validate=True aplicam acelasi
    # filtru ca genereaza_situatie, ca formularele F sa nu publice cantitati
    # nevalidate si sa se reconcilieze cu valoare_totala_luna.
    q_luna = CantitateExecutataLunara.query.filter(
        CantitateExecutataLunara.pozitie_boq_id.in_([p.id for p in pozitii]),
        CantitateExecutataLunara.an == situatie.an,
        CantitateExecutataLunara.luna == situatie.luna,
    )
    if doar_validate:
        q_luna = q_luna.filter(CantitateExecutataLunara.validat == True)
    cantitati_luna_map = {}
    for cant in q_luna.all():
        cantitati_luna_map[cant.pozitie_boq_id] = cant

    # Cumul la zi: sum cantitati <= (an, luna) per pozitie (acelasi filtru validat)
    q_cumul = CantitateExecutataLunara.query.filter(
        CantitateExecutataLunara.pozitie_boq_id.in_([p.id for p in pozitii]),
        db.or_(
            CantitateExecutataLunara.an < situatie.an,
            db.and_(
                CantitateExecutataLunara.an == situatie.an,
                CantitateExecutataLunara.luna <= situatie.luna,
            ),
        ),
    )
    if doar_validate:
        q_cumul = q_cumul.filter(CantitateExecutataLunara.validat == True)
    cumul_map: dict[int, Decimal] = {}
    for cant in q_cumul.all():
        cumul_map[cant.pozitie_boq_id] = (
            cumul_map.get(cant.pozitie_boq_id, Decimal('0'))
            + (cant.cantitate_executata or Decimal('0'))
        )

    rows = []
    for p in pozitii:
        cant_luna = cantitati_luna_map.get(p.id)
        cant_cumul = cumul_map.get(p.id, Decimal('0'))
        val_luna = ((cant_luna.cantitate_executata or Decimal('0'))
                    * (p.pret_unitar or Decimal('0'))) if cant_luna else Decimal('0')
        val_cumul = cant_cumul * (p.pret_unitar or Decimal('0'))
        val_oferta = (p.cantitate_oferta or Decimal('0')) * (p.pret_unitar or Decimal('0'))
        rows.append({
            'cod_articol': p.cod_articol,
            'cod_capitol': p.cod_capitol or '',
            'denumire': p.denumire,
            'um': p.um,
            'cant_oferta': p.cantitate_oferta or Decimal('0'),
            'pret_unitar': p.pret_unitar or Decimal('0'),
            'val_oferta': val_oferta,
            'cant_luna': cant_luna.cantitate_executata if cant_luna else Decimal('0'),
            'val_luna': val_luna,
            'cant_cumul': cant_cumul,
            'val_cumul': val_cumul,
            'categorie': p.categorie,
            'categorie_lucrare': p.categorie_lucrare,
        })

    return {
        'situatie': situatie,
        'contract': contract,
        'proiect': proiect,
        'rows': rows,
        'luna_text': LUNI_RO.get(situatie.luna, str(situatie.luna)),
    }


def export_situatie_xlsx(situatie_id: int) -> str:
    """Export Excel pentru SituatieLunara. Returneaza path-ul fisierului."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    situatie = SituatieLunara.query.get(situatie_id)
    if situatie is None:
        raise ValueError(f'Situatie id={situatie_id} nu exista.')

    data = _get_situatie_data(situatie)
    rows = data['rows']
    contract = data['contract']
    proiect = data['proiect']

    wb = Workbook()
    ws = wb.active
    ws.title = f'Situatie {situatie.an}-{situatie.luna:02d}'

    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    gold_fill = PatternFill('solid', fgColor='C9A961')
    thin = Side(border_style='thin', color='888888')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)

    # ---- Antet ----
    ws['A1'] = 'SITUATIE DE LUCRARI'
    ws['A1'].font = Font(bold=True, size=14, color='0B1426')
    ws.merge_cells('A1:K1')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    ws['A2'] = f'Luna: {data["luna_text"]} {situatie.an}'
    ws['A2'].font = bold
    ws.merge_cells('A2:E2')
    ws['F2'] = f'Numar situatie: {situatie.numar_situatie or "-"}'
    ws.merge_cells('F2:K2')
    ws['A3'] = f'Proiect: {proiect.cod_proiect} - {proiect.nume}'
    ws.merge_cells('A3:K3')
    ws['A4'] = f'Contract: {contract.nr_contract}'
    ws['F4'] = f'Beneficiar: {contract.beneficiar or "-"}'
    ws.merge_cells('A4:E4')
    ws.merge_cells('F4:K4')

    # ---- Header tabel ----
    hdr_row = 6
    headers = [
        'Cod articol', 'Capitol', 'Denumire', 'UM',
        'Cant. ofertă', 'Preț unitar', 'Val. ofertă',
        'Cant. lună', 'Val. lună', 'Cant. cumul.', 'Val. cumul.',
    ]
    for i, h in enumerate(headers, start=1):
        cell = ws.cell(row=hdr_row, column=i, value=h)
        cell.font = bold
        cell.fill = gold_fill
        cell.alignment = center
        cell.border = bord
    ws.row_dimensions[hdr_row].height = 30

    # ---- Pozitii ----
    r = hdr_row + 1
    total_luna = Decimal('0')
    total_cumul = Decimal('0')
    total_oferta = Decimal('0')
    for row in rows:
        # Skip pozitiile care nu au activitate lunara SI nu au cumul
        # (filtru optional, pastram pentru transparenta - aici lasam toate)
        ws.cell(row=r, column=1, value=row['cod_articol']).border = bord
        ws.cell(row=r, column=2, value=row['cod_capitol']).border = bord
        ws.cell(row=r, column=3, value=row['denumire']).border = bord
        ws.cell(row=r, column=4, value=row['um']).alignment = center
        ws.cell(row=r, column=4).border = bord
        ws.cell(row=r, column=5, value=float(row['cant_oferta'])).border = bord
        ws.cell(row=r, column=6, value=float(row['pret_unitar'])).border = bord
        ws.cell(row=r, column=7, value=float(row['val_oferta'])).border = bord
        ws.cell(row=r, column=8, value=float(row['cant_luna'])).border = bord
        ws.cell(row=r, column=9, value=float(row['val_luna'])).border = bord
        ws.cell(row=r, column=10, value=float(row['cant_cumul'])).border = bord
        ws.cell(row=r, column=11, value=float(row['val_cumul'])).border = bord
        total_luna += row['val_luna']
        total_cumul += row['val_cumul']
        total_oferta += row['val_oferta']
        r += 1

    # ---- Totaluri ----
    ws.cell(row=r, column=1, value='TOTAL').font = bold
    ws.cell(row=r, column=7, value=float(total_oferta)).font = bold
    ws.cell(row=r, column=9, value=float(total_luna)).font = bold
    ws.cell(row=r, column=11, value=float(total_cumul)).font = bold
    for col in range(1, 12):
        ws.cell(row=r, column=col).fill = gold_fill
        ws.cell(row=r, column=col).border = bord

    # ---- Recapitulatie ----
    r += 2
    ws.cell(row=r, column=1, value='Procent avans total:').font = bold
    ws.cell(row=r, column=2,
            value=f'{situatie.procent_avans_total}%' if situatie.procent_avans_total else '-')
    ws.cell(row=r + 1, column=1, value='Status:').font = bold
    ws.cell(row=r + 1, column=2, value=situatie.status)

    # Latimi coloane
    widths = [16, 16, 40, 6, 12, 12, 14, 12, 14, 12, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w

    # ---- Salvez ----
    upload_dir = _get_upload_dir('situatii')
    filename = f'situatie_{situatie.proiect_id}_{situatie.an}_{situatie.luna:02d}_{datetime.utcnow():%Y%m%d%H%M%S}.xlsx'
    path = os.path.join(upload_dir, filename)
    wb.save(path)
    situatie.fisier_export_xlsx_path = path
    db.session.commit()
    return path


def export_situatie_pdf(situatie_id: int) -> str:
    """Export PDF pentru SituatieLunara. Returneaza path-ul fisierului."""
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )
    from reportlab.lib.units import mm

    situatie = SituatieLunara.query.get(situatie_id)
    if situatie is None:
        raise ValueError(f'Situatie id={situatie_id} nu exista.')

    data = _get_situatie_data(situatie)
    rows = data['rows']
    contract = data['contract']
    proiect = data['proiect']

    upload_dir = _get_upload_dir('situatii')
    filename = f'situatie_{situatie.proiect_id}_{situatie.an}_{situatie.luna:02d}_{datetime.utcnow():%Y%m%d%H%M%S}.pdf'
    path = os.path.join(upload_dir, filename)

    # Rapoarte Faza 2: branding Cinzel/header gated pe flag. Cu OFF, PDF-ul
    # ramane identic cu cel istoric (Helvetica, fara header brandat).
    try:
        from services.feature_flags import is_enabled
        branding_on = is_enabled('rapoarte-pdf-cinzel')
    except Exception:
        branding_on = False

    # Fontul header-ului tabelului: serif brandat cu ON, Helvetica-Bold cu OFF.
    header_font = 'Helvetica-Bold'
    if branding_on:
        try:
            from rapoarte import brand
            _serif, header_font = brand.get_pdf_fonts()
        except Exception:
            header_font = 'Helvetica-Bold'

    doc = SimpleDocTemplate(
        path, pagesize=landscape(A4),
        leftMargin=10 * mm, rightMargin=10 * mm,
        topMargin=10 * mm, bottomMargin=10 * mm,
    )
    elems = []
    styles = getSampleStyleSheet()
    if branding_on:
        # Header brandat Edifico (logo daca exista + wordmark Cinzel + titlu).
        from rapoarte import brand
        elems += brand.pdf_header_elements('SITUATIE DE LUCRARI', cu_logo=True)
    else:
        title_style = ParagraphStyle(
            'title', parent=styles['Title'],
            fontSize=14, textColor=colors.HexColor('#0B1426'),
            alignment=1,
        )
        elems.append(Paragraph('SITUATIE DE LUCRARI', title_style))
        elems.append(Spacer(1, 4 * mm))
    elems.append(Paragraph(
        f'<b>Luna:</b> {data["luna_text"]} {situatie.an} &nbsp;&nbsp; '
        f'<b>Nr. situatie:</b> {situatie.numar_situatie or "-"}',
        styles['Normal']))
    elems.append(Paragraph(
        f'<b>Proiect:</b> {proiect.cod_proiect} - {proiect.nume}',
        styles['Normal']))
    elems.append(Paragraph(
        f'<b>Contract:</b> {contract.nr_contract} &nbsp;&nbsp; '
        f'<b>Beneficiar:</b> {contract.beneficiar or "-"}',
        styles['Normal']))
    elems.append(Spacer(1, 4 * mm))

    # Tabel pozitii (compact)
    table_data = [[
        'Cod', 'Capitol', 'Denumire', 'UM',
        'Cant.of.', 'P.unit.', 'Val.of.',
        'Cant.luna', 'Val.luna', 'Cant.cum.', 'Val.cum.',
    ]]
    total_luna = Decimal('0')
    total_cumul = Decimal('0')
    total_oferta = Decimal('0')
    for row in rows:
        table_data.append([
            row['cod_articol'],
            row['cod_capitol'][:14] if row['cod_capitol'] else '',
            row['denumire'][:55] + ('...' if len(row['denumire']) > 55 else ''),
            row['um'],
            f'{row["cant_oferta"]:.2f}',
            f'{row["pret_unitar"]:.2f}',
            f'{row["val_oferta"]:.2f}',
            f'{row["cant_luna"]:.2f}',
            f'{row["val_luna"]:.2f}',
            f'{row["cant_cumul"]:.2f}',
            f'{row["val_cumul"]:.2f}',
        ])
        total_luna += row['val_luna']
        total_cumul += row['val_cumul']
        total_oferta += row['val_oferta']
    table_data.append([
        'TOTAL', '', '', '', '', '',
        f'{total_oferta:.2f}', '', f'{total_luna:.2f}', '', f'{total_cumul:.2f}',
    ])
    tbl = Table(table_data, repeatRows=1)
    tbl.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#C9A961')),
        ('TEXTCOLOR', (0, 0), (-1, 0),
         colors.HexColor('#0B1426') if branding_on else colors.white),
        ('FONTNAME', (0, 0), (-1, 0), header_font),
        ('FONTSIZE', (0, 0), (-1, 0), 8),
        ('FONTSIZE', (0, 1), (-1, -1), 7),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('ALIGN', (4, 1), (-1, -1), 'RIGHT'),
        ('GRID', (0, 0), (-1, -1), 0.25, colors.HexColor('#888888')),
        ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#F5F1E8')),
        ('FONTNAME', (0, -1), (-1, -1), header_font),
    ]))
    elems.append(tbl)
    elems.append(Spacer(1, 6 * mm))

    elems.append(Paragraph(
        f'<b>Procent avans total:</b> '
        f'{situatie.procent_avans_total or 0}% &nbsp;&nbsp; '
        f'<b>Status:</b> {situatie.status}',
        styles['Normal']))

    doc.build(elems)

    situatie.fisier_export_pdf_path = path
    db.session.commit()
    return path


# ============================================================
# FORMULARE F1 / F2 / F3 (HG 907/2016) pe situatie de lucrari (dz-4)
#
# Recapitulatii ierarhice obiect (disciplina) -> categorie de lucrare ->
# articol, peste cantitatile EXECUTATE in luna situatiei. Reutilizeaza
# services.centralizator.recapitulatie_ierarhica. Toate valorile sunt FARA
# TVA (pozitiile); TVA (21%) se adauga o singura data la final pe F1
# (playbook deviz §2.6). Verificare matematica: Sigma articole == val luna.
#
# Semantica formularelor (situatie de lucrari):
#   F3 - Lista cu cantitati de lucrari: articolele detaliate, grupate pe
#        obiect -> categorie de lucrare (cant. luna, pret unitar, valoare).
#   F2 - Centralizatorul cheltuielilor pe categorii de lucrari: recapitulatie
#        obiect -> categorie (subtotaluri pe obiect), fara articole.
#   F1 - Centralizatorul cheltuielilor pe obiectiv: un rand per obiect +
#        TOTAL fara TVA + TVA + TOTAL cu TVA.
#
# Gating: rutele de export sunt gated pe flag 'situatii-f-forms'; serviciul in
# sine e pur (nu citeste flagul), ca sa fie testabil direct.
# ============================================================

COTA_TVA_F = Decimal('21')  # % standard RO (formularele F)
_Q2_F = Decimal('0.01')


def _f_linii_situatie(situatie_id: int) -> dict:
    """
    Construieste liniile (articolele) lunii situatiei pentru formularele F.

    Returneaza:
        {situatie, contract, proiect, luna_text,
         linii: [{disciplina, categorie, cod_articol, denumire, um,
                  cant_luna, pret_unitar, valoare}],
         total_luna: Decimal}

    Filtreaza pozitiile FARA activitate in luna (cant_luna == 0): formularele F
    reflecta strict lucrarea executata si valorificata in luna respectiva.
    Valoarea articolului = cant_luna * pret_unitar (FARA TVA).

    Foloseste doar_validate=True: doar cantitatile validate (aceeasi baza ca
    genereaza_situatie), ca Sigma articole F == situatie.valoare_totala_luna.
    """
    situatie = SituatieLunara.query.get(situatie_id)
    if situatie is None:
        raise ValueError(f'Situatie id={situatie_id} nu exista.')

    from services.deviz_pricing import deduce_disciplina

    data = _get_situatie_data(situatie, doar_validate=True)
    linii = []
    total_luna = Decimal('0')
    for row in data['rows']:
        cant_luna = row['cant_luna'] or Decimal('0')
        if cant_luna <= 0:
            continue  # doar lucrarea executata in luna
        val = row['val_luna'] or Decimal('0')
        linii.append({
            'disciplina': deduce_disciplina(row['cod_capitol']),
            'categorie': row['categorie_lucrare'] or 'neclasificat',
            'cod_articol': row['cod_articol'],
            'cod_capitol': row['cod_capitol'],
            'denumire': row['denumire'],
            'um': row['um'],
            'cant_luna': cant_luna,
            'pret_unitar': row['pret_unitar'] or Decimal('0'),
            'valoare': val,
        })
        total_luna += val

    return {
        'situatie': situatie,
        'contract': data['contract'],
        'proiect': data['proiect'],
        'luna_text': data['luna_text'],
        'linii': linii,
        'total_luna': total_luna.quantize(_Q2_F),
    }


def _antet_f(ws, titlu: str, ctx: dict, ultima_col: str):
    """Scrie antetul comun al formularelor F (titlu + identificare situatie)."""
    from openpyxl.styles import Font, Alignment
    situatie = ctx['situatie']
    contract = ctx['contract']
    proiect = ctx['proiect']
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws['A1'] = titlu
    ws['A1'].font = Font(bold=True, size=14, name='Arial', color='0B1426')
    ws.merge_cells(f'A1:{ultima_col}1')
    ws['A1'].alignment = center
    ws.row_dimensions[1].height = 26

    ws['A2'] = f'Proiect: {proiect.cod_proiect} - {proiect.nume}'
    ws.merge_cells(f'A2:{ultima_col}2')
    ws['A3'] = (f'Contract: {contract.nr_contract}  |  '
                f'Situatie: {situatie.numar_situatie or "-"}  |  '
                f'Luna: {ctx["luna_text"]} {situatie.an}')
    ws.merge_cells(f'A3:{ultima_col}3')


def genereaza_f3(situatie_id: int) -> dict:
    """
    F3 - Lista cu cantitati de lucrari (articole detaliate, grupate pe obiect ->
    categorie de lucrare). Reutilizeaza recapitulatie_ierarhica pentru subtotaluri.

    Returneaza:
        {grupe: [{disciplina, subtotal,
                  categorii: [{categorie, valoare, nr, articole: [...]}]}],
         total_general, nr_pozitii, ctx}
    """
    from services.centralizator import recapitulatie_ierarhica

    ctx = _f_linii_situatie(situatie_id)
    linii = ctx['linii']
    recap = recapitulatie_ierarhica(linii)

    # Atasez articolele detaliate la fiecare (disciplina, categorie).
    art_idx: dict[tuple, list] = {}
    for ln in linii:
        art_idx.setdefault((ln['disciplina'], ln['categorie'] or 'neclasificat'), []).append(ln)

    for grup in recap['grupe']:
        for cat in grup['categorii']:
            arts = art_idx.get((grup['disciplina'], cat['categorie']), [])
            arts_sortate = sorted(arts, key=lambda x: x['cod_articol'] or '')
            articole = [
                {'cod_articol': a['cod_articol'],
                 'denumire': a['denumire'],
                 'um': a['um'],
                 'cant_luna': a['cant_luna'],
                 'pret_unitar': a['pret_unitar'],
                 'valoare': (a['valoare'] or Decimal('0')).quantize(_Q2_F)}
                for a in arts_sortate
            ]
            # Reconciliere rotunjire: suma valorilor de articol AFISATE trebuie
            # sa fie EXACT subtotalul categoriei (cat['valoare'], rotunjit din
            # suma RAW). Daca rotunjirea per-articol a introdus o diferenta de
            # bani, o absorbim in ultimul articol al categoriei, ca beneficiarul
            # care aduna articolele tiparite sa obtina fix subtotalul.
            if articole:
                suma_afisata = sum((a['valoare'] for a in articole), Decimal('0'))
                rest = cat['valoare'] - suma_afisata
                if rest != 0:
                    articole[-1]['valoare'] = (
                        articole[-1]['valoare'] + rest).quantize(_Q2_F)
            cat['articole'] = articole
    recap['ctx'] = ctx
    return recap


def genereaza_f2(situatie_id: int) -> dict:
    """
    F2 - Centralizatorul cheltuielilor pe categorii de lucrari (recapitulatie
    obiect -> categorie, subtotaluri pe obiect, fara articole).
    """
    from services.centralizator import recapitulatie_ierarhica
    ctx = _f_linii_situatie(situatie_id)
    recap = recapitulatie_ierarhica(ctx['linii'])
    recap['ctx'] = ctx
    return recap


def genereaza_f1(situatie_id: int, cota_tva: Decimal = COTA_TVA_F) -> dict:
    """
    F1 - Centralizatorul cheltuielilor pe obiectiv (un rand per obiect/disciplina
    + TOTAL fara TVA + TVA + TOTAL cu TVA). TVA aplicat o singura data la final.

    Sursa unica de adevar pentru totaluri: recapitulatie_ierarhica (acelasi
    quantize GLOBAL ca F2/F3). Randurile per-obiect = subtotalurile recap (deja
    quantize din suma RAW), iar total_fara_tva = recap['total_general'].
    Asa F1.total_fara_tva == F2.total_general == F3.total_general, fara
    acumulare de rotunjire din subtotalurile per-obiect.
    """
    from decimal import ROUND_HALF_UP
    from services.centralizator import recapitulatie_ierarhica
    cota_tva = Decimal(str(cota_tva))
    ctx = _f_linii_situatie(situatie_id)

    recap = recapitulatie_ierarhica(ctx['linii'])
    randuri = [
        {'obiect': grup['disciplina'], 'valoare': grup['subtotal']}
        for grup in sorted(recap['grupe'], key=lambda g: g['disciplina'])
    ]
    total_fara_tva = recap['total_general']
    tva = (total_fara_tva * cota_tva / 100).quantize(_Q2_F, rounding=ROUND_HALF_UP)
    total_cu_tva = total_fara_tva + tva
    return {
        'randuri': randuri,
        'total_fara_tva': total_fara_tva.quantize(_Q2_F),
        'cota_tva': cota_tva,
        'tva': tva,
        'total_cu_tva': total_cu_tva.quantize(_Q2_F),
        'ctx': ctx,
    }


def export_f3(situatie_id: int) -> str:
    """Export Excel F3 (Lista cu cantitati de lucrari). Returneaza path absolut."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = genereaza_f3(situatie_id)
    ctx = data['ctx']
    situatie = ctx['situatie']

    wb = Workbook()
    ws = wb.active
    ws.title = 'F3'
    _antet_f(ws, 'FORMULAR F3 - LISTA CU CANTITATI DE LUCRARI', ctx, 'G')

    # IMPORTANT: subtotalurile obiect/categorie merg intr-o coloana SEPARATA
    # ('Subtotal') ca sa NU se amestece cu valorile de articol. Asa un SUM pe
    # coloana 'Valoare (fara TVA)' (col 6, doar articole) da EXACT total_general,
    # iar un beneficiar care selecteaza coloana nu obtine ~4x totalul (din
    # dublarea subtotaluri + articole pe aceeasi coloana).
    r = 5
    _styled_header_f(ws, r, ['Cod articol', 'Denumire', 'UM', 'Cant. luna',
                             'Pret unitar', 'Valoare (fara TVA)',
                             'Subtotal (fara TVA)'])
    r += 1
    money = '#,##0.00'
    bold = Font(bold=True, name='Arial')
    for grup in data['grupe']:
        ws.cell(row=r, column=1, value='OBIECT: ' + grup['disciplina'].upper()).font = bold
        vc = ws.cell(row=r, column=7, value=float(grup['subtotal']))
        vc.font = bold; vc.number_format = money
        r += 1
        for cat in grup['categorii']:
            ws.cell(row=r, column=2, value='  ' + cat['categorie']).font = Font(italic=True, name='Arial')
            cc = ws.cell(row=r, column=7, value=float(cat['valoare']))
            cc.number_format = money
            r += 1
            for a in cat['articole']:
                ws.cell(row=r, column=1, value=a['cod_articol'])
                ws.cell(row=r, column=2, value=a['denumire'])
                ws.cell(row=r, column=3, value=a['um'])
                ws.cell(row=r, column=4, value=float(a['cant_luna'])).number_format = money
                ws.cell(row=r, column=5, value=float(a['pret_unitar'])).number_format = money
                ws.cell(row=r, column=6, value=float(a['valoare'])).number_format = money
                r += 1

    r += 1
    ws.cell(row=r, column=1, value='TOTAL F3 (fara TVA)').font = Font(bold=True, size=12, name='Arial')
    tc = ws.cell(row=r, column=7, value=float(data['total_general']))
    tc.font = Font(bold=True, size=12, name='Arial'); tc.number_format = money

    for col, w in zip('ABCDEFG', [18, 46, 8, 14, 14, 18, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A6'  # sub randul de antet real al tabelului
    wb.calculation.fullCalcOnLoad = True
    return _salveaza_f(wb, situatie, 'f3')


def export_f2(situatie_id: int) -> str:
    """Export Excel F2 (Centralizator categorii de lucrari). Returneaza path absolut."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = genereaza_f2(situatie_id)
    ctx = data['ctx']
    situatie = ctx['situatie']

    wb = Workbook()
    ws = wb.active
    ws.title = 'F2'
    _antet_f(ws, 'FORMULAR F2 - CENTRALIZATOR CATEGORII DE LUCRARI', ctx, 'D')

    r = 5
    _styled_header_f(ws, r, ['Obiect', 'Categorie de lucrare', 'Nr. articole',
                             'Valoare (fara TVA)'])
    r += 1
    money = '#,##0.00'
    bold = Font(bold=True, name='Arial')
    for grup in data['grupe']:
        ws.cell(row=r, column=1, value=grup['disciplina'].upper()).font = bold
        vc = ws.cell(row=r, column=4, value=float(grup['subtotal']))
        vc.font = bold; vc.number_format = money
        r += 1
        for cat in grup['categorii']:
            ws.cell(row=r, column=2, value=cat['categorie'])
            ws.cell(row=r, column=3, value=cat['nr'])
            cc = ws.cell(row=r, column=4, value=float(cat['valoare']))
            cc.number_format = money
            r += 1

    r += 1
    ws.cell(row=r, column=1, value='TOTAL F2 (fara TVA)').font = Font(bold=True, size=12, name='Arial')
    tc = ws.cell(row=r, column=4, value=float(data['total_general']))
    tc.font = Font(bold=True, size=12, name='Arial'); tc.number_format = money

    for col, w in zip('ABCD', [22, 32, 12, 18]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A6'
    wb.calculation.fullCalcOnLoad = True
    return _salveaza_f(wb, situatie, 'f2')


def export_f1(situatie_id: int, cota_tva: Decimal = COTA_TVA_F) -> str:
    """Export Excel F1 (Centralizator pe obiectiv + TVA). Returneaza path absolut."""
    from openpyxl import Workbook
    from openpyxl.styles import Font

    data = genereaza_f1(situatie_id, cota_tva)
    ctx = data['ctx']
    situatie = ctx['situatie']

    wb = Workbook()
    ws = wb.active
    ws.title = 'F1'
    _antet_f(ws, 'FORMULAR F1 - CENTRALIZATOR PE OBIECTIV', ctx, 'B')

    r = 5
    _styled_header_f(ws, r, ['Obiect', 'Valoare (fara TVA)'])
    r += 1
    money = '#,##0.00'
    for rand in data['randuri']:
        ws.cell(row=r, column=1, value=rand['obiect'].upper())
        ws.cell(row=r, column=2, value=float(rand['valoare'])).number_format = money
        r += 1

    bold = Font(bold=True, name='Arial')
    ws.cell(row=r, column=1, value='TOTAL (fara TVA)').font = bold
    c = ws.cell(row=r, column=2, value=float(data['total_fara_tva']))
    c.number_format = money; c.font = bold
    r += 1
    ws.cell(row=r, column=1, value=f'TVA ({data["cota_tva"]}%)').font = bold
    c = ws.cell(row=r, column=2, value=float(data['tva']))
    c.number_format = money; c.font = bold
    r += 1
    ws.cell(row=r, column=1, value='TOTAL (cu TVA)').font = Font(bold=True, size=12, name='Arial')
    tc = ws.cell(row=r, column=2, value=float(data['total_cu_tva']))
    tc.font = Font(bold=True, size=12, name='Arial'); tc.number_format = money

    for col, w in zip('AB', [34, 20]):
        ws.column_dimensions[col].width = w
    ws.freeze_panes = 'A6'
    wb.calculation.fullCalcOnLoad = True
    return _salveaza_f(wb, situatie, 'f1')


def _styled_header_f(ws, row, headers, gold='C9A961'):
    """Antet de tabel brandat (gold) pentru formularele F."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    bold = Font(bold=True, color='0B1426', name='Arial')
    fill = PatternFill('solid', fgColor=gold)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin = Side(border_style='thin', color='888888')
    bord = Border(left=thin, right=thin, top=thin, bottom=thin)
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=i, value=h)
        c.font = bold; c.fill = fill; c.alignment = center; c.border = bord
    ws.row_dimensions[row].height = 28


def _salveaza_f(wb, situatie, eticheta: str) -> str:
    """Salveaza workbook-ul F intr-un fisier in uploads/situatii_f/ si intoarce path-ul."""
    upload_dir = _get_upload_dir('situatii_f')
    filename = (f'{eticheta}_situatie_{situatie.id}_{situatie.an}_'
                f'{situatie.luna:02d}_{datetime.utcnow():%Y%m%d%H%M%S}.xlsx')
    path = os.path.join(upload_dir, filename)
    wb.save(path)
    return path
