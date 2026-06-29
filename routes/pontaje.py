"""
Rute pentru gestionarea pontajelor - Modul Complet
Timekeeping pentru constructii cu calcul ore, aprobare, export Excel
"""

import calendar
from datetime import datetime, date, timedelta
from decimal import Decimal
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, abort
from flask_login import login_required, current_user
from werkzeug.exceptions import HTTPException
from models import db, Pontaj, Angajat, Proiect, SarbatoareLegala
from forms.pontaje_forms import PontajForm
from services.security.tenant_access import (
    get_project_or_404,
    get_tenant_mode,
    get_timesheet_or_404,
    query_for_tenant,
    query_timesheets_for_tenant,
    require_timesheet_inputs_same_tenant,
    tenant_id_for_new_record_or_403,
)
from services.timesheet_service import (
    approve_timesheet,
    calculate_timesheet_hours,
    check_timesheet_duplicate,
    create_multiple_timesheets_from_form_data,
    create_timesheet_from_form_data,
    get_daily_timesheet_rows,
    get_project_employees_for_timesheet,
    get_timesheet_approval_context,
    get_timesheet_calendar_context,
    get_timesheet_list_context,
    reject_timesheet,
    submit_timesheet_for_approval,
    update_timesheet_from_form_data,
)

pontaje_bp = Blueprint('pontaje', __name__)


# ============================================================
# LOGICA CALCUL ORE
# ============================================================

def calculate_hours(ora_start, ora_sfarsit, tip_zi, data_pontaj=None):
    """Calculeaza orele lucrate conform legislatiei constructiilor.

    Wrapper subtire peste timesheet_service.calculate_timesheet_hours, pastrat
    pentru compatibilitatea apelantilor existenti din ruta.
    """
    return calculate_timesheet_hours(
        ora_start=ora_start, ora_sfarsit=ora_sfarsit,
        tip_zi=tip_zi, data_pontaj=data_pontaj,
    )


def _detect_tip_zi(data_pontaj):
    """Detecteaza automat tipul zilei."""
    if not data_pontaj:
        return 'lucratoare'
    is_sarb = SarbatoareLegala.query.filter_by(data=data_pontaj).first()
    if is_sarb:
        return 'sarbatoare_legala'
    dow = data_pontaj.weekday()
    if dow == 5:
        return 'sambata'
    elif dow == 6:
        return 'duminica'
    return 'lucratoare'


def _form_int(name, default=0):
    try:
        return int(request.form.get(name) or default)
    except (TypeError, ValueError):
        return default


# ============================================================
# PANOU PRINCIPAL PONTAJ
# ============================================================

@pontaje_bp.route('/')
@login_required
def lista():
    today = date.today()
    # Filtrele vin din query string; contextul (read-only, tenant-safe) este
    # construit in timesheet_service. Ruta ramane responsabila doar de HTTP.
    filters = {
        'luna': request.args.get('luna', today.month, type=int),
        'anul': request.args.get('anul', today.year, type=int),
        'data': request.args.get('data', ''),
        'proiect_id': request.args.get('proiect_id', ''),
        'angajat_id': request.args.get('angajat_id', ''),
        'status': request.args.get('status', ''),
    }
    context = get_timesheet_list_context(filters=filters)
    return render_template('pontaje/panou.html', **context)


# ============================================================
# ADAUGA PONTAJ INDIVIDUAL
# ============================================================

@pontaje_bp.route('/adauga', methods=['GET', 'POST'])
@login_required
def adauga():
    form = PontajForm()

    if request.method == 'POST':
        tenant_id_curent = tenant_id_for_new_record_or_403()
        proiect_id_form = _form_int('proiect_id')
        angajat_id_form = _form_int('angajat_id')
        if proiect_id_form and angajat_id_form:
            require_timesheet_inputs_same_tenant(
                proiect_id=proiect_id_form,
                angajat_id=angajat_id_form,
                tenant_id=tenant_id_curent,
            )

    if form.validate_on_submit():
        try:
            rezultat = create_timesheet_from_form_data(
                form_data=form,
                current_user=current_user,
            )
        except HTTPException:
            db.session.rollback()
            raise
        except Exception:
            db.session.rollback()
            raise

        if rezultat['duplicate']:
            flash('Exista deja un pontaj pentru acest angajat in aceasta zi!', 'danger')
            return render_template('pontaje/formular_individual.html', form=form)

        status = rezultat['timesheet'].status
        if status == 'trimis':
            flash('Pontajul a fost inregistrat si trimis spre aprobare!', 'success')
        else:
            flash('Pontajul a fost salvat ca draft.', 'success')
        return redirect(url_for('pontaje.lista'))

    return render_template('pontaje/formular_individual.html', form=form)


# ============================================================
# ADAUGA PONTAJ IN MASA
# ============================================================

@pontaje_bp.route('/adauga-multiplu', methods=['GET', 'POST'])
@login_required
def adauga_multiplu():
    if request.method == 'POST':
        try:
            rezultat = create_multiple_timesheets_from_form_data(
                form_data=request.form,
                current_user=current_user,
            )
        except HTTPException:
            db.session.rollback()
            raise
        except Exception:
            db.session.rollback()
            raise

        count_ok = rezultat['created_count']
        count_skip = rezultat['skipped_count']
        flash(f'{count_ok} pontaje inregistrate cu succes. {count_skip} duplicate omise.', 'success')
        return redirect(url_for('pontaje.lista'))

    proiecte = query_for_tenant(Proiect).filter(
        Proiect.status.in_(['activ', 'planificat'])
    ).order_by(Proiect.cod_proiect).all()

    return render_template('pontaje/formular_masa.html',
                           proiecte=proiecte,
                           tipuri_zi=Pontaj.TIPURI_ZI,
                           today=date.today())


# ============================================================
# ANGAJATI PE PROIECT (AJAX - pentru formular masa)
# ============================================================

@pontaje_bp.route('/angajati-proiect/<int:proiect_id>')
@login_required
def angajati_proiect(proiect_id):
    """Returneaza angajatii activi pe un proiect (AJAX)."""
    return jsonify(get_project_employees_for_timesheet(project_id=proiect_id))


# ============================================================
# CALCUL ORE AJAX
# ============================================================

@pontaje_bp.route('/calcul-ore')
@login_required
def calcul_ore():
    """Calculeaza orele in timp real (AJAX)."""
    ora_start = request.args.get('ora_start', '08:00')
    ora_sfarsit = request.args.get('ora_sfarsit', '16:00')
    tip_zi = request.args.get('tip_zi', 'lucratoare')
    data_str = request.args.get('data', '')

    data_pontaj = None
    if data_str:
        try:
            data_pontaj = datetime.strptime(data_str, '%Y-%m-%d').date()
        except ValueError:
            pass

    result = calculate_hours(ora_start, ora_sfarsit, tip_zi, data_pontaj)
    return jsonify(result)


# ============================================================
# VERIFICARE DUPLICAT (AJAX)
# ============================================================

@pontaje_bp.route('/verificare-duplicat')
@login_required
def verificare_duplicat():
    """Verifica daca exista deja pontaj pentru angajat in ziua respectiva."""
    angajat_id = request.args.get('angajat_id', 0, type=int)
    data_str = request.args.get('data', '')
    pontaj_id = request.args.get('pontaj_id', 0, type=int)

    if not angajat_id or not data_str:
        return jsonify({'exists': False})

    try:
        data_pontaj = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'exists': False})

    return jsonify(check_timesheet_duplicate(
        employee_id=angajat_id,
        date_value=data_pontaj,
        exclude_timesheet_id=pontaj_id or None,
    ))


# ============================================================
# SITUATIE ZILNICA (AJAX)
# ============================================================

@pontaje_bp.route('/situatie-zilnica')
@login_required
def situatie_zilnica():
    """Returneaza pontajele dintr-o zi (AJAX pentru calendar click)."""
    data_str = request.args.get('data', '')
    if not data_str:
        return jsonify([])

    try:
        data_pontaj = datetime.strptime(data_str, '%Y-%m-%d').date()
    except ValueError:
        return jsonify([])

    return jsonify(get_daily_timesheet_rows(date_value=data_pontaj))


# ============================================================
# CALENDAR ANGAJAT
# ============================================================

@pontaje_bp.route('/calendar')
@login_required
def calendar_view():
    """Vedere calendar lunar per angajat."""
    angajat_id = request.args.get('angajat_id', 0, type=int)
    luna = request.args.get('luna', date.today().month, type=int)
    anul = request.args.get('anul', date.today().year, type=int)

    context = get_timesheet_calendar_context(angajat_id=angajat_id, luna=luna, anul=anul)
    return render_template('pontaje/calendar_angajat.html', **context)


# ============================================================
# APROBARE PONTAJE
# ============================================================

@pontaje_bp.route('/aprobare')
@login_required
def aprobare():
    if not current_user.is_manager:
        flash('Nu aveti permisiunea de a accesa aceasta pagina.', 'danger')
        return redirect(url_for('pontaje.lista'))

    filters = {
        'proiect_id': request.args.get('proiect_id', '', type=str),
        'angajat_id': request.args.get('angajat_id', '', type=str),
        'data_start': request.args.get('data_start', ''),
        'data_sfarsit': request.args.get('data_sfarsit', ''),
    }
    context = get_timesheet_approval_context(filters=filters)
    return render_template('pontaje/aprobare.html', **context)


# ============================================================
# APROBA PONTAJ
# ============================================================

@pontaje_bp.route('/<int:id>/aproba', methods=['POST'])
@login_required
def aproba(id):
    if not current_user.is_manager:
        flash('Nu aveti permisiunea de a aproba pontaje.', 'danger')
        return redirect(url_for('pontaje.lista'))

    pontaj = get_timesheet_or_404(id)
    try:
        approve_timesheet(timesheet=pontaj, current_user=current_user)
    except HTTPException:
        db.session.rollback()
        raise
    except Exception:
        db.session.rollback()
        raise
    flash('Pontajul a fost aprobat!', 'success')

    next_url = request.form.get('next', url_for('pontaje.aprobare'))
    return redirect(next_url)


# ============================================================
# RESPINGE PONTAJ
# ============================================================

@pontaje_bp.route('/<int:id>/respinge', methods=['POST'])
@login_required
def respinge(id):
    if not current_user.is_manager:
        flash('Nu aveti permisiunea de a respinge pontaje.', 'danger')
        return redirect(url_for('pontaje.lista'))

    pontaj = get_timesheet_or_404(id)
    motiv = request.form.get('motiv', '')
    try:
        reject_timesheet(timesheet=pontaj, current_user=current_user, reason=motiv)
    except HTTPException:
        db.session.rollback()
        raise
    except Exception:
        db.session.rollback()
        raise
    flash('Pontajul a fost respins.', 'warning')

    next_url = request.form.get('next', url_for('pontaje.aprobare'))
    return redirect(next_url)


# ============================================================
# APROBARE MULTIPLA
# ============================================================

@pontaje_bp.route('/aproba-multiplu', methods=['POST'])
@login_required
def aproba_multiplu():
    if not current_user.is_manager:
        flash('Nu aveti permisiunea de a aproba pontaje.', 'danger')
        return redirect(url_for('pontaje.lista'))

    ids = request.form.getlist('pontaj_ids')
    count = 0
    if get_tenant_mode() == 'off':
        pontaje = []
        for pid_str in ids:
            pontaj = Pontaj.query.get(int(pid_str))
            if pontaj and pontaj.status == 'trimis':
                pontaje.append(pontaj)
    else:
        try:
            ids_int = [int(pid_str) for pid_str in ids]
        except (TypeError, ValueError):
            ids_int = []
        ids_int = [pid for pid in ids_int if pid > 0]
        if len(ids_int) != len(ids) or not ids_int:
            abort(404)
        pontaje = query_timesheets_for_tenant().filter(
            Pontaj.id.in_(ids_int)
        ).all()
        if len({p.id for p in pontaje}) != len(set(ids_int)):
            abort(404)
        pontaje = [p for p in pontaje if p.status == 'trimis']

    for pontaj in pontaje:
        pontaj.status = 'aprobat'
        pontaj.aprobat_de = current_user.id
        pontaj.data_aprobare = datetime.utcnow()
        count += 1

    db.session.commit()
    flash(f'{count} pontaje aprobate cu succes!', 'success')
    return redirect(url_for('pontaje.aprobare'))


# ============================================================
# TRIMITE PONTAJ
# ============================================================

@pontaje_bp.route('/<int:id>/trimite', methods=['POST'])
@login_required
def trimite(id):
    pontaj = get_timesheet_or_404(id)
    try:
        rezultat = submit_timesheet_for_approval(timesheet=pontaj)
    except HTTPException:
        db.session.rollback()
        raise
    except Exception:
        db.session.rollback()
        raise

    if rezultat['submitted']:
        flash('Pontajul a fost trimis spre aprobare.', 'info')
    return redirect(url_for('pontaje.lista'))


# ============================================================
# EDITEAZA PONTAJ
# ============================================================

@pontaje_bp.route('/<int:id>/editeaza', methods=['GET', 'POST'])
@login_required
def editeaza(id):
    pontaj = get_timesheet_or_404(id)

    if pontaj.status not in ('draft', 'respins'):
        flash('Doar pontajele draft sau respinse pot fi editate.', 'danger')
        return redirect(url_for('pontaje.lista'))

    form = PontajForm(obj=pontaj)

    if request.method == 'POST':
        tenant_id_curent = tenant_id_for_new_record_or_403()
        proiect_id_form = _form_int('proiect_id')
        angajat_id_form = _form_int('angajat_id')
        if proiect_id_form and angajat_id_form:
            require_timesheet_inputs_same_tenant(
                proiect_id=proiect_id_form,
                angajat_id=angajat_id_form,
                tenant_id=tenant_id_curent,
            )

    if form.validate_on_submit():
        try:
            rezultat = update_timesheet_from_form_data(
                timesheet=pontaj,
                form_data=form,
                current_user=current_user,
            )
        except HTTPException:
            db.session.rollback()
            raise
        except Exception:
            db.session.rollback()
            raise

        if rezultat['duplicate']:
            flash('Exista deja un pontaj pentru acest angajat in aceasta zi!', 'danger')
            return render_template('pontaje/formular_individual.html', form=form, pontaj=pontaj)
        flash('Pontajul a fost actualizat!', 'success')
        return redirect(url_for('pontaje.lista'))

    form.pontaj_id.data = pontaj.id
    return render_template('pontaje/formular_individual.html', form=form, pontaj=pontaj)


# ============================================================
# STERGE PONTAJ
# ============================================================

@pontaje_bp.route('/<int:id>/sterge', methods=['POST'])
@login_required
def sterge(id):
    pontaj = get_timesheet_or_404(id)
    if pontaj.status in ('draft', 'respins'):
        db.session.delete(pontaj)
        db.session.commit()
        flash('Pontajul a fost sters.', 'warning')
    else:
        flash('Doar pontajele draft sau respinse pot fi sterse.', 'danger')
    return redirect(url_for('pontaje.lista'))


# ============================================================
# EXPORT EXCEL PONTAJ LUNAR
# ============================================================

@pontaje_bp.route('/export-lunar')
@login_required
def export_lunar():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    luna = request.args.get('luna', date.today().month, type=int)
    anul = request.args.get('anul', date.today().year, type=int)
    proiect_id = request.args.get('proiect_id', 0, type=int)

    _, days_in_month = calendar.monthrange(anul, luna)
    month_names = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
                   'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    # Get pontaje
    query = query_timesheets_for_tenant().filter(
        db.extract('month', Pontaj.data) == luna,
        db.extract('year', Pontaj.data) == anul
    )
    proiect_export = None
    if proiect_id:
        proiect_export = get_project_or_404(proiect_id)
        query = query.filter(Pontaj.proiect_id == proiect_id)

    pontaje = query.all()

    # Group by angajat
    angajat_data = {}
    for p in pontaje:
        if p.angajat_id not in angajat_data:
            angajat_data[p.angajat_id] = {
                'angajat': p.angajat,
                'zile': {},
                'total_ore': 0,
                'ore_normale': 0,
                'ore_supl_50': 0,
                'ore_supl_100': 0
            }
        ad = angajat_data[p.angajat_id]
        ad['zile'][p.data.day] = p
        ad['total_ore'] += float(p.ore_lucrate or 0)
        ad['ore_normale'] += float(p.ore_normale or 0)
        ad['ore_supl_50'] += float(p.ore_suplimentare_50 or 0)
        ad['ore_supl_100'] += float(p.ore_suplimentare_100 or 0)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    weekend_fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
    sarb_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sarbatori luna
    sarbatori = {s.data.day for s in SarbatoareLegala.query.filter(
        db.extract('month', SarbatoareLegala.data) == luna,
        db.extract('year', SarbatoareLegala.data) == anul
    ).all()}

    # --- Sheet 1: Foaie Colectiva ---
    ws1 = wb.active
    ws1.title = 'Foaie Colectiva'
    ws1.sheet_properties.pageSetUpPr.fitToPage = True

    proiect_info = ''
    if proiect_id:
        pr = proiect_export
        if pr:
            proiect_info = f' - {pr.cod_proiect} {pr.nume}'

    # Title
    ws1.merge_cells(start_row=1, start_column=1, end_row=1, end_column=days_in_month + 5)
    title_cell = ws1.cell(row=1, column=1, value=f'FOAIA COLECTIVA DE PREZENTA - {month_names[luna]} {anul}{proiect_info}')
    title_cell.font = Font(bold=True, size=14, color='1a237e')
    title_cell.alignment = Alignment(horizontal='center')

    ws1.merge_cells(start_row=2, start_column=1, end_row=2, end_column=days_in_month + 5)
    ws1.cell(row=2, column=1, value='EDIFICO WORKFORCE - Management Forta de Munca in Constructii').font = Font(size=10, color='666666')
    ws1.cell(row=2, column=1).alignment = Alignment(horizontal='center')

    # Headers row 4
    row_h = 4
    headers = ['Nr.', 'Nume si Prenume']
    for d in range(1, days_in_month + 1):
        dt = date(anul, luna, d)
        day_names = ['Lu', 'Ma', 'Mi', 'Jo', 'Vi', 'Sa', 'Du']
        headers.append(f'{d}\n{day_names[dt.weekday()]}')
    headers.extend(['Total\nOre', 'Ore\nSupl.', 'Semn.'])

    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=row_h, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ws1.column_dimensions['A'].width = 5
    ws1.column_dimensions['B'].width = 25
    for col in range(3, days_in_month + 3):
        ws1.column_dimensions[get_column_letter(col)].width = 4.5
    ws1.column_dimensions[get_column_letter(days_in_month + 3)].width = 8
    ws1.column_dimensions[get_column_letter(days_in_month + 4)].width = 8
    ws1.column_dimensions[get_column_letter(days_in_month + 5)].width = 10

    # Data rows
    row_idx = row_h + 1
    sorted_angajati = sorted(angajat_data.values(), key=lambda x: x['angajat'].nume_complet)
    for idx, ad in enumerate(sorted_angajati, 1):
        ws1.cell(row=row_idx, column=1, value=idx).border = thin_border
        ws1.cell(row=row_idx, column=1).alignment = Alignment(horizontal='center')
        ws1.cell(row=row_idx, column=2, value=ad['angajat'].nume_complet).border = thin_border

        for d in range(1, days_in_month + 1):
            col = d + 2
            cell = ws1.cell(row=row_idx, column=col)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            dt = date(anul, luna, d)
            if d in ad['zile']:
                p = ad['zile'][d]
                if p.tip_zi == 'co':
                    cell.value = 'CO'
                elif p.tip_zi == 'cm':
                    cell.value = 'CM'
                elif p.tip_zi == 'invoiere':
                    cell.value = 'I'
                else:
                    cell.value = float(p.ore_lucrate) if p.ore_lucrate else ''
            elif dt.weekday() >= 5:
                cell.fill = weekend_fill
            elif d in sarbatori:
                cell.value = 'SL'
                cell.fill = sarb_fill

        # Totals
        ws1.cell(row=row_idx, column=days_in_month + 3, value=round(ad['total_ore'], 1)).border = thin_border
        ws1.cell(row=row_idx, column=days_in_month + 3).alignment = Alignment(horizontal='center')
        ws1.cell(row=row_idx, column=days_in_month + 3).font = Font(bold=True)

        ore_supl = round(ad['ore_supl_50'] + ad['ore_supl_100'], 1)
        ws1.cell(row=row_idx, column=days_in_month + 4, value=ore_supl).border = thin_border
        ws1.cell(row=row_idx, column=days_in_month + 4).alignment = Alignment(horizontal='center')

        ws1.cell(row=row_idx, column=days_in_month + 5).border = thin_border  # Semnatura

        row_idx += 1

    ws1.freeze_panes = f'C{row_h + 1}'

    # --- Sheet 2: Centralizator ---
    ws2 = wb.create_sheet('Centralizator')
    cent_headers = ['Nr.', 'Nume si Prenume', 'Ore Normale', 'Ore Supl. 50%',
                    'Ore Supl. 100%', 'Total Ore', 'Tarif (RON/h)', 'Cost Normal',
                    'Cost Supl. 50%', 'Cost Supl. 100%', 'Cost Total']
    for col, h in enumerate(cent_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    row_idx = 2
    total_cost = 0
    for idx, ad in enumerate(sorted_angajati, 1):
        ang = ad['angajat']
        tarif = float(ang.tarif_orar) if ang.tarif_orar else 0

        cost_normal = round(ad['ore_normale'] * tarif, 2)
        cost_50 = round(ad['ore_supl_50'] * tarif * 1.5, 2)
        cost_100 = round(ad['ore_supl_100'] * tarif * 2, 2)
        cost_total = cost_normal + cost_50 + cost_100
        total_cost += cost_total

        row = [idx, ang.nume_complet,
               round(ad['ore_normale'], 1), round(ad['ore_supl_50'], 1),
               round(ad['ore_supl_100'], 1), round(ad['total_ore'], 1),
               tarif, cost_normal, cost_50, cost_100, round(cost_total, 2)]

        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            if col >= 3:
                cell.alignment = Alignment(horizontal='right')
        row_idx += 1

    # Total row
    ws2.cell(row=row_idx, column=1, value='').border = thin_border
    total_cell = ws2.cell(row=row_idx, column=2, value='TOTAL')
    total_cell.font = Font(bold=True)
    total_cell.border = thin_border
    for col in range(3, 12):
        ws2.cell(row=row_idx, column=col).border = thin_border
    ws2.cell(row=row_idx, column=11, value=round(total_cost, 2)).font = Font(bold=True)
    ws2.cell(row=row_idx, column=11).border = thin_border

    for col in range(1, 12):
        ws2.column_dimensions[get_column_letter(col)].width = 16
    ws2.column_dimensions['A'].width = 5
    ws2.column_dimensions['B'].width = 25
    ws2.auto_filter.ref = f'A1:K{row_idx}'
    ws2.freeze_panes = 'A2'

    # --- Sheet 3: Absente ---
    ws3 = wb.create_sheet('Absente')
    abs_headers = ['Nr.', 'Angajat', 'Data', 'Tip Absenta', 'Observatii']
    for col, h in enumerate(abs_headers, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row_idx = 2
    nr = 1
    for p in pontaje:
        if p.tip_zi in ('co', 'cm', 'invoiere'):
            tip_display = {'co': 'Concediu odihna', 'cm': 'Concediu medical', 'invoiere': 'Invoiere'}.get(p.tip_zi, p.tip_zi)
            ws3.cell(row=row_idx, column=1, value=nr).border = thin_border
            ws3.cell(row=row_idx, column=2, value=p.angajat.nume_complet if p.angajat else '-').border = thin_border
            ws3.cell(row=row_idx, column=3, value=p.data.strftime('%d.%m.%Y') if p.data else '-').border = thin_border
            ws3.cell(row=row_idx, column=4, value=tip_display).border = thin_border
            ws3.cell(row=row_idx, column=5, value=p.observatii or '').border = thin_border
            row_idx += 1
            nr += 1

    for col in [1, 2, 3, 4, 5]:
        ws3.column_dimensions[get_column_letter(col)].width = [5, 25, 12, 20, 30][col - 1]
    ws3.freeze_panes = 'A2'

    # Save
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'pontaj_{month_names[luna]}_{anul}.xlsx'
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# TEMPLATE IMPORT EXCEL
# ============================================================

@pontaje_bp.route('/template-import')
@login_required
def template_import():
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Sheet 1 - Template
    ws = wb.active
    ws.title = 'Pontaje'
    headers = ['CNP Angajat*', 'Cod Proiect*', 'Data (ZZ.LL.AAAA)*',
               'Ora Start (HH:MM)*', 'Ora Sfarsit (HH:MM)*', 'Tip Zi', 'Observatii']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    # Example row
    example = ['1234567890123', 'PRJ-2026-001', '15.03.2026', '08:00', '16:30', 'lucratoare', '']
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val).border = thin_border

    # Validation tip_zi
    dv = DataValidation(type='list', formula1='"lucratoare,sambata,duminica,sarbatoare_legala,co,cm,invoiere"')
    dv.error = 'Selectati un tip de zi valid'
    ws.add_data_validation(dv)
    dv.add(f'F2:F1000')

    for col in range(1, 8):
        ws.column_dimensions[chr(64 + col)].width = [18, 18, 18, 18, 18, 18, 25][col - 1]

    # Sheet 2 - Instructiuni
    ws2 = wb.create_sheet('Instructiuni')
    instructions = [
        ['INSTRUCTIUNI IMPORT PONTAJE'],
        [''],
        ['Campuri obligatorii (marcate cu *):'],
        ['- CNP Angajat: 13 cifre, trebuie sa existe in sistem'],
        ['- Cod Proiect: cod exact din sistem (ex: PRJ-2026-001)'],
        ['- Data: format ZZ.LL.AAAA (ex: 15.03.2026)'],
        ['- Ora Start: format HH:MM (ex: 08:00)'],
        ['- Ora Sfarsit: format HH:MM (ex: 16:30)'],
        [''],
        ['Campuri optionale:'],
        ['- Tip Zi: lucratoare, sambata, duminica, sarbatoare_legala, co, cm, invoiere'],
        ['  (daca nu se completeaza, se detecteaza automat din data)'],
        ['- Observatii: text liber'],
        [''],
        ['Note:'],
        ['- Orele se calculeaza automat (normale, suplimentare 50%, 100%)'],
        ['- Pauza de masa (30 min) se deduce automat daca > 6h'],
        ['- Duplicate (acelasi angajat + aceeasi zi) sunt omise'],
        ['- Pontajele se importa cu status "draft"'],
    ]
    for row_idx, row in enumerate(instructions, 1):
        ws2.cell(row=row_idx, column=1, value=row[0] if row else '')
    ws2.cell(row=1, column=1).font = Font(bold=True, size=14, color='1a237e')
    ws2.column_dimensions['A'].width = 70

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(buffer, as_attachment=True, download_name='template_import_pontaje.xlsx',
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# IMPORT EXCEL
# ============================================================

@pontaje_bp.route('/import-excel', methods=['POST'])
@login_required
def import_excel():
    from openpyxl import load_workbook

    fisier = request.files.get('fisier')
    if not fisier or not fisier.filename.endswith('.xlsx'):
        flash('Selectati un fisier Excel (.xlsx) valid.', 'danger')
        return redirect(url_for('pontaje.lista'))

    try:
        wb = load_workbook(fisier, data_only=True)
        ws = wb.active
    except Exception:
        flash('Eroare la citirea fisierului Excel.', 'danger')
        return redirect(url_for('pontaje.lista'))

    count_ok = 0
    count_err = 0
    errors = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not row[0]:
            continue

        try:
            cnp = str(row[0]).strip()
            cod_proiect = str(row[1]).strip()
            data_str = str(row[2]).strip()
            ora_start = str(row[3]).strip()
            ora_sfarsit = str(row[4]).strip()
            tip_zi = str(row[5]).strip() if row[5] else ''
            obs = str(row[6]).strip() if len(row) > 6 and row[6] else ''

            # Find angajat by CNP
            angajat = query_for_tenant(Angajat).filter_by(cnp=cnp).first()
            if not angajat:
                errors.append(f'Rand {row_idx}: CNP {cnp} nu exista in sistem')
                count_err += 1
                continue

            # Find proiect
            proiect = query_for_tenant(Proiect).filter_by(cod_proiect=cod_proiect).first()
            if not proiect:
                errors.append(f'Rand {row_idx}: Proiect {cod_proiect} nu exista')
                count_err += 1
                continue

            # Parse date
            try:
                if '.' in data_str:
                    data_pontaj = datetime.strptime(data_str, '%d.%m.%Y').date()
                else:
                    data_pontaj = datetime.strptime(data_str, '%Y-%m-%d').date()
            except ValueError:
                errors.append(f'Rand {row_idx}: Format data invalid ({data_str})')
                count_err += 1
                continue

            # Check duplicate
            exista = query_timesheets_for_tenant().filter_by(
                angajat_id=angajat.id,
                data=data_pontaj,
            ).first()
            if exista:
                errors.append(f'Rand {row_idx}: Duplicat - {angajat.nume_complet} are deja pontaj in {data_str}')
                count_err += 1
                continue

            if not tip_zi:
                tip_zi = _detect_tip_zi(data_pontaj)

            result = calculate_hours(ora_start, ora_sfarsit, tip_zi, data_pontaj)

            pontaj = Pontaj(
                angajat_id=angajat.id,
                proiect_id=proiect.id,
                data=data_pontaj,
                ora_start=ora_start,
                ora_sfarsit=ora_sfarsit,
                ore_lucrate=result['ore_lucrate'],
                ore_normale=result['ore_normale'],
                ore_suplimentare_50=result['ore_supl_50'],
                ore_suplimentare_100=result['ore_supl_100'],
                tip_zi=result['tip_zi'],
                status='draft',
                observatii=obs,
                introdus_de=current_user.id
            )
            db.session.add(pontaj)
            count_ok += 1

        except Exception as e:
            errors.append(f'Rand {row_idx}: Eroare - {str(e)}')
            count_err += 1

    db.session.commit()

    if count_ok > 0:
        flash(f'{count_ok} pontaje importate cu succes!', 'success')
    if count_err > 0:
        flash(f'{count_err} erori la import: ' + '; '.join(errors[:5]), 'warning')

    return redirect(url_for('pontaje.lista'))
