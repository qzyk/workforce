"""
INNOVA WORKFORCE - Modul Documente (Complet)
Gestionarea documentelor angajatilor: upload, expirare, alerte, export Excel.
"""

import os
import uuid
from datetime import datetime, date, timedelta
from flask import (
    Blueprint, render_template, redirect, url_for, flash, request,
    current_app, send_file, jsonify, abort
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from models import db, Document, Angajat, Proiect
from forms.documente_forms import (
    DocumentUploadForm, DocumentEditForm,
    DURATA_EXPIRARE, DOCUMENTE_OBLIGATORII
)

documente_bp = Blueprint('documente', __name__)


# ============================================================
# HELPERS
# ============================================================

ALLOWED_EXT = {'pdf', 'jpg', 'jpeg', 'png', 'docx'}
MAX_SIZE = 10 * 1024 * 1024  # 10 MB

TIPURI_DICT = dict(Document.TIPURI)


def _fisier_permis(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXT


def _get_angajat_folder(angajat_id):
    """Creaza si returneaza folderul per angajat: uploads/angajat_<id>/"""
    folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'angajat_{angajat_id}')
    os.makedirs(folder, exist_ok=True)
    return folder


def _secure_save(fisier, angajat_id):
    """Salveaza fisierul cu nume unic in folderul angajatului. Returneaza (path, marime)."""
    ext = fisier.filename.rsplit('.', 1)[1].lower()
    unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.{ext}"
    folder = _get_angajat_folder(angajat_id)
    filepath = os.path.join(folder, unique_name)
    fisier.save(filepath)
    marime = os.path.getsize(filepath)
    return filepath, marime


def _make_thumbnail(filepath):
    """Creaza thumbnail pentru imagini folosind Pillow (optional)."""
    try:
        from PIL import Image
        ext = filepath.rsplit('.', 1)[1].lower()
        if ext in ('jpg', 'jpeg', 'png'):
            thumb_dir = os.path.join(os.path.dirname(filepath), 'thumbnails')
            os.makedirs(thumb_dir, exist_ok=True)
            thumb_path = os.path.join(thumb_dir, os.path.basename(filepath))
            img = Image.open(filepath)
            img.thumbnail((200, 200))
            img.save(thumb_path)
            return thumb_path
    except ImportError:
        pass
    except Exception:
        pass
    return None


def _update_all_statuses():
    """Actualizeaza statusul tuturor documentelor."""
    docs = Document.query.filter(Document.data_expirare.isnot(None)).all()
    for doc in docs:
        new_status = doc.status_calculat
        if doc.status != new_status:
            doc.status = new_status
    db.session.commit()


def _get_documente_obligatorii(functie):
    """Returneaza lista tipuri obligatorii per functie."""
    return DOCUMENTE_OBLIGATORII.get(functie, DOCUMENTE_OBLIGATORII['default'])


def _get_stats():
    """Returneaza statistici documente."""
    total = Document.query.count()
    expirate = Document.query.filter_by(status='expirat').count()
    in_curand = Document.query.filter_by(status='in_curand').count()
    valabile = Document.query.filter_by(status='valabil').count()
    return {
        'total': total,
        'expirate': expirate,
        'in_curand': in_curand,
        'valabile': valabile,
    }


# ============================================================
# 1. PANOU PRINCIPAL (Dashboard Documente)
# ============================================================

@documente_bp.route('/')
@login_required
def panou():
    _update_all_statuses()
    stats = _get_stats()

    # Documente expirate (top 10)
    expirate = Document.query.filter_by(status='expirat')\
        .order_by(Document.data_expirare.asc()).limit(10).all()

    # Documente care expira curand (top 10)
    expira_curand = Document.query.filter_by(status='in_curand')\
        .order_by(Document.data_expirare.asc()).limit(10).all()

    # Ultimele documente incarcate
    recente = Document.query.order_by(Document.data_upload.desc()).limit(10).all()

    # Angajati cu documente lipsa
    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    angajati_alerte = []
    for ang in angajati:
        obligatorii = _get_documente_obligatorii(ang.functie)
        existente = [d.tip for d in ang.documente.filter(
            Document.status.in_(['valabil', 'in_curand'])
        ).all()]
        lipsa = [t for t in obligatorii if t not in existente]
        if lipsa:
            angajati_alerte.append({
                'angajat': ang,
                'lipsa': lipsa,
                'nr_lipsa': len(lipsa),
            })

    return render_template('documente/panou.html',
                           stats=stats,
                           expirate=expirate,
                           expira_curand=expira_curand,
                           recente=recente,
                           angajati_alerte=angajati_alerte,
                           tipuri_dict=TIPURI_DICT,
                           today=date.today())


# ============================================================
# 2. LISTA DOCUMENTE PER ANGAJAT
# ============================================================

@documente_bp.route('/angajat/<int:angajat_id>')
@login_required
def lista_angajat(angajat_id):
    angajat = Angajat.query.get_or_404(angajat_id)
    _update_all_statuses()

    documente = Document.query.filter_by(angajat_id=angajat_id)\
        .order_by(Document.tip, Document.data_upload.desc()).all()

    # Grupare per tip
    docs_per_tip = {}
    for doc in documente:
        if doc.tip not in docs_per_tip:
            docs_per_tip[doc.tip] = []
        docs_per_tip[doc.tip].append(doc)

    # Documente obligatorii
    obligatorii = _get_documente_obligatorii(angajat.functie)
    existente_tipuri = set(d.tip for d in documente if d.status in ('valabil', 'in_curand'))
    lipsa = [t for t in obligatorii if t not in existente_tipuri]

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()

    return render_template('documente/lista_angajat.html',
                           angajat=angajat,
                           documente=documente,
                           docs_per_tip=docs_per_tip,
                           obligatorii=obligatorii,
                           lipsa=lipsa,
                           tipuri_dict=TIPURI_DICT,
                           angajati=angajati,
                           today=date.today())


# ============================================================
# 3. UPLOAD DOCUMENT (Drag & Drop + Auto-Expiry)
# ============================================================

@documente_bp.route('/upload', methods=['GET', 'POST'])
@login_required
def upload():
    form = DocumentUploadForm()

    if form.validate_on_submit():
        fisier = form.fisier.data

        # Validare dimensiune
        fisier.seek(0, 2)
        size = fisier.tell()
        fisier.seek(0)
        if size > MAX_SIZE:
            flash('Fisierul depaseste limita de 10 MB!', 'danger')
            return redirect(request.url)

        angajat_id = form.angajat_id.data
        filepath, marime = _secure_save(fisier, angajat_id)

        # Thumbnail pt imagini
        _make_thumbnail(filepath)

        doc = Document(
            angajat_id=angajat_id,
            proiect_id=form.proiect_id.data if form.proiect_id.data else None,
            tip=form.tip.data,
            nume_document=form.nume_document.data.strip(),
            fisier_path=filepath,
            marime_fisier=marime,
            data_emitere=form.data_emitere.data,
            data_expirare=form.data_expirare.data,
            emitent=form.emitent.data.strip() if form.emitent.data else '',
            serie_numar=form.serie_numar.data.strip() if form.serie_numar.data else '',
            observatii=form.observatii.data.strip() if form.observatii.data else '',
            incarcat_de=current_user.id,
        )
        doc.status = doc.status_calculat
        db.session.add(doc)
        db.session.commit()

        flash('Documentul a fost incarcat cu succes!', 'success')
        return redirect(url_for('documente.lista_angajat', angajat_id=angajat_id))

    # Pre-fill angajat daca vine din lista angajat
    angajat_id_prefill = request.args.get('angajat_id', type=int)
    if angajat_id_prefill and not form.angajat_id.data:
        form.angajat_id.data = angajat_id_prefill

    return render_template('documente/upload.html', form=form,
                           durata_expirare=DURATA_EXPIRARE)


# ============================================================
# 4. EDITARE METADATA DOCUMENT
# ============================================================

@documente_bp.route('/<int:id>/editeaza', methods=['GET', 'POST'])
@login_required
def editeaza(id):
    doc = Document.query.get_or_404(id)
    form = DocumentEditForm(obj=doc)

    if form.validate_on_submit():
        doc.tip = form.tip.data
        doc.nume_document = form.nume_document.data.strip()
        doc.data_emitere = form.data_emitere.data
        doc.data_expirare = form.data_expirare.data
        doc.emitent = form.emitent.data.strip() if form.emitent.data else ''
        doc.serie_numar = form.serie_numar.data.strip() if form.serie_numar.data else ''
        doc.observatii = form.observatii.data.strip() if form.observatii.data else ''
        doc.status = doc.status_calculat
        db.session.commit()
        flash('Documentul a fost actualizat!', 'success')
        if doc.angajat_id:
            return redirect(url_for('documente.lista_angajat', angajat_id=doc.angajat_id))
        return redirect(url_for('documente.panou'))

    return render_template('documente/editeaza.html', form=form, doc=doc)


# ============================================================
# 5. DESCARCARE FISIER
# ============================================================

@documente_bp.route('/<int:id>/descarca')
@login_required
def descarca(id):
    doc = Document.query.get_or_404(id)
    if doc.fisier_path and os.path.exists(doc.fisier_path):
        ext = doc.fisier_path.rsplit('.', 1)[1].lower() if '.' in doc.fisier_path else ''
        download_name = f"{doc.nume_document}.{ext}" if ext else doc.nume_document
        return send_file(doc.fisier_path, as_attachment=True, download_name=download_name)
    flash('Fisierul nu a fost gasit pe server!', 'danger')
    return redirect(url_for('documente.panou'))


# ============================================================
# 6. PREVIZUALIZARE FISIER (inline)
# ============================================================

@documente_bp.route('/<int:id>/preview')
@login_required
def preview(id):
    doc = Document.query.get_or_404(id)
    if doc.fisier_path and os.path.exists(doc.fisier_path):
        ext = doc.fisier_path.rsplit('.', 1)[1].lower() if '.' in doc.fisier_path else ''
        mime_types = {
            'pdf': 'application/pdf',
            'jpg': 'image/jpeg',
            'jpeg': 'image/jpeg',
            'png': 'image/png',
        }
        mimetype = mime_types.get(ext, 'application/octet-stream')
        return send_file(doc.fisier_path, mimetype=mimetype)
    flash('Fisierul nu a fost gasit!', 'danger')
    return redirect(url_for('documente.panou'))


# ============================================================
# 7. STERGERE DOCUMENT
# ============================================================

@documente_bp.route('/<int:id>/sterge', methods=['POST'])
@login_required
def sterge(id):
    if not current_user.is_manager:
        flash('Nu aveti permisiunea de a sterge documente.', 'danger')
        return redirect(url_for('documente.panou'))

    doc = Document.query.get_or_404(id)
    angajat_id = doc.angajat_id

    # Stergere fisier fizic
    if doc.fisier_path and os.path.exists(doc.fisier_path):
        os.remove(doc.fisier_path)
        # Stergere thumbnail daca exista
        thumb_dir = os.path.join(os.path.dirname(doc.fisier_path), 'thumbnails')
        thumb_path = os.path.join(thumb_dir, os.path.basename(doc.fisier_path))
        if os.path.exists(thumb_path):
            os.remove(thumb_path)

    db.session.delete(doc)
    db.session.commit()
    flash('Documentul a fost sters.', 'warning')

    if angajat_id:
        return redirect(url_for('documente.lista_angajat', angajat_id=angajat_id))
    return redirect(url_for('documente.panou'))


# ============================================================
# 8. RAPORT DOCUMENTE EXPIRATE
# ============================================================

@documente_bp.route('/expirate')
@login_required
def expirate():
    _update_all_statuses()

    # Toate expirate
    docs_expirate = Document.query.filter_by(status='expirat')\
        .join(Angajat).order_by(Angajat.nume, Document.data_expirare.asc()).all()

    # Toate care expira in 30 zile
    docs_in_curand = Document.query.filter_by(status='in_curand')\
        .join(Angajat).order_by(Document.data_expirare.asc()).all()

    # Grupare per angajat
    per_angajat = {}
    for doc in docs_expirate + docs_in_curand:
        ang_id = doc.angajat_id
        if ang_id not in per_angajat:
            per_angajat[ang_id] = {
                'angajat': doc.angajat,
                'expirate': [],
                'in_curand': [],
            }
        if doc.status == 'expirat':
            per_angajat[ang_id]['expirate'].append(doc)
        else:
            per_angajat[ang_id]['in_curand'].append(doc)

    return render_template('documente/expirate.html',
                           docs_expirate=docs_expirate,
                           docs_in_curand=docs_in_curand,
                           per_angajat=per_angajat,
                           tipuri_dict=TIPURI_DICT,
                           today=date.today())


# ============================================================
# 9. EXPORT EXCEL (2 sheets)
# ============================================================

@documente_bp.route('/export-excel')
@login_required
def export_excel():
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    except ImportError:
        flash('Modulul openpyxl nu este instalat!', 'danger')
        return redirect(url_for('documente.panou'))

    _update_all_statuses()

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    red_fill = PatternFill(start_color='FFCDD2', end_color='FFCDD2', fill_type='solid')
    yellow_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')
    green_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')

    # === Sheet 1: Situatie Completa ===
    ws1 = wb.active
    ws1.title = 'Situatie Completa'

    headers1 = ['Nr.', 'Angajat', 'Functie', 'Tip Document', 'Denumire', 'Serie/Numar',
                 'Emitent', 'Data Emitere', 'Data Expirare', 'Status', 'Zile Ramase']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border
    ws1.row_dimensions[1].height = 30

    documente = Document.query.join(Angajat)\
        .order_by(Angajat.nume, Document.tip).all()

    for i, doc in enumerate(documente, 1):
        ang = doc.angajat
        zile_ramase = ''
        if doc.data_expirare:
            zile_ramase = (doc.data_expirare - date.today()).days

        row_data = [
            i,
            ang.nume_complet if ang else '-',
            ang.functie if ang else '-',
            TIPURI_DICT.get(doc.tip, doc.tip),
            doc.nume_document,
            doc.serie_numar or '-',
            doc.emitent or '-',
            doc.data_emitere.strftime('%d.%m.%Y') if doc.data_emitere else '-',
            doc.data_expirare.strftime('%d.%m.%Y') if doc.data_expirare else 'Permanent',
            doc.status.upper().replace('_', ' '),
            zile_ramase if zile_ramase != '' else '-',
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws1.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        # Conditional coloring
        status_cell = ws1.cell(row=i + 1, column=10)
        if doc.status == 'expirat':
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=i + 1, column=c).fill = red_fill
        elif doc.status == 'in_curand':
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=i + 1, column=c).fill = yellow_fill
        else:
            for c in range(1, len(headers1) + 1):
                ws1.cell(row=i + 1, column=c).fill = green_fill

    # Auto-width
    for col in ws1.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws1.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)
    ws1.freeze_panes = 'A2'

    # === Sheet 2: Rezumat per Angajat ===
    ws2 = wb.create_sheet('Rezumat per Angajat')
    tipuri_all = [t[0] for t in Document.TIPURI]
    headers2 = ['Nr.', 'Angajat', 'Functie'] + [TIPURI_DICT.get(t, t) for t in tipuri_all] + ['Total', 'Expirate', 'Lipsa']

    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
    ws2.row_dimensions[1].height = 40

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    for i, ang in enumerate(angajati, 1):
        docs_ang = Document.query.filter_by(angajat_id=ang.id).all()
        docs_per_tip = {}
        for d in docs_ang:
            docs_per_tip[d.tip] = d

        obligatorii = _get_documente_obligatorii(ang.functie)
        nr_total = len(docs_ang)
        nr_expirate = sum(1 for d in docs_ang if d.status == 'expirat')
        existente_tipuri = set(d.tip for d in docs_ang if d.status in ('valabil', 'in_curand'))
        nr_lipsa = sum(1 for t in obligatorii if t not in existente_tipuri)

        row_data = [i, ang.nume_complet, ang.functie]
        for tip in tipuri_all:
            if tip in docs_per_tip:
                d = docs_per_tip[tip]
                if d.status == 'expirat':
                    row_data.append('EXPIRAT')
                elif d.status == 'in_curand':
                    row_data.append('EXP.CURAND')
                else:
                    row_data.append('OK')
            elif tip in obligatorii:
                row_data.append('LIPSA')
            else:
                row_data.append('-')
        row_data += [nr_total, nr_expirate, nr_lipsa]

        for col, val in enumerate(row_data, 1):
            cell = ws2.cell(row=i + 1, column=col, value=val)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

            # Color per cell value
            if val == 'EXPIRAT' or val == 'LIPSA':
                cell.fill = red_fill
                cell.font = Font(bold=True, color='C62828')
            elif val == 'EXP.CURAND':
                cell.fill = yellow_fill
                cell.font = Font(bold=True, color='E65100')
            elif val == 'OK':
                cell.fill = green_fill
                cell.font = Font(bold=True, color='2E7D32')

    for col in ws2.columns:
        max_len = 0
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 3, 25)
    ws2.freeze_panes = 'D2'

    # Salvare
    export_dir = current_app.config.get('EXPORT_FOLDER', os.path.join(current_app.root_path, 'exports'))
    os.makedirs(export_dir, exist_ok=True)
    filename = f"Documente_Angajati_{date.today().strftime('%Y%m%d')}.xlsx"
    filepath = os.path.join(export_dir, filename)
    wb.save(filepath)

    return send_file(filepath, as_attachment=True, download_name=filename)


# ============================================================
# 10. AJAX: ALERTE DOCUMENTE
# ============================================================

@documente_bp.route('/api/alerte')
@login_required
def api_alerte():
    _update_all_statuses()
    expirate = Document.query.filter_by(status='expirat').count()
    in_curand = Document.query.filter_by(status='in_curand').count()

    docs_urgente = Document.query.filter(
        Document.status.in_(['expirat', 'in_curand'])
    ).join(Angajat).order_by(Document.data_expirare.asc()).limit(5).all()

    alerte = []
    for doc in docs_urgente:
        zile = (doc.data_expirare - date.today()).days if doc.data_expirare else 0
        alerte.append({
            'id': doc.id,
            'angajat': doc.angajat.nume_complet if doc.angajat else '-',
            'tip': TIPURI_DICT.get(doc.tip, doc.tip),
            'data_expirare': doc.data_expirare.strftime('%d.%m.%Y') if doc.data_expirare else '-',
            'zile_ramase': zile,
            'status': doc.status,
        })

    return jsonify({
        'expirate': expirate,
        'in_curand': in_curand,
        'alerte': alerte,
    })


# ============================================================
# 11. AJAX: DURATA EXPIRARE PER TIP
# ============================================================

@documente_bp.route('/api/durata-expirare')
@login_required
def api_durata_expirare():
    tip = request.args.get('tip', '')
    durata = DURATA_EXPIRARE.get(tip, 0)
    return jsonify({'tip': tip, 'durata_zile': durata})


# ============================================================
# 12. LISTA GENERALA (pastreaza compatibilitatea)
# ============================================================

@documente_bp.route('/lista')
@login_required
def lista():
    tip = request.args.get('tip', '')
    status = request.args.get('status', '')
    angajat_id = request.args.get('angajat_id', '')

    query = Document.query
    if tip:
        query = query.filter_by(tip=tip)
    if angajat_id:
        query = query.filter_by(angajat_id=int(angajat_id))

    page = request.args.get('page', 1, type=int)
    per_page = current_app.config.get('ITEMS_PER_PAGE', 25)
    pagination = query.order_by(Document.data_upload.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    documente = pagination.items

    # Actualizare status
    for doc in documente:
        new_status = doc.status_calculat
        if doc.status != new_status:
            doc.status = new_status
    db.session.commit()

    # Filtrare dupa status post-actualizare
    if status:
        documente = [d for d in documente if d.status == status]

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()

    return render_template('documente/lista.html',
                           documente=documente,
                           angajati=angajati,
                           tipuri=Document.TIPURI,
                           tip_filtru=tip,
                           status_filtru=status,
                           angajat_id_filtru=angajat_id,
                           pagination=pagination)
