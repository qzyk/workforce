"""
EDIFICO WORKFORCE - Rute Documente Proiecte
Gestiune documente per proiect, organizate pe tipuri de instalatii.
"""

import os
import io
from datetime import datetime, date, timedelta
from werkzeug.utils import secure_filename

from flask import (
    Blueprint, render_template, redirect, url_for, flash,
    request, jsonify, send_file, current_app, abort
)
from flask_login import login_required, current_user
from models import (
    db, Proiect, Utilizator, TipInstalatie, TipDocumentProiect,
    DocumentProiect, RevizieDocument
)
from services.security.tenant_access import (
    get_project_document_or_404,
    get_project_document_revision_or_404,
    get_project_or_404,
    query_project_documents_for_tenant,
)

doc_proiecte_bp = Blueprint('doc_proiecte', __name__,
                             url_prefix='/proiecte/<int:proiect_id>/documente')


# ============================================================
# HELPERS
# ============================================================

def _get_proiect_doc_folder(proiect_id):
    """Returneaza calea directorului pentru documentele unui proiect."""
    folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'proiect_{proiect_id}')
    os.makedirs(folder, exist_ok=True)
    return folder


def _allowed_file(filename):
    """Verifica daca extensia fisierului este permisa."""
    if '.' not in filename:
        return False
    ext = filename.rsplit('.', 1)[1].lower()
    return ext in current_app.config.get('ALLOWED_EXTENSIONS_PROIECT',
                                          {'pdf', 'dwg', 'dxf', 'docx', 'xlsx', 'jpg', 'jpeg', 'png', 'zip'})


def _secure_save(file, proiect_id, prefix=''):
    """Salveaza un fisier in directorul proiectului, returneaza (path, size, ext)."""
    if not file or not file.filename:
        return None, 0, ''

    filename = secure_filename(file.filename)
    if not filename:
        filename = 'document.pdf'

    ext = filename.rsplit('.', 1)[1].lower() if '.' in filename else 'pdf'
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    if prefix:
        new_filename = f'{prefix}_{timestamp}.{ext}'
    else:
        new_filename = f'{timestamp}_{filename}'

    folder = _get_proiect_doc_folder(proiect_id)
    filepath = os.path.join(folder, new_filename)
    file.save(filepath)
    size = os.path.getsize(filepath)

    # Calea relativa pentru DB
    rel_path = os.path.join(f'proiect_{proiect_id}', new_filename)
    return rel_path, size, ext


def _query_documente_proiect(proiect_id):
    """Documente proiect vizibile tenantului curent pentru proiectul validat."""
    return query_project_documents_for_tenant().filter_by(proiect_id=proiect_id)


def _get_document_proiect_in_project_or_404(proiect_id, doc_id):
    """Lookup document proiect tenant-safe si verificat pe URL parent."""
    doc = get_project_document_or_404(doc_id)
    if doc.proiect_id != proiect_id:
        abort(404)
    return doc


def _get_completitudine(proiect_id):
    """Calculeaza completitudinea documentelor obligatorii per instalatie."""
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()
    result = []

    for inst in instalatii:
        tipuri_obligatorii = TipDocumentProiect.query.filter_by(
            tip_instalatie_id=inst.id, obligatoriu=True
        ).all()

        total_obligatorii = len(tipuri_obligatorii)
        if total_obligatorii == 0:
            continue

        prezente = 0
        expirate = 0
        lipsa = []

        for tip_doc in tipuri_obligatorii:
            doc = _query_documente_proiect(proiect_id).filter_by(
                proiect_id=proiect_id,
                tip_document_id=tip_doc.id,
                versiune_curenta=True
            ).filter(DocumentProiect.status.notin_(['anulat', 'arhivat'])).first()

            if doc:
                prezente += 1
                if doc.is_expirat:
                    expirate += 1
            else:
                lipsa.append(tip_doc)

        total_docs = _query_documente_proiect(proiect_id).filter_by(
            proiect_id=proiect_id,
            tip_instalatie_id=inst.id,
            versiune_curenta=True
        ).filter(DocumentProiect.status.notin_(['anulat'])).count()

        procent = int((prezente / total_obligatorii) * 100) if total_obligatorii > 0 else 0

        result.append({
            'instalatie': inst,
            'total_obligatorii': total_obligatorii,
            'prezente': prezente,
            'expirate': expirate,
            'lipsa': lipsa,
            'total_docs': total_docs,
            'procent': procent,
        })

    return result


# ============================================================
# INDEX DOCUMENTE PROIECT
# ============================================================

@doc_proiecte_bp.route('/')
@login_required
def index(proiect_id):
    """Pagina principala documente proiect cu tab-uri per instalatie."""
    proiect = get_project_or_404(proiect_id)
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()

    # Tab activ
    tab_activ = request.args.get('tab', '')
    etapa_activa = request.args.get('etapa', '')

    # Completitudine per instalatie
    completitudine = _get_completitudine(proiect_id)

    # Documente per instalatie
    documente_per_inst = {}
    for inst in instalatii:
        query = _query_documente_proiect(proiect_id).filter_by(
            proiect_id=proiect_id,
            tip_instalatie_id=inst.id,
            versiune_curenta=True
        ).filter(DocumentProiect.status != 'anulat')

        if etapa_activa:
            query = query.filter_by(etapa_proiect=etapa_activa)

        docs = query.order_by(DocumentProiect.etapa_proiect, DocumentProiect.data_upload.desc()).all()
        documente_per_inst[inst.cod] = docs

    # Documente obligatorii lipsa per instalatie
    lipsa_per_inst = {}
    for comp in completitudine:
        if comp['lipsa']:
            lipsa_per_inst[comp['instalatie'].cod] = comp['lipsa']

    return render_template('proiecte/documente/index.html',
                           proiect=proiect,
                           instalatii=instalatii,
                           completitudine=completitudine,
                           documente_per_inst=documente_per_inst,
                           lipsa_per_inst=lipsa_per_inst,
                           tab_activ=tab_activ,
                           etapa_activa=etapa_activa)


# ============================================================
# DOCUMENTE FILTRATE PER TIP INSTALATIE
# ============================================================

@doc_proiecte_bp.route('/<string:tip_instalatie_cod>')
@login_required
def per_instalatie(proiect_id, tip_instalatie_cod):
    """Documente filtrate pe un singur tip de instalatie."""
    proiect = get_project_or_404(proiect_id)
    instalatie = TipInstalatie.query.filter_by(cod=tip_instalatie_cod).first_or_404()
    etapa = request.args.get('etapa', '')

    query = _query_documente_proiect(proiect_id).filter_by(
        proiect_id=proiect_id,
        tip_instalatie_id=instalatie.id,
        versiune_curenta=True
    ).filter(DocumentProiect.status != 'anulat')

    if etapa:
        query = query.filter_by(etapa_proiect=etapa)

    documente = query.order_by(DocumentProiect.etapa_proiect, DocumentProiect.data_upload.desc()).all()

    # Documente obligatorii lipsa
    tipuri_obligatorii = TipDocumentProiect.query.filter_by(
        tip_instalatie_id=instalatie.id, obligatoriu=True
    ).all()
    doc_tip_ids = {d.tip_document_id for d in documente if d.tip_document_id}
    lipsa = [t for t in tipuri_obligatorii if t.id not in doc_tip_ids]

    return render_template('proiecte/documente/per_instalatie.html',
                           proiect=proiect,
                           instalatie=instalatie,
                           documente=documente,
                           lipsa=lipsa,
                           etapa=etapa)


# ============================================================
# ADAUGA DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/adauga', methods=['GET', 'POST'])
@login_required
def adauga(proiect_id):
    """Formular upload document nou."""
    proiect = get_project_or_404(proiect_id)
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()

    if request.method == 'POST':
        tip_instalatie_id = request.form.get('tip_instalatie_id', type=int)
        tip_document_id = request.form.get('tip_document_id', type=int) or None
        denumire = request.form.get('denumire_document', '').strip()
        nr_document = request.form.get('nr_document', '').strip()
        revizie = request.form.get('revizie', 'Rev.0').strip()
        emitent = request.form.get('emitent', '').strip()
        elaborat_de = request.form.get('elaborat_de', '').strip()
        etapa_proiect = request.form.get('etapa_proiect', 'executie')
        status = request.form.get('status', 'draft')
        observatii = request.form.get('observatii', '').strip()

        data_emitere = None
        data_aprobare = None
        data_expirare = None
        try:
            if request.form.get('data_emitere'):
                data_emitere = datetime.strptime(request.form['data_emitere'], '%Y-%m-%d').date()
            if request.form.get('data_aprobare'):
                data_aprobare = datetime.strptime(request.form['data_aprobare'], '%Y-%m-%d').date()
            if request.form.get('data_expirare'):
                data_expirare = datetime.strptime(request.form['data_expirare'], '%Y-%m-%d').date()
        except ValueError:
            flash('Format data invalid.', 'danger')
            return render_template('proiecte/documente/formular.html',
                                   proiect=proiect, instalatii=instalatii, doc=None)

        errors = []
        if not tip_instalatie_id:
            errors.append('Selectati tipul de instalatie.')
        if not denumire:
            errors.append('Denumirea documentului este obligatorie.')

        # Validare fisier
        fisier = request.files.get('fisier')
        fisier_path = None
        marime = 0
        ext = ''

        if fisier and fisier.filename:
            if not _allowed_file(fisier.filename):
                errors.append('Tip fisier nepermis. Formate acceptate: PDF, DWG, DXF, DOCX, XLSX, JPG, PNG, ZIP.')
            else:
                # Prefix descriptiv
                inst = TipInstalatie.query.get(tip_instalatie_id)
                prefix = f'{inst.cod}_{secure_filename(denumire[:30])}' if inst else ''
                fisier_path, marime, ext = _secure_save(fisier, proiect_id, prefix)

        if errors:
            for e in errors:
                flash(e, 'danger')
            return render_template('proiecte/documente/formular.html',
                                   proiect=proiect, instalatii=instalatii, doc=None)

        doc = DocumentProiect(
            proiect_id=proiect_id,
            tip_instalatie_id=tip_instalatie_id,
            tip_document_id=tip_document_id,
            denumire_document=denumire,
            nr_document=nr_document,
            revizie=revizie,
            emitent=emitent,
            elaborat_de=elaborat_de,
            data_emitere=data_emitere,
            data_aprobare=data_aprobare,
            data_expirare=data_expirare,
            etapa_proiect=etapa_proiect,
            status=status,
            fisier_path=fisier_path,
            marime_fisier=marime,
            tip_fisier=ext,
            observatii=observatii,
            versiune_curenta=True,
            incarcat_de=current_user.id
        )
        db.session.add(doc)
        db.session.commit()

        flash(f'Documentul "{denumire}" a fost adaugat cu succes!', 'success')

        inst = TipInstalatie.query.get(tip_instalatie_id)
        return redirect(url_for('doc_proiecte.index', proiect_id=proiect_id,
                                tab=inst.cod if inst else ''))

    # GET - pre-selectare instalatie si tip document daca sunt in query params
    pre_instalatie = request.args.get('instalatie', type=int)
    pre_tip_doc = request.args.get('tip_doc', type=int)

    return render_template('proiecte/documente/formular.html',
                           proiect=proiect,
                           instalatii=instalatii,
                           doc=None,
                           pre_instalatie=pre_instalatie,
                           pre_tip_doc=pre_tip_doc)


# ============================================================
# DETALIU DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>')
@login_required
def detaliu(proiect_id, doc_id):
    """Detaliu document cu istoric revizii."""
    proiect = get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)

    revizii = doc.revizii.order_by(RevizieDocument.nr_revizie.desc()).all()

    return render_template('proiecte/documente/detaliu.html',
                           proiect=proiect,
                           doc=doc,
                           revizii=revizii)


# ============================================================
# EDITEAZA DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>/editeaza', methods=['GET', 'POST'])
@login_required
def editeaza(proiect_id, doc_id):
    """Editare metadate document."""
    proiect = get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)

    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()

    if request.method == 'POST':
        doc.tip_instalatie_id = request.form.get('tip_instalatie_id', type=int) or doc.tip_instalatie_id
        doc.tip_document_id = request.form.get('tip_document_id', type=int) or None
        doc.denumire_document = request.form.get('denumire_document', '').strip() or doc.denumire_document
        doc.nr_document = request.form.get('nr_document', '').strip()
        doc.emitent = request.form.get('emitent', '').strip()
        doc.elaborat_de = request.form.get('elaborat_de', '').strip()
        doc.etapa_proiect = request.form.get('etapa_proiect', doc.etapa_proiect)
        doc.status = request.form.get('status', doc.status)
        doc.observatii = request.form.get('observatii', '').strip()

        try:
            doc.data_emitere = datetime.strptime(request.form['data_emitere'], '%Y-%m-%d').date() if request.form.get('data_emitere') else None
            doc.data_aprobare = datetime.strptime(request.form['data_aprobare'], '%Y-%m-%d').date() if request.form.get('data_aprobare') else None
            doc.data_expirare = datetime.strptime(request.form['data_expirare'], '%Y-%m-%d').date() if request.form.get('data_expirare') else None
        except ValueError:
            pass

        # Daca se incarca un fisier nou, inlocuieste
        fisier = request.files.get('fisier')
        if fisier and fisier.filename and _allowed_file(fisier.filename):
            inst = TipInstalatie.query.get(doc.tip_instalatie_id)
            prefix = f'{inst.cod}_{secure_filename(doc.denumire_document[:30])}' if inst else ''
            fisier_path, marime, ext = _secure_save(fisier, proiect_id, prefix)
            doc.fisier_path = fisier_path
            doc.marime_fisier = marime
            doc.tip_fisier = ext

        db.session.commit()
        flash('Documentul a fost actualizat!', 'success')
        return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))

    return render_template('proiecte/documente/formular.html',
                           proiect=proiect,
                           instalatii=instalatii,
                           doc=doc)


# ============================================================
# ADAUGA REVIZIE
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>/revizie', methods=['POST'])
@login_required
def adauga_revizie(proiect_id, doc_id):
    """Upload revizie noua pentru un document."""
    proiect = get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)

    motiv = request.form.get('motiv_revizie', '').strip()
    fisier = request.files.get('fisier')

    if not fisier or not fisier.filename:
        flash('Fisierul este obligatoriu pentru o noua revizie.', 'danger')
        return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))

    if not _allowed_file(fisier.filename):
        flash('Tip fisier nepermis.', 'danger')
        return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))

    # Determina numarul reviziei
    ultima_revizie = doc.revizii.order_by(RevizieDocument.nr_revizie.desc()).first()
    nr_nou = (ultima_revizie.nr_revizie + 1) if ultima_revizie else 1

    # Salveaza fisier
    inst = TipInstalatie.query.get(doc.tip_instalatie_id)
    prefix = f'{inst.cod}_Rev{nr_nou}' if inst else f'Rev{nr_nou}'
    fisier_path, marime, ext = _secure_save(fisier, proiect_id, prefix)

    # Creeaza revizia
    rev = RevizieDocument(
        document_proiect_id=doc.id,
        nr_revizie=nr_nou,
        motiv_revizie=motiv,
        fisier_path=fisier_path,
        realizat_de=current_user.id
    )
    db.session.add(rev)

    # Actualizeaza documentul principal
    doc.revizie = f'Rev.{nr_nou}'
    doc.fisier_path = fisier_path
    doc.marime_fisier = marime
    doc.tip_fisier = ext
    doc.status = 'in_revizie'
    doc.data_upload = datetime.utcnow()

    db.session.commit()

    flash(f'Revizia Rev.{nr_nou} a fost adaugata cu succes!', 'success')
    return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))


# ============================================================
# DESCARCA DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>/descarca')
@login_required
def descarca(proiect_id, doc_id):
    """Descarca fisierul documentului."""
    get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)
    if not doc.fisier_path:
        abort(404)

    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.fisier_path)
    if not os.path.exists(filepath):
        flash('Fisierul nu a fost gasit pe disc.', 'danger')
        return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))

    ext = doc.tip_fisier or 'pdf'
    download_name = f'{doc.denumire_document}.{ext}'
    return send_file(filepath, as_attachment=True, download_name=download_name)


# ============================================================
# PREVIEW DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>/preview')
@login_required
def preview(proiect_id, doc_id):
    """Preview inline a fisierului (PDF, imagini)."""
    get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)
    if not doc.fisier_path:
        abort(404)

    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.fisier_path)
    if not os.path.exists(filepath):
        abort(404)

    mime_map = {
        'pdf': 'application/pdf',
        'jpg': 'image/jpeg', 'jpeg': 'image/jpeg',
        'png': 'image/png',
    }
    ext = doc.tip_fisier or ''
    mimetype = mime_map.get(ext, 'application/octet-stream')

    return send_file(filepath, mimetype=mimetype)


# ============================================================
# STERGE DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>/sterge', methods=['POST'])
@login_required
def sterge(proiect_id, doc_id):
    """Sterge un document proiect."""
    get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)

    denumire = doc.denumire_document

    # Sterge reviziile
    RevizieDocument.query.filter_by(document_proiect_id=doc.id).delete()

    # Sterge fisierul de pe disc
    if doc.fisier_path:
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], doc.fisier_path)
        if os.path.exists(filepath):
            try:
                os.remove(filepath)
            except OSError:
                pass

    db.session.delete(doc)
    db.session.commit()

    flash(f'Documentul "{denumire}" a fost sters.', 'success')
    return redirect(url_for('doc_proiecte.index', proiect_id=proiect_id))


# ============================================================
# APROBARE DOCUMENT
# ============================================================

@doc_proiecte_bp.route('/<int:doc_id>/aprobare', methods=['POST'])
@login_required
def aprobare(proiect_id, doc_id):
    """Aprobare document de catre manager/admin."""
    if not current_user.is_manager:
        flash('Nu aveti permisiunea de a aproba documente.', 'danger')
        return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))

    get_project_or_404(proiect_id)
    doc = _get_document_proiect_in_project_or_404(proiect_id, doc_id)

    actiune = request.form.get('actiune', 'aproba')

    if actiune == 'aproba':
        doc.status = 'aprobat'
        doc.aprobat_de_id = current_user.id
        doc.data_aprobare = date.today()
        flash(f'Documentul "{doc.denumire_document}" a fost aprobat!', 'success')
    elif actiune == 'respinge':
        doc.status = 'draft'
        flash(f'Documentul "{doc.denumire_document}" a fost respins.', 'warning')

    db.session.commit()
    return redirect(url_for('doc_proiecte.detaliu', proiect_id=proiect_id, doc_id=doc_id))


# ============================================================
# EXPORT INDEX EXCEL
# ============================================================

@doc_proiecte_bp.route('/export-index')
@login_required
def export_index(proiect_id):
    """Export Excel index complet documente proiect."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    proiect = get_project_or_404(proiect_id)
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(bold=True, color='FFFFFF', size=10)
    title_font = Font(bold=True, size=14, color='1A237E')

    # === Sheet 1: Index Complet ===
    ws = wb.active
    ws.title = 'Index Documente'

    # Header proiect
    ws.merge_cells('A1:I1')
    ws['A1'] = f'INDEX DOCUMENTE - {proiect.cod_proiect} {proiect.nume}'
    ws['A1'].font = title_font
    ws['A2'] = f'Generat: {datetime.now().strftime("%d.%m.%Y %H:%M")} | Beneficiar: {proiect.beneficiar or "-"}'
    ws['A2'].font = Font(italic=True, color='666666')

    headers = ['Nr.', 'Tip Instalatie', 'Tip Document', 'Denumire', 'Nr. Doc', 'Rev.',
               'Data Emitere', 'Status', 'Emitent', 'Completitudine']
    row_idx = 4
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = thin_border

    nr = 0
    for inst in instalatii:
        docs = _query_documente_proiect(proiect_id).filter_by(
            proiect_id=proiect_id,
            tip_instalatie_id=inst.id,
            versiune_curenta=True
        ).filter(DocumentProiect.status != 'anulat').order_by(DocumentProiect.etapa_proiect).all()

        if not docs:
            continue

        for doc in docs:
            nr += 1
            row_idx += 1
            vals = [
                nr,
                inst.denumire,
                doc.tip_document.denumire if doc.tip_document else '-',
                doc.denumire_document,
                doc.nr_document or '-',
                doc.revizie or 'Rev.0',
                doc.data_emitere.strftime('%d.%m.%Y') if doc.data_emitere else '-',
                doc.status.replace('_', ' ').title(),
                doc.emitent or '-',
                'Da' if doc.fisier_path else 'Nu'
            ]
            for col, v in enumerate(vals, 1):
                cell = ws.cell(row=row_idx, column=col, value=v)
                cell.border = thin_border

    # Auto width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[chr(64 + col) if col <= 26 else 'A'].width = 18
    ws.column_dimensions['D'].width = 40
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 35
    ws.auto_filter.ref = f'A4:J{row_idx}'
    ws.freeze_panes = 'A5'

    # === Sheet 2: Documente Obligatorii Lipsa ===
    ws2 = wb.create_sheet('Documente Obligatorii Lipsa')
    ws2.merge_cells('A1:E1')
    ws2['A1'] = 'DOCUMENTE OBLIGATORII LIPSA'
    ws2['A1'].font = title_font

    headers2 = ['Nr.', 'Tip Instalatie', 'Tip Document', 'Obligatoriu', 'Status']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = PatternFill(start_color='B71C1C', end_color='B71C1C', fill_type='solid')
        cell.border = thin_border

    nr_lipsa = 0
    row_lipsa = 3
    completitudine = _get_completitudine(proiect_id)
    for comp in completitudine:
        for tip_doc in comp['lipsa']:
            nr_lipsa += 1
            row_lipsa += 1
            vals = [nr_lipsa, comp['instalatie'].denumire, tip_doc.denumire, 'Da', 'LIPSA']
            for col, v in enumerate(vals, 1):
                cell = ws2.cell(row=row_lipsa, column=col, value=v)
                cell.border = thin_border
                if col == 5:
                    cell.font = Font(bold=True, color='B71C1C')

    for col in range(1, 6):
        ws2.column_dimensions[chr(64 + col)].width = 30

    # === Sheets per instalatie ===
    for inst in instalatii:
        docs = _query_documente_proiect(proiect_id).filter_by(
            proiect_id=proiect_id,
            tip_instalatie_id=inst.id,
            versiune_curenta=True
        ).filter(DocumentProiect.status != 'anulat').order_by(DocumentProiect.etapa_proiect).all()

        if not docs:
            continue

        sheet_name = inst.cod[:31]  # Excel max 31 chars
        wsi = wb.create_sheet(sheet_name)
        wsi.merge_cells('A1:H1')
        wsi['A1'] = f'{inst.denumire}'
        wsi['A1'].font = Font(bold=True, size=12)

        color_hex = inst.culoare_hex.replace('#', '')
        inst_fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type='solid')

        h_row = ['Nr.', 'Denumire', 'Nr. Doc', 'Rev.', 'Etapa', 'Data Emitere', 'Status', 'Emitent']
        for col, h in enumerate(h_row, 1):
            cell = wsi.cell(row=3, column=col, value=h)
            cell.font = Font(bold=True, color='FFFFFF')
            cell.fill = inst_fill
            cell.border = thin_border

        for idx, doc in enumerate(docs, 1):
            r = 3 + idx
            vals = [
                idx, doc.denumire_document, doc.nr_document or '-',
                doc.revizie or 'Rev.0', doc.etapa_proiect.title() if doc.etapa_proiect else '-',
                doc.data_emitere.strftime('%d.%m.%Y') if doc.data_emitere else '-',
                doc.status.replace('_', ' ').title(), doc.emitent or '-'
            ]
            for col, v in enumerate(vals, 1):
                cell = wsi.cell(row=r, column=col, value=v)
                cell.border = thin_border

        for col in range(1, 9):
            wsi.column_dimensions[chr(64 + col)].width = 20
        wsi.column_dimensions['B'].width = 40

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'Index_Documente_{proiect.cod_proiect}_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# VERIFICARE COMPLETITUDINE
# ============================================================

@doc_proiecte_bp.route('/verificare-completitudine')
@login_required
def verificare_completitudine(proiect_id):
    """Raport documente obligatorii lipsa per instalatie."""
    proiect = get_project_or_404(proiect_id)
    completitudine = _get_completitudine(proiect_id)

    total_obligatorii = sum(c['total_obligatorii'] for c in completitudine)
    total_prezente = sum(c['prezente'] for c in completitudine)
    total_lipsa = sum(len(c['lipsa']) for c in completitudine)
    total_expirate = sum(c['expirate'] for c in completitudine)
    procent_general = int((total_prezente / total_obligatorii) * 100) if total_obligatorii > 0 else 0

    return render_template('proiecte/documente/completitudine.html',
                           proiect=proiect,
                           completitudine=completitudine,
                           total_obligatorii=total_obligatorii,
                           total_prezente=total_prezente,
                           total_lipsa=total_lipsa,
                           total_expirate=total_expirate,
                           procent_general=procent_general)


# ============================================================
# API: TIPURI DOCUMENTE PER INSTALATIE (AJAX)
# ============================================================

@doc_proiecte_bp.route('/api/tipuri-documente/<int:tip_instalatie_id>')
@login_required
def api_tipuri_documente(proiect_id, tip_instalatie_id):
    """Returneaza tipurile de documente pentru o instalatie (AJAX)."""
    get_project_or_404(proiect_id)
    tipuri = TipDocumentProiect.query.filter_by(
        tip_instalatie_id=tip_instalatie_id
    ).order_by(TipDocumentProiect.ordine).all()

    return jsonify([{
        'id': t.id,
        'cod': t.cod,
        'denumire': t.denumire,
        'obligatoriu': t.obligatoriu,
    } for t in tipuri])


# ============================================================
# DESCARCA REVIZIE
# ============================================================

@doc_proiecte_bp.route('/revizie/<int:rev_id>/descarca')
@login_required
def descarca_revizie(proiect_id, rev_id):
    """Descarca fisierul unei revizii anterioare."""
    get_project_or_404(proiect_id)
    rev = get_project_document_revision_or_404(rev_id)
    doc = rev.document_proiect

    if not doc or doc.proiect_id != proiect_id or not rev.fisier_path:
        abort(404)

    filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], rev.fisier_path)
    if not os.path.exists(filepath):
        abort(404)

    return send_file(filepath, as_attachment=True,
                     download_name=f'{doc.denumire_document}_Rev{rev.nr_revizie}.pdf')
