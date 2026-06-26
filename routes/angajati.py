"""
Rute pentru gestionarea angajatilor - Modul complet
"""

import os
import io
import re
from datetime import datetime, date, timedelta
from decimal import Decimal

from flask import (
    Blueprint, render_template, redirect, url_for, flash,
    request, current_app, jsonify, send_file
)
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from sqlalchemy import extract, func

from models import db, Angajat, AngajatProiect, Pontaj, Document, Concediu, Proiect
from forms.angajati_forms import AngajatForm
from services.security.tenant_access import (
    get_employee_or_404,
    get_project_or_404,
    query_employees_for_tenant,
    query_for_tenant,
    query_leave_requests_for_tenant,
    query_legacy_documents_for_tenant,
    query_timesheets_for_tenant,
    tenant_id_for_new_record_or_403,
)

angajati_bp = Blueprint('angajati', __name__)


# ============================================================
# LISTA ANGAJATI - paginata, cu filtre si sortare
# ============================================================

@angajati_bp.route('/')
@login_required
def lista():
    page = request.args.get('page', 1, type=int)
    cautare = request.args.get('cautare', '').strip()
    status = request.args.get('status', '')
    functie = request.args.get('functie', '')
    proiect = request.args.get('proiect', '', type=str)
    sort = request.args.get('sort', 'nume_asc')

    query = query_employees_for_tenant()

    # Filtre text
    if cautare:
        query = query.filter(
            db.or_(
                Angajat.nume.ilike(f'%{cautare}%'),
                Angajat.prenume.ilike(f'%{cautare}%'),
                Angajat.cnp.ilike(f'%{cautare}%')
            )
        )
    if status:
        query = query.filter_by(status=status)
    if functie:
        query = query.filter_by(functie=functie)

    # Filtru proiect activ
    if proiect:
        get_project_or_404(int(proiect))
        query = query.filter(
            Angajat.id.in_(
                db.session.query(AngajatProiect.angajat_id).filter(
                    AngajatProiect.proiect_id == int(proiect),
                    db.or_(
                        AngajatProiect.data_sfarsit.is_(None),
                        AngajatProiect.data_sfarsit >= date.today()
                    )
                )
            )
        )

    # Sortare
    sort_map = {
        'nume_asc': [Angajat.nume.asc(), Angajat.prenume.asc()],
        'nume_desc': [Angajat.nume.desc(), Angajat.prenume.desc()],
        'data_angajare_asc': [Angajat.data_angajare.asc()],
        'data_angajare_desc': [Angajat.data_angajare.desc()],
        'functie': [Angajat.functie.asc(), Angajat.nume.asc()],
    }
    for col in sort_map.get(sort, sort_map['nume_asc']):
        query = query.order_by(col)

    pagination = query.paginate(page=page, per_page=20, error_out=False)

    # Statistici
    total_activi = query_employees_for_tenant().filter_by(status='activ').count()
    total_inactivi = query_employees_for_tenant().filter_by(status='inactiv').count()
    total_suspendati = query_employees_for_tenant().filter_by(status='suspendat').count()
    stats_functii = dict(
        query_employees_for_tenant()
        .with_entities(Angajat.functie, func.count())
        .filter_by(status='activ')
        .group_by(Angajat.functie).all()
    )

    # Proiecte active (pentru dropdown filtru)
    proiecte_active = query_for_tenant(Proiect).filter_by(status='activ').order_by(Proiect.nume).all()

    return render_template('angajati/lista.html',
                           pagination=pagination,
                           angajati=pagination.items,
                           cautare=cautare,
                           status_filtru=status,
                           functie_filtru=functie,
                           proiect_filtru=proiect,
                           sort=sort,
                           functii=Angajat.FUNCTII,
                           proiecte_active=proiecte_active,
                           total_activi=total_activi,
                           total_inactivi=total_inactivi,
                           total_suspendati=total_suspendati,
                           stats_functii=stats_functii)


# ============================================================
# ADAUGA ANGAJAT
# ============================================================

@angajati_bp.route('/adauga', methods=['GET', 'POST'])
@login_required
def adauga():
    form = AngajatForm()
    if form.validate_on_submit():
        angajat = Angajat(
            tenant_id=tenant_id_for_new_record_or_403(),
            nume=form.nume.data.strip(),
            prenume=form.prenume.data.strip(),
            cnp=form.cnp.data.strip() if form.cnp.data else None,
            telefon=form.telefon.data.strip() if form.telefon.data else '',
            email=form.email.data.strip() if form.email.data else '',
            adresa=form.adresa.data.strip() if form.adresa.data else '',
            functie=form.functie.data,
            specializari=form.specializari.data.strip() if form.specializari.data else '',
            tip_contract=form.tip_contract.data,
            salariu_baza=form.salariu_baza.data if form.salariu_baza.data else None,
            nr_contract=form.nr_contract.data.strip() if form.nr_contract.data else '',
            serie_bi=form.serie_bi.data.strip() if form.serie_bi.data else '',
            nr_bi=form.nr_bi.data.strip() if form.nr_bi.data else '',
            observatii=form.observatii.data.strip() if form.observatii.data else '',
            data_angajare=form.data_angajare.data,
            data_nasterii=form.data_nasterii.data,
        )

        # Poza profil
        _handle_photo_upload(angajat, form.poza_profil.data)

        db.session.add(angajat)
        db.session.flush()  # ca sa avem angajat.id pentru AngajatProiect

        # Asignare proiecte (santiere) selectate
        _sync_angajat_proiecte(angajat, form.proiecte_asignate.data or [])

        db.session.commit()
        flash(f'Angajatul {angajat.nume_complet} a fost adaugat cu succes!', 'success')

        if 'salveaza_si_adauga' in request.form:
            return redirect(url_for('angajati.adauga'))
        return redirect(url_for('angajati.detalii', id=angajat.id))

    return render_template('angajati/formular.html', form=form, angajat=None)


# ============================================================
# FISA ANGAJAT (detalii cu tabs)
# ============================================================

@angajati_bp.route('/<int:id>')
@login_required
def detalii(id):
    angajat = get_employee_or_404(id)
    proiecte_vizibile = query_for_tenant(Proiect).with_entities(Proiect.id)
    proiecte_asoc = AngajatProiect.query.filter(
        AngajatProiect.angajat_id == id,
        AngajatProiect.proiect_id.in_(proiecte_vizibile),
    ).all()
    pontaje = query_timesheets_for_tenant().filter_by(angajat_id=id).order_by(Pontaj.data.desc()).limit(30).all()
    documente = query_legacy_documents_for_tenant().filter_by(angajat_id=id).order_by(Document.data_upload.desc()).all()
    concedii = query_leave_requests_for_tenant().filter_by(angajat_id=id).order_by(Concediu.data_start.desc()).limit(10).all()

    # Ore luna curenta
    luna_start = date.today().replace(day=1)
    ore_luna = query_timesheets_for_tenant().with_entities(func.sum(Pontaj.ore_lucrate)).filter(
        Pontaj.angajat_id == id,
        Pontaj.data >= luna_start
    ).scalar()
    ore_luna = float(ore_luna) if ore_luna else 0

    # Documente expirate / expira curand
    doc_expirate = query_legacy_documents_for_tenant().filter(
        Document.angajat_id == id,
        Document.data_expirare.isnot(None),
        Document.data_expirare < date.today()
    ).count()
    doc_expira_curand = query_legacy_documents_for_tenant().filter(
        Document.angajat_id == id,
        Document.data_expirare.isnot(None),
        Document.data_expirare >= date.today(),
        Document.data_expirare <= date.today() + timedelta(days=30)
    ).count()

    return render_template('angajati/fisa.html',
                           angajat=angajat,
                           proiecte_asoc=proiecte_asoc,
                           pontaje=pontaje,
                           documente=documente,
                           concedii=concedii,
                           ore_luna=ore_luna,
                           doc_expirate=doc_expirate,
                           doc_expira_curand=doc_expira_curand)


# ============================================================
# EDITEAZA ANGAJAT
# ============================================================

@angajati_bp.route('/<int:id>/editeaza', methods=['GET', 'POST'])
@login_required
def editeaza(id):
    angajat = get_employee_or_404(id)
    form = AngajatForm(obj=angajat)
    form.angajat_id.data = str(angajat.id)

    # Pre-populez proiectele asignate active (data_sfarsit NULL sau >= today)
    if request.method == 'GET':
        asocieri_active = AngajatProiect.query.filter(
            AngajatProiect.angajat_id == angajat.id,
            AngajatProiect.proiect_id.in_(query_for_tenant(Proiect).with_entities(Proiect.id)),
            db.or_(
                AngajatProiect.data_sfarsit.is_(None),
                AngajatProiect.data_sfarsit >= date.today(),
            )
        ).all()
        form.proiecte_asignate.data = [a.proiect_id for a in asocieri_active]

    if form.validate_on_submit():
        angajat.nume = form.nume.data.strip()
        angajat.prenume = form.prenume.data.strip()
        angajat.cnp = form.cnp.data.strip() if form.cnp.data else None
        angajat.telefon = form.telefon.data.strip() if form.telefon.data else ''
        angajat.email = form.email.data.strip() if form.email.data else ''
        angajat.adresa = form.adresa.data.strip() if form.adresa.data else ''
        angajat.functie = form.functie.data
        angajat.specializari = form.specializari.data.strip() if form.specializari.data else ''
        angajat.tip_contract = form.tip_contract.data
        angajat.salariu_baza = form.salariu_baza.data if form.salariu_baza.data else None
        angajat.nr_contract = form.nr_contract.data.strip() if form.nr_contract.data else ''
        angajat.serie_bi = form.serie_bi.data.strip() if form.serie_bi.data else ''
        angajat.nr_bi = form.nr_bi.data.strip() if form.nr_bi.data else ''
        angajat.status = form.status.data
        angajat.observatii = form.observatii.data.strip() if form.observatii.data else ''
        angajat.data_angajare = form.data_angajare.data
        angajat.data_nasterii = form.data_nasterii.data
        angajat.data_incetare = form.data_incetare.data

        _handle_photo_upload(angajat, form.poza_profil.data)

        # Sync proiecte (santiere) - diff cu cele existente
        proiecte_noi = form.proiecte_asignate.data or []
        adaugate, eliminate = _sync_angajat_proiecte(angajat, proiecte_noi)

        db.session.commit()

        msg = 'Datele angajatului au fost actualizate!'
        if adaugate or eliminate:
            parti = []
            if adaugate:
                parti.append(f'{adaugate} santiere adaugate')
            if eliminate:
                parti.append(f'{eliminate} santiere eliminate')
            msg += f' ({", ".join(parti)})'
        flash(msg, 'success')
        return redirect(url_for('angajati.detalii', id=id))

    return render_template('angajati/formular.html', form=form, angajat=angajat)


def _sync_angajat_proiecte(angajat, proiecte_ids_noi):
    """
    Sincronizeaza asocierile angajat-proiect cu lista noua de id-uri.

    Strategie:
    - Pentru fiecare proiect_id NOU care nu exista in asocierile active:
      - Daca exista o asociere dezalocata recent (data_sfarsit setat), o re-activez
        (data_sfarsit = NULL)
      - Altfel creez o asociere noua cu data_start = today
    - Pentru fiecare asociere ACTIVA al carei proiect NU e in lista noua:
      - Setez data_sfarsit = today (dezalocare blanda, pastreaza istoricul)

    Returneaza tuple (adaugate, eliminate) - counturi pentru flash message.
    """
    proiecte_ids_noi = set(int(pid) for pid in proiecte_ids_noi if pid)
    for pid in proiecte_ids_noi:
        get_project_or_404(pid)
    today_d = date.today()

    # Asocieri active curente
    asocieri_active = AngajatProiect.query.filter(
        AngajatProiect.angajat_id == angajat.id,
        AngajatProiect.proiect_id.in_(query_for_tenant(Proiect).with_entities(Proiect.id)),
        db.or_(
            AngajatProiect.data_sfarsit.is_(None),
            AngajatProiect.data_sfarsit >= today_d,
        )
    ).all()
    proiecte_active = {a.proiect_id: a for a in asocieri_active}

    adaugate = 0
    eliminate = 0

    # ADAUGA: proiecte din lista noua care nu sunt active
    for pid in proiecte_ids_noi:
        if pid in proiecte_active:
            continue  # deja asignat
        # Verific daca exista o asociere dezalocata in trecut
        veche = (AngajatProiect.query
                 .filter_by(angajat_id=angajat.id, proiect_id=pid)
                 .order_by(AngajatProiect.data_sfarsit.desc()).first())
        if veche and veche.data_sfarsit:
            # Re-activez asocierea dezalocata
            veche.data_sfarsit = None
            veche.data_start = today_d
        else:
            ap = AngajatProiect(
                angajat_id=angajat.id, proiect_id=pid,
                data_start=today_d,
                functie_pe_proiect=angajat.functie,
            )
            db.session.add(ap)
        adaugate += 1

    # ELIMINA: asocieri active care nu mai sunt in lista noua
    for pid, asoc in proiecte_active.items():
        if pid not in proiecte_ids_noi:
            asoc.data_sfarsit = today_d
            eliminate += 1

    return adaugate, eliminate


# ============================================================
# DEZACTIVEAZA / ACTIVEAZA ANGAJAT
# ============================================================

@angajati_bp.route('/<int:id>/dezactiveaza', methods=['POST'])
@login_required
def dezactiveaza(id):
    angajat = get_employee_or_404(id)
    if angajat.status == 'activ':
        angajat.status = 'inactiv'
        angajat.data_incetare = date.today()
        flash(f'Angajatul {angajat.nume_complet} a fost dezactivat.', 'warning')
    else:
        angajat.status = 'activ'
        angajat.data_incetare = None
        flash(f'Angajatul {angajat.nume_complet} a fost reactivat.', 'success')
    db.session.commit()
    return redirect(url_for('angajati.lista'))


# ============================================================
# UPLOAD POZA PROFIL (AJAX + standard)
# ============================================================

@angajati_bp.route('/<int:id>/poza', methods=['POST'])
@login_required
def upload_poza(id):
    angajat = get_employee_or_404(id)

    if 'poza' not in request.files:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'Niciun fisier selectat.'}), 400
        flash('Niciun fisier selectat.', 'danger')
        return redirect(url_for('angajati.detalii', id=id))

    fisier = request.files['poza']
    if not fisier.filename:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'Fisier invalid.'}), 400
        flash('Fisier invalid.', 'danger')
        return redirect(url_for('angajati.detalii', id=id))

    ext = fisier.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('jpg', 'jpeg', 'png'):
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': 'Doar fisiere JPG sau PNG.'}), 400
        flash('Doar fisiere JPG sau PNG.', 'danger')
        return redirect(url_for('angajati.detalii', id=id))

    try:
        from PIL import Image
        img = Image.open(fisier.stream)
        # Center crop to square
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        img = img.resize((300, 300), Image.LANCZOS)

        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        filename = f"profil_{angajat.id}.jpg"
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        img.save(filepath, 'JPEG', quality=85)
        angajat.poza_profil = filename
        db.session.commit()

        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({
                'success': True,
                'url': url_for('static', filename=f'uploads/{filename}', _external=False)
            })
        flash('Poza de profil a fost actualizata!', 'success')
    except Exception as e:
        if request.is_json or request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            return jsonify({'success': False, 'error': str(e)}), 500
        flash(f'Eroare la procesarea imaginii: {e}', 'danger')

    return redirect(url_for('angajati.detalii', id=id))


# ============================================================
# EXPORT EXCEL
# ============================================================

@angajati_bp.route('/export-excel')
@login_required
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = Workbook()

    # --- Sheet 1: Lista Angajati ---
    ws = wb.active
    ws.title = 'Angajati'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Nr.', 'Nume', 'Prenume', 'CNP', 'Telefon', 'Email', 'Functie',
               'Tip Contract', 'Salariu Baza', 'Data Angajare', 'Status',
               'Specializari', 'Nr. Contract']

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    angajati = query_employees_for_tenant().order_by(Angajat.nume).all()
    for row_idx, ang in enumerate(angajati, 2):
        values = [
            row_idx - 1,
            ang.nume,
            ang.prenume,
            ang.cnp or '',
            ang.telefon or '',
            ang.email or '',
            dict(Angajat.FUNCTII).get(ang.functie, ang.functie),
            ang.tip_contract or '',
            float(ang.salariu_baza) if ang.salariu_baza else '',
            ang.data_angajare.strftime('%d.%m.%Y') if ang.data_angajare else '',
            ang.status,
            ang.specializari or '',
            ang.nr_contract or '',
        ]
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = ws.dimensions

    # --- Sheet 2: Statistici ---
    ws2 = wb.create_sheet('Statistici')
    ws2.cell(row=1, column=1, value='Functie').font = header_font
    ws2.cell(row=1, column=1).fill = header_fill
    ws2.cell(row=1, column=2, value='Activi').font = header_font
    ws2.cell(row=1, column=2).fill = header_fill
    ws2.cell(row=1, column=3, value='Inactivi').font = header_font
    ws2.cell(row=1, column=3).fill = header_fill
    ws2.cell(row=1, column=4, value='Suspendati').font = header_font
    ws2.cell(row=1, column=4).fill = header_fill
    ws2.cell(row=1, column=5, value='Total').font = header_font
    ws2.cell(row=1, column=5).fill = header_fill

    stats = db.session.query(
        Angajat.functie, Angajat.status, func.count()
    ).filter(
        Angajat.id.in_(query_employees_for_tenant().with_entities(Angajat.id))
    ).group_by(Angajat.functie, Angajat.status).all()

    pivot = {}
    for f, s, c in stats:
        if f not in pivot:
            pivot[f] = {'activ': 0, 'inactiv': 0, 'suspendat': 0}
        pivot[f][s] = c

    for row_idx, (f, counts) in enumerate(sorted(pivot.items()), 2):
        label = dict(Angajat.FUNCTII).get(f, f)
        total = sum(counts.values())
        ws2.cell(row=row_idx, column=1, value=label)
        ws2.cell(row=row_idx, column=2, value=counts['activ'])
        ws2.cell(row=row_idx, column=3, value=counts['inactiv'])
        ws2.cell(row=row_idx, column=4, value=counts['suspendat'])
        ws2.cell(row=row_idx, column=5, value=total)

    for col in ws2.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(c.value or '')) for c in col)
        ws2.column_dimensions[col_letter].width = max_len + 3

    ws2.freeze_panes = 'A2'

    # Trimite fisier
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'angajati_export_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(
        buffer,
        as_attachment=True,
        download_name=filename,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============================================================
# TEMPLATE IMPORT EXCEL
# ============================================================

@angajati_bp.route('/template-import')
@login_required
def template_import():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation

    wb = Workbook()

    # Sheet 1: Date Angajati
    ws = wb.active
    ws.title = 'Date Angajati'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')

    headers = ['Nume*', 'Prenume*', 'CNP', 'Telefon', 'Email',
               'Functie*', 'Specializari', 'Data Angajare* (ZZ.LL.AAAA)',
               'Tip Contract', 'Salariu Baza', 'Nr. Contract',
               'Serie BI', 'Nr. BI', 'Observatii']

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    # Exemplu
    example = ['Popescu', 'Ion', '1850412345678', '0721000001',
               'ion@email.ro', 'Muncitor', 'Zidarie, Tencuiala',
               '15.03.2024', 'nedeterminat', '4500', 'CIM-2024-001',
               'XZ', '123456', '']
    for col, val in enumerate(example, 1):
        ws.cell(row=2, column=col, value=val)

    # Data validation pe functie
    functii_str = ','.join([f[0] for f in Angajat.FUNCTII])
    dv_functie = DataValidation(type='list', formula1=f'"{functii_str}"', allow_blank=True)
    dv_functie.error = 'Selectati o functie din lista'
    dv_functie.prompt = 'Alegeti functia'
    ws.add_data_validation(dv_functie)
    dv_functie.add(f'F2:F1000')

    dv_contract = DataValidation(type='list', formula1='"nedeterminat,determinat,zilier"')
    ws.add_data_validation(dv_contract)
    dv_contract.add(f'I2:I1000')

    for col in ws.columns:
        col_letter = col[0].column_letter
        max_len = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col_letter].width = max_len + 3

    # Sheet 2: Instructiuni
    ws2 = wb.create_sheet('Instructiuni')
    instructions = [
        'INSTRUCTIUNI IMPORT ANGAJATI',
        '',
        'Campuri obligatorii (marcate cu *):',
        '  - Nume',
        '  - Prenume',
        '  - Functie (din lista: Muncitor, Maistru, Sef_echipa, Inginer, Tehnician, Conducator_auto, Macaragiu, Sudor, Electrician, Alte)',
        '  - Data Angajare (format: ZZ.LL.AAAA, ex: 15.03.2024)',
        '',
        'Campuri optionale:',
        '  - CNP: exact 13 cifre',
        '  - Telefon: format 07xxxxxxxx sau +407xxxxxxxx',
        '  - Email: adresa valida de email',
        '  - Specializari: separate prin virgula',
        '  - Tip Contract: nedeterminat, determinat, zilier (implicit: nedeterminat)',
        f'  - Salariu Baza: minim 3700 RON',
        '  - Nr. Contract, Serie BI, Nr. BI: text liber',
        '',
        'NOTE:',
        '  - Primul rand contine antetul (nu se importa)',
        '  - CNP-ul trebuie sa fie unic - duplicatele vor fi raportate ca erori',
        '  - Completati datele in sheet-ul "Date Angajati"',
    ]
    for row, text in enumerate(instructions, 1):
        cell = ws2.cell(row=row, column=1, value=text)
        if row == 1:
            cell.font = Font(bold=True, size=14)
    ws2.column_dimensions['A'].width = 80

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    return send_file(
        buffer,
        as_attachment=True,
        download_name='template_import_angajati.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )


# ============================================================
# IMPORT EXCEL
# ============================================================

@angajati_bp.route('/import-excel', methods=['POST'])
@login_required
def import_excel():
    if 'fisier' not in request.files:
        flash('Niciun fisier selectat.', 'danger')
        return redirect(url_for('angajati.lista'))

    fisier = request.files['fisier']
    if not fisier.filename or not fisier.filename.endswith('.xlsx'):
        flash('Selectati un fisier Excel (.xlsx).', 'danger')
        return redirect(url_for('angajati.lista'))

    try:
        from openpyxl import load_workbook
        wb = load_workbook(fisier.stream)
        ws = wb.worksheets[0]
    except Exception:
        flash('Fisierul Excel nu a putut fi citit.', 'danger')
        return redirect(url_for('angajati.lista'))

    errors = []
    angajati_noi = []
    cnp_seen = set()
    functii_valide = [f[0] for f in Angajat.FUNCTII]

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or not any(row):
            continue

        row = list(row) + [None] * (14 - len(row))  # pad
        (nume, prenume, cnp, telefon, email, functie, specializari,
         data_ang_str, tip_contract, salariu, nr_contract,
         serie_bi, nr_bi, observatii) = row[:14]

        row_errors = []

        # Required
        if not nume or not str(nume).strip():
            row_errors.append('Numele este obligatoriu')
        if not prenume or not str(prenume).strip():
            row_errors.append('Prenumele este obligatoriu')
        if not functie or str(functie).strip() not in functii_valide:
            row_errors.append(f'Functie invalida: {functie}')
        if not data_ang_str:
            row_errors.append('Data angajarii este obligatorie')

        # Parse date
        data_angajare = None
        if data_ang_str:
            data_ang_str = str(data_ang_str).strip()
            for fmt in ('%d.%m.%Y', '%Y-%m-%d', '%d/%m/%Y'):
                try:
                    data_angajare = datetime.strptime(data_ang_str, fmt).date()
                    break
                except ValueError:
                    continue
            if not data_angajare:
                # Try if openpyxl already parsed it as datetime
                if isinstance(data_ang_str, (datetime, date)):
                    data_angajare = data_ang_str if isinstance(data_ang_str, date) else data_ang_str.date()
                else:
                    row_errors.append(f'Format data invalid: {data_ang_str}')

        # CNP
        if cnp:
            cnp = str(cnp).strip()
            if not re.match(r'^\d{13}$', cnp):
                row_errors.append('CNP invalid (trebuie 13 cifre)')
            elif cnp in cnp_seen:
                row_errors.append('CNP duplicat in fisier')
            elif Angajat.query.filter_by(cnp=cnp).first():
                row_errors.append('CNP exista deja in baza de date')
            else:
                cnp_seen.add(cnp)

        # Telefon
        if telefon:
            telefon = str(telefon).strip()
            if not re.match(r'^(\+40|0)[0-9]{9}$', telefon):
                row_errors.append('Format telefon invalid')

        # Email
        if email:
            email = str(email).strip()
            if '@' not in email:
                row_errors.append('Email invalid')

        # Salariu
        salariu_val = None
        if salariu:
            try:
                salariu_val = float(salariu)
                if salariu_val < 3700:
                    row_errors.append('Salariul este sub minimul pe economie (3700 RON)')
            except (ValueError, TypeError):
                row_errors.append('Salariu invalid')

        if row_errors:
            errors.append({'rand': row_idx, 'erori': row_errors})
        else:
            angajati_noi.append(Angajat(
                tenant_id=tenant_id_for_new_record_or_403(),
                nume=str(nume).strip(),
                prenume=str(prenume).strip(),
                cnp=str(cnp).strip() if cnp else None,
                telefon=str(telefon).strip() if telefon else '',
                email=str(email).strip() if email else '',
                functie=str(functie).strip(),
                specializari=str(specializari).strip() if specializari else '',
                data_angajare=data_angajare,
                tip_contract=str(tip_contract).strip() if tip_contract else 'nedeterminat',
                salariu_baza=Decimal(str(salariu_val)) if salariu_val else None,
                nr_contract=str(nr_contract).strip() if nr_contract else '',
                serie_bi=str(serie_bi).strip() if serie_bi else '',
                nr_bi=str(nr_bi).strip() if nr_bi else '',
                observatii=str(observatii).strip() if observatii else '',
            ))

    if errors:
        error_details = '; '.join(
            f'Rand {e["rand"]}: {", ".join(e["erori"])}' for e in errors[:10]
        )
        flash(f'Import esuat - {len(errors)} randuri cu erori. {error_details}', 'danger')
    elif not angajati_noi:
        flash('Fisierul nu contine date de importat.', 'warning')
    else:
        for ang in angajati_noi:
            db.session.add(ang)
        db.session.commit()
        flash(f'{len(angajati_noi)} angajati importati cu succes!', 'success')

    return redirect(url_for('angajati.lista'))


# ============================================================
# PONTAJE JSON (AJAX pentru calendar)
# ============================================================

@angajati_bp.route('/<int:id>/pontaje-json')
@login_required
def pontaje_json(id):
    get_employee_or_404(id)
    luna = request.args.get('luna', date.today().month, type=int)
    an = request.args.get('an', date.today().year, type=int)

    pontaje = query_timesheets_for_tenant().filter(
        Pontaj.angajat_id == id,
        extract('month', Pontaj.data) == luna,
        extract('year', Pontaj.data) == an
    ).all()

    concedii = query_leave_requests_for_tenant().filter(
        Concediu.angajat_id == id,
        Concediu.status == 'aprobat',
        extract('month', Concediu.data_start) <= luna,
        extract('month', Concediu.data_sfarsit) >= luna,
        extract('year', Concediu.data_start) <= an,
    ).all()

    events = []
    for p in pontaje:
        events.append({
            'data': p.data.isoformat(),
            'ore': float(p.ore_lucrate) if p.ore_lucrate else 0,
            'tip_zi': p.tip_zi,
            'status': p.status,
            'proiect': p.proiect.cod_proiect if p.proiect else '',
            'ora_start': p.ora_start or '',
            'ora_sfarsit': p.ora_sfarsit or '',
        })

    # Statistici lunare
    total_ore = sum(e['ore'] for e in events)
    ore_normale = sum(float(p.ore_normale or 0) for p in pontaje)
    ore_supl_50 = sum(float(p.ore_suplimentare_50 or 0) for p in pontaje)
    ore_supl_100 = sum(float(p.ore_suplimentare_100 or 0) for p in pontaje)

    return jsonify({
        'events': events,
        'stats': {
            'total_ore': round(total_ore, 1),
            'ore_normale': round(ore_normale, 1),
            'ore_suplimentare_50': round(ore_supl_50, 1),
            'ore_suplimentare_100': round(ore_supl_100, 1),
            'zile_lucrate': len([e for e in events if e['ore'] > 0]),
        },
        'concedii': [
            {
                'tip': c.tip,
                'data_start': c.data_start.isoformat(),
                'data_sfarsit': c.data_sfarsit.isoformat(),
            } for c in concedii
        ]
    })


# ============================================================
# CAUTARE AJAX
# ============================================================

@angajati_bp.route('/cauta')
@login_required
def cauta():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])

    results = query_employees_for_tenant().filter(
        db.or_(
            Angajat.nume.ilike(f'%{q}%'),
            Angajat.prenume.ilike(f'%{q}%'),
            Angajat.cnp.ilike(f'%{q}%')
        )
    ).limit(10).all()

    return jsonify([{
        'id': a.id,
        'nume_complet': a.nume_complet,
        'cnp': a.cnp or '',
        'functie': dict(Angajat.FUNCTII).get(a.functie, a.functie),
        'status': a.status,
    } for a in results])


# ============================================================
# HELPER: Upload poza profil
# ============================================================

def _handle_photo_upload(angajat, file_data):
    """Proceseaza upload poza profil cu crop si resize."""
    if not file_data or not hasattr(file_data, 'filename') or not file_data.filename:
        return

    ext = file_data.filename.rsplit('.', 1)[-1].lower()
    if ext not in ('jpg', 'jpeg', 'png'):
        return

    try:
        from PIL import Image
        img = Image.open(file_data.stream if hasattr(file_data, 'stream') else file_data)
        w, h = img.size
        side = min(w, h)
        left = (w - side) // 2
        top = (h - side) // 2
        img = img.crop((left, top, left + side, top + side))
        img = img.resize((300, 300), Image.LANCZOS)

        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        # Use ID if available, else temp name
        ident = angajat.id or 'new'
        filename = f"profil_{ident}.jpg"
        filepath = os.path.join(current_app.config['UPLOAD_FOLDER'], filename)
        img.save(filepath, 'JPEG', quality=85)
        angajat.poza_profil = filename
    except Exception:
        pass  # Fail silently for photo, don't break the main operation
