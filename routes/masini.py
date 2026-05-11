"""
EDIFICO WORKFORCE - Routes Masini (Flota Auto)
Modul complet de gestiune masini de serviciu si atribuire angajati.
"""

import os
from datetime import datetime, date, timedelta
from functools import wraps

from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, abort
from flask_login import login_required, current_user

from models import (
    db, Masina, DocumentMasina, AtribuireMasina, ConducereMasina,
    DefectiuneMasina, Angajat, Proiect
)

masini_bp = Blueprint('masini', __name__, url_prefix='/masini')


def manager_or_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_manager:
            abort(403)
        return f(*args, **kwargs)
    return decorated


# ============================================================
# LISTA MASINI (cu filtre + card/table view)
# ============================================================

@masini_bp.route('/')
@login_required
def lista():
    page = request.args.get('page', 1, type=int)
    per_page = 20
    view = request.args.get('view', 'card')  # card sau table

    # Filtre
    cautare = request.args.get('cautare', '').strip()
    status = request.args.get('status', '')
    tip_vehicul = request.args.get('tip_vehicul', '')
    proiect_id = request.args.get('proiect_id', '', type=str)
    combustibil = request.args.get('combustibil', '')
    sort = request.args.get('sort', 'nr_asc')

    query = Masina.query

    if cautare:
        like = f'%{cautare}%'
        query = query.filter(
            db.or_(
                Masina.numar_inmatriculare.ilike(like),
                Masina.marca.ilike(like),
                Masina.model.ilike(like),
                Masina.vin.ilike(like),
            )
        )
    if status:
        query = query.filter(Masina.status == status)
    if tip_vehicul:
        query = query.filter(Masina.tip_vehicul == tip_vehicul)
    if proiect_id:
        query = query.filter(Masina.proiect_id == int(proiect_id))
    if combustibil:
        query = query.filter(Masina.tip_combustibil == combustibil)

    # Sort
    sort_map = {
        'nr_asc': Masina.numar_inmatriculare.asc(),
        'nr_desc': Masina.numar_inmatriculare.desc(),
        'marca_asc': Masina.marca.asc(),
        'marca_desc': Masina.marca.desc(),
        'an_desc': Masina.an_fabricatie.desc(),
        'km_desc': Masina.km_bord.desc(),
    }
    query = query.order_by(sort_map.get(sort, Masina.numar_inmatriculare.asc()))

    pagination = query.paginate(page=page, per_page=per_page, error_out=False)
    masini = pagination.items

    # Stats
    total = Masina.query.count()
    disponibile = Masina.query.filter_by(status='disponibila').count()
    atribuite = Masina.query.filter_by(status='atribuita').count()
    in_service = Masina.query.filter_by(status='service').count()

    # Alerte documente
    today = date.today()
    docs_expirate = Masina.query.filter(
        db.or_(
            db.and_(Masina.data_itp_expirare.isnot(None), Masina.data_itp_expirare < today),
            db.and_(Masina.data_rca_expirare.isnot(None), Masina.data_rca_expirare < today),
        )
    ).count()

    proiecte = Proiect.query.filter_by(status='activ').order_by(Proiect.cod_proiect).all()

    return render_template('masini/lista.html',
        masini=masini, pagination=pagination, view=view,
        cautare=cautare, status=status, tip_vehicul=tip_vehicul,
        proiect_id=proiect_id, combustibil=combustibil, sort=sort,
        total=total, disponibile=disponibile, atribuite=atribuite,
        in_service=in_service, docs_expirate=docs_expirate,
        proiecte=proiecte)


# ============================================================
# ADAUGA MASINA
# ============================================================

@masini_bp.route('/adauga', methods=['GET', 'POST'])
@login_required
@manager_or_admin
def adauga():
    if request.method == 'POST':
        return _salveaza_masina(None)

    proiecte = Proiect.query.filter_by(status='activ').order_by(Proiect.cod_proiect).all()
    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    return render_template('masini/formular.html', masina=None, proiecte=proiecte, angajati=angajati)


# ============================================================
# FISA MASINA (profil cu 6 tab-uri)
# ============================================================

@masini_bp.route('/<int:id>')
@login_required
def fisa(id):
    masina = Masina.query.get_or_404(id)

    # Tab 1: Info generala (direct din masina)
    # Tab 2: Documente
    documente = masina.documente_masina.order_by(DocumentMasina.data_expirare.desc()).all()

    # Tab 3: Atribuiri
    atribuiri = masina.atribuiri.order_by(AtribuireMasina.data_atribuire.desc()).all()

    # Tab 4: Foaie de parcurs (conduceri)
    luna = request.args.get('luna', date.today().month, type=int)
    an = request.args.get('an', date.today().year, type=int)
    conduceri = masina.conduceri.filter(
        db.extract('month', ConducereMasina.data) == luna,
        db.extract('year', ConducereMasina.data) == an,
    ).order_by(ConducereMasina.data.desc()).all()

    total_km_luna = sum(c.km_parcursi for c in conduceri if c.km_parcursi)
    total_combustibil = sum(float(c.combustibil_alimentat or 0) for c in conduceri)
    total_cost_combustibil = sum(float(c.cost_combustibil or 0) for c in conduceri)

    # Tab 5: Defectiuni
    defectiuni = masina.defectiuni.order_by(DefectiuneMasina.data_raportare.desc()).all()

    # Tab 6: Istoric
    proiecte = Proiect.query.filter_by(status='activ').order_by(Proiect.cod_proiect).all()
    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()

    return render_template('masini/fisa.html',
        masina=masina, documente=documente, atribuiri=atribuiri,
        conduceri=conduceri, defectiuni=defectiuni,
        total_km_luna=total_km_luna, total_combustibil=total_combustibil,
        total_cost_combustibil=total_cost_combustibil,
        luna=luna, an=an, proiecte=proiecte, angajati=angajati)


# ============================================================
# EDITEAZA MASINA
# ============================================================

@masini_bp.route('/<int:id>/editeaza', methods=['GET', 'POST'])
@login_required
@manager_or_admin
def editeaza(id):
    masina = Masina.query.get_or_404(id)
    if request.method == 'POST':
        return _salveaza_masina(masina)

    proiecte = Proiect.query.filter_by(status='activ').order_by(Proiect.cod_proiect).all()
    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    return render_template('masini/formular.html', masina=masina, proiecte=proiecte, angajati=angajati)


def _salveaza_masina(masina):
    """Shared save logic for add/edit."""
    f = request.form
    errors = []

    nr_inm = f.get('numar_inmatriculare', '').strip().upper()
    marca = f.get('marca', '').strip()
    model_v = f.get('model', '').strip()

    if not nr_inm:
        errors.append('Numarul de inmatriculare este obligatoriu.')
    if not marca:
        errors.append('Marca este obligatorie.')
    if not model_v:
        errors.append('Modelul este obligatoriu.')

    # Check uniqueness
    if nr_inm:
        existing = Masina.query.filter_by(numar_inmatriculare=nr_inm).first()
        if existing and (masina is None or existing.id != masina.id):
            errors.append(f'Numarul de inmatriculare {nr_inm} exista deja.')

    vin = f.get('vin', '').strip().upper()
    if vin:
        existing_vin = Masina.query.filter_by(vin=vin).first()
        if existing_vin and (masina is None or existing_vin.id != masina.id):
            errors.append(f'VIN-ul {vin} exista deja.')

    if errors:
        for e in errors:
            flash(e, 'danger')
        proiecte = Proiect.query.filter_by(status='activ').order_by(Proiect.cod_proiect).all()
        angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
        return render_template('masini/formular.html', masina=masina, proiecte=proiecte, angajati=angajati)

    if masina is None:
        masina = Masina()
        db.session.add(masina)

    masina.numar_inmatriculare = nr_inm
    masina.marca = marca
    masina.model = model_v
    masina.an_fabricatie = int(f.get('an_fabricatie')) if f.get('an_fabricatie') else None
    masina.vin = vin or None
    masina.culoare = f.get('culoare', '').strip() or None
    masina.tip_combustibil = f.get('tip_combustibil', 'motorina')
    masina.capacitate_cilindrica = int(f.get('capacitate_cilindrica')) if f.get('capacitate_cilindrica') else None
    masina.putere_kw = int(f.get('putere_kw')) if f.get('putere_kw') else None
    masina.tip_vehicul = f.get('tip_vehicul', 'autoturism')
    masina.nr_locuri = int(f.get('nr_locuri', 5)) if f.get('nr_locuri') else 5
    masina.masa_maxima = int(f.get('masa_maxima')) if f.get('masa_maxima') else None
    masina.categorie_permis = f.get('categorie_permis', 'B')
    masina.km_bord = int(f.get('km_bord', 0)) if f.get('km_bord') else 0
    masina.consum_mediu = float(f.get('consum_mediu')) if f.get('consum_mediu') else None
    masina.serie_civ = f.get('serie_civ', '').strip() or None
    masina.nr_carte_identitate = f.get('nr_carte_identitate', '').strip() or None
    masina.proiect_id = int(f.get('proiect_id')) if f.get('proiect_id') else None
    masina.angajat_responsabil_id = int(f.get('angajat_responsabil_id')) if f.get('angajat_responsabil_id') else None
    masina.status = f.get('status', 'disponibila')
    masina.observatii = f.get('observatii', '').strip() or None

    # Date
    for field in ['data_achizitie', 'data_prima_inmatriculare', 'data_itp_expirare',
                   'data_rca_expirare', 'data_casco_expirare', 'data_rovinieta_expirare']:
        val = f.get(field, '').strip()
        setattr(masina, field, datetime.strptime(val, '%Y-%m-%d').date() if val else None)

    db.session.commit()
    flash(f'Masina {masina.denumire_completa} a fost salvata cu succes!', 'success')
    return redirect(url_for('masini.fisa', id=masina.id))


# ============================================================
# ATRIBUIRE MASINA -> ANGAJAT
# ============================================================

@masini_bp.route('/<int:id>/atribuie', methods=['POST'])
@login_required
@manager_or_admin
def atribuie(id):
    masina = Masina.query.get_or_404(id)
    f = request.form

    angajat_id = f.get('angajat_id', type=int)
    if not angajat_id:
        flash('Selectati un angajat.', 'danger')
        return redirect(url_for('masini.fisa', id=id))

    angajat = Angajat.query.get_or_404(angajat_id)

    atribuire = AtribuireMasina(
        masina_id=masina.id,
        angajat_id=angajat.id,
        proiect_id=int(f.get('proiect_id')) if f.get('proiect_id') else None,
        data_atribuire=datetime.strptime(f.get('data_atribuire'), '%Y-%m-%d').date() if f.get('data_atribuire') else date.today(),
        km_preluare=int(f.get('km_preluare')) if f.get('km_preluare') else masina.km_bord,
        stare_preluare=f.get('stare_preluare', 'buna'),
        motiv=f.get('motiv', '').strip() or None,
        observatii=f.get('observatii_atribuire', '').strip() or None,
        atribuit_de=current_user.id,
    )
    db.session.add(atribuire)

    masina.status = 'atribuita'
    masina.angajat_responsabil_id = angajat.id
    if f.get('proiect_id'):
        masina.proiect_id = int(f.get('proiect_id'))

    db.session.commit()
    flash(f'Masina {masina.numar_inmatriculare} a fost atribuita lui {angajat.nume_complet}.', 'success')
    return redirect(url_for('masini.fisa', id=id))


# ============================================================
# RETURNARE MASINA
# ============================================================

@masini_bp.route('/<int:id>/returneaza', methods=['POST'])
@login_required
@manager_or_admin
def returneaza(id):
    masina = Masina.query.get_or_404(id)
    f = request.form

    # Find active assignment
    atribuire = AtribuireMasina.query.filter_by(
        masina_id=masina.id, data_returnare=None
    ).order_by(AtribuireMasina.data_atribuire.desc()).first()

    if atribuire:
        atribuire.data_returnare = date.today()
        atribuire.km_returnare = int(f.get('km_returnare')) if f.get('km_returnare') else masina.km_bord
        atribuire.stare_returnare = f.get('stare_returnare', 'buna')

    masina.status = 'disponibila'
    if f.get('km_returnare'):
        masina.km_bord = int(f.get('km_returnare'))
    masina.angajat_responsabil_id = None

    db.session.commit()
    flash(f'Masina {masina.numar_inmatriculare} a fost returnata.', 'success')
    return redirect(url_for('masini.fisa', id=id))


# ============================================================
# ADAUGA DOCUMENT MASINA
# ============================================================

@masini_bp.route('/<int:id>/document', methods=['POST'])
@login_required
@manager_or_admin
def adauga_document(id):
    masina = Masina.query.get_or_404(id)
    f = request.form

    doc = DocumentMasina(
        masina_id=masina.id,
        tip=f.get('tip_document', 'altele'),
        nume_document=f.get('nume_document', '').strip() or None,
        numar_document=f.get('numar_document', '').strip() or None,
        emitent=f.get('emitent', '').strip() or None,
        data_emitere=datetime.strptime(f.get('data_emitere'), '%Y-%m-%d').date() if f.get('data_emitere') else None,
        data_expirare=datetime.strptime(f.get('data_expirare_doc'), '%Y-%m-%d').date() if f.get('data_expirare_doc') else None,
        cost=float(f.get('cost_document')) if f.get('cost_document') else None,
        observatii=f.get('observatii_document', '').strip() or None,
    )
    db.session.add(doc)

    # Auto-update masina expiry dates based on document type
    if doc.tip == 'itp' and doc.data_expirare:
        masina.data_itp_expirare = doc.data_expirare
    elif doc.tip == 'rca' and doc.data_expirare:
        masina.data_rca_expirare = doc.data_expirare
    elif doc.tip == 'casco' and doc.data_expirare:
        masina.data_casco_expirare = doc.data_expirare
    elif doc.tip == 'rovinieta' and doc.data_expirare:
        masina.data_rovinieta_expirare = doc.data_expirare

    db.session.commit()
    flash(f'Document {doc.tip.upper()} adaugat cu succes.', 'success')
    return redirect(url_for('masini.fisa', id=id))


# ============================================================
# ADAUGA CONDUCERE (FOAIE DE PARCURS)
# ============================================================

@masini_bp.route('/<int:id>/conducere', methods=['POST'])
@login_required
def adauga_conducere(id):
    masina = Masina.query.get_or_404(id)
    f = request.form

    angajat_id = f.get('angajat_id_conducere', type=int)
    if not angajat_id:
        flash('Selectati un angajat.', 'danger')
        return redirect(url_for('masini.fisa', id=id))

    km_start = int(f.get('km_start')) if f.get('km_start') else masina.km_bord
    km_sfarsit = int(f.get('km_sfarsit')) if f.get('km_sfarsit') else None

    conducere = ConducereMasina(
        masina_id=masina.id,
        angajat_id=angajat_id,
        proiect_id=int(f.get('proiect_id_conducere')) if f.get('proiect_id_conducere') else None,
        data=datetime.strptime(f.get('data_conducere'), '%Y-%m-%d').date() if f.get('data_conducere') else date.today(),
        km_start=km_start,
        km_sfarsit=km_sfarsit,
        ruta=f.get('ruta', '').strip() or None,
        scop=f.get('scop', '').strip() or None,
        combustibil_alimentat=float(f.get('combustibil_alimentat')) if f.get('combustibil_alimentat') else None,
        cost_combustibil=float(f.get('cost_combustibil')) if f.get('cost_combustibil') else None,
        observatii=f.get('observatii_conducere', '').strip() or None,
    )
    db.session.add(conducere)

    # Update km on car
    if km_sfarsit and km_sfarsit > masina.km_bord:
        masina.km_bord = km_sfarsit

    db.session.commit()
    flash('Inregistrare foaie de parcurs adaugata cu succes.', 'success')
    return redirect(url_for('masini.fisa', id=id))


# ============================================================
# ADAUGA DEFECTIUNE
# ============================================================

@masini_bp.route('/<int:id>/defectiune', methods=['POST'])
@login_required
def adauga_defectiune(id):
    masina = Masina.query.get_or_404(id)
    f = request.form

    defectiune = DefectiuneMasina(
        masina_id=masina.id,
        raportat_de=int(f.get('raportat_de')) if f.get('raportat_de') else None,
        data_raportare=datetime.strptime(f.get('data_raportare'), '%Y-%m-%d').date() if f.get('data_raportare') else date.today(),
        descriere=f.get('descriere_defectiune', '').strip(),
        gravitate=f.get('gravitate', 'medie'),
        observatii=f.get('observatii_defectiune', '').strip() or None,
    )

    if not defectiune.descriere:
        flash('Descrierea defectiunii este obligatorie.', 'danger')
        return redirect(url_for('masini.fisa', id=id))

    db.session.add(defectiune)
    db.session.commit()
    flash('Defectiune raportata cu succes.', 'success')
    return redirect(url_for('masini.fisa', id=id))


# ============================================================
# ACTUALIZEAZA STATUS DEFECTIUNE
# ============================================================

@masini_bp.route('/defectiune/<int:def_id>/status', methods=['POST'])
@login_required
@manager_or_admin
def update_defectiune(def_id):
    defectiune = DefectiuneMasina.query.get_or_404(def_id)
    f = request.form

    defectiune.status = f.get('status_defectiune', defectiune.status)
    if defectiune.status == 'rezolvata':
        defectiune.data_rezolvare = date.today()
        defectiune.cost_reparatie = float(f.get('cost_reparatie')) if f.get('cost_reparatie') else None
        defectiune.service_extern = f.get('service_extern', '').strip() or None
        defectiune.detalii_reparatie = f.get('detalii_reparatie', '').strip() or None

    db.session.commit()
    flash('Status defectiune actualizat.', 'success')
    return redirect(url_for('masini.fisa', id=defectiune.masina_id))


# ============================================================
# SCHIMBA STATUS MASINA (service/disponibila)
# ============================================================

@masini_bp.route('/<int:id>/status', methods=['POST'])
@login_required
@manager_or_admin
def schimba_status(id):
    masina = Masina.query.get_or_404(id)
    new_status = request.form.get('new_status', '')

    valid = [s[0] for s in Masina.STATUSURI]
    if new_status in valid:
        masina.status = new_status
        db.session.commit()
        flash(f'Status masina actualizat: {new_status.title()}', 'success')
    else:
        flash('Status invalid.', 'danger')

    return redirect(url_for('masini.fisa', id=id))


# ============================================================
# EXPORT EXCEL FLOTA
# ============================================================

@masini_bp.route('/export-excel')
@login_required
@manager_or_admin
def export_excel():
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import tempfile

    wb = Workbook()

    # Sheet 1: Lista masini
    ws1 = wb.active
    ws1.title = 'Flota Auto'

    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    headers = ['Nr. Inmatriculare', 'Marca', 'Model', 'An', 'Tip', 'Combustibil',
               'KM Bord', 'Status', 'Sofer Responsabil', 'Proiect', 'ITP Expira', 'RCA Expira']
    for col, h in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    masini = Masina.query.order_by(Masina.numar_inmatriculare).all()
    for row, m in enumerate(masini, 2):
        vals = [
            m.numar_inmatriculare, m.marca, m.model, m.an_fabricatie,
            m.tip_vehicul, m.tip_combustibil, m.km_bord, m.status,
            m.angajat_responsabil.nume_complet if m.angajat_responsabil else '',
            m.proiect.cod_proiect if m.proiect else '',
            m.data_itp_expirare.strftime('%d.%m.%Y') if m.data_itp_expirare else '',
            m.data_rca_expirare.strftime('%d.%m.%Y') if m.data_rca_expirare else '',
        ]
        for col, v in enumerate(vals, 1):
            cell = ws1.cell(row=row, column=col, value=v)
            cell.border = thin_border

    # Auto width
    for col in ws1.columns:
        max_len = max(len(str(cell.value or '')) for cell in col) + 2
        ws1.column_dimensions[col[0].column_letter].width = min(max_len, 30)
    ws1.auto_filter.ref = ws1.dimensions
    ws1.freeze_panes = 'A2'

    # Sheet 2: Atribuiri active
    ws2 = wb.create_sheet('Atribuiri Active')
    headers2 = ['Nr. Masina', 'Marca Model', 'Angajat', 'Data Atribuire', 'KM Preluare', 'Proiect', 'Motiv']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    atribuiri = AtribuireMasina.query.filter_by(data_returnare=None).order_by(AtribuireMasina.data_atribuire.desc()).all()
    for row, a in enumerate(atribuiri, 2):
        ws2.cell(row=row, column=1, value=a.masina.numar_inmatriculare)
        ws2.cell(row=row, column=2, value=f'{a.masina.marca} {a.masina.model}')
        ws2.cell(row=row, column=3, value=a.angajat.nume_complet if a.angajat else '')
        ws2.cell(row=row, column=4, value=a.data_atribuire.strftime('%d.%m.%Y') if a.data_atribuire else '')
        ws2.cell(row=row, column=5, value=a.km_preluare)
        ws2.cell(row=row, column=6, value=a.proiect.cod_proiect if a.proiect else '')
        ws2.cell(row=row, column=7, value=a.motiv or '')

    # Sheet 3: Defectiuni nerezolvate
    ws3 = wb.create_sheet('Defectiuni Active')
    headers3 = ['Nr. Masina', 'Data Raportare', 'Descriere', 'Gravitate', 'Status', 'Raportat de']
    for col, h in enumerate(headers3, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    defectiuni = DefectiuneMasina.query.filter(
        DefectiuneMasina.status.in_(['raportata', 'in_lucru'])
    ).order_by(DefectiuneMasina.data_raportare.desc()).all()
    for row, d in enumerate(defectiuni, 2):
        ws3.cell(row=row, column=1, value=d.masina.numar_inmatriculare)
        ws3.cell(row=row, column=2, value=d.data_raportare.strftime('%d.%m.%Y') if d.data_raportare else '')
        ws3.cell(row=row, column=3, value=d.descriere[:200] if d.descriere else '')
        ws3.cell(row=row, column=4, value=d.gravitate)
        ws3.cell(row=row, column=5, value=d.status)
        ws3.cell(row=row, column=6, value=d.angajat_raportor.nume_complet if d.angajat_raportor else '')

    # Sheet 4: Alerte documente
    ws4 = wb.create_sheet('Alerte Documente')
    headers4 = ['Nr. Masina', 'Document', 'Data Expirare', 'Zile Ramase', 'Status']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    alert_row = 2
    today = date.today()
    for m in Masina.query.all():
        for alerta in m.alerte_documente:
            ws4.cell(row=alert_row, column=1, value=m.numar_inmatriculare)
            ws4.cell(row=alert_row, column=2, value=alerta['doc'])
            exp_date_map = {'ITP': m.data_itp_expirare, 'RCA': m.data_rca_expirare,
                           'CASCO': m.data_casco_expirare, 'Rovinieta': m.data_rovinieta_expirare}
            exp = exp_date_map.get(alerta['doc'])
            ws4.cell(row=alert_row, column=3, value=exp.strftime('%d.%m.%Y') if exp else '')
            ws4.cell(row=alert_row, column=4, value=alerta['zile'])
            ws4.cell(row=alert_row, column=5, value=alerta['tip'].replace('_', ' ').title())
            alert_row += 1

    # Save
    filepath = os.path.join(tempfile.gettempdir(), f'flota_auto_{today.strftime("%Y%m%d")}.xlsx')
    wb.save(filepath)
    return send_file(filepath, as_attachment=True, download_name=f'flota_auto_{today.strftime("%Y%m%d")}.xlsx')


# ============================================================
# CAUTARE AJAX
# ============================================================

@masini_bp.route('/cauta')
@login_required
def cauta():
    q = request.args.get('q', '').strip()
    if len(q) < 2:
        return jsonify([])

    like = f'%{q}%'
    masini = Masina.query.filter(
        db.or_(
            Masina.numar_inmatriculare.ilike(like),
            Masina.marca.ilike(like),
            Masina.model.ilike(like),
        )
    ).limit(10).all()

    return jsonify([{
        'id': m.id,
        'numar': m.numar_inmatriculare,
        'denumire': m.denumire_completa,
        'status': m.status,
        'tip': m.tip_vehicul,
    } for m in masini])


# ============================================================
# API: Conduceri pe luna (JSON for chart)
# ============================================================

@masini_bp.route('/<int:id>/conduceri-json')
@login_required
def conduceri_json(id):
    masina = Masina.query.get_or_404(id)
    luna = request.args.get('luna', date.today().month, type=int)
    an = request.args.get('an', date.today().year, type=int)

    conduceri = masina.conduceri.filter(
        db.extract('month', ConducereMasina.data) == luna,
        db.extract('year', ConducereMasina.data) == an,
    ).order_by(ConducereMasina.data).all()

    return jsonify({
        'conduceri': [{
            'id': c.id,
            'data': c.data.strftime('%d.%m.%Y'),
            'angajat': c.angajat.nume_complet if c.angajat else '',
            'km_start': c.km_start,
            'km_sfarsit': c.km_sfarsit,
            'km_parcursi': c.km_parcursi,
            'ruta': c.ruta or '',
            'scop': c.scop or '',
            'combustibil': float(c.combustibil_alimentat or 0),
            'cost': float(c.cost_combustibil or 0),
        } for c in conduceri],
        'total_km': sum(c.km_parcursi for c in conduceri if c.km_parcursi),
        'total_combustibil': sum(float(c.combustibil_alimentat or 0) for c in conduceri),
    })


# ============================================================
# ALERTE DOCUMENTE (pagina dedicata)
# ============================================================

@masini_bp.route('/alerte')
@login_required
def alerte():
    masini = Masina.query.filter(
        Masina.status.notin_(['casata', 'vanduta'])
    ).all()

    alerte_list = []
    for m in masini:
        for a in m.alerte_documente:
            alerte_list.append({
                'masina': m,
                'document': a['doc'],
                'zile': a['zile'],
                'tip': a['tip'],
            })

    # Sort by urgency (most expired first)
    alerte_list.sort(key=lambda x: x['zile'])

    return render_template('masini/alerte.html', alerte=alerte_list)


# ============================================================
# STERGE DOCUMENT MASINA
# ============================================================

@masini_bp.route('/document/<int:doc_id>/sterge', methods=['POST'])
@login_required
@manager_or_admin
def sterge_document(doc_id):
    doc = DocumentMasina.query.get_or_404(doc_id)
    masina_id = doc.masina_id
    db.session.delete(doc)
    db.session.commit()
    flash('Document sters cu succes.', 'success')
    return redirect(url_for('masini.fisa', id=masina_id))
