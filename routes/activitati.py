"""
EDIFICO WORKFORCE - Modul Rapoarte Activitate Zilnica
Blueprint: /activitati
"""

import json
import os
from datetime import datetime, date, timedelta
from decimal import Decimal
from functools import wraps
from io import BytesIO

from flask import (
    Blueprint, render_template, request, redirect, url_for,
    flash, jsonify, send_file, current_app
)
from flask_login import login_required, current_user

from models import (
    db, Angajat, Proiect, Pontaj, TipInstalatie,
    AngajatProiect, SarbatoareLegala,
    RaportActivitate, CategorieActivitate,
    Santier, Cladire, ElementBIM, Spatiu, Zona,
)

activitati_bp = Blueprint('activitati', __name__, url_prefix='/activitati')


# ============================================================
# DECORATORI
# ============================================================

def manager_or_admin(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or current_user.rol not in ('admin', 'manager'):
            flash('Acces permis doar managerilor si administratorilor.', 'danger')
            return redirect(url_for('activitati.panou'))
        return f(*args, **kwargs)
    return decorated


# ============================================================
# HELPERI
# ============================================================

def _get_zile_lucratoare(an, luna):
    """Returneaza numarul de zile lucratoare din luna."""
    import calendar
    sarbatori = set()
    sarb_query = SarbatoareLegala.query.filter_by(an=an).all()
    for s in sarb_query:
        if s.data.month == luna:
            sarbatori.add(s.data.day)

    cal = calendar.monthcalendar(an, luna)
    zile = 0
    for week in cal:
        for i in range(5):  # Luni-Vineri
            day = week[i]
            if day != 0 and day not in sarbatori:
                zile += 1
    return zile


def _get_saptamana_bounds(an, saptamana):
    """Returneaza (luni, duminica) pentru saptamana ISO data."""
    from datetime import date as d_date
    jan4 = d_date(an, 1, 4)
    start_year = jan4 - timedelta(days=jan4.isoweekday() - 1)
    luni = start_year + timedelta(weeks=saptamana - 1)
    duminica = luni + timedelta(days=6)
    return luni, duminica


def _get_angajat_for_user(user):
    """Incearca sa gaseasca angajatul asociat utilizatorului logat (dupa email)."""
    if user and user.email:
        return Angajat.query.filter_by(email=user.email, status='activ').first()
    return None


# ============================================================
# RUTA PRINCIPALA - PANOU ACTIVITATI
# ============================================================

@activitati_bp.route('/')
@login_required
def panou():
    """Panou principal activitati."""
    today = date.today()
    angajat_curent = _get_angajat_for_user(current_user)

    # Date personale
    activitate_azi = None
    activitati_saptamana = []
    activitati_luna = []
    zile_cu_activitate_sapt = 0
    zile_cu_activitate_luna = 0
    zile_lucratoare_luna = _get_zile_lucratoare(today.year, today.month)
    proiecte_saptamana = set()
    activitati_aprobate_luna = 0
    activitati_pending_luna = 0

    if angajat_curent:
        # Activitate azi
        activitate_azi = RaportActivitate.query.filter_by(
            angajat_id=angajat_curent.id, data=today
        ).first()

        # Saptamana curenta
        luni = today - timedelta(days=today.weekday())
        duminica = luni + timedelta(days=6)
        activitati_saptamana = RaportActivitate.query.filter(
            RaportActivitate.angajat_id == angajat_curent.id,
            RaportActivitate.data >= luni,
            RaportActivitate.data <= duminica
        ).all()
        zile_cu_activitate_sapt = len(set(a.data for a in activitati_saptamana))
        proiecte_saptamana = set(a.proiect_id for a in activitati_saptamana)

        # Luna curenta
        prima_zi = today.replace(day=1)
        activitati_luna = RaportActivitate.query.filter(
            RaportActivitate.angajat_id == angajat_curent.id,
            RaportActivitate.data >= prima_zi,
            RaportActivitate.data <= today
        ).all()
        zile_cu_activitate_luna = len(set(a.data for a in activitati_luna))
        activitati_aprobate_luna = sum(1 for a in activitati_luna if a.status == 'aprobat')
        activitati_pending_luna = sum(1 for a in activitati_luna if a.status in ('draft', 'trimis'))

    # Tabel activitati recente (manageri vad tot, operatorii doar ale lor)
    query = RaportActivitate.query

    # Filtre
    f_angajat = request.args.get('angajat_id', '', type=int) or None
    f_proiect = request.args.get('proiect_id', '', type=int) or None
    f_instalatie = request.args.get('tip_instalatie_id', '', type=int) or None
    f_status = request.args.get('status', '')
    f_tip = request.args.get('tip', '')  # tip_activitate: zilnica/saptamanala/lunara
    f_status_executie = request.args.get('status_executie', '')
    f_data_start = request.args.get('data_start', '')
    f_data_end = request.args.get('data_end', '')
    # Filtre BIM
    f_santier = request.args.get('santier_id', type=int) or None
    f_cladire = request.args.get('cladire_id', type=int) or None
    f_element_bim = request.args.get('element_bim_id', type=int) or None
    f_tip_element = request.args.get('tip_element', '').strip()

    if current_user.rol == 'operator' and angajat_curent:
        query = query.filter_by(angajat_id=angajat_curent.id)
    elif current_user.rol == 'operator':
        query = query.filter_by(id=-1)  # no results

    if f_angajat:
        query = query.filter_by(angajat_id=f_angajat)
    if f_proiect:
        query = query.filter_by(proiect_id=f_proiect)
    if f_instalatie:
        query = query.filter_by(tip_instalatie_id=f_instalatie)
    if f_status:
        query = query.filter_by(status=f_status)
    if f_tip in ('zilnica', 'saptamanala', 'lunara'):
        query = query.filter_by(tip_activitate=f_tip)
    if f_status_executie in ('planificata', 'in_desfasurare', 'finalizata'):
        query = query.filter_by(status_executie=f_status_executie)
    if f_data_start:
        try:
            ds = datetime.strptime(f_data_start, '%Y-%m-%d').date()
            query = query.filter(RaportActivitate.data >= ds)
        except ValueError:
            pass
    if f_data_end:
        try:
            de = datetime.strptime(f_data_end, '%Y-%m-%d').date()
            query = query.filter(RaportActivitate.data <= de)
        except ValueError:
            pass

    # Filtre BIM (un singur JOIN cu ElementBIM ca sa evit dublarea)
    if f_element_bim:
        query = query.filter_by(element_bim_id=f_element_bim)
    if f_tip_element or f_cladire or f_santier:
        query = query.join(ElementBIM, RaportActivitate.element_bim_id == ElementBIM.id)
        if f_tip_element:
            query = query.filter(ElementBIM.tip_element == f_tip_element)
        if f_cladire:
            query = query.filter(ElementBIM.cladire_id == f_cladire)
        if f_santier:
            query = query.join(Cladire, ElementBIM.cladire_id == Cladire.id) \
                         .filter(Cladire.santier_id == f_santier)

    activitati_recente = query.order_by(RaportActivitate.data.desc(), RaportActivitate.introdus_la.desc()).limit(50).all()

    # Aprobare count (pentru manageri)
    pending_aprobare = 0
    if current_user.rol in ('admin', 'manager'):
        pending_aprobare = RaportActivitate.query.filter_by(status='trimis').count()

    # Dropdown-uri filtre
    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    proiecte = Proiect.query.filter(Proiect.status.in_(['activ', 'planificat'])).order_by(Proiect.cod_proiect).all()
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()
    bim_santiere = Santier.query.order_by(Santier.cod).all()
    bim_cladiri = Cladire.query.order_by(Cladire.cod).all()
    bim_tipuri_element = ElementBIM.TIPURI

    return render_template('activitati/panou.html',
        today=today,
        angajat_curent=angajat_curent,
        activitate_azi=activitate_azi,
        activitati_saptamana=activitati_saptamana,
        zile_cu_activitate_sapt=zile_cu_activitate_sapt,
        proiecte_saptamana=proiecte_saptamana,
        activitati_luna=activitati_luna,
        zile_cu_activitate_luna=zile_cu_activitate_luna,
        zile_lucratoare_luna=zile_lucratoare_luna,
        activitati_aprobate_luna=activitati_aprobate_luna,
        activitati_pending_luna=activitati_pending_luna,
        activitati_recente=activitati_recente,
        pending_aprobare=pending_aprobare,
        angajati=angajati,
        proiecte=proiecte,
        instalatii=instalatii,
        f_angajat=f_angajat,
        f_proiect=f_proiect,
        f_instalatie=f_instalatie,
        f_status=f_status,
        f_tip=f_tip,
        f_status_executie=f_status_executie,
        f_data_start=f_data_start,
        f_data_end=f_data_end,
        bim_santiere=bim_santiere,
        bim_cladiri=bim_cladiri,
        bim_tipuri_element=bim_tipuri_element,
        f_santier=f_santier,
        f_cladire=f_cladire,
        f_element_bim=f_element_bim,
        f_tip_element=f_tip_element,
    )


# ============================================================
# ADAUGA ACTIVITATE
# ============================================================

@activitati_bp.route('/adauga', methods=['GET', 'POST'])
@login_required
def adauga():
    """Formular adaugare activitate noua."""
    if request.method == 'POST':
        return _salveaza_activitate(None)

    today = date.today()
    angajat_curent = _get_angajat_for_user(current_user)

    # Pre-completare din pontaj
    pontaj_azi = None
    if angajat_curent:
        pontaj_azi = Pontaj.query.filter_by(angajat_id=angajat_curent.id, data=today).first()

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    proiecte = Proiect.query.filter(Proiect.status.in_(['activ', 'planificat'])).order_by(Proiect.cod_proiect).all()
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()
    categorii = CategorieActivitate.query.filter_by(activa=True).order_by(CategorieActivitate.ordine).all()
    santiere = Santier.query.order_by(Santier.cod).all()

    return render_template('activitati/formular.html',
        activitate=None,
        today=today,
        angajat_curent=angajat_curent,
        pontaj_azi=pontaj_azi,
        angajati=angajati,
        proiecte=proiecte,
        instalatii=instalatii,
        categorii=categorii,
        santiere=santiere,
    )


@activitati_bp.route('/adauga-rapida', methods=['GET', 'POST'])
@login_required
def adauga_rapida():
    """Formular simplificat (modal AJAX)."""
    if request.method == 'POST':
        return _salveaza_activitate(None, rapida=True)

    angajat_curent = _get_angajat_for_user(current_user)
    proiecte = Proiect.query.filter(Proiect.status.in_(['activ', 'planificat'])).order_by(Proiect.cod_proiect).all()
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()

    return render_template('activitati/formular_rapid.html',
        angajat_curent=angajat_curent,
        proiecte=proiecte,
        instalatii=instalatii,
    )


# ============================================================
# DETALIU + EDITARE + ACTIUNI
# ============================================================

@activitati_bp.route('/<int:id>')
@login_required
def detaliu(id):
    """Detaliu raport activitate."""
    activitate = RaportActivitate.query.get_or_404(id)
    angajat_curent = _get_angajat_for_user(current_user)

    # Verificare acces
    if current_user.rol == 'operator' and angajat_curent:
        if activitate.angajat_id != angajat_curent.id:
            flash('Nu aveti acces la aceasta activitate.', 'danger')
            return redirect(url_for('activitati.panou'))

    # Pontaj din aceeasi zi
    pontaj_zi = Pontaj.query.filter_by(
        angajat_id=activitate.angajat_id, data=activitate.data
    ).first()

    return render_template('activitati/detaliu.html',
        activitate=activitate,
        pontaj_zi=pontaj_zi,
    )


@activitati_bp.route('/<int:id>/editeaza', methods=['GET', 'POST'])
@login_required
def editeaza(id):
    """Editare activitate existenta."""
    activitate = RaportActivitate.query.get_or_404(id)

    # Doar draft-urile pot fi editate de autori
    angajat_curent = _get_angajat_for_user(current_user)
    if current_user.rol == 'operator':
        if not angajat_curent or activitate.angajat_id != angajat_curent.id:
            flash('Nu aveti acces la aceasta activitate.', 'danger')
            return redirect(url_for('activitati.panou'))
        if activitate.status not in ('draft', 'respins'):
            flash('Activitatea nu mai poate fi editata.', 'warning')
            return redirect(url_for('activitati.detaliu', id=id))

    if request.method == 'POST':
        return _salveaza_activitate(activitate)

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    proiecte = Proiect.query.filter(Proiect.status.in_(['activ', 'planificat'])).order_by(Proiect.cod_proiect).all()
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()
    categorii = CategorieActivitate.query.filter_by(activa=True).order_by(CategorieActivitate.ordine).all()
    santiere = Santier.query.order_by(Santier.cod).all()

    return render_template('activitati/formular.html',
        activitate=activitate,
        today=date.today(),
        angajat_curent=angajat_curent,
        pontaj_azi=None,
        angajati=angajati,
        proiecte=proiecte,
        instalatii=instalatii,
        categorii=categorii,
        santiere=santiere,
    )


def _salveaza_activitate(activitate, rapida=False):
    """Logica comuna salvare activitate (add/edit)."""
    try:
        angajat_id = request.form.get('angajat_id', type=int)
        # Multi-proiect: accepta atat 'proiect_ids[]' (multi) cat si 'proiect_id' (legacy)
        proiect_ids_raw = request.form.getlist('proiect_ids[]') or request.form.getlist('proiect_ids')
        proiecte_ids_clean = []
        for v in proiect_ids_raw:
            try:
                n = int(v)
                if n > 0 and n not in proiecte_ids_clean:
                    proiecte_ids_clean.append(n)
            except (ValueError, TypeError):
                continue
        if not proiecte_ids_clean:
            # Fallback la single proiect_id (forma veche)
            single = request.form.get('proiect_id', type=int)
            if single:
                proiecte_ids_clean = [single]
        proiect_id = proiecte_ids_clean[0] if proiecte_ids_clean else None
        proiecte_json = json.dumps(proiecte_ids_clean) if proiecte_ids_clean else None
        data_str = request.form.get('data', '')
        data_sfarsit_str = request.form.get('data_sfarsit', '').strip()
        tip_activitate = request.form.get('tip_activitate', 'zilnica').strip() or 'zilnica'
        if tip_activitate not in ('zilnica', 'saptamanala', 'lunara'):
            tip_activitate = 'zilnica'
        supervisor_id = request.form.get('supervisor_id', type=int) or None
        subordonati_raw = request.form.getlist('subordonati_ids[]') or request.form.getlist('subordonati_ids')
        ore_lucrate_str = request.form.get('ore_lucrate', '').strip()
        status_executie = request.form.get('status_executie', 'planificata').strip() or 'planificata'
        if status_executie not in ('planificata', 'in_desfasurare', 'finalizata'):
            status_executie = 'planificata'
        tip_instalatie_id = request.form.get('tip_instalatie_id', type=int) or None
        categorie_id = request.form.get('categorie_activitate_id', type=int) or None
        zona_lucru = request.form.get('zona_lucru', '').strip()
        activitate_principala = request.form.get('activitate_principala', '').strip()
        activitate_detaliata = request.form.get('activitate_detaliata', '').strip()
        cantitate = request.form.get('cantitate_executata', '')
        um = request.form.get('unitate_masura', '').strip()
        procent = request.form.get('procent_realizare', type=int)
        probleme = request.form.get('probleme_intampinate', '').strip()
        solutii = request.form.get('solutii_aplicate', '').strip()
        observatii = request.form.get('observatii', '').strip()
        necesita_aprobare = bool(request.form.get('necesita_aprobare_tehnica'))
        include_sambata = bool(request.form.get('include_sambata'))
        include_duminica = bool(request.form.get('include_duminica'))
        actiune = request.form.get('actiune', 'draft')  # draft / trimite / alta

        # BIM context (toate optionale)
        bim_element_id = request.form.get('bim_element_id', type=int) or None
        bim_spatiu_id = request.form.get('bim_spatiu_id', type=int) or None
        # Zona se ia din spatiu daca exista, altfel din formular
        bim_zona_id = request.form.get('bim_zona_id', type=int) or None
        if not bim_zona_id and bim_spatiu_id:
            sp = Spatiu.query.get(bim_spatiu_id)
            if sp and sp.zona_id:
                bim_zona_id = sp.zona_id

        # Detalii pe zi (pentru saptamanala/lunara): liste paralele
        det_data_list = request.form.getlist('detaliu_data[]')
        det_proiect_list = request.form.getlist('detaliu_proiect[]')
        det_text_list = request.form.getlist('detaliu_text[]')
        det_ore_list = request.form.getlist('detaliu_ore[]')

        # Validare
        if not angajat_id or not proiect_id:
            flash('Angajatul si cel putin un proiect sunt obligatorii.', 'danger')
            return redirect(request.url)
        if not activitate_principala:
            flash('Titlul activitatii este obligatoriu.', 'danger')
            return redirect(request.url)
        if not data_str:
            flash('Data de inceput este obligatorie.', 'danger')
            return redirect(request.url)

        data_val = datetime.strptime(data_str, '%Y-%m-%d').date()
        data_sfarsit_val = None
        if data_sfarsit_str:
            try:
                data_sfarsit_val = datetime.strptime(data_sfarsit_str, '%Y-%m-%d').date()
            except ValueError:
                data_sfarsit_val = None

        # subordonati_ids - parseaza si curata
        subordonati_ids_clean = []
        for raw_val in subordonati_raw:
            try:
                v = int(raw_val)
                if v != angajat_id and v not in subordonati_ids_clean:
                    subordonati_ids_clean.append(v)
            except (ValueError, TypeError):
                continue
        subordonati_json = json.dumps(subordonati_ids_clean) if subordonati_ids_clean else None

        # ore lucrate
        try:
            ore_lucrate_val = Decimal(ore_lucrate_str) if ore_lucrate_str else None
        except (ValueError, TypeError):
            ore_lucrate_val = None

        # Materiale JSON
        materiale_json = '[]'
        mat_denumiri = request.form.getlist('mat_denumire[]')
        mat_cantitati = request.form.getlist('mat_cantitate[]')
        mat_um = request.form.getlist('mat_um[]')
        if mat_denumiri:
            materiale = []
            for i in range(len(mat_denumiri)):
                if mat_denumiri[i].strip():
                    materiale.append({
                        'denumire': mat_denumiri[i].strip(),
                        'cantitate': mat_cantitati[i].strip() if i < len(mat_cantitati) else '',
                        'um': mat_um[i].strip() if i < len(mat_um) else ''
                    })
            materiale_json = json.dumps(materiale, ensure_ascii=False)

        # Construire JSON detalii pe zi (doar randuri cu macar un camp completat)
        detalii_json = None
        if tip_activitate in ('saptamanala', 'lunara') and det_data_list:
            detalii_curate = []
            for i, d_str in enumerate(det_data_list):
                d_str = (d_str or '').strip()
                if not d_str:
                    continue
                proi_str = det_proiect_list[i] if i < len(det_proiect_list) else ''
                text = (det_text_list[i] if i < len(det_text_list) else '').strip()
                ore_str = (det_ore_list[i] if i < len(det_ore_list) else '').strip()
                # Sare peste randuri complet goale (dar pastreaza data ca placeholder)
                if not text and not ore_str and not proi_str:
                    continue
                item = {'data': d_str}
                if proi_str:
                    try:
                        item['proiect_id'] = int(proi_str)
                    except (ValueError, TypeError):
                        pass
                if text:
                    item['text'] = text[:500]
                if ore_str:
                    try:
                        item['ore'] = float(ore_str)
                    except (ValueError, TypeError):
                        pass
                detalii_curate.append(item)
            if detalii_curate:
                detalii_json = json.dumps(detalii_curate, ensure_ascii=False)

        echipamente = request.form.get('echipamente_folosite', '').strip()

        if activitate is None:
            activitate = RaportActivitate()
            db.session.add(activitate)

        activitate.angajat_id = angajat_id
        activitate.proiect_id = proiect_id
        activitate.proiecte_ids = proiecte_json
        activitate.element_bim_id = bim_element_id
        activitate.spatiu_id = bim_spatiu_id
        activitate.zona_id = bim_zona_id
        activitate.data = data_val
        activitate.data_sfarsit = data_sfarsit_val
        activitate.tip_activitate = tip_activitate
        activitate.supervisor_id = supervisor_id if supervisor_id != angajat_id else None
        activitate.subordonati_ids = subordonati_json
        activitate.ore_lucrate = ore_lucrate_val
        activitate.status_executie = status_executie
        activitate.tip_instalatie_id = tip_instalatie_id
        activitate.categorie_activitate_id = categorie_id
        activitate.zona_lucru = zona_lucru
        activitate.activitate_principala = activitate_principala[:500]
        activitate.activitate_detaliata = activitate_detaliata[:2000] if activitate_detaliata else None
        activitate.materiale_folosite = materiale_json
        activitate.echipamente_folosite = echipamente or None
        activitate.cantitate_executata = Decimal(cantitate) if cantitate else None
        activitate.unitate_masura = um or None
        activitate.procent_realizare = min(100, max(0, procent)) if procent is not None else None
        activitate.probleme_intampinate = probleme or None
        activitate.solutii_aplicate = solutii or None
        activitate.observatii = observatii or None
        activitate.necesita_aprobare_tehnica = necesita_aprobare
        activitate.include_sambata = include_sambata
        activitate.include_duminica = include_duminica
        activitate.detalii_pe_zi = detalii_json

        # Auto-completare numar_saptamana / luna_an din tip si data inceput
        activitate.calculeaza_perioada()

        if actiune == 'trimite':
            activitate.status = 'trimis'
        elif activitate.status == 'respins':
            activitate.status = 'draft'

        db.session.commit()

        if actiune == 'trimite':
            flash('Activitatea a fost trimisa spre aprobare.', 'success')
        else:
            flash('Activitatea a fost salvata ca draft.', 'success')

        if actiune == 'alta':
            return redirect(url_for('activitati.adauga'))

        if rapida:
            return jsonify({'success': True, 'id': activitate.id})

        return redirect(url_for('activitati.detaliu', id=activitate.id))

    except Exception as e:
        db.session.rollback()
        flash(f'Eroare la salvare: {str(e)}', 'danger')
        return redirect(request.url)


@activitati_bp.route('/<int:id>/trimite', methods=['POST'])
@login_required
def trimite(id):
    """Schimba status draft -> trimis."""
    activitate = RaportActivitate.query.get_or_404(id)
    if activitate.status != 'draft':
        flash('Doar activitatile draft pot fi trimise.', 'warning')
    else:
        activitate.status = 'trimis'
        db.session.commit()
        flash('Activitatea a fost trimisa spre aprobare.', 'success')
    return redirect(url_for('activitati.detaliu', id=id))


@activitati_bp.route('/<int:id>/aproba', methods=['POST'])
@login_required
@manager_or_admin
def aproba(id):
    """Aprobare activitate (manager/admin)."""
    activitate = RaportActivitate.query.get_or_404(id)
    activitate.status = 'aprobat'
    activitate.aprobat_de_id = current_user.id
    activitate.data_aprobare = datetime.utcnow()
    db.session.commit()
    flash('Activitatea a fost aprobata.', 'success')
    return redirect(url_for('activitati.detaliu', id=id))


@activitati_bp.route('/<int:id>/respinge', methods=['POST'])
@login_required
@manager_or_admin
def respinge(id):
    """Respingere activitate cu motiv."""
    activitate = RaportActivitate.query.get_or_404(id)
    motiv = request.form.get('motiv_respingere', '').strip()
    activitate.status = 'respins'
    activitate.motiv_respingere = motiv or 'Fara motiv specificat'
    db.session.commit()
    flash('Activitatea a fost respinsa.', 'warning')
    return redirect(url_for('activitati.detaliu', id=id))


@activitati_bp.route('/<int:id>/sterge', methods=['POST'])
@login_required
def sterge(id):
    """Sterge activitate (doar draft)."""
    activitate = RaportActivitate.query.get_or_404(id)
    angajat_curent = _get_angajat_for_user(current_user)

    if current_user.rol == 'operator':
        if not angajat_curent or activitate.angajat_id != angajat_curent.id:
            flash('Nu aveti acces.', 'danger')
            return redirect(url_for('activitati.panou'))
    if activitate.status not in ('draft', 'respins') and current_user.rol != 'admin':
        flash('Doar activitatile draft/respinse pot fi sterse.', 'warning')
        return redirect(url_for('activitati.detaliu', id=id))

    db.session.delete(activitate)
    db.session.commit()
    flash('Activitatea a fost stearsa.', 'info')
    return redirect(url_for('activitati.panou'))


# ============================================================
# ACTIVITATILE MELE
# ============================================================

@activitati_bp.route('/ale-mele')
@login_required
def ale_mele():
    """Activitatile angajatului curent logat."""
    angajat_curent = _get_angajat_for_user(current_user)
    if not angajat_curent:
        flash('Nu sunteti asociat unui angajat.', 'warning')
        return redirect(url_for('activitati.panou'))

    luna = request.args.get('luna', date.today().month, type=int)
    an = request.args.get('an', date.today().year, type=int)

    activitati = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_curent.id,
        db.extract('month', RaportActivitate.data) == luna,
        db.extract('year', RaportActivitate.data) == an,
    ).order_by(RaportActivitate.data.desc()).all()

    return render_template('activitati/ale_mele.html',
        angajat=angajat_curent,
        activitati=activitati,
        luna=luna,
        an=an,
    )


@activitati_bp.route('/ale-mele/saptamana')
@login_required
def ale_mele_saptamana():
    """View saptamanal personal."""
    angajat_curent = _get_angajat_for_user(current_user)
    if not angajat_curent:
        flash('Nu sunteti asociat unui angajat.', 'warning')
        return redirect(url_for('activitati.panou'))

    today = date.today()
    an = request.args.get('an', today.year, type=int)
    sapt = request.args.get('saptamana', today.isocalendar()[1], type=int)
    luni, duminica = _get_saptamana_bounds(an, sapt)

    activitati = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_curent.id,
        RaportActivitate.data >= luni,
        RaportActivitate.data <= duminica,
    ).order_by(RaportActivitate.data).all()

    # Group by day
    zile = {}
    for i in range(7):
        zi = luni + timedelta(days=i)
        zile[zi] = [a for a in activitati if a.data == zi]

    return render_template('activitati/saptamana.html',
        angajat=angajat_curent,
        an=an, saptamana=sapt,
        luni=luni, duminica=duminica,
        zile=zile,
        activitati=activitati,
    )


@activitati_bp.route('/ale-mele/luna')
@login_required
def ale_mele_luna():
    """View lunar personal."""
    return redirect(url_for('activitati.ale_mele',
                            luna=request.args.get('luna', date.today().month),
                            an=request.args.get('an', date.today().year)))


# ============================================================
# APROBARE
# ============================================================

@activitati_bp.route('/aprobare')
@login_required
@manager_or_admin
def aprobare():
    """Lista activitati in asteptare de aprobare."""
    activitati = RaportActivitate.query.filter_by(status='trimis').order_by(
        RaportActivitate.data.desc()
    ).all()

    return render_template('activitati/aprobare.html',
        activitati=activitati,
    )


@activitati_bp.route('/aprobare/masa', methods=['POST'])
@login_required
@manager_or_admin
def aprobare_masa():
    """Aprobare in masa."""
    ids = request.form.getlist('activitate_ids[]')
    actiune = request.form.get('actiune', 'aproba')  # aproba / respinge
    motiv = request.form.get('motiv_respingere', '').strip()

    count = 0
    for aid in ids:
        try:
            a = RaportActivitate.query.get(int(aid))
            if a and a.status == 'trimis':
                if actiune == 'aproba':
                    a.status = 'aprobat'
                    a.aprobat_de_id = current_user.id
                    a.data_aprobare = datetime.utcnow()
                else:
                    a.status = 'respins'
                    a.motiv_respingere = motiv or 'Respins in masa'
                count += 1
        except (ValueError, TypeError):
            continue

    db.session.commit()

    if actiune == 'aproba':
        flash(f'{count} activitati aprobate.', 'success')
    else:
        flash(f'{count} activitati respinse.', 'warning')
    return redirect(url_for('activitati.aprobare'))


# ============================================================
# CALENDAR
# ============================================================

@activitati_bp.route('/calendar')
@login_required
def calendar_view():
    """Calendar activitati."""
    today = date.today()
    luna = request.args.get('luna', today.month, type=int)
    an = request.args.get('an', today.year, type=int)
    f_angajat = request.args.get('angajat_id', type=int) or None
    f_proiect = request.args.get('proiect_id', type=int) or None
    view_type = request.args.get('view', 'lunar')  # lunar / saptamanal

    angajati = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()
    proiecte = Proiect.query.filter(Proiect.status.in_(['activ', 'planificat'])).order_by(Proiect.cod_proiect).all()
    instalatii = TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()

    return render_template('activitati/calendar.html',
        today=today,
        luna=luna, an=an,
        f_angajat=f_angajat,
        f_proiect=f_proiect,
        view_type=view_type,
        angajati=angajati,
        proiecte=proiecte,
        instalatii=instalatii,
    )


@activitati_bp.route('/api/calendar-data')
@login_required
def api_calendar_data():
    """Date JSON pentru calendar."""
    an = request.args.get('an', date.today().year, type=int)
    luna = request.args.get('luna', date.today().month, type=int)
    f_angajat = request.args.get('angajat_id', type=int) or None
    f_proiect = request.args.get('proiect_id', type=int) or None

    import calendar
    _, last_day = calendar.monthrange(an, luna)
    prima_zi = date(an, luna, 1)
    ultima_zi = date(an, luna, last_day)

    query = RaportActivitate.query.filter(
        RaportActivitate.data >= prima_zi,
        RaportActivitate.data <= ultima_zi,
    )

    if f_angajat:
        query = query.filter_by(angajat_id=f_angajat)
    if f_proiect:
        query = query.filter_by(proiect_id=f_proiect)

    # Operators see only their own
    if current_user.rol == 'operator':
        angajat_curent = _get_angajat_for_user(current_user)
        if angajat_curent:
            query = query.filter_by(angajat_id=angajat_curent.id)
        else:
            query = query.filter_by(id=-1)

    activitati = query.order_by(RaportActivitate.data).all()

    # Sarbatori
    sarbatori = set()
    for s in SarbatoareLegala.query.filter_by(an=an).all():
        if s.data.month == luna:
            sarbatori.add(s.data.day)

    # Group by day
    days = {}
    for a in activitati:
        day_str = a.data.isoformat()
        if day_str not in days:
            days[day_str] = []
        days[day_str].append({
            'id': a.id,
            'angajat': a.angajat.nume_complet if a.angajat else '?',
            'proiect': a.proiect.cod_proiect if a.proiect else '?',
            'activitate': a.activitate_scurta,
            'status': a.status,
            'instalatie_cod': a.tip_instalatie.cod if a.tip_instalatie else None,
            'instalatie_culoare': a.tip_instalatie.culoare_hex if a.tip_instalatie else '#546E7A',
        })

    return jsonify({
        'an': an,
        'luna': luna,
        'zile': days,
        'sarbatori': list(sarbatori),
    })


@activitati_bp.route('/api/categorii/<int:tip_instalatie_id>')
@login_required
def api_categorii(tip_instalatie_id):
    """Categorii activitate filtrate pe tip instalatie (AJAX)."""
    categorii = CategorieActivitate.query.filter(
        db.or_(
            CategorieActivitate.tip_instalatie_id == tip_instalatie_id,
            CategorieActivitate.tip_instalatie_id.is_(None)
        ),
        CategorieActivitate.activa == True
    ).order_by(CategorieActivitate.ordine).all()

    return jsonify([{
        'id': c.id,
        'denumire': c.denumire,
        'um_default': c.unitate_masura_default,
        'universal': c.tip_instalatie_id is None,
    } for c in categorii])


# ============================================================
# EXPORTURI EXCEL
# ============================================================

@activitati_bp.route('/raport/saptamanal')
@login_required
def raport_saptamanal():
    """Generator export saptamanal."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    today = date.today()
    an = request.args.get('an', today.year, type=int)
    sapt = request.args.get('saptamana', today.isocalendar()[1], type=int)
    luni, duminica = _get_saptamana_bounds(an, sapt)

    activitati = RaportActivitate.query.filter(
        RaportActivitate.data >= luni,
        RaportActivitate.data <= duminica,
    ).order_by(RaportActivitate.angajat_id, RaportActivitate.data).all()

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')

    # Sheet 1: Rezumat Saptamana
    ws = wb.active
    ws.title = 'Rezumat Saptamana'

    ws.merge_cells('A1:H1')
    ws['A1'] = f'Raport Activitati - Saptamana {sapt} ({luni.strftime("%d.%m.%Y")} - {duminica.strftime("%d.%m.%Y")})'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1A237E')

    # Headers
    zile_labels = ['Luni', 'Marti', 'Miercuri', 'Joi', 'Vineri', 'Sambata', 'Duminica']
    headers = ['Angajat'] + [f'{zl}\n{(luni + timedelta(days=i)).strftime("%d.%m")}' for i, zl in enumerate(zile_labels)]
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    # Group by angajat
    from collections import defaultdict
    angajat_activitati = defaultdict(lambda: defaultdict(list))
    for a in activitati:
        angajat_activitati[a.angajat_id][a.data.weekday()].append(a)

    row = 4
    for ang_id, zile_dict in angajat_activitati.items():
        ang = Angajat.query.get(ang_id)
        ws.cell(row=row, column=1, value=ang.nume_complet if ang else '?').font = Font(name='Arial', bold=True, size=9)
        ws.cell(row=row, column=1).border = thin_border

        for day_idx in range(7):
            acts = zile_dict.get(day_idx, [])
            if acts:
                text = '\n'.join(f'{a.proiect.cod_proiect}: {a.activitate_scurta}' for a in acts)
            else:
                text = '-'
            cell = ws.cell(row=row, column=day_idx + 2, value=text)
            cell.alignment = Alignment(wrap_text=True, vertical='top')
            cell.font = Font(name='Arial', size=8)
            cell.border = thin_border

        row += 1

    ws.column_dimensions['A'].width = 22
    for col_idx in range(2, 9):
        ws.column_dimensions[chr(64 + col_idx)].width = 20

    # Sheet 2: Detalii
    ws2 = wb.create_sheet('Detalii')
    detail_headers = ['Data', 'Angajat', 'Proiect', 'Instalatie', 'Zona', 'Activitate',
                       'Cantitate', 'U.M.', 'Materiale', 'Probleme', 'Status']
    for col_idx, h in enumerate(detail_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    for row_idx, a in enumerate(activitati, 2):
        ws2.cell(row=row_idx, column=1, value=a.data.strftime('%d.%m.%Y')).border = thin_border
        ws2.cell(row=row_idx, column=2, value=a.angajat.nume_complet if a.angajat else '').border = thin_border
        ws2.cell(row=row_idx, column=3, value=a.proiect.cod_proiect if a.proiect else '').border = thin_border
        ws2.cell(row=row_idx, column=4, value=a.tip_instalatie.cod if a.tip_instalatie else '').border = thin_border
        ws2.cell(row=row_idx, column=5, value=a.zona_lucru or '').border = thin_border
        ws2.cell(row=row_idx, column=6, value=a.activitate_principala).border = thin_border
        ws2.cell(row=row_idx, column=7, value=float(a.cantitate_executata) if a.cantitate_executata else '').border = thin_border
        ws2.cell(row=row_idx, column=8, value=a.unitate_masura or '').border = thin_border
        ws2.cell(row=row_idx, column=9, value=', '.join(m['denumire'] for m in a.materiale_lista)).border = thin_border
        ws2.cell(row=row_idx, column=10, value=a.probleme_intampinate or '').border = thin_border
        ws2.cell(row=row_idx, column=11, value=a.status.title()).border = thin_border

    for col_idx in range(1, 12):
        ws2.column_dimensions[chr(64 + col_idx)].width = 18

    # Sheet 3: Sinteza per Proiect
    ws3 = wb.create_sheet('Sinteza per Proiect')
    proiect_acts = defaultdict(list)
    for a in activitati:
        proiect_acts[a.proiect_id].append(a)

    row = 1
    for pid, acts in proiect_acts.items():
        p = Proiect.query.get(pid)
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
        ws3.cell(row=row, column=1, value=f'{p.cod_proiect} - {p.nume}' if p else '?')
        ws3.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=11, color='1A237E')
        row += 1

        for col_idx, h in enumerate(['Data', 'Angajat', 'Activitate', 'Cantitate', 'U.M.', 'Status'], 1):
            cell = ws3.cell(row=row, column=col_idx, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
            cell.border = thin_border
        row += 1

        for a in acts:
            ws3.cell(row=row, column=1, value=a.data.strftime('%d.%m.%Y')).border = thin_border
            ws3.cell(row=row, column=2, value=a.angajat.nume_complet if a.angajat else '').border = thin_border
            ws3.cell(row=row, column=3, value=a.activitate_principala).border = thin_border
            ws3.cell(row=row, column=4, value=float(a.cantitate_executata) if a.cantitate_executata else '').border = thin_border
            ws3.cell(row=row, column=5, value=a.unitate_masura or '').border = thin_border
            ws3.cell(row=row, column=6, value=a.status.title()).border = thin_border
            row += 1
        row += 1

    for col_idx in range(1, 7):
        ws3.column_dimensions[chr(64 + col_idx)].width = 20

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Activitati_Saptamana_{sapt}_{an}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@activitati_bp.route('/raport/lunar')
@login_required
def raport_lunar():
    """Generator export lunar."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    today = date.today()
    luna = request.args.get('luna', today.month, type=int)
    an = request.args.get('an', today.year, type=int)

    import calendar
    _, last_day = calendar.monthrange(an, luna)
    prima_zi = date(an, luna, 1)
    ultima_zi = date(an, luna, last_day)
    luni_ro = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
               'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    activitati = RaportActivitate.query.filter(
        RaportActivitate.data >= prima_zi,
        RaportActivitate.data <= ultima_zi,
    ).order_by(RaportActivitate.angajat_id, RaportActivitate.data).all()

    # Pontaje din aceeasi perioada
    pontaje = Pontaj.query.filter(
        Pontaj.data >= prima_zi, Pontaj.data <= ultima_zi
    ).all()
    pontaj_map = {}
    for p in pontaje:
        pontaj_map[(p.angajat_id, p.data.isoformat())] = float(p.ore_lucrate) if p.ore_lucrate else 0

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')

    # Sheet 1: Foaie Activitate Lunara
    ws = wb.active
    ws.title = 'Foaie Activitate Lunara'
    ws.merge_cells('A1:G1')
    ws['A1'] = f'Raport Activitati - {luni_ro[luna]} {an}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1A237E')

    from collections import defaultdict
    angajat_acts = defaultdict(list)
    for a in activitati:
        angajat_acts[a.angajat_id].append(a)

    row = 3
    for ang_id, acts in angajat_acts.items():
        ang = Angajat.query.get(ang_id)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=7)
        ws.cell(row=row, column=1, value=ang.nume_complet if ang else '?')
        ws.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=12, color='1A237E')
        row += 1

        for col_idx, h in enumerate(['Data', 'Proiect', 'Tip Instalatie', 'Activitate', 'Cantitate', 'Ore Pontate', 'Obs'], 1):
            cell = ws.cell(row=row, column=col_idx, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        row += 1

        for a in acts:
            ore = pontaj_map.get((a.angajat_id, a.data.isoformat()), '')
            ws.cell(row=row, column=1, value=a.data.strftime('%d.%m.%Y')).border = thin_border
            ws.cell(row=row, column=2, value=a.proiect.cod_proiect if a.proiect else '').border = thin_border
            ws.cell(row=row, column=3, value=a.tip_instalatie.cod if a.tip_instalatie else '').border = thin_border
            ws.cell(row=row, column=4, value=a.activitate_principala).border = thin_border
            ws.cell(row=row, column=5, value=f'{float(a.cantitate_executata)} {a.unitate_masura or ""}' if a.cantitate_executata else '').border = thin_border
            ws.cell(row=row, column=6, value=ore).border = thin_border
            ws.cell(row=row, column=7, value=a.observatii or '').border = thin_border
            row += 1
        row += 2

    for col_idx, w in enumerate([12, 15, 14, 40, 14, 12, 20], 1):
        ws.column_dimensions[chr(64 + col_idx)].width = w

    # Sheet 2: Centralizator
    ws2 = wb.create_sheet('Centralizator')
    inst_codes = [t.cod for t in TipInstalatie.query.filter_by(activ=True).order_by(TipInstalatie.ordine).all()]
    cent_headers = ['Angajat'] + inst_codes + ['Total Zile']
    for col_idx, h in enumerate(cent_headers, 1):
        cell = ws2.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for ang_id, acts in angajat_acts.items():
        ang = Angajat.query.get(ang_id)
        ws2.cell(row=row, column=1, value=ang.nume_complet if ang else '?').border = thin_border

        inst_zile = defaultdict(set)
        for a in acts:
            cod = a.tip_instalatie.cod if a.tip_instalatie else 'ALTA'
            inst_zile[cod].add(a.data)

        total_zile = len(set(a.data for a in acts))
        for col_idx, ic in enumerate(inst_codes, 2):
            ws2.cell(row=row, column=col_idx, value=len(inst_zile.get(ic, set()))).border = thin_border
        ws2.cell(row=row, column=len(inst_codes) + 2, value=total_zile).border = thin_border
        row += 1

    for col_idx in range(1, len(cent_headers) + 1):
        ws2.column_dimensions[chr(64 + col_idx)].width = 14

    # Sheet 3: Per Proiect
    ws3 = wb.create_sheet('Per Proiect')
    proiect_acts = defaultdict(list)
    for a in activitati:
        proiect_acts[a.proiect_id].append(a)

    row = 1
    for pid, acts in proiect_acts.items():
        p = Proiect.query.get(pid)
        ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws3.cell(row=row, column=1, value=f'{p.cod_proiect} - {p.nume}' if p else '?')
        ws3.cell(row=row, column=1).font = Font(name='Arial', bold=True, size=11, color='1A237E')
        row += 1
        for col_idx, h in enumerate(['Categorie', 'Total Activitati', 'Cantitate Totala', 'U.M.', 'Angajati'], 1):
            cell = ws3.cell(row=row, column=col_idx, value=h)
            cell.font = Font(name='Arial', bold=True, size=9)
            cell.fill = PatternFill(start_color='E8EAF6', end_color='E8EAF6', fill_type='solid')
            cell.border = thin_border
        row += 1

        cat_stats = defaultdict(lambda: {'count': 0, 'cantitate': 0, 'um': '', 'angajati': set()})
        for a in acts:
            cat_name = a.categorie_activitate.denumire if a.categorie_activitate else 'Altele'
            cat_stats[cat_name]['count'] += 1
            if a.cantitate_executata:
                cat_stats[cat_name]['cantitate'] += float(a.cantitate_executata)
                cat_stats[cat_name]['um'] = a.unitate_masura or ''
            cat_stats[cat_name]['angajati'].add(a.angajat_id)

        for cat_name, stats in cat_stats.items():
            ws3.cell(row=row, column=1, value=cat_name).border = thin_border
            ws3.cell(row=row, column=2, value=stats['count']).border = thin_border
            ws3.cell(row=row, column=3, value=round(stats['cantitate'], 2) if stats['cantitate'] else '').border = thin_border
            ws3.cell(row=row, column=4, value=stats['um']).border = thin_border
            ws3.cell(row=row, column=5, value=len(stats['angajati'])).border = thin_border
            row += 1
        row += 1

    for col_idx in range(1, 6):
        ws3.column_dimensions[chr(64 + col_idx)].width = 20

    # Sheet 4: Probleme Raportate
    ws4 = wb.create_sheet('Probleme Raportate')
    prob_headers = ['Data', 'Angajat', 'Proiect', 'Problema', 'Solutie Aplicata']
    for col_idx, h in enumerate(prob_headers, 1):
        cell = ws4.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    row = 2
    for a in activitati:
        if a.probleme_intampinate:
            ws4.cell(row=row, column=1, value=a.data.strftime('%d.%m.%Y')).border = thin_border
            ws4.cell(row=row, column=2, value=a.angajat.nume_complet if a.angajat else '').border = thin_border
            ws4.cell(row=row, column=3, value=a.proiect.cod_proiect if a.proiect else '').border = thin_border
            ws4.cell(row=row, column=4, value=a.probleme_intampinate).border = thin_border
            ws4.cell(row=row, column=5, value=a.solutii_aplicate or '').border = thin_border
            row += 1

    for col_idx in range(1, 6):
        ws4.column_dimensions[chr(64 + col_idx)].width = 25

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Activitati_Luna_{luni_ro[luna]}_{an}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@activitati_bp.route('/raport/anual')
@login_required
def raport_anual():
    """Generator export anual."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference

    an = request.args.get('an', date.today().year, type=int)
    luni_ro = ['', 'Ian', 'Feb', 'Mar', 'Apr', 'Mai', 'Iun', 'Iul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    luni_full = ['', 'Ianuarie', 'Februarie', 'Martie', 'Aprilie', 'Mai', 'Iunie',
                 'Iulie', 'August', 'Septembrie', 'Octombrie', 'Noiembrie', 'Decembrie']

    prima_zi = date(an, 1, 1)
    ultima_zi = date(an, 12, 31)

    activitati = RaportActivitate.query.filter(
        RaportActivitate.data >= prima_zi,
        RaportActivitate.data <= ultima_zi,
    ).order_by(RaportActivitate.angajat_id, RaportActivitate.data).all()

    wb = Workbook()
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_font = Font(name='Arial', bold=True, color='FFFFFF', size=10)
    header_fill = PatternFill(start_color='1A237E', end_color='1A237E', fill_type='solid')

    # Sheet 1: Sumar Anual
    ws = wb.active
    ws.title = 'Sumar Anual'
    ws.merge_cells('A1:N1')
    ws['A1'] = f'Raport Activitati Anual - {an}'
    ws['A1'].font = Font(name='Arial', bold=True, size=14, color='1A237E')

    headers = ['Angajat'] + luni_ro[1:] + ['Total']
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_idx, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border

    from collections import defaultdict
    angajat_lunar = defaultdict(lambda: defaultdict(int))
    for a in activitati:
        angajat_lunar[a.angajat_id][a.data.month] += 1

    row = 4
    for ang_id, lunar in angajat_lunar.items():
        ang = Angajat.query.get(ang_id)
        ws.cell(row=row, column=1, value=ang.nume_complet if ang else '?').border = thin_border
        total = 0
        for m in range(1, 13):
            val = lunar.get(m, 0)
            ws.cell(row=row, column=m + 1, value=val).border = thin_border
            total += val
        ws.cell(row=row, column=14, value=total).border = thin_border
        ws.cell(row=row, column=14).font = Font(name='Arial', bold=True)
        row += 1

    ws.column_dimensions['A'].width = 22
    for c in range(2, 15):
        ws.column_dimensions[chr(64 + c)].width = 8

    # Add chart
    if row > 4:
        chart = BarChart()
        chart.type = 'col'
        chart.title = f'Activitati per Angajat - {an}'
        chart.y_axis.title = 'Nr. Activitati'
        data = Reference(ws, min_col=2, max_col=13, min_row=3, max_row=row - 1)
        cats = Reference(ws, min_col=1, min_row=4, max_row=row - 1)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(cats)
        chart.shape = 4
        chart.width = 30
        chart.height = 15
        ws.add_chart(chart, f'A{row + 2}')

    # Sheet 2: Statistici
    ws2 = wb.create_sheet('Statistici')
    ws2.merge_cells('A1:D1')
    ws2['A1'] = f'Statistici Activitati - {an}'
    ws2['A1'].font = Font(name='Arial', bold=True, size=14, color='1A237E')

    # Top activitati ca frecventa
    ws2.cell(row=3, column=1, value='Top Categorii Activitati').font = Font(name='Arial', bold=True, size=11)
    cat_freq = defaultdict(int)
    for a in activitati:
        cat_name = a.categorie_activitate.denumire if a.categorie_activitate else 'Altele'
        cat_freq[cat_name] += 1

    ws2.cell(row=4, column=1, value='Categorie').font = Font(name='Arial', bold=True)
    ws2.cell(row=4, column=2, value='Frecventa').font = Font(name='Arial', bold=True)
    row = 5
    for cat, freq in sorted(cat_freq.items(), key=lambda x: -x[1])[:20]:
        ws2.cell(row=row, column=1, value=cat).border = thin_border
        ws2.cell(row=row, column=2, value=freq).border = thin_border
        row += 1

    # Distributie per tip instalatie
    ws2.cell(row=row + 1, column=1, value='Distributie per Tip Instalatie').font = Font(name='Arial', bold=True, size=11)
    inst_freq = defaultdict(int)
    for a in activitati:
        cod = a.tip_instalatie.cod if a.tip_instalatie else 'N/A'
        inst_freq[cod] += 1

    row += 2
    ws2.cell(row=row, column=1, value='Instalatie').font = Font(name='Arial', bold=True)
    ws2.cell(row=row, column=2, value='Nr. Activitati').font = Font(name='Arial', bold=True)
    row += 1
    for inst, freq in sorted(inst_freq.items(), key=lambda x: -x[1]):
        ws2.cell(row=row, column=1, value=inst).border = thin_border
        ws2.cell(row=row, column=2, value=freq).border = thin_border
        row += 1

    ws2.column_dimensions['A'].width = 35
    ws2.column_dimensions['B'].width = 15

    # Sheet 3: Raport Management
    ws3 = wb.create_sheet('Raport Management')
    ws3.merge_cells('A1:D1')
    ws3['A1'] = f'Raport Management - {an}'
    ws3['A1'].font = Font(name='Arial', bold=True, size=14, color='1A237E')

    ws3.cell(row=3, column=1, value='Total activitati raportate:').font = Font(bold=True)
    ws3.cell(row=3, column=2, value=len(activitati))
    ws3.cell(row=4, column=1, value='Angajati cu rapoarte:').font = Font(bold=True)
    ws3.cell(row=4, column=2, value=len(angajat_lunar))
    ws3.cell(row=5, column=1, value='Proiecte active:').font = Font(bold=True)
    proiecte_set = set(a.proiect_id for a in activitati)
    ws3.cell(row=5, column=2, value=len(proiecte_set))

    # Probleme recurente
    ws3.cell(row=7, column=1, value='Probleme Raportate').font = Font(name='Arial', bold=True, size=11, color='C62828')
    probleme = [a for a in activitati if a.probleme_intampinate]
    ws3.cell(row=8, column=1, value='Total probleme raportate:').font = Font(bold=True)
    ws3.cell(row=8, column=2, value=len(probleme))

    row = 10
    for col_idx, h in enumerate(['Luna', 'Nr. Probleme', 'Proiecte Afectate'], 1):
        ws3.cell(row=row, column=col_idx, value=h).font = Font(bold=True)
    row += 1
    prob_lunar = defaultdict(lambda: {'count': 0, 'proiecte': set()})
    for a in probleme:
        prob_lunar[a.data.month]['count'] += 1
        prob_lunar[a.data.month]['proiecte'].add(a.proiect_id)
    for m in range(1, 13):
        if m in prob_lunar:
            ws3.cell(row=row, column=1, value=luni_full[m]).border = thin_border
            ws3.cell(row=row, column=2, value=prob_lunar[m]['count']).border = thin_border
            ws3.cell(row=row, column=3, value=len(prob_lunar[m]['proiecte'])).border = thin_border
            row += 1

    ws3.column_dimensions['A'].width = 30
    ws3.column_dimensions['B'].width = 18
    ws3.column_dimensions['C'].width = 18

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f'Activitati_Anual_{an}.xlsx'
    return send_file(output, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


@activitati_bp.route('/raport/proiect')
@login_required
def raport_proiect():
    """Raport activitati per proiect."""
    proiect_id = request.args.get('proiect_id', type=int)
    if not proiect_id:
        flash('Selectati un proiect.', 'warning')
        return redirect(url_for('activitati.panou'))

    proiect = Proiect.query.get_or_404(proiect_id)
    activitati = RaportActivitate.query.filter_by(proiect_id=proiect_id).order_by(
        RaportActivitate.data.desc()
    ).all()

    return render_template('activitati/raport_proiect.html',
        proiect=proiect,
        activitati=activitati,
    )


# ============================================================
# EXPORT EDIFICO - Structura xlsx exacta dupa template-ul referinta
# Un sheet per angajat, grupare pe saptamani in luna
# ============================================================

LUNI_RO = [
    '', 'ianuarie', 'februarie', 'martie', 'aprilie', 'mai', 'iunie',
    'iulie', 'august', 'septembrie', 'octombrie', 'noiembrie', 'decembrie'
]


def _get_company_name():
    """Citeste numele firmei din config-ul aplicatiei (fallback EDIFICO)."""
    try:
        from routes.setari import _load_config
        cfg = _load_config()
        nume = cfg.get('firma_nume', '').strip()
        if nume:
            # Extrage doar prima parte (ex: "EDIFICO CONSTRUCT SRL" -> "EDIFICO")
            short = nume.split()[0] if nume else 'EDIFICO'
            return short, nume
    except Exception:
        pass
    return 'EDIFICO', 'EDIFICO CONSTRUCT SRL'


def _saptamani_din_luna(an, luna, sarbatori_set, zile_extra_lucrate=None):
    """
    Returneaza lista de saptamani din luna sub forma:
    [(numar_sapt_local, numar_sapt_iso, [zile_lucratoare]), ...]
    - Default: include doar Luni-Vineri (weekday 0-4)
    - zile_extra_lucrate: set de date() suplimentare (de ex sambete cu activitate) care
      vor fi incluse in saptamana corespunzatoare.
    """
    import calendar
    if zile_extra_lucrate is None:
        zile_extra_lucrate = set()

    _, last_day = calendar.monthrange(an, luna)

    saptamani = {}  # iso_week -> list of dates
    for d in range(1, last_day + 1):
        zi = date(an, luna, d)
        # Includem Luni-Vineri sau zilele extra (sambete/duminici cu activitate)
        if zi.weekday() <= 4 or zi in zile_extra_lucrate:
            iso_w = zi.isocalendar()[1]
            saptamani.setdefault(iso_w, []).append(zi)

    sorted_weeks = sorted(saptamani.items())
    rezultat = []
    for idx, (iso_w, zile) in enumerate(sorted_weeks, start=1):
        # Sorteaza zilele crescator
        zile_sortate = sorted(zile)
        rezultat.append((idx, iso_w, zile_sortate))
    return rezultat


def _zile_extra_lucrate_pentru_angajat(angajat_id, an, luna):
    """
    Returneaza set de date (sambata/duminica) care apar ca zile lucrate pentru angajat in luna.
    Reguli:
    - Activitate zilnica cu data in sambata/duminica -> include acea zi
    - Activitate saptamanala/lunara cu include_sambata=True -> include sambetele
    - Activitate saptamanala/lunara cu include_duminica=True -> include duminicile
    """
    import calendar
    _, last_day = calendar.monthrange(an, luna)
    prima = date(an, luna, 1)
    ultima = date(an, luna, last_day)

    extra = set()

    # 1. Activitati zilnice in weekend
    daily_weekends = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate == 'zilnica',
        RaportActivitate.data >= prima,
        RaportActivitate.data <= ultima,
    ).all()
    for a in daily_weekends:
        if a.data and a.data.weekday() >= 5:
            extra.add(a.data)

    # 2. Activitati saptamanale/lunare cu include_sambata sau include_duminica
    luna_an = f'{an:04d}-{luna:02d}'
    flagged = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate.in_(['saptamanala', 'lunara']),
        db.or_(
            RaportActivitate.include_sambata == True,
            RaportActivitate.include_duminica == True,
        ),
        db.or_(
            RaportActivitate.luna_an == luna_an,
            db.and_(
                RaportActivitate.data <= ultima,
                db.or_(
                    RaportActivitate.data_sfarsit.is_(None),
                    RaportActivitate.data_sfarsit >= prima,
                ),
            ),
        ),
    ).all()

    for a in flagged:
        # Determina intervalul activitatii in luna data
        ds = max(a.data, prima) if a.data else prima
        df = min(a.data_sfarsit, ultima) if a.data_sfarsit else ultima
        if a.tip_activitate == 'lunara' and a.luna_an == luna_an:
            ds, df = prima, ultima
        cur = ds
        while cur <= df:
            if a.include_sambata and cur.weekday() == 5:
                extra.add(cur)
            if a.include_duminica and cur.weekday() == 6:
                extra.add(cur)
            cur += timedelta(days=1)

    return extra


def _detalii_pe_zi_pentru_saptamana(angajat_id, zile_saptamana, luna, an):
    """
    Returneaza dict {data: text} cu detalii per zi din activitatile saptamanale/lunare
    care au detalii_pe_zi populate.
    """
    if not zile_saptamana:
        return {}

    prima_zi = min(zile_saptamana)
    ultima_zi = max(zile_saptamana)
    iso_week = prima_zi.isocalendar()[1]
    luna_an = f'{an:04d}-{luna:02d}'

    rapoarte = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate.in_(['saptamanala', 'lunara']),
        RaportActivitate.detalii_pe_zi.isnot(None),
        db.or_(
            RaportActivitate.numar_saptamana == iso_week,
            RaportActivitate.luna_an == luna_an,
            db.and_(
                RaportActivitate.data <= ultima_zi,
                db.or_(
                    RaportActivitate.data_sfarsit.is_(None),
                    RaportActivitate.data_sfarsit >= prima_zi,
                ),
            ),
        ),
    ).all()

    rezultat = {}
    zile_set = set(zile_saptamana)
    for r in rapoarte:
        for det in r.detalii_pe_zi_lista:
            d_obj = det.get('_data_obj')
            if d_obj and d_obj in zile_set:
                text_parts = []
                if det.get('text'):
                    text_parts.append(det['text'])
                if det.get('proiect_id'):
                    p = Proiect.query.get(det['proiect_id'])
                    if p:
                        text_parts.append(f'[{p.cod_proiect}]')
                if det.get('ore'):
                    text_parts.append(f"({det['ore']}h)")
                if text_parts:
                    rezultat.setdefault(d_obj, []).append(' '.join(text_parts))

    # Concateneaza listele cu newline
    return {d: '\n'.join(texte) for d, texte in rezultat.items()}


def _activitati_pentru_saptamana(angajat_id, zile_saptamana, luna, an):
    """
    Returneaza lista de texte de activitate pentru un angajat in saptamana data.
    Include: activitati zilnice (cu data in zile_saptamana) + saptamanale + lunare relevante.
    NU include activitatile saptamanale/lunare care au detalii_pe_zi (pentru ca acelea sunt
    afisate per zi separata).
    """
    if not zile_saptamana:
        return []

    prima_zi = min(zile_saptamana)
    ultima_zi = max(zile_saptamana)
    iso_week = prima_zi.isocalendar()[1]
    luna_an = f'{an:04d}-{luna:02d}'

    # Daily: data intre prima si ultima zi
    daily = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate == 'zilnica',
        RaportActivitate.data >= prima_zi,
        RaportActivitate.data <= ultima_zi,
    ).all()

    # Weekly: same iso week sau interval suprapus (activitatea = principala+detaliata)
    weekly = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate == 'saptamanala',
        db.or_(
            RaportActivitate.numar_saptamana == iso_week,
            db.and_(
                RaportActivitate.data <= ultima_zi,
                db.or_(
                    RaportActivitate.data_sfarsit.is_(None),
                    RaportActivitate.data_sfarsit >= prima_zi,
                ),
            ),
        ),
    ).all()

    # Monthly: aceeasi luna_an (activitatea = principala+detaliata)
    monthly = RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate == 'lunara',
        db.or_(
            RaportActivitate.luna_an == luna_an,
            db.and_(
                RaportActivitate.data >= date(an, luna, 1),
                RaportActivitate.data < (date(an + (1 if luna == 12 else 0), 1 if luna == 12 else luna + 1, 1)),
            ),
        ),
    ).all()

    texte = []
    seen_ids = set()
    for grup in (weekly, daily, monthly):
        for a in grup:
            if a.id in seen_ids:
                continue
            seen_ids.add(a.id)
            t = (a.activitate_principala or '').strip()
            if a.activitate_detaliata:
                detalii = a.activitate_detaliata.strip()
                if detalii and detalii not in t:
                    t = t + ('\n' + detalii if t else detalii)
            t = _curata_activitate(t)
            if t:
                texte.append(t)
    return texte


# === STILURI XLSX (cu paleta consistenta) ===
# Paleta Edifico (premium): navy obsidian + champagne gold + cream
COLOR_NAVY = '0B1426'        # navy obsidian - titluri, header luna, total
COLOR_GOLD = 'C9A961'        # champagne gold - accente, text pe navy
COLOR_PRIMARY = '0B1426'     # border accent (navy)
COLOR_TITLE_RED = '0B1426'   # (nume pastrat) titlu raport -> navy
COLOR_HEADER_BG = '0B1426'   # header luna: fundal navy
COLOR_HEADER_FG = 'C9A961'   # text gold pe navy
COLOR_COLHDR_BG = 'C9A961'   # header coloane: fundal gold
COLOR_TEXT = '2B2B2B'        # text celule (gri inchis, lizibil)
COLOR_SAT_BG = 'EFE7D2'      # gold-tint deschis (sambata)
COLOR_SUN_BG = 'E6D9BC'      # gold-tint cald (duminica)
COLOR_HOLIDAY_BG = 'F3E1B0'  # gold soft (sarbatoare)
COLOR_ZEBRA_BG = 'F5F1E8'    # cream (zebra / hartie)
COLOR_TOTAL_BG = '0B1426'    # navy (total) cu text gold
COLOR_BORDER = 'D8CBA8'      # gold-grey deschis (border)


def _stiluri_xlsx():
    """Returneaza un dict cu stilurile reutilizabile pentru exportul xlsx."""
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    thin = Side(style='thin', color=COLOR_BORDER)
    medium = Side(style='medium', color=COLOR_PRIMARY)
    return {
        'titlu_font': Font(name='Calibri', size=16, bold=True, color=COLOR_NAVY),
        'subtitlu_font': Font(name='Calibri', size=10, italic=True, color='7A7A7A'),
        'nume_font': Font(name='Calibri', size=12, bold=True, color=COLOR_NAVY),
        'luna_font': Font(name='Calibri', size=11, bold=True, color=COLOR_GOLD),
        'saptamana_font': Font(name='Calibri', size=10, bold=True, color=COLOR_NAVY),
        'cell_font': Font(name='Calibri', size=10, color=COLOR_TEXT),
        'cell_font_bold': Font(name='Calibri', size=10, bold=True, color=COLOR_NAVY),
        'sat_font': Font(name='Calibri', size=10, bold=True, color=COLOR_NAVY),
        'sun_font': Font(name='Calibri', size=10, bold=True, color=COLOR_NAVY),
        'header_font': Font(name='Calibri', size=10, bold=True, color=COLOR_NAVY),
        'total_font': Font(name='Calibri', size=10, bold=True, color=COLOR_GOLD),
        'sat_fill': PatternFill(start_color=COLOR_SAT_BG, end_color=COLOR_SAT_BG, fill_type='solid'),
        'sun_fill': PatternFill(start_color=COLOR_SUN_BG, end_color=COLOR_SUN_BG, fill_type='solid'),
        'holiday_fill': PatternFill(start_color=COLOR_HOLIDAY_BG, end_color=COLOR_HOLIDAY_BG, fill_type='solid'),
        'header_fill': PatternFill(start_color=COLOR_HEADER_BG, end_color=COLOR_HEADER_BG, fill_type='solid'),
        'colhdr_fill': PatternFill(start_color=COLOR_COLHDR_BG, end_color=COLOR_COLHDR_BG, fill_type='solid'),
        'zebra_fill': PatternFill(start_color=COLOR_ZEBRA_BG, end_color=COLOR_ZEBRA_BG, fill_type='solid'),
        'total_fill': PatternFill(start_color=COLOR_TOTAL_BG, end_color=COLOR_TOTAL_BG, fill_type='solid'),
        'border_thin': Border(left=thin, right=thin, top=thin, bottom=thin),
        'border_medium_top': Border(left=thin, right=thin, top=medium, bottom=thin),
        'align_center': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'align_left': Alignment(horizontal='left', vertical='center', wrap_text=True),
    }


def _curata_activitate(text):
    """
    Text activitate curat pentru export: fara bullet '•', fara marcatori de lista
    ('- ', '– ', '* ') la inceput de rand, fara paranteze/paranteze patrate (se
    scot doar caracterele, continutul ramane). Punctuatia normala se pastreaza.
    """
    import re
    if not text:
        return ''
    linii = []
    for raw in str(text).replace('•', ' ').splitlines():
        s = re.sub(r'^[\-–—•\*·]+\s*', '', raw.strip()).strip()
        if s:
            linii.append(s)
    txt = '\n'.join(linii)
    for ch in '()[]':
        txt = txt.replace(ch, '')
    txt = re.sub(r'[ \t]{2,}', ' ', txt)
    return txt.strip()


def _nume_proiect(a):
    """Numele proiectului (nu codul), trunchiat la 60 caractere."""
    if not a or not a.proiect or not a.proiect.nume:
        return None
    return a.proiect.nume.strip()[:60]


ZILE_RO_SCURT = ['Lu', 'Ma', 'Mi', 'Jo', 'Vi', 'Sa', 'Du']
LUNI_RO_SCURT = ['', 'ian', 'feb', 'mar', 'apr', 'mai', 'iun',
                 'iul', 'aug', 'sep', 'oct', 'nov', 'dec']


def _data_ro(zi):
    """Data in romana, fara engleza si fara paranteze. Ex: 'Lu 05 ian'."""
    return f'{ZILE_RO_SCURT[zi.weekday()]} {zi.day:02d} {LUNI_RO_SCURT[zi.month]}'


def _activitate_text(a):
    """Mereu: activitate principala + activitate detaliata (daca exista), curatat."""
    t = (a.activitate_principala or '').strip()
    det = (a.activitate_detaliata or '').strip()
    if det and det not in t:
        t = (t + '\n' + det).strip()
    return _curata_activitate(t)


def _info_zile(angajat_id, zile):
    """
    Per-zi: {data: {'proiect': nume|None, 'ore': float, 'activitati': [texte]}}.
    Proiectul = cel SELECTAT in ziua respectiva (raportul zilnic pe acea data);
    daca ziua n-are raport zilnic, se completeaza cu proiectul raportului
    saptamanal/lunar care o acopera. Activitatea = principala + detaliata.
    """
    if not zile:
        return {}
    prima, ultima = min(zile), max(zile)
    info = {z: {'proiect': None, 'ore': 0.0, 'activitati': []} for z in zile}

    # 1. ZILNICE - proiectul selectat in ziua respectiva (prioritar)
    for a in RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate == 'zilnica',
        RaportActivitate.data >= prima, RaportActivitate.data <= ultima,
    ).all():
        if a.data not in info:
            continue
        if info[a.data]['proiect'] is None:
            info[a.data]['proiect'] = _nume_proiect(a)
        if a.ore_lucrate:
            try:
                info[a.data]['ore'] += float(a.ore_lucrate)
            except (TypeError, ValueError):
                pass
        t = _activitate_text(a)
        if t and t not in info[a.data]['activitati']:
            info[a.data]['activitati'].append(t)

    # 2. SPAN (saptamanal/lunar) - completeaza DOAR proiectul pe zilele fara zilnic
    for a in RaportActivitate.query.filter(
        RaportActivitate.angajat_id == angajat_id,
        RaportActivitate.tip_activitate.in_(['saptamanala', 'lunara']),
    ).all():
        proj = _nume_proiect(a)
        if not proj:
            continue
        ds = a.data or prima
        df = a.data_sfarsit or ultima
        for z in zile:
            if ds <= z <= df and info[z]['proiect'] is None:
                info[z]['proiect'] = proj
    return info


def _adauga_sectiune_luna(ws, angajat, an, luna, company_short, start_row, S, zile_extra=None):
    """
    Adauga o sectiune luna. Coloane: B=Luna | C=Saptamana | D=Data |
    E=Proiect | F=Ore | G=Activitati desfasurate (cu activitatea principala).
    """
    if zile_extra is None:
        zile_extra = _zile_extra_lucrate_pentru_angajat(angajat.id, an, luna)

    luna_text = LUNI_RO[luna].capitalize()

    sarbatori = set()
    for s in SarbatoareLegala.query.filter_by(an=an).all():
        if s.data.month == luna:
            sarbatori.add(s.data)

    saptamani = _saptamani_din_luna(an, luna, sarbatori, zile_extra)
    if not saptamani:
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=7)
        c = ws.cell(row=start_row, column=2, value=f'{luna_text} {an} — Nicio zi lucratoare')
        c.font = S['subtitlu_font']
        c.alignment = S['align_center']
        return start_row + 2

    # === Header luna (rand titlu) ===
    ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row, end_column=7)
    titlu_cell = ws.cell(row=start_row, column=2, value=f'{luna_text.upper()} {an}')
    titlu_cell.font = S['luna_font']
    titlu_cell.fill = S['header_fill']
    titlu_cell.alignment = S['align_center']
    titlu_cell.border = S['border_thin']
    for col in range(2, 8):
        ws.cell(row=start_row, column=col).fill = S['header_fill']
        ws.cell(row=start_row, column=col).border = S['border_thin']
    ws.row_dimensions[start_row].height = 24

    # === Rand header coloane ===
    header_row = start_row + 1
    headers = ['Luna', 'Saptamana', 'Data', 'Proiect', 'Ore', 'Activitati desfasurate']
    for col_idx, h in enumerate(headers, start=2):
        c = ws.cell(row=header_row, column=col_idx, value=h)
        c.font = S['header_font']
        c.fill = S['colhdr_fill']
        c.alignment = S['align_center']
        c.border = S['border_thin']
    ws.row_dimensions[header_row].height = 22

    current_row = header_row + 1
    primul_rand_sapt = current_row
    ultimul_rand_sapt = None
    total_zile_lucrate = 0
    total_ore = 0.0

    for idx_sapt, (sapt_idx, iso_week, zile) in enumerate(saptamani):
        nr_zile = len(zile)
        if nr_zile == 0:
            continue
        sr, er = current_row, current_row + nr_zile - 1
        ultimul_rand_sapt = er
        zebra = (idx_sapt % 2 == 1)

        info_zi = _info_zile(angajat.id, zile)
        texte_general = _activitati_pentru_saptamana(angajat.id, zile, luna, an)
        # Avem activitati zilnice (principala+detaliata) per zi?
        are_per_zi = any(info_zi[z]['activitati'] for z in zile)

        # Coloana C: numele saptamanii (merged)
        if nr_zile > 1:
            ws.merge_cells(start_row=sr, start_column=3, end_row=er, end_column=3)
        c_cell = ws.cell(row=sr, column=3, value=f'Saptamana {sapt_idx}\n(S{iso_week})')
        c_cell.font = S['saptamana_font']
        c_cell.alignment = S['align_center']
        c_cell.border = S['border_thin']

        # Coloana G (activitati): merged pe saptamana DOAR daca nu avem continut per-zi
        if not are_per_zi:
            if nr_zile > 1:
                ws.merge_cells(start_row=sr, start_column=7, end_row=er, end_column=7)
            text_g = '\n'.join(texte_general) if texte_general else '—'
            g_cell = ws.cell(row=sr, column=7, value=text_g)
            g_cell.font = S['cell_font']
            g_cell.alignment = S['align_left']
            g_cell.border = S['border_thin']

        total_ore += sum(info_zi[z]['ore'] for z in zile)

        # Randuri per zi: D=data, E=proiect, F=ore, G=activitate (daca per-zi)
        for off, zi in enumerate(zile):
            r = sr + off
            total_zile_lucrate += 1
            ws.row_dimensions[r].height = 16.5

            # Fill-ul zilei (dupa tip)
            if zi in sarbatori:
                fill, font_zi = S['holiday_fill'], S['cell_font_bold']
            elif zi.weekday() == 5:
                fill, font_zi = S['sat_fill'], S['sat_font']
            elif zi.weekday() == 6:
                fill, font_zi = S['sun_fill'], S['sun_font']
            elif zebra:
                fill, font_zi = S['zebra_fill'], S['cell_font']
            else:
                fill, font_zi = None, S['cell_font']

            # D: data (in romana, ca text - fara engleza/paranteze)
            d_cell = ws.cell(row=r, column=4, value=_data_ro(zi))
            d_cell.alignment = S['align_center']
            d_cell.border = S['border_thin']
            d_cell.font = font_zi
            if fill:
                d_cell.fill = fill

            # E: proiectul selectat in ziua respectiva
            e_cell = ws.cell(row=r, column=5, value=info_zi[zi]['proiect'] or '—')
            e_cell.font = S['cell_font']
            e_cell.alignment = S['align_center']
            e_cell.border = S['border_thin']
            if fill:
                e_cell.fill = fill

            # F: ore
            ore_zi = info_zi[zi]['ore']
            f_cell = ws.cell(row=r, column=6, value=(round(ore_zi, 1) if ore_zi else None))
            f_cell.alignment = S['align_center']
            f_cell.border = S['border_thin']
            f_cell.font = S['cell_font']
            if fill:
                f_cell.fill = fill

            # G: activitate pe zi = activitate principala + detaliata
            if are_per_zi:
                parts = list(info_zi[zi]['activitati'])
                txt = '\n'.join(parts) if parts else ''
                g_cell = ws.cell(row=r, column=7, value=txt)
                g_cell.font = S['cell_font']
                g_cell.alignment = S['align_left']
                g_cell.border = S['border_thin']
            else:
                ws.cell(row=r, column=7).border = S['border_thin']

        current_row = er + 1

    # === Coloana B: numele lunii (merged pe toate sapt) ===
    if ultimul_rand_sapt and ultimul_rand_sapt >= primul_rand_sapt:
        ws.merge_cells(start_row=primul_rand_sapt, start_column=2,
                       end_row=ultimul_rand_sapt, end_column=2)
        b_cell = ws.cell(row=primul_rand_sapt, column=2, value=luna_text)
        b_cell.font = S['nume_font']
        b_cell.alignment = S['align_center']
        b_cell.fill = S['zebra_fill']
        b_cell.border = S['border_thin']
        for r in range(primul_rand_sapt, ultimul_rand_sapt + 1):
            ws.cell(row=r, column=2).border = S['border_thin']

    # === Rand total luna === (B:C label | D zile | F ore)
    total_row = (ultimul_rand_sapt or current_row) + 1
    ws.merge_cells(start_row=total_row, start_column=2, end_row=total_row, end_column=3)
    for col in range(2, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.fill = S['total_fill']
        cell.border = S['border_thin']
    t1 = ws.cell(row=total_row, column=2, value=f'TOTAL {luna_text.upper()}')
    t1.font = S['total_font']
    t1.alignment = S['align_center']

    t2 = ws.cell(row=total_row, column=4, value=f'{total_zile_lucrate} zile')
    t2.font = S['total_font']
    t2.alignment = S['align_center']

    t3 = ws.cell(row=total_row, column=6,
                 value=(f'{total_ore:.1f} ore' if total_ore > 0 else '—'))
    t3.font = S['total_font']
    t3.alignment = S['align_center']
    ws.row_dimensions[total_row].height = 22

    return total_row + 2  # 1 rand gol intre luni


def _construieste_sheet_angajat(wb, angajat, perioade, company_short, sheet_index=0):
    """
    Construieste un sheet EDIFICO pentru un angajat, pe o lista de perioade [(an, luna), ...].
    """
    S = _stiluri_xlsx()

    # Titlu sheet (max 31 caractere, evita caractere ilegale Excel)
    base_title = angajat.nume_complet
    illegal = '[]:*?/\\'
    safe_title = ''.join(c for c in base_title if c not in illegal)[:31]
    if not safe_title:
        safe_title = f'Angajat {angajat.id}'

    existing_titles = {ws.title for ws in wb.worksheets}
    final_title = safe_title
    suffix = 2
    while final_title in existing_titles:
        final_title = f'{safe_title[:28]}_{suffix}'[:31]
        suffix += 1

    if sheet_index == 0 and len(wb.worksheets) == 1 and wb.active.title == 'Sheet':
        ws = wb.active
        ws.title = final_title
    else:
        ws = wb.create_sheet(title=final_title)

    # Latimi coloane optimizate (B=Luna C=Sapt D=Data E=Proiect F=Ore G=Activitati)
    ws.column_dimensions['A'].width = 4.0
    ws.column_dimensions['B'].width = 14.0
    ws.column_dimensions['C'].width = 15.0
    ws.column_dimensions['D'].width = 16.0
    ws.column_dimensions['E'].width = 22.0
    ws.column_dimensions['F'].width = 8.0
    ws.column_dimensions['G'].width = 55.0

    # === Rand 2: Titlu mare cu numele firmei + perioada ===
    if perioade:
        prima_p = perioade[0]
        ultima_p = perioade[-1]
        if prima_p == ultima_p:
            perioada_text = f'{LUNI_RO[prima_p[1]].capitalize()} {prima_p[0]}'
        else:
            perioada_text = f'{LUNI_RO[prima_p[1]].capitalize()} {prima_p[0]} — {LUNI_RO[ultima_p[1]].capitalize()} {ultima_p[0]}'
    else:
        perioada_text = ''

    ws.merge_cells('B2:G2')
    titlu_cell = ws.cell(row=2, column=2, value=f'RAPORT DE ACTIVITATE {company_short.upper()}')
    titlu_cell.font = S['titlu_font']
    titlu_cell.alignment = S['align_center']
    ws.row_dimensions[2].height = 28

    ws.merge_cells('B3:G3')
    sub_cell = ws.cell(row=3, column=2, value=perioada_text)
    sub_cell.font = S['subtitlu_font']
    sub_cell.alignment = S['align_center']
    ws.row_dimensions[3].height = 18

    # === Rand 5: Numele angajatului ===
    ws.merge_cells('B5:G5')
    nume_cell = ws.cell(row=5, column=2, value=f'Angajat: {angajat.nume_complet}    |    Functie: {angajat.functie}')
    nume_cell.font = S['nume_font']
    nume_cell.alignment = S['align_center']
    nume_cell.fill = S['zebra_fill']
    nume_cell.border = S['border_thin']
    for col in range(2, 8):
        ws.cell(row=5, column=col).fill = S['zebra_fill']
        ws.cell(row=5, column=col).border = S['border_thin']
    ws.row_dimensions[5].height = 24

    # === Sectiuni pe luna ===
    current_row = 7
    for an, luna in perioade:
        zile_extra = _zile_extra_lucrate_pentru_angajat(angajat.id, an, luna)
        current_row = _adauga_sectiune_luna(ws, angajat, an, luna, company_short, current_row, S, zile_extra)

    # Freeze panes pentru a tine titlul vizibil la scroll
    ws.freeze_panes = 'A6'

    return ws


@activitati_bp.route('/raport/edifico')
@activitati_bp.route('/export')
@login_required
def export_edifico():
    """
    Export xlsx cu structura EDIFICO exacta:
    - Un sheet per angajat
    - Titlu rosu bold "Raport de activitate <COMPANY> - <luna> <an>"
    - Numele angajatului in B6:E6
    - Saptamani grupate cu activitati concatenate

    Parametri suportati (query string):
    - ?angajat_id=X         exporta doar acel angajat
    - ?luna=YYYY-MM         filtreaza luna (default: luna curenta)
    - ?tip=zilnica|saptamanala|lunara  filtreaza tip activitati incluse
    """
    from openpyxl import Workbook

    today = date.today()

    # === Parametri perioada ===
    # Suporta atat luna_start/luna_end (interval) cat si luna (compat)
    luna_start_param = request.args.get('luna_start', '').strip()
    luna_end_param = request.args.get('luna_end', '').strip()
    luna_param = request.args.get('luna', '').strip()  # legacy

    if not luna_start_param and luna_param:
        luna_start_param = luna_param
    if not luna_end_param and luna_param:
        luna_end_param = luna_param
    if not luna_start_param:
        luna_start_param = today.strftime('%Y-%m')
    if not luna_end_param:
        luna_end_param = luna_start_param

    def _parse_luna(s):
        try:
            y, m = s.split('-')
            y, m = int(y), int(m)
            if not (1 <= m <= 12):
                raise ValueError
            return y, m
        except (ValueError, AttributeError):
            return None

    p_start = _parse_luna(luna_start_param)
    p_end = _parse_luna(luna_end_param)
    if not p_start or not p_end:
        flash('Format perioada invalid (folositi YYYY-MM).', 'danger')
        return redirect(url_for('activitati.panou'))

    # Asigura ordinea corecta start <= end
    if (p_end[0], p_end[1]) < (p_start[0], p_start[1]):
        p_start, p_end = p_end, p_start

    # Construieste lista de luni in interval
    perioade = []
    cy, cm = p_start
    while (cy, cm) <= (p_end[0], p_end[1]):
        perioade.append((cy, cm))
        cm += 1
        if cm > 12:
            cm = 1
            cy += 1
        if len(perioade) > 36:
            break

    # Pentru compatibilitate cu logica veche (selectia angajatilor cu activitati)
    an = p_start[0]
    luna = p_start[1]

    # Suporta atat ?angajat_id=X cat si ?angajat_id=X&angajat_id=Y (multi-select)
    f_angajat_ids_raw = request.args.getlist('angajat_id')
    f_angajat_ids = []
    for v in f_angajat_ids_raw:
        try:
            n = int(v)
            if n > 0 and n not in f_angajat_ids:
                f_angajat_ids.append(n)
        except (ValueError, TypeError):
            continue
    f_tip = request.args.get('tip', '').strip()  # zilnica/saptamanala/lunara

    # === Determinare angajati de exportat ===
    operator_angajat = _get_angajat_for_user(current_user)
    if current_user.rol == 'operator':
        if not operator_angajat:
            flash('Nu sunteti asociat unui angajat.', 'warning')
            return redirect(url_for('activitati.panou'))
        angajati = [operator_angajat]
    elif f_angajat_ids:
        angajati = Angajat.query.filter(Angajat.id.in_(f_angajat_ids)).order_by(
            Angajat.nume, Angajat.prenume
        ).all()
        if not angajati:
            flash('Niciunul din angajatii selectati nu exista.', 'danger')
            return redirect(url_for('activitati.panou'))
    else:
        # Toti angajatii care au activitati in oricare luna din interval
        from calendar import monthrange
        prima_zi = date(p_start[0], p_start[1], 1)
        ultima_zi = date(p_end[0], p_end[1], monthrange(p_end[0], p_end[1])[1])

        q = db.session.query(RaportActivitate.angajat_id).filter(
            RaportActivitate.data >= prima_zi,
            RaportActivitate.data <= ultima_zi,
        )
        if f_tip in ('zilnica', 'saptamanala', 'lunara'):
            q = q.filter(RaportActivitate.tip_activitate == f_tip)
        angajati_ids = [r[0] for r in q.distinct().all() if r[0]]
        if angajati_ids:
            angajati = Angajat.query.filter(Angajat.id.in_(angajati_ids)).order_by(
                Angajat.nume, Angajat.prenume
            ).all()
        else:
            # Fallback: toti angajatii activi
            angajati = Angajat.query.filter_by(status='activ').order_by(
                Angajat.nume, Angajat.prenume
            ).all()

    if not angajati:
        flash('Niciun angajat de exportat.', 'warning')
        return redirect(url_for('activitati.panou'))

    # === Construire workbook ===
    company_short, _ = _get_company_name()
    wb = Workbook()
    # Sterge sheet-ul implicit; il vom recrea cu primul angajat
    default_ws = wb.active
    default_ws.title = 'Sheet'

    for idx, ang in enumerate(angajati):
        _construieste_sheet_angajat(wb, ang, perioade, company_short, sheet_index=idx)

    # === Salvare si raspuns ===
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Numele fisierului
    if p_start == p_end:
        perioada_text = f'{LUNI_RO[p_start[1]]}_{p_start[0]}'
    else:
        perioada_text = f'{LUNI_RO[p_start[1]]}_{p_start[0]}_-_{LUNI_RO[p_end[1]]}_{p_end[0]}'

    if len(angajati) == 1:
        ang = angajati[0]
        nume_curat = f'{ang.nume}_{ang.prenume}'.replace(' ', '_')
        filename = f'Raport_activitate_{nume_curat}_{perioada_text}.xlsx'
    else:
        filename = f'Raport_activitate_{company_short}_{perioada_text}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
