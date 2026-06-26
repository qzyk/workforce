"""
Rute pentru gestionarea proiectelor - Modul Complet
"""

import os
import uuid
from datetime import datetime, date, timedelta
from decimal import Decimal
from flask import Blueprint, render_template, redirect, url_for, flash, request, jsonify, send_file, current_app, abort
from flask_login import login_required, current_user
from werkzeug.utils import secure_filename
from models import db, Proiect, Angajat, AngajatProiect, Pontaj, Document, Utilizator
from forms.proiecte_forms import ProiectForm
from services.security.tenant_access import (
    get_legacy_document_or_404,
    get_project_or_404,
    query_for_tenant,
    query_legacy_documents_for_tenant,
    query_project_documents_for_tenant,
    tenant_id_for_new_record_or_403,
)

ALLOWED_EXT_PROIECT = {'pdf', 'dwg', 'dxf', 'docx', 'xlsx', 'jpg', 'jpeg', 'png', 'zip'}

proiecte_bp = Blueprint('proiecte', __name__)


# ============================================================
# LISTA PROIECTE (cu paginare, filtre, stats)
# ============================================================

@proiecte_bp.route('/')
@login_required
def lista():
    page = request.args.get('page', 1, type=int)
    status_filtru = request.args.get('status', '')
    cautare = request.args.get('cautare', '').strip()
    manager_filtru = request.args.get('manager', '', type=str)
    sort = request.args.get('sort', 'data_start_desc')

    query = query_for_tenant(Proiect)

    if status_filtru:
        query = query.filter_by(status=status_filtru)
    if cautare:
        query = query.filter(
            db.or_(
                Proiect.nume.ilike(f'%{cautare}%'),
                Proiect.cod_proiect.ilike(f'%{cautare}%'),
                Proiect.beneficiar.ilike(f'%{cautare}%')
            )
        )
    if manager_filtru:
        query = query.filter_by(manager_id=int(manager_filtru))

    # Sortare
    if sort == 'nume_asc':
        query = query.order_by(Proiect.nume.asc())
    elif sort == 'nume_desc':
        query = query.order_by(Proiect.nume.desc())
    elif sort == 'data_start_asc':
        query = query.order_by(Proiect.data_start.asc())
    elif sort == 'buget_desc':
        query = query.order_by(Proiect.buget_total.desc().nullslast())
    elif sort == 'status':
        query = query.order_by(Proiect.status.asc())
    else:
        query = query.order_by(Proiect.data_start.desc())

    pagination = query.paginate(page=page, per_page=12, error_out=False)
    proiecte = pagination.items

    # Statistici
    stats_query = query_for_tenant(Proiect)
    total_active = stats_query.filter_by(status='activ').count()
    total_planificate = stats_query.filter_by(status='planificat').count()
    total_finalizate = stats_query.filter_by(status='finalizat').count()
    total_suspendate = stats_query.filter_by(status='suspendat').count()
    buget_total_all = stats_query.with_entities(db.func.sum(Proiect.buget_total)).filter(
        Proiect.status.in_(['activ', 'planificat'])
    ).scalar() or 0

    # Manager list for filter
    manageri = Utilizator.query.filter(
        Utilizator.rol.in_(['admin', 'manager']),
        Utilizator.activ == True
    ).all()

    view_mode = request.args.get('view', 'cards')

    return render_template('proiecte/lista.html',
                           proiecte=proiecte,
                           pagination=pagination,
                           status_filtru=status_filtru,
                           cautare=cautare,
                           manager_filtru=manager_filtru,
                           sort=sort,
                           view_mode=view_mode,
                           total_active=total_active,
                           total_planificate=total_planificate,
                           total_finalizate=total_finalizate,
                           total_suspendate=total_suspendate,
                           buget_total_all=buget_total_all,
                           manageri=manageri)


# ============================================================
# ADAUGA PROIECT
# ============================================================

@proiecte_bp.route('/adauga', methods=['GET', 'POST'])
@login_required
def adauga():
    tenant_id_nou = tenant_id_for_new_record_or_403()
    form = ProiectForm()

    if form.validate_on_submit():
        locatie = ''
        if form.judet.data:
            locatie = form.judet.data
            if form.localitate.data:
                locatie = f"{form.localitate.data}, {form.judet.data}"

        proiect = Proiect(
            cod_proiect=form.cod_proiect.data.strip(),
            nume=form.nume.data.strip(),
            descriere=form.descriere.data or '',
            locatie=locatie,
            adresa_santier=form.adresa_santier.data or '',
            beneficiar=form.beneficiar.data or '',
            nr_contract_beneficiar=form.nr_contract_beneficiar.data or '',
            data_start=form.data_start.data,
            data_sfarsit_planificat=form.data_sfarsit_planificat.data,
            status=form.status.data,
            manager_id=form.manager_id.data if form.manager_id.data else None,
            buget_total=form.buget_total.data,
            buget_manopera=form.buget_manopera.data,
            tenant_id=tenant_id_nou,
        )
        db.session.add(proiect)
        db.session.commit()
        flash(f'Proiectul {proiect.cod_proiect} a fost creat cu succes!', 'success')
        return redirect(url_for('proiecte.detalii', id=proiect.id))

    # Auto-generate cod_proiect for new projects
    if not form.cod_proiect.data:
        year = date.today().year
        count = Proiect.query.filter(
            Proiect.cod_proiect.like(f'PRJ-{year}-%')
        ).count()
        form.cod_proiect.data = f'PRJ-{year}-{count + 1:03d}'

    return render_template('proiecte/formular.html', form=form, proiect=None)


# ============================================================
# DETALII PROIECT (5 tab-uri)
# ============================================================

@proiecte_bp.route('/<int:id>')
@login_required
def detalii(id):
    proiect = get_project_or_404(id)

    # Tab Echipa
    angajati_asoc = AngajatProiect.query.filter_by(proiect_id=id).order_by(
        AngajatProiect.data_sfarsit.asc().nullsfirst(),
        AngajatProiect.data_start.desc()
    ).all()

    # Activ = data_sfarsit IS NULL. Cand se dezaloca, set data_sfarsit = today,
    # deci dispare imediat din "activi". Filter simplu si clar.
    angajati_activi = [a for a in angajati_asoc if not a.data_sfarsit]
    angajati_disponibili = Angajat.query.filter_by(status='activ').order_by(Angajat.nume).all()

    # Distributie functii pentru chart
    dist_functii = {}
    for ap in angajati_activi:
        fn = ap.functie_pe_proiect or ap.angajat.functie or 'Necunoscut'
        dist_functii[fn] = dist_functii.get(fn, 0) + 1

    # Tab Pontaje
    luna = request.args.get('luna', date.today().month, type=int)
    anul = request.args.get('anul', date.today().year, type=int)

    pontaje = Pontaj.query.filter(
        Pontaj.proiect_id == id,
        db.extract('month', Pontaj.data) == luna,
        db.extract('year', Pontaj.data) == anul
    ).order_by(Pontaj.data.desc()).all()

    total_ore = proiect.get_total_ore()

    # Ore per angajat (luna curenta)
    ore_per_angajat = db.session.query(
        Angajat.nume, Angajat.prenume,
        db.func.sum(Pontaj.ore_lucrate).label('total_ore')
    ).join(Pontaj, Angajat.id == Pontaj.angajat_id).filter(
        Pontaj.proiect_id == id,
        db.extract('month', Pontaj.data) == luna,
        db.extract('year', Pontaj.data) == anul
    ).group_by(Angajat.id).all()

    # Ore saptamanale (ultimele 12 saptamani)
    ore_saptamanale = _get_ore_saptamanale(id)

    # Tab Financiar
    cost_manopera = _calculeaza_cost_manopera(id)
    cost_lunar = _get_cost_lunar(id)

    # Tab Documente
    documente = query_legacy_documents_for_tenant().filter_by(proiect_id=id).order_by(
        Document.data_upload.desc()
    ).all()

    return render_template('proiecte/detalii.html',
                           proiect=proiect,
                           angajati_asoc=angajati_asoc,
                           angajati_activi=angajati_activi,
                           angajati_disponibili=angajati_disponibili,
                           dist_functii=dist_functii,
                           pontaje=pontaje,
                           total_ore=total_ore,
                           ore_per_angajat=ore_per_angajat,
                           ore_saptamanale=ore_saptamanale,
                           cost_manopera=cost_manopera,
                           cost_lunar=cost_lunar,
                           documente=documente,
                           luna=luna,
                           anul=anul)


# ============================================================
# EDITEAZA PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/hub')
@login_required
def hub(id):
    """Proiect 360: agrega cross-modul (BIM, contracte, oferte, planuri Gantt,
    situatii, locatie, angajati, documente) pe FK-urile existente."""
    from sqlalchemy import func
    proiect = get_project_or_404(id)

    def _safe(fn, dflt=None):
        try:
            return fn()
        except Exception:
            return dflt

    def _suma(model, camp):
        return float(db.session.query(func.coalesce(func.sum(getattr(model, camp)), 0))
                     .filter(model.proiect_id == id).scalar() or 0)

    from models import (Contract, OfertaContract, SituatieLunara, LocatieProiect,
                        GanttPlan, DocumentProiect, ModelBIM, Cladire, ElementBIM,
                        Santier, ProiectSantier)

    h = {}
    h['contracte'] = _safe(lambda: {'nr': Contract.query.filter_by(proiect_id=id).count(),
                                    'valoare': _suma(Contract, 'valoare_totala')}, {'nr': 0, 'valoare': 0})
    h['oferte'] = _safe(lambda: {'nr': OfertaContract.query.filter_by(proiect_id=id).count(),
                                 'valoare': _suma(OfertaContract, 'valoare_totala')}, {'nr': 0, 'valoare': 0})
    h['situatie'] = _safe(lambda: (lambda s: {'procent': float(s.procent_avans_total or 0),
                                              'cumulat': float(s.valoare_cumulat_la_zi or 0)} if s else None)(
        SituatieLunara.query.filter_by(proiect_id=id).order_by(SituatieLunara.id.desc()).first()))
    h['gantt'] = _safe(lambda: {'nr': GanttPlan.query.filter_by(proiect_id=id).count(),
                                'cost': _suma(GanttPlan, 'cost_total')}, {'nr': 0, 'cost': 0})
    h['locatie'] = _safe(lambda: (lambda l: {'lat': float(l.latitudine), 'lng': float(l.longitudine)}
                                  if l and l.latitudine is not None else None)(
        LocatieProiect.query.filter_by(proiect_id=id).first()))
    h['angajati'] = _safe(lambda: AngajatProiect.query.filter_by(proiect_id=id)
                          .filter(AngajatProiect.data_sfarsit.is_(None)).count(), 0)
    h['documente'] = _safe(lambda: query_project_documents_for_tenant().filter_by(
        proiect_id=id
    ).count(), 0)

    # santiere BIM legate (many-to-many) -> agregam BIM peste toate
    santiere = _safe(lambda: [ls.santier for ls in proiect.legaturi_santiere
                              if ls.santier is not None], []) or []
    if santiere:
        sids = [s.id for s in santiere]
        h['bim'] = _safe(lambda: {
            'nr_santiere': len(santiere),
            'modele': ModelBIM.query.filter(ModelBIM.santier_id.in_(sids)).count(),
            'elemente': (ElementBIM.query.join(Cladire, ElementBIM.cladire_id == Cladire.id)
                         .filter(Cladire.santier_id.in_(sids)).count()),
        }, {'nr_santiere': len(santiere), 'modele': 0, 'elemente': 0})
    # santiere disponibile pentru legare (neasociate inca)
    legate_ids = {s.id for s in santiere}
    disponibile = _safe(lambda: [s for s in Santier.query.order_by(Santier.cod).all()
                                 if s.id not in legate_ids], []) or []

    # ---- Parcurs ghidat (U2): etapele fluxului cu stare + pasul urmator ----
    from models import GanttWbsNod, ConsumUtilaj
    plan = _safe(lambda: GanttPlan.query.filter_by(proiect_id=id)
                 .order_by(GanttPlan.data_creare.desc()).first())
    are_wbs = _safe(lambda: db.session.query(GanttWbsNod.id)
                    .join(GanttPlan, GanttWbsNod.plan_id == GanttPlan.id)
                    .filter(GanttPlan.proiect_id == id).first() is not None, False)
    are_utilaj = _safe(lambda: ConsumUtilaj.query.filter_by(proiect_id=id).first() is not None, False)
    contracte_on = _safe(lambda: __import__('services.feature_flags', fromlist=['is_enabled'])
                         .is_enabled('controale-contract',
                                     tenant_id=getattr(current_user, 'tenant_id', None)), False)
    u_plan = (lambda f, d: url_for(f, id_=plan.id) if plan else url_for(d))
    parcurs = [
        {'nume': 'Deviz / oferta', 'icon': 'calculator', 'gata': h['oferte']['nr'] > 0,
         'hint': 'Pretuieste lista de cantitati (F3)',
         'url': url_for('contracte.lista') if contracte_on else url_for('proiecte.detalii', id=id)},
        {'nume': 'Plan Gantt', 'icon': 'diagram-project', 'gata': h['gantt']['nr'] > 0,
         'hint': 'Genereaza planul din F3', 'url': u_plan('gantt.plan', 'gantt.index')},
        {'nume': 'WBS', 'icon': 'sitemap', 'gata': are_wbs,
         'hint': 'Structureaza / editeaza WBS-ul', 'url': u_plan('gantt.plan_wbs', 'gantt.index')},
        {'nume': 'Cost 5D', 'icon': 'coins', 'gata': (h['gantt']['cost'] or 0) > 0,
         'hint': 'Cost pe categorii + resurse', 'url': u_plan('gantt.plan', 'gantt.index')},
        {'nume': 'Utilaje', 'icon': 'truck-arrow-right', 'gata': are_utilaj,
         'hint': 'Consum real de utilaj', 'url': url_for('proiecte.utilaje', id=id)},
        {'nume': 'EVM / avans', 'icon': 'chart-line', 'gata': h['situatie'] is not None,
         'hint': 'Planificat vs realizat (SPI/CPI)', 'url': url_for('proiecte.evm', id=id)},
    ]
    next_idx = next((i for i, e in enumerate(parcurs) if not e['gata']), None)

    return render_template('proiecte/hub.html', proiect=proiect, h=h,
                           santiere=santiere, santiere_disponibile=disponibile,
                           parcurs=parcurs, next_idx=next_idx)


@proiecte_bp.route('/<int:id>/leaga-santier', methods=['POST'])
@login_required
def leaga_santier(id):
    from models import ProiectSantier, Santier
    Proiect.query.get_or_404(id)
    try:
        sid = int(request.form['santier_id'])
    except (KeyError, ValueError, TypeError):
        flash('Alege un santier BIM.', 'warning')
        return redirect(url_for('proiecte.hub', id=id))
    if not Santier.query.get(sid):
        abort(404)
    if not ProiectSantier.query.filter_by(proiect_id=id, santier_id=sid).first():
        ps = ProiectSantier(proiect_id=id, santier_id=sid,
                            tenant_id=getattr(current_user, 'tenant_id', None),
                            creat_de_id=getattr(current_user, 'id', None))
        db.session.add(ps)
        db.session.commit()
        try:
            from services import audit
            audit.log('create', 'proiect_santier', ps.id,
                      new_values={'proiect_id': id, 'santier_id': sid}, commit=True)
        except Exception:
            pass
        flash('Santier BIM legat de proiect.', 'success')
    return redirect(url_for('proiecte.hub', id=id))


@proiecte_bp.route('/<int:id>/dezleaga-santier/<int:santier_id>', methods=['POST'])
@login_required
def dezleaga_santier(id, santier_id):
    from models import ProiectSantier
    ps = ProiectSantier.query.filter_by(proiect_id=id, santier_id=santier_id).first()
    if ps:
        psid = ps.id
        db.session.delete(ps)
        db.session.commit()
        try:
            from services import audit
            audit.log('delete', 'proiect_santier', psid, commit=True)
        except Exception:
            pass
        flash('Santier dezlegat.', 'success')
    return redirect(url_for('proiecte.hub', id=id))


@proiecte_bp.route('/<int:id>/evm')
@login_required
def evm(id):
    """Earned Value: plan (curba S Gantt) vs realizat (situatii lunare) -> SPI/CPI."""
    from services.evm import evm_proiect
    proiect = Proiect.query.get_or_404(id)
    data = None
    try:
        data = evm_proiect(id, getattr(current_user, 'tenant_id', None))
    except Exception:
        data = None
    return render_template('proiecte/evm.html', proiect=proiect, evm=data)


@proiecte_bp.route('/<int:id>/utilaje')
@login_required
def utilaje(id):
    """Consum real de utilaj pe proiect (Faza 3 - C) + comparatie cu planificatul."""
    from models import ConsumUtilaj, Masina
    from services.evm import evm_proiect
    proiect = Proiect.query.get_or_404(id)
    randuri = (ConsumUtilaj.query.filter_by(proiect_id=id)
               .order_by(ConsumUtilaj.data.desc(), ConsumUtilaj.id.desc()).all())
    masini = Masina.query.order_by(Masina.numar_inmatriculare).all()
    real_total = sum(r.calc_cost() for r in randuri)
    planificat = 0.0
    try:
        ev = evm_proiect(id, getattr(current_user, 'tenant_id', None))
        if ev and ev.get('utilaj'):
            planificat = ev['utilaj']['planificat']
    except Exception:
        pass
    return render_template('proiecte/utilaje.html', proiect=proiect, randuri=randuri,
                           masini=masini, real_total=real_total, planificat=planificat,
                           today=date.today())


@proiecte_bp.route('/<int:id>/utilaje/adauga', methods=['POST'])
@login_required
def utilaje_adauga(id):
    from models import ConsumUtilaj
    Proiect.query.get_or_404(id)
    denumire = (request.form.get('denumire') or '').strip()
    if not denumire:
        flash('Completeaza denumirea utilajului.', 'danger')
        return redirect(url_for('proiecte.utilaje', id=id))

    def _f(camp):
        try:
            return float((request.form.get(camp) or '0').replace(',', '.'))
        except ValueError:
            return 0.0
    ore, tarif = _f('ore'), _f('tarif_ora')
    cost = _f('cost') or (ore * tarif)
    d = date.today()
    ds = (request.form.get('data') or '').strip()
    if ds:
        try:
            d = date.fromisoformat(ds)
        except ValueError:
            pass
    try:
        mid = int(request.form.get('masina_id') or 0) or None
    except ValueError:
        mid = None
    db.session.add(ConsumUtilaj(
        proiect_id=id, masina_id=mid, denumire=denumire[:150], data=d,
        ore=ore, tarif_ora=tarif, cost=cost,
        categorie_lucrare=(request.form.get('categorie_lucrare') or '').strip()[:60] or None,
        observatii=(request.form.get('observatii') or '').strip() or None,
        introdus_de=getattr(current_user, 'id', None),
        tenant_id=getattr(current_user, 'tenant_id', None)))
    db.session.commit()
    flash(f'Consum utilaj adaugat: {denumire} ({cost:,.0f} lei).', 'success')
    return redirect(url_for('proiecte.utilaje', id=id))


@proiecte_bp.route('/<int:id>/utilaje/<int:cid>/sterge', methods=['POST'])
@login_required
def utilaje_sterge(id, cid):
    from models import ConsumUtilaj
    r = db.session.get(ConsumUtilaj, cid)
    if r and r.proiect_id == id:
        db.session.delete(r)
        db.session.commit()
        flash('Consum sters.', 'success')
    return redirect(url_for('proiecte.utilaje', id=id))


# ---- Extrase de resurse C6/C7/C8 (Faza 2) ----
@proiecte_bp.route('/<int:id>/resurse')
@login_required
def resurse(id):
    """Extrase de resurse din deviz (C6 materiale / C7 manopera / C8 utilaje)."""
    from models import ExtrasResursa
    proiect = Proiect.query.get_or_404(id)
    grupuri = {'material': [], 'manopera': [], 'utilaj': []}
    for e in (ExtrasResursa.query.filter_by(proiect_id=id)
              .order_by(ExtrasResursa.valoare.desc()).all()):
        grupuri.setdefault(e.tip, []).append(e)
    totaluri = {t: sum(float(x.valoare or 0) for x in lst) for t, lst in grupuri.items()}
    ore = {t: sum(float(x.cantitate or 0) for x in grupuri.get(t, []))
           for t in ('manopera', 'utilaj')}
    # D: reconciliere F3 (plan) vs extrase. C: furnizori (aprovizionare)
    from services.deviz_extras import reconciliere
    recon = reconciliere(id)
    furnizori = {}
    for m in grupuri.get('material', []):
        f = (m.furnizor or '—')
        furnizori[f] = round(furnizori.get(f, 0.0) + float(m.valoare or 0), 2)
    return render_template('proiecte/resurse.html', proiect=proiect,
                           grupuri=grupuri, totaluri=totaluri, ore=ore,
                           recon=recon, furnizori=furnizori)


@proiecte_bp.route('/<int:id>/resurse/conexiune')
@login_required
def resurse_conexiune(id):
    """Conexiunea reala F3 <-> C pe cod de resursa: reconciliere la nivel de articol,
    necesar esalonat in timp (per resursa) si drill-down resursa -> activitati."""
    from services.deviz_extras import legatura_resurse
    proiect = Proiect.query.get_or_404(id)
    leg = legatura_resurse(id)
    sumar = {'ok': 0, 'atentie': 0, 'critic': 0, 'lipsa': 0}
    for r in leg.get('resurse', []):
        sumar[r['status']] = sumar.get(r['status'], 0) + 1
    return render_template('proiecte/legatura_resurse.html', proiect=proiect,
                           leg=leg, sumar=sumar)


@proiecte_bp.route('/<int:id>/resurse/aprovizionare.csv')
@login_required
def resurse_aprovizionare_csv(id):
    """Necesar materiale (C6) pentru comanda - CSV grupat pe furnizor."""
    import csv as _csv
    from io import StringIO, BytesIO
    from models import ExtrasResursa
    Proiect.query.get_or_404(id)
    mat = (ExtrasResursa.query.filter_by(proiect_id=id, tip='material')
           .order_by(ExtrasResursa.furnizor, ExtrasResursa.denumire).all())
    out = StringIO()
    w = _csv.writer(out, delimiter=';')
    w.writerow(['Furnizor', 'Cod', 'Denumire', 'UM', 'Cantitate', 'Pret unitar', 'Valoare'])
    for m in mat:
        w.writerow([m.furnizor or '', m.cod or '', m.denumire, m.um or '',
                    float(m.cantitate or 0), float(m.tarif_unitar or 0), float(m.valoare or 0)])
    return send_file(BytesIO(out.getvalue().encode('utf-8-sig')), mimetype='text/csv',
                     as_attachment=True, download_name=f'necesar_materiale_proiect_{id}.csv')


@proiecte_bp.route('/<int:id>/resurse/upload', methods=['POST'])
@login_required
def resurse_upload(id):
    from models import ExtrasResursa
    from services.deviz_extras import parse_extras
    Proiect.query.get_or_404(id)
    fisier = request.files.get('fisier')
    if not fisier or not fisier.filename:
        flash('Selecteaza un fisier C6/C7/C8 (.xls/.xlsx).', 'warning')
        return redirect(url_for('proiecte.resurse', id=id))
    ext = os.path.splitext(fisier.filename)[1].lower()
    try:
        tip, rows = parse_extras(fisier.read(), ext)
    except Exception as e:
        flash(f'Nu pot citi fisierul: {e}', 'danger')
        return redirect(url_for('proiecte.resurse', id=id))
    if not tip or not rows:
        flash('Fisierul nu pare un extras C6/C7/C8 valid (Formular C6/C7/C8).', 'warning')
        return redirect(url_for('proiecte.resurse', id=id))
    # re-import = inlocuieste extrasul de acelasi tip
    ExtrasResursa.query.filter_by(proiect_id=id, tip=tip).delete()
    for r in rows:
        db.session.add(ExtrasResursa(
            proiect_id=id, tip=tip, cod=(r['cod'][:60] or None), denumire=r['denumire'],
            um=r['um'], cantitate=r['cantitate'], tarif_unitar=r['tarif_unitar'],
            valoare=r['valoare'], furnizor=r['furnizor'],
            nume_fisier=fisier.filename[:255],
            introdus_de=getattr(current_user, 'id', None),
            tenant_id=getattr(current_user, 'tenant_id', None)))
    db.session.commit()
    etichete = {'material': 'C6 materiale', 'manopera': 'C7 manopera', 'utilaj': 'C8 utilaje'}
    flash(f'{etichete.get(tip, tip)}: {len(rows)} resurse importate.', 'success')
    return redirect(url_for('proiecte.resurse', id=id))


@proiecte_bp.route('/<int:id>/resurse/sterge/<tip>', methods=['POST'])
@login_required
def resurse_sterge(id, tip):
    from models import ExtrasResursa
    n = ExtrasResursa.query.filter_by(proiect_id=id, tip=tip).delete()
    db.session.commit()
    flash(f'{n} resurse sterse.', 'success')
    return redirect(url_for('proiecte.resurse', id=id))


@proiecte_bp.route('/<int:id>/bim-deviz')
@login_required
def bim_deviz(id):
    """Puntea BIM (model 3D) <-> deviz F3 <-> resurse C, pe categorie de lucrare:
    reconciliere cantitati model vs deviz + resursele implicate."""
    from services.legatura_bim import legatura_bim
    proiect = Proiect.query.get_or_404(id)
    leg = legatura_bim(id)
    return render_template('proiecte/bim_deviz.html', proiect=proiect, leg=leg)


@proiecte_bp.route('/<int:id>/editeaza', methods=['GET', 'POST'])
@login_required
def editeaza(id):
    proiect = get_project_or_404(id)
    form = ProiectForm(obj=proiect)

    if form.validate_on_submit():
        locatie = ''
        if form.judet.data:
            locatie = form.judet.data
            if form.localitate.data:
                locatie = f"{form.localitate.data}, {form.judet.data}"

        proiect.cod_proiect = form.cod_proiect.data.strip()
        proiect.nume = form.nume.data.strip()
        proiect.descriere = form.descriere.data or ''
        proiect.locatie = locatie
        proiect.adresa_santier = form.adresa_santier.data or ''
        proiect.beneficiar = form.beneficiar.data or ''
        proiect.nr_contract_beneficiar = form.nr_contract_beneficiar.data or ''
        proiect.data_start = form.data_start.data
        proiect.data_sfarsit_planificat = form.data_sfarsit_planificat.data
        proiect.data_sfarsit_real = form.data_sfarsit_real.data
        proiect.status = form.status.data
        proiect.manager_id = form.manager_id.data if form.manager_id.data else None
        proiect.buget_total = form.buget_total.data
        proiect.buget_manopera = form.buget_manopera.data

        db.session.commit()
        flash('Proiectul a fost actualizat cu succes!', 'success')
        return redirect(url_for('proiecte.detalii', id=id))

    # Pre-populate judet/localitate from locatie
    if request.method == 'GET' and proiect.locatie:
        parts = proiect.locatie.split(', ')
        if len(parts) == 2:
            form.localitate.data = parts[0]
            form.judet.data = parts[1]
        elif len(parts) == 1:
            form.judet.data = parts[0]

    form.proiect_id.data = proiect.id

    return render_template('proiecte/formular.html', form=form, proiect=proiect)


# ============================================================
# SCHIMBA STATUS (AJAX)
# ============================================================

@proiecte_bp.route('/<int:id>/schimba-status', methods=['POST'])
@login_required
def schimba_status(id):
    proiect = get_project_or_404(id)
    data = request.get_json()
    new_status = data.get('status', '')

    valid_statuses = [s[0] for s in Proiect.STATUSURI]
    if new_status not in valid_statuses:
        return jsonify({'success': False, 'error': 'Status invalid'}), 400

    proiect.status = new_status
    if new_status == 'finalizat' and not proiect.data_sfarsit_real:
        proiect.data_sfarsit_real = date.today()

    db.session.commit()
    return jsonify({'success': True, 'status': new_status})


# ============================================================
# ADAUGA ANGAJAT PE PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/adauga-angajat', methods=['POST'])
@login_required
def adauga_angajat(id):
    proiect = Proiect.query.get_or_404(id)
    angajat_id = int(request.form['angajat_id'])
    functie = request.form.get('functie_pe_proiect', '')
    tarif = float(request.form['tarif_negociat']) if request.form.get('tarif_negociat') else None

    exista = AngajatProiect.query.filter_by(
        angajat_id=angajat_id, proiect_id=id
    ).filter(AngajatProiect.data_sfarsit.is_(None)).first()

    if exista:
        flash('Angajatul este deja asignat pe acest proiect.', 'warning')
    else:
        ap = AngajatProiect(
            angajat_id=angajat_id, proiect_id=id,
            data_start=date.today(),
            functie_pe_proiect=functie,
            tarif_negociat=tarif
        )
        db.session.add(ap)
        db.session.commit()
        flash('Angajat adaugat pe proiect cu succes!', 'success')

    return redirect(url_for('proiecte.detalii', id=id))


# ============================================================
# ELIMINA ANGAJAT DE PE PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/elimina-angajat/<int:ap_id>', methods=['POST'])
@login_required
def elimina_angajat(id, ap_id):
    """Dezalocare soft: seteaza data_sfarsit = today, pastreaza istoricul."""
    ap = AngajatProiect.query.get_or_404(ap_id)
    if ap.proiect_id != id:
        flash('Asociere invalida.', 'danger')
        return redirect(url_for('proiecte.detalii', id=id))

    nume = ap.angajat.nume_complet
    ap.data_sfarsit = date.today()
    db.session.commit()
    flash(f'Angajatul {nume} a fost dezalocat de pe proiect. '
          f'Apare in tab Istoric si poate fi re-alocat oricand.', 'success')
    return redirect(url_for('proiecte.detalii', id=id))


@proiecte_bp.route('/<int:id>/realoca-angajat/<int:ap_id>', methods=['POST'])
@login_required
def realoca_angajat(id, ap_id):
    """Re-activare asociere dezalocata: data_sfarsit = NULL, data_start = today."""
    ap = AngajatProiect.query.get_or_404(ap_id)
    if ap.proiect_id != id:
        flash('Asociere invalida.', 'danger')
        return redirect(url_for('proiecte.detalii', id=id))

    if not ap.data_sfarsit:
        flash(f'Angajatul {ap.angajat.nume_complet} e deja activ pe proiect.', 'info')
        return redirect(url_for('proiecte.detalii', id=id))

    ap.data_sfarsit = None
    ap.data_start = date.today()
    db.session.commit()
    flash(f'Angajatul {ap.angajat.nume_complet} a fost re-alocat pe proiect.', 'success')
    return redirect(url_for('proiecte.detalii', id=id))


@proiecte_bp.route('/<int:id>/sterge-asignare/<int:ap_id>', methods=['POST'])
@login_required
def sterge_asignare(id, ap_id):
    """
    Stergere definitiva a asocierii angajat-proiect din DB.
    NU sterge angajatul, doar randul din AngajatProiect.
    Permis doar de manager / admin (asociere e istoric audit-relevant).
    """
    if current_user.rol not in ('admin', 'manager'):
        flash('Doar admin / manager poate sterge definitiv o asociere.', 'danger')
        return redirect(url_for('proiecte.detalii', id=id))

    ap = AngajatProiect.query.get_or_404(ap_id)
    if ap.proiect_id != id:
        flash('Asociere invalida.', 'danger')
        return redirect(url_for('proiecte.detalii', id=id))

    nume = ap.angajat.nume_complet
    db.session.delete(ap)
    db.session.commit()
    flash(f'Asocierea cu {nume} a fost stearsa definitiv din istoric.', 'info')
    return redirect(url_for('proiecte.detalii', id=id))


# ============================================================
# RAPORT PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/raport')
@login_required
def raport(id):
    proiect = Proiect.query.get_or_404(id)

    angajati_asoc = AngajatProiect.query.filter_by(proiect_id=id).all()
    # Activ = data_sfarsit IS NULL. Cand se dezaloca, set data_sfarsit = today,
    # deci dispare imediat din "activi". Filter simplu si clar.
    angajati_activi = [a for a in angajati_asoc if not a.data_sfarsit]

    total_ore = proiect.get_total_ore()
    cost_manopera = _calculeaza_cost_manopera(id)
    cost_lunar = _get_cost_lunar(id)

    # Ore per angajat (toate lunile)
    ore_per_angajat = db.session.query(
        Angajat.nume, Angajat.prenume,
        db.func.sum(Pontaj.ore_lucrate).label('total_ore'),
        db.func.sum(Pontaj.ore_suplimentare_50).label('ore_supl_50'),
        db.func.sum(Pontaj.ore_suplimentare_100).label('ore_supl_100')
    ).join(Pontaj, Angajat.id == Pontaj.angajat_id).filter(
        Pontaj.proiect_id == id
    ).group_by(Angajat.id).all()

    return render_template('proiecte/raport.html',
                           proiect=proiect,
                           angajati_asoc=angajati_asoc,
                           angajati_activi=angajati_activi,
                           total_ore=total_ore,
                           cost_manopera=cost_manopera,
                           cost_lunar=cost_lunar,
                           ore_per_angajat=ore_per_angajat)


# ============================================================
# EXPORT EXCEL
# ============================================================

@proiecte_bp.route('/<int:id>/export-excel')
@login_required
def export_excel(id):
    import io
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    proiect = get_project_or_404(id)

    wb = Workbook()
    header_font = Font(bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1a237e', end_color='1a237e', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # --- Sheet 1: Informatii proiect ---
    ws1 = wb.active
    ws1.title = 'Informatii Proiect'
    info_data = [
        ['Cod Proiect', proiect.cod_proiect],
        ['Nume', proiect.nume],
        ['Status', proiect.status.capitalize()],
        ['Beneficiar', proiect.beneficiar or '-'],
        ['Locatie', proiect.locatie or '-'],
        ['Adresa Santier', proiect.adresa_santier or '-'],
        ['Nr. Contract', proiect.nr_contract_beneficiar or '-'],
        ['Manager', proiect.manager.get_full_name() if proiect.manager else '-'],
        ['Data Start', proiect.data_start.strftime('%d.%m.%Y') if proiect.data_start else '-'],
        ['Data Sfarsit Planificat', proiect.data_sfarsit_planificat.strftime('%d.%m.%Y') if proiect.data_sfarsit_planificat else '-'],
        ['Buget Total (RON)', float(proiect.buget_total) if proiect.buget_total else 0],
        ['Buget Manopera (RON)', float(proiect.buget_manopera) if proiect.buget_manopera else 0],
        ['Total Ore Lucrate', proiect.get_total_ore()],
        ['Descriere', proiect.descriere or '-'],
    ]
    for row_idx, (label, value) in enumerate(info_data, 1):
        ws1.cell(row=row_idx, column=1, value=label).font = Font(bold=True)
        ws1.cell(row=row_idx, column=2, value=value)
    ws1.column_dimensions['A'].width = 25
    ws1.column_dimensions['B'].width = 50

    # --- Sheet 2: Echipa ---
    ws2 = wb.create_sheet('Echipa')
    echipa_headers = ['Nr.', 'Angajat', 'Functie pe Proiect', 'Tarif (RON/ora)', 'Data Start', 'Data Sfarsit', 'Status']
    for col, header in enumerate(echipa_headers, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    angajati_asoc = AngajatProiect.query.filter_by(proiect_id=id).all()
    for idx, ap in enumerate(angajati_asoc, 1):
        activ = not ap.data_sfarsit or ap.data_sfarsit >= date.today()
        row = [
            idx,
            ap.angajat.nume_complet,
            ap.functie_pe_proiect or '-',
            float(ap.tarif_negociat) if ap.tarif_negociat else '-',
            ap.data_start.strftime('%d.%m.%Y') if ap.data_start else '-',
            ap.data_sfarsit.strftime('%d.%m.%Y') if ap.data_sfarsit else '-',
            'Activ' if activ else 'Inactiv'
        ]
        for col, val in enumerate(row, 1):
            cell = ws2.cell(row=idx + 1, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(echipa_headers) + 1):
        ws2.column_dimensions[chr(64 + col)].width = 20
    ws2.auto_filter.ref = f'A1:G{len(angajati_asoc) + 1}'
    ws2.freeze_panes = 'A2'

    # --- Sheet 3: Pontaje ---
    ws3 = wb.create_sheet('Pontaje')
    pontaj_headers = ['Data', 'Angajat', 'Ora Start', 'Ora Sfarsit', 'Ore Lucrate', 'Ore Normale', 'Ore Supl. 50%', 'Ore Supl. 100%', 'Tip Zi', 'Status']
    for col, header in enumerate(pontaj_headers, 1):
        cell = ws3.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thin_border

    pontaje = Pontaj.query.filter_by(proiect_id=id).order_by(Pontaj.data.desc()).all()
    for idx, p in enumerate(pontaje, 1):
        row = [
            p.data.strftime('%d.%m.%Y') if p.data else '-',
            p.angajat.nume_complet if p.angajat else '-',
            p.ora_start or '-',
            p.ora_sfarsit or '-',
            float(p.ore_lucrate) if p.ore_lucrate else 0,
            float(p.ore_normale) if p.ore_normale else 0,
            float(p.ore_suplimentare_50) if p.ore_suplimentare_50 else 0,
            float(p.ore_suplimentare_100) if p.ore_suplimentare_100 else 0,
            p.tip_zi or '-',
            p.status or '-'
        ]
        for col, val in enumerate(row, 1):
            cell = ws3.cell(row=idx + 1, column=col, value=val)
            cell.border = thin_border

    for col in range(1, len(pontaj_headers) + 1):
        ws3.column_dimensions[chr(64 + col)].width = 16
    ws3.auto_filter.ref = f'A1:J{len(pontaje) + 1}'
    ws3.freeze_panes = 'A2'

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f'proiect_{proiect.cod_proiect}_{date.today().strftime("%Y%m%d")}.xlsx'
    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')


# ============================================================
# UPLOAD DOCUMENT PE PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/upload-document', methods=['POST'])
@login_required
def upload_document(id):
    proiect = get_project_or_404(id)

    fisier = request.files.get('fisier')
    nume_document = request.form.get('nume_document', '').strip()
    tip = request.form.get('tip', 'alte')
    data_emitere_str = request.form.get('data_emitere', '')
    data_expirare_str = request.form.get('data_expirare', '')
    observatii = request.form.get('observatii', '').strip()

    if not fisier or not fisier.filename:
        flash('Selectati un fisier pentru incarcare.', 'warning')
        return redirect(url_for('proiecte.detalii', id=id, _anchor='tab-documente'))

    ext = fisier.filename.rsplit('.', 1)[-1].lower() if '.' in fisier.filename else ''
    if ext not in ALLOWED_EXT_PROIECT:
        flash(f'Tipul de fisier .{ext} nu este permis. Formate acceptate: {", ".join(sorted(ALLOWED_EXT_PROIECT))}.', 'danger')
        return redirect(url_for('proiecte.detalii', id=id, _anchor='tab-documente'))

    # Salveaza fisierul
    folder = os.path.join(current_app.config['UPLOAD_FOLDER'], f'proiect_{id}')
    os.makedirs(folder, exist_ok=True)
    unique_name = f"{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.{ext}"
    filepath = os.path.join(folder, unique_name)
    fisier.save(filepath)
    marime = os.path.getsize(filepath)

    # Parseaza datele
    data_emitere = None
    data_expirare = None
    try:
        if data_emitere_str:
            data_emitere = date.fromisoformat(data_emitere_str)
        if data_expirare_str:
            data_expirare = date.fromisoformat(data_expirare_str)
    except ValueError:
        pass

    doc = Document(
        proiect_id=id,
        angajat_id=None,
        tip=tip,
        nume_document=nume_document or secure_filename(fisier.filename),
        fisier_path=filepath,
        marime_fisier=marime,
        data_emitere=data_emitere,
        data_expirare=data_expirare,
        emitent=proiect.beneficiar or '',
        observatii=observatii,
        status='valabil',
        incarcat_de=current_user.id
    )
    db.session.add(doc)
    db.session.commit()
    flash(f'Documentul "{doc.nume_document}" a fost incarcat cu succes!', 'success')
    return redirect(url_for('proiecte.detalii', id=id) + '#tab-documente')


# ============================================================
# DOWNLOAD DOCUMENT PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/documente/<int:doc_id>/download')
@login_required
def download_document(id, doc_id):
    get_project_or_404(id)
    doc = get_legacy_document_or_404(doc_id)
    if doc.proiect_id != id:
        abort(404)
    if not doc.fisier_path or not os.path.exists(doc.fisier_path):
        flash('Fisierul nu mai este disponibil pe server.', 'danger')
        return redirect(url_for('proiecte.detalii', id=id))
    return send_file(doc.fisier_path, as_attachment=True, download_name=doc.nume_document)


# ============================================================
# STERGE DOCUMENT PROIECT
# ============================================================

@proiecte_bp.route('/<int:id>/documente/<int:doc_id>/sterge', methods=['POST'])
@login_required
def sterge_document(id, doc_id):
    if not current_user.is_manager:
        abort(403)
    get_project_or_404(id)
    doc = get_legacy_document_or_404(doc_id)
    if doc.proiect_id != id:
        abort(404)
    # Sterge fisierul de pe disk
    if doc.fisier_path and os.path.exists(doc.fisier_path):
        try:
            os.remove(doc.fisier_path)
        except OSError:
            pass
    db.session.delete(doc)
    db.session.commit()
    flash(f'Documentul "{doc.nume_document}" a fost sters.', 'success')
    return redirect(url_for('proiecte.detalii', id=id) + '#tab-documente')


# ============================================================
# HELPER: Calculeaza cost manopera
# ============================================================

def _calculeaza_cost_manopera(proiect_id):
    """Calculeaza costul total al manoperei pe proiect."""
    result = db.session.query(
        db.func.sum(
            Pontaj.ore_normale * db.func.coalesce(AngajatProiect.tarif_negociat, 0) +
            Pontaj.ore_suplimentare_50 * db.func.coalesce(AngajatProiect.tarif_negociat, 0) * 1.5 +
            Pontaj.ore_suplimentare_100 * db.func.coalesce(AngajatProiect.tarif_negociat, 0) * 2
        )
    ).join(AngajatProiect, db.and_(
        AngajatProiect.angajat_id == Pontaj.angajat_id,
        AngajatProiect.proiect_id == Pontaj.proiect_id
    )).filter(
        Pontaj.proiect_id == proiect_id
    ).scalar()
    return float(result) if result else 0


def _get_ore_saptamanale(proiect_id, weeks=12):
    """Returneaza orele lucrate pe saptamana pentru ultimele N saptamani."""
    results = []
    today = date.today()
    for i in range(weeks - 1, -1, -1):
        start = today - timedelta(days=today.weekday() + 7 * i)
        end = start + timedelta(days=6)
        ore = db.session.query(
            db.func.sum(Pontaj.ore_lucrate)
        ).filter(
            Pontaj.proiect_id == proiect_id,
            Pontaj.data >= start,
            Pontaj.data <= end
        ).scalar()
        results.append({
            'label': f'S{start.isocalendar()[1]}',
            'start': start.strftime('%d.%m'),
            'end': end.strftime('%d.%m'),
            'ore': float(ore) if ore else 0
        })
    return results


def _get_cost_lunar(proiect_id, months=6):
    """Returneaza costul manoperei pe luna pentru ultimele N luni."""
    results = []
    today = date.today()
    for i in range(months - 1, -1, -1):
        m = today.month - i
        y = today.year
        while m <= 0:
            m += 12
            y -= 1

        cost = db.session.query(
            db.func.sum(
                Pontaj.ore_lucrate * db.func.coalesce(AngajatProiect.tarif_negociat, 0)
            )
        ).join(AngajatProiect, db.and_(
            AngajatProiect.angajat_id == Pontaj.angajat_id,
            AngajatProiect.proiect_id == Pontaj.proiect_id
        )).filter(
            Pontaj.proiect_id == proiect_id,
            db.extract('month', Pontaj.data) == m,
            db.extract('year', Pontaj.data) == y
        ).scalar()

        month_names = ['', 'Ian', 'Feb', 'Mar', 'Apr', 'Mai', 'Iun',
                       'Iul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        results.append({
            'label': f'{month_names[m]} {y}',
            'cost': float(cost) if cost else 0
        })
    return results
